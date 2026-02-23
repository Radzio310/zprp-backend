# app/province_travel.py
from __future__ import annotations

from datetime import datetime, timezone
import json
import logging
import traceback
from typing import Any, Dict, Optional, List

from fastapi import APIRouter, HTTPException, Query, Path, BackgroundTasks
from fastapi.responses import FileResponse
from pathlib import Path as SysPath
from tempfile import NamedTemporaryFile
from openpyxl import load_workbook
from sqlalchemy import select, insert, update

from app.db import database, province_travel
from app.schemas import (
    ProvinceTravelUpsertAllRequest,
    ProvinceTravelUpsertSeasonRequest,
    ProvinceTravelItem,
    GetProvinceTravelResponse,
    ListProvinceTravelResponse,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/province-travel", tags=["Province Travel"])


def _now_utc() -> datetime:
    return datetime.now(timezone.utc)


def _normalize_province(p: str) -> str:
    return (p or "").strip().upper()


def _parse_json(raw: Any) -> dict:
    if raw is None:
        return {}
    if isinstance(raw, dict):
        return raw
    try:
        return json.loads(raw)
    except Exception:
        return {}


def _normalize_root_payload(data_json: Any) -> Dict[str, Any]:
    """
    Trzymamy ustandaryzowaną strukturę, ale nie blokujemy rozszerzeń:
      {
        "seasons": {
           "2024/2025": {...},
           "2025/2026": {...}
        },
        ... (inne pola opcjonalnie)
      }
    """
    base = _parse_json(data_json)
    seasons = base.get("seasons")
    if not isinstance(seasons, dict):
        seasons = {}
    # klucze sezonów jako string
    clean_seasons: Dict[str, Any] = {}
    for k, v in seasons.items():
        kk = str(k).strip()
        if not kk:
            continue
        clean_seasons[kk] = v
    base["seasons"] = clean_seasons
    return base


def _row_to_item(row: Any) -> ProvinceTravelItem:
    return ProvinceTravelItem(
        judge_id=str(row["judge_id"]),
        full_name=row["full_name"],
        province=row["province"],
        data_json=_normalize_root_payload(row["data_json"]),
        updated_at=row["updated_at"],
    )


# ───────────────────────── EXPORT XLSX (statystyki dojazdów) ─────────────────────────

BUCKET_KEYS = ["0-15", "16-30", "31-45", "46-60", "61-80", "81-100", ">100"]

def _season_start_year(key: str) -> int:
    try:
        if isinstance(key, str) and len(key) >= 4:
            return int(key.split("/")[0])
    except Exception:
        pass
    return 0

def _extract_season_payload_and_updated_at(season_entry: Any) -> tuple[dict, Optional[str]]:
    if not isinstance(season_entry, dict):
        return {}, None

    # new: { data: {...}, updated_at: "..." }
    if "data" in season_entry and season_entry.get("data") is not None:
        payload = season_entry.get("data")
        if not isinstance(payload, dict):
            payload = {}
        upd = season_entry.get("updated_at")
        return payload, (str(upd) if upd else None)

    # old: bezpośrednio payload
    return season_entry, None

def _safe_num(x: Any) -> Optional[float]:
    try:
        n = float(x)
        if n != n or n in (float("inf"), float("-inf")):
            return None
        return n
    except Exception:
        return None

def _safe_int(x: Any) -> int:
    n = _safe_num(x)
    if n is None:
        return 0
    return int(round(n))

@router.get(
    "/export-xlsx",
    summary="Eksport XLSX statystyk dojazdów (szablon app/templates/statystyki_dojazdow.xlsx)",
)
async def export_travel_xlsx(
    background_tasks: BackgroundTasks,
    province: Optional[str] = Query(None, description="Województwo (opcjonalnie), np. ŚLĄSKIE"),
    q: Optional[str] = Query(None, description="Szukaj po judge_id lub full_name (opcjonalnie)"),
    limit: int = Query(2000, ge=1, le=20000),
):
    # 1) Pobierz rekordy
    stmt = select(province_travel)

    if province:
        stmt = stmt.where(province_travel.c.province == _normalize_province(province))

    if q and q.strip():
        needle = f"%{q.strip()}%"
        stmt = stmt.where(
            (province_travel.c.judge_id.ilike(needle)) | (province_travel.c.full_name.ilike(needle))
        )

    stmt = stmt.order_by(province_travel.c.updated_at.desc()).limit(limit)
    rows = await database.fetch_all(stmt)

    # 2) Załaduj template
    base_dir = SysPath(__file__).resolve().parent  # .../app
    tpl_path = base_dir / "templates" / "statystyki_dojazdow.xlsx"
    if not tpl_path.exists():
        raise HTTPException(status_code=404, detail="Brak szablonu: app/templates/statystyki_dojazdow.xlsx")

    wb = load_workbook(filename=str(tpl_path))

    if "Dane" not in wb.sheetnames:
        raise HTTPException(status_code=422, detail="Szablon nie zawiera arkusza 'Dane'")
    if "Listy" not in wb.sheetnames:
        raise HTTPException(status_code=422, detail="Szablon nie zawiera arkusza 'Listy'")

    ws_dane = wb["Dane"]
    ws_listy = wb["Listy"]

    # 3) Wyczyść stare dane
    max_clear_rows = max(2000, ws_dane.max_row)
    for r in range(2, max_clear_rows + 1):
        for c in range(1, 16 + 1):  # A..P
            ws_dane.cell(row=r, column=c).value = None

    # 4) Wypełnij dane
    all_seasons: set[str] = set()
    out_row = 2

    for row in rows:
        judge_id = str(row["judge_id"] or "").strip()
        full_name = str(row["full_name"] or "").strip()
        prov = str(row["province"] or "").strip()

        root = _normalize_root_payload(row["data_json"])
        seasons_obj = root.get("seasons") if isinstance(root.get("seasons"), dict) else {}

        for season_key, season_entry in seasons_obj.items():
            sk = str(season_key or "").strip()
            if not sk:
                continue

            payload, season_upd = _extract_season_payload_and_updated_at(season_entry)

            stats = payload.get("stats") if isinstance(payload, dict) else None
            if not isinstance(stats, dict):
                stats = {}

            total_matches = _safe_int(stats.get("totalMatches"))
            avg_km = _safe_num(stats.get("avgKm"))
            min_km = _safe_num(stats.get("minKm"))
            max_km = _safe_num(stats.get("maxKm"))

            buckets_raw = stats.get("buckets")
            if not isinstance(buckets_raw, dict):
                buckets_raw = {}

            updated_at = row["updated_at"]
            upd_val = season_upd or (updated_at.isoformat() if updated_at else "")

            ws_dane.cell(out_row, 1).value = sk
            ws_dane.cell(out_row, 2).value = prov
            ws_dane.cell(out_row, 3).value = judge_id
            ws_dane.cell(out_row, 4).value = full_name
            ws_dane.cell(out_row, 5).value = total_matches
            ws_dane.cell(out_row, 6).value = (avg_km if avg_km is not None else None)
            ws_dane.cell(out_row, 7).value = (min_km if min_km is not None else None)
            ws_dane.cell(out_row, 8).value = (max_km if max_km is not None else None)

            for idx, bk in enumerate(BUCKET_KEYS):
                ws_dane.cell(out_row, 9 + idx).value = _safe_int(buckets_raw.get(bk))

            ws_dane.cell(out_row, 16).value = upd_val

            all_seasons.add(sk)
            out_row += 1

    # 5) Listy -> sezony
    for r in range(2, 200 + 1):
        ws_listy.cell(r, 1).value = None

    seasons_sorted = sorted(all_seasons, key=_season_start_year, reverse=True)
    for i, s in enumerate(seasons_sorted, start=2):
        ws_listy.cell(i, 1).value = s

    # 6) Zapis do tmp + FileResponse (attachment)
    with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    wb.save(tmp_path)

    def _cleanup(path: str):
        try:
            import os
            os.remove(path)
        except Exception:
            pass

    background_tasks.add_task(_cleanup, tmp_path)

    filename = "statystyki_dojazdow.xlsx"

    headers = {
        # wymusza zachowanie jak download (a nie inline preview)
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Cache-Control": "no-store",
    }

    return FileResponse(
        path=tmp_path,
        filename=filename,  # nadal warto zostawić
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )

# ───────────────────────── END EXPORT XLSX ─────────────────────────

@router.get(
    "/{judge_id}",
    response_model=GetProvinceTravelResponse,
    summary="Pobierz zapis przejazdów dla sędziego (wszystkie sezony)",
)
async def get_travel_for_judge(judge_id: str = Path(..., description="ID sędziego")):
    jid = str(judge_id).strip()
    if not jid:
        raise HTTPException(400, "Brak judge_id")

    row = await database.fetch_one(select(province_travel).where(province_travel.c.judge_id == jid))
    if not row:
        return GetProvinceTravelResponse(record=None)

    return GetProvinceTravelResponse(record=_row_to_item(row))


@router.get(
    "/",
    response_model=ListProvinceTravelResponse,
    summary="Lista zapisów przejazdów (admin/okręg) – filtrowanie po province",
)
async def list_travel_records(
    province: Optional[str] = Query(None, description="Województwo, np. ŚLĄSKIE"),
    q: Optional[str] = Query(None, description="Szukaj po judge_id lub full_name"),
    limit: int = Query(200, ge=1, le=2000),
):
    stmt = select(province_travel)

    if province:
        stmt = stmt.where(province_travel.c.province == _normalize_province(province))

    if q and q.strip():
        needle = f"%{q.strip()}%"
        # databases + SQLAlchemy: prosto po full_name/judge_id
        stmt = stmt.where(
            (province_travel.c.judge_id.ilike(needle)) | (province_travel.c.full_name.ilike(needle))
        )

    stmt = stmt.order_by(province_travel.c.updated_at.desc()).limit(limit)

    rows = await database.fetch_all(stmt)
    return ListProvinceTravelResponse(records=[_row_to_item(r) for r in rows])


@router.put(
    "/{judge_id}",
    response_model=dict,
    summary="Upsert CAŁOŚCI danych przejazdów (wszystkie sezony) – replace",
)
async def upsert_travel_all(judge_id: str, body: ProvinceTravelUpsertAllRequest):
    """
    UWAGA: to jest replace data_json (pełny stan).
    Do bezpiecznej aktualizacji tylko jednego sezonu użyj PATCH /season.
    """
    jid_path = str(judge_id).strip()
    jid = str(body.judge_id).strip()

    if not jid_path or jid_path != jid:
        raise HTTPException(400, "judge_id w path i body musi być identyczny")

    province = _normalize_province(body.province)
    if not province:
        raise HTTPException(400, "Brak province")

    full_name = (body.full_name or "").strip()
    if not full_name:
        raise HTTPException(400, "Brak full_name")

    data = _normalize_root_payload(body.data_json)

    now = _now_utc()
    existing = await database.fetch_one(select(province_travel.c.judge_id).where(province_travel.c.judge_id == jid))

    try:
        if not existing:
            await database.execute(
                insert(province_travel).values(
                    judge_id=jid,
                    full_name=full_name,
                    province=province,
                    data_json=data,
                    updated_at=now,
                )
            )
            return {"success": True, "created": True}
        else:
            await database.execute(
                update(province_travel)
                .where(province_travel.c.judge_id == jid)
                .values(
                    full_name=full_name,
                    province=province,
                    data_json=data,
                    updated_at=now,
                )
            )
            return {"success": True, "created": False}
    except Exception as e:
        logger.error("upsert_travel_all failed: %s\n%s", e, traceback.format_exc())
        raise HTTPException(500, f"upsert_travel_all failed: {e}")


@router.patch(
    "/{judge_id}/season",
    response_model=dict,
    summary="Upsert tylko JEDNEGO sezonu – merge bez usuwania innych sezonów",
)
async def upsert_travel_season(judge_id: str, body: ProvinceTravelUpsertSeasonRequest):
    jid_path = str(judge_id).strip()
    jid = str(body.judge_id).strip()

    if not jid_path or jid_path != jid:
        raise HTTPException(400, "judge_id w path i body musi być identyczny")

    province = _normalize_province(body.province)
    if not province:
        raise HTTPException(400, "Brak province")

    full_name = (body.full_name or "").strip()
    if not full_name:
        raise HTTPException(400, "Brak full_name")

    season_key = (body.season_key or "").strip()
    if not season_key:
        raise HTTPException(400, "Brak season_key")

    season_json = body.season_json if body.season_json is not None else {}
    season_updated_at = body.season_updated_at or _now_utc()

    now = _now_utc()
    row = await database.fetch_one(select(province_travel).where(province_travel.c.judge_id == jid))

    try:
        if not row:
            # create z jednym sezonem
            data = {"seasons": {season_key: {"data": season_json, "updated_at": season_updated_at.isoformat()}}}
            await database.execute(
                insert(province_travel).values(
                    judge_id=jid,
                    full_name=full_name,
                    province=province,
                    data_json=data,
                    updated_at=now,
                )
            )
            return {"success": True, "created": True, "season_key": season_key}

        # merge: zachowaj inne sezony
        data = _normalize_root_payload(row["data_json"])
        seasons = data.get("seasons") if isinstance(data.get("seasons"), dict) else {}
        seasons[season_key] = {"data": season_json, "updated_at": season_updated_at.isoformat()}
        data["seasons"] = seasons

        await database.execute(
            update(province_travel)
            .where(province_travel.c.judge_id == jid)
            .values(
                full_name=full_name,
                province=province,
                data_json=data,
                updated_at=now,
            )
        )
        return {"success": True, "created": False, "season_key": season_key}
    except Exception as e:
        logger.error("upsert_travel_season failed: %s\n%s", e, traceback.format_exc())
        raise HTTPException(500, f"upsert_travel_season failed: {e}")
