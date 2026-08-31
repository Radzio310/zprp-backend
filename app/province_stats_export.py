"""Eksport statystyk okregowych: CSV, XLSX oraz raporty PDF.

Dane przychodza GOTOWE z przegladarki. Caly rachunek - obsady, kilometry,
ryczalty, szczeble - dzieje sie po stronie BAZA_web, a ten modul wylacznie
nadaje im forme pliku. To swiadome: przeniesienie logiki tutaj oznaczaloby
druga, rownolegla implementacje tych samych regul w Pythonie, ktora rozjezdza
sie z pierwsza przy kazdej zmianie stawek.

Wszystkie trzy eksporty jada tym samym opisem tresci (`ExportSection`), wiec
dolozenie nowej tabeli po stronie przegladarki nie wymaga zmian w backendzie.
"""

from __future__ import annotations

import csv
import io
import os
import re
import tempfile
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional
from zoneinfo import ZoneInfo

from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel, Field
from starlette.background import BackgroundTask

router = APIRouter(tags=["Statystyki okregowe: eksport"])

TEMPLATE_DIR = Path(__file__).resolve().parent / "templates"
TEMPLATE_NAME = "okreg_statystyki_raport.html"
DOWNLOAD_DIR = "/tmp/okreg_stats_downloads"

# Zabezpieczenie przed ladunkiem, ktory zabilby proces pamiecia. Sezon okregu
# to rzad 1500 meczow i ~4000 obsad; 50 000 wierszy zostawia zapas na eksport
# wszystkich pozycji obsady razem z meczami spoza okregu.
MAX_ROWS = 50_000
MAX_SECTIONS = 40


class ExportColumn(BaseModel):
    key: str
    label: str
    # "text" | "int" | "dec1" | "dec2" | "pct" | "pln" | "km" | "date"
    format: str = "text"
    align: Optional[str] = None


class ExportSection(BaseModel):
    title: str
    subtitle: Optional[str] = None
    columns: List[ExportColumn] = Field(default_factory=list)
    rows: List[Dict[str, Any]] = Field(default_factory=list)
    """Krotkie liczby nad tabela - KPI sekcji."""
    tiles: List[Dict[str, Any]] = Field(default_factory=list)
    note: Optional[str] = None


class ExportMeta(BaseModel):
    province: str = ""
    season: str = ""
    title: str = ""
    subtitle: Optional[str] = None
    """Opis aktywnych zawezen. Musi trafic do pliku - inaczej za tydzien nikt
    nie odtworzy, jakiego wycinka dotyczyly te liczby."""
    filters: List[str] = Field(default_factory=list)
    generated_by: Optional[str] = None
    footnotes: List[str] = Field(default_factory=list)


class TableExportRequest(BaseModel):
    meta: ExportMeta
    """Kilka tabel naraz: w XLSX kazda dostaje wlasny arkusz, w CSV ida jedna
    pod druga. Zakladka ekranu ma po kilka zestawien i pobieranie ich osobno
    znaczyloby piec klikniec zamiast jednego."""
    sections: List[ExportSection] = Field(default_factory=list)
    format: str = "xlsx"  # "xlsx" | "csv"
    filename: Optional[str] = None


class ReportExportRequest(BaseModel):
    meta: ExportMeta
    sections: List[ExportSection] = Field(default_factory=list)
    filename: Optional[str] = None


# ---------------------------------------------------------------- formatowanie


def _fmt(value: Any, fmt: str) -> str:
    """Liczba w postaci, w jakiej widzi ja uzytkownik na ekranie."""
    if value is None or value == "":
        return ""
    if fmt == "text":
        return str(value)
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value)

    if fmt == "int":
        return f"{round(number):,}".replace(",", " ")
    if fmt == "dec1":
        return f"{number:,.1f}".replace(",", " ").replace(".", ",")
    if fmt == "dec2":
        return f"{number:,.2f}".replace(",", " ").replace(".", ",")
    if fmt == "pct":
        return f"{number:,.1f}".replace(".", ",") + "%"
    if fmt == "pln":
        return f"{round(number):,}".replace(",", " ") + " zl"
    if fmt == "km":
        return f"{round(number):,}".replace(",", " ") + " km"
    if fmt == "date":
        try:
            return datetime.fromtimestamp(number / 1000).strftime("%d.%m.%Y")
        except (OverflowError, OSError, ValueError):
            return ""
    return str(value)


def _safe_name(name: str, fallback: str) -> str:
    """Nazwa pliku bez znakow, ktore psuja naglowek Content-Disposition."""
    cleaned = re.sub(r"[^A-Za-z0-9_.-]+", "_", (name or "").strip())
    cleaned = cleaned.strip("_.")
    return cleaned or fallback


def _now_label() -> str:
    return datetime.now(ZoneInfo("Europe/Warsaw")).strftime("%d.%m.%Y %H:%M")


def _check_size(sections: List[ExportSection]) -> None:
    if len(sections) > MAX_SECTIONS:
        raise HTTPException(413, f"Za duzo sekcji (max {MAX_SECTIONS}).")
    total = sum(len(s.rows) for s in sections)
    if total > MAX_ROWS:
        raise HTTPException(413, f"Za duzo wierszy: {total} (max {MAX_ROWS}).")


# ------------------------------------------------------------------- CSV/XLSX


@router.post("/zprp/statystyki/okreg/export/table")
async def export_table(payload: TableExportRequest):
    """Tabele do CSV albo XLSX - dokladnie to, co widac po zawezeniu."""
    _check_size(payload.sections)
    sections = [s for s in payload.sections if s.columns]
    if not sections:
        raise HTTPException(422, "Zadna sekcja nie ma kolumn.")

    fmt = (payload.format or "xlsx").lower()
    base = _safe_name(
        payload.filename or payload.meta.title or sections[0].title,
        "statystyki_okregu",
    )

    if fmt == "csv":
        return _csv_response(payload.meta, sections, base)
    if fmt == "xlsx":
        return _xlsx_response(payload.meta, sections, base)
    raise HTTPException(422, "format musi byc 'csv' albo 'xlsx'.")


def _meta_lines(meta: ExportMeta) -> List[str]:
    lines = [
        f"{meta.title or 'Statystyki okregowe'}",
        f"Okreg: {meta.province}   Sezon: {meta.season}",
        f"Wygenerowano: {_now_label()}",
    ]
    if meta.filters:
        lines.append("Zawezenia: " + "; ".join(meta.filters))
    else:
        lines.append("Zawezenia: brak - caly sezon")
    return lines


def _csv_response(
    meta: ExportMeta, sections: List[ExportSection], base: str
) -> StreamingResponse:
    buffer = io.StringIO()
    # Srednik, bo polski Excel dzieli kolumny wlasnie nim; przecinek wrzucilby
    # caly wiersz do jednej komorki.
    writer = csv.writer(buffer, delimiter=";", quoting=csv.QUOTE_MINIMAL)

    for line in _meta_lines(meta):
        writer.writerow([line])

    for section in sections:
        writer.writerow([])
        writer.writerow([section.title])
        if section.subtitle:
            writer.writerow([section.subtitle])
        writer.writerow([c.label for c in section.columns])
        for row in section.rows:
            writer.writerow([_fmt(row.get(c.key), c.format) for c in section.columns])
        if section.note:
            writer.writerow([section.note])

    if meta.footnotes:
        writer.writerow([])
        for note in meta.footnotes:
            writer.writerow([note])

    # BOM, inaczej Excel czyta polskie znaki jako krzaki.
    data = ("﻿" + buffer.getvalue()).encode("utf-8")
    return StreamingResponse(
        io.BytesIO(data),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{base}.csv"'},
    )


def _sheet_title(raw: str, used: set) -> str:
    """Nazwa arkusza: max 31 znakow, bez znakow zakazanych przez Excela i unikalna."""
    cleaned = re.sub(r"[\\/*?:\[\]]", "-", raw or "Arkusz").strip() or "Arkusz"
    cleaned = cleaned[:31]
    candidate = cleaned
    suffix = 2
    while candidate.lower() in used:
        tail = f" {suffix}"
        candidate = cleaned[: 31 - len(tail)] + tail
        suffix += 1
    used.add(candidate.lower())
    return candidate


def _xlsx_response(
    meta: ExportMeta, sections: List[ExportSection], base: str
) -> StreamingResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    used_titles: set = set()

    for section in sections:
        _write_sheet(wb, meta, section, _sheet_title(section.title, used_titles))

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{base}.xlsx"'},
    )


def _write_sheet(wb, meta: ExportMeta, section: ExportSection, title: str) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(title=title)

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="1F2430")
    header_font = Font(bold=True, color="FFFFFF")

    row_i = 1
    for line in _meta_lines(meta):
        ws.cell(row=row_i, column=1, value=line).font = bold if row_i == 1 else Font()
        row_i += 1
    ws.cell(row=row_i, column=1, value=section.title).font = bold
    row_i += 1
    if section.subtitle:
        ws.cell(row=row_i, column=1, value=section.subtitle)
        row_i += 1
    row_i += 1

    header_row = row_i
    for col_i, column in enumerate(section.columns, start=1):
        cell = ws.cell(row=header_row, column=col_i, value=column.label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    row_i += 1

    # Liczby ida do XLSX jako LICZBY, nie jako sformatowany tekst - inaczej
    # arkusz nie da sie ani posortowac, ani zsumowac, a po to sie go pobiera.
    numeric = {"int", "dec1", "dec2", "pct", "pln", "km"}
    for row in section.rows:
        for col_i, column in enumerate(section.columns, start=1):
            raw = row.get(column.key)
            if column.format in numeric and raw is not None and raw != "":
                try:
                    value: Any = float(raw)
                    if column.format == "int":
                        value = int(round(value))
                except (TypeError, ValueError):
                    value = _fmt(raw, column.format)
            else:
                value = _fmt(raw, column.format)
            ws.cell(row=row_i, column=col_i, value=value)
        row_i += 1

    if section.note or meta.footnotes:
        row_i += 1
        for note in ([section.note] if section.note else []) + list(meta.footnotes):
            ws.cell(row=row_i, column=1, value=note)
            row_i += 1

    for col_i, column in enumerate(section.columns, start=1):
        longest = max(
            [len(column.label)]
            + [len(_fmt(r.get(column.key), column.format)) for r in section.rows[:400]]
        )
        ws.column_dimensions[get_column_letter(col_i)].width = min(46, max(10, longest + 3))

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


# ------------------------------------------------------------------------ PDF


def _ensure_download_dir() -> None:
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)


def _render_pdf(payload: ReportExportRequest, default_name: str) -> FileResponse:
    from jinja2 import Environment, FileSystemLoader, select_autoescape
    import weasyprint

    _check_size(payload.sections)
    if not payload.sections:
        raise HTTPException(422, "Raport nie ma zadnej sekcji.")

    template_path = TEMPLATE_DIR / TEMPLATE_NAME
    if not template_path.exists():
        raise HTTPException(500, detail=f"Brak szablonu: {TEMPLATE_NAME}")

    # Autoescape wlaczone: nazwy druzyn i hal pochodza z ZPRP, wiec do szablonu
    # trafia tekst spoza naszej kontroli.
    env = Environment(
        loader=FileSystemLoader(str(TEMPLATE_DIR)),
        autoescape=select_autoescape(["html"]),
    )
    env.filters["cell"] = _fmt
    template = env.get_template(TEMPLATE_NAME)

    html_str = template.render(
        meta=payload.meta,
        sections=payload.sections,
        generated_at=_now_label(),
    )

    tmp_dir = tempfile.mkdtemp()
    html_path = os.path.join(tmp_dir, "raport.html")
    pdf_path = os.path.join(tmp_dir, "raport.pdf")
    with open(html_path, "w", encoding="utf-8") as handle:
        handle.write(html_str)
    weasyprint.HTML(filename=html_path).write_pdf(pdf_path)

    _ensure_download_dir()
    base = _safe_name(payload.filename or default_name, default_name)
    final_path = os.path.join(DOWNLOAD_DIR, f"{base}_{uuid.uuid4().hex[:8]}.pdf")
    os.replace(pdf_path, final_path)

    return FileResponse(
        final_path,
        media_type="application/pdf",
        filename=f"{base}.pdf",
        # Plik znika zaraz po wyslaniu - katalog na Railwayu jest ulotny,
        # a raport i tak da sie wygenerowac ponownie jednym kliknieciem.
        background=BackgroundTask(
            lambda: os.remove(final_path) if os.path.exists(final_path) else None
        ),
    )


@router.post("/zprp/statystyki/okreg/export/season-pdf")
async def export_season_pdf(payload: ReportExportRequest):
    """Raport sezonu okregu - wielostronicowy dokument do archiwum i zarzadu."""
    return _render_pdf(payload, "raport_sezonu_okregu")


@router.post("/zprp/statystyki/okreg/export/referee-pdf")
async def export_referee_pdf(payload: ReportExportRequest):
    """Karta sezonu pojedynczego sedziego."""
    return _render_pdf(payload, "karta_sedziego")
