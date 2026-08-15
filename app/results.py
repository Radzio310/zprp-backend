# app/results.py

import base64
import gzip
import hashlib
import json
import logging
import random
import re
import secrets
import time
import unicodedata
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import parse_qs, urlencode

from bs4 import BeautifulSoup
from fastapi import (
    APIRouter,
    Depends,
    File,
    Header,
    HTTPException,
    Query,
    Request,
    UploadFile,
    Path as ApiPath,
)
# Tożsamość aktora — ta sama zależność co przy zapisach ProEl, żeby wgląd w
# dziennik protokołów miał dokładnie tych samych adminów co reszta systemu.
from app.proel_auth import header_text, proel_actor
from pathlib import Path as SysPath
from httpx import AsyncClient
from pydantic import BaseModel

from cryptography.hazmat.primitives.asymmetric import padding

from app.deps import Settings, get_rsa_keys, get_settings
from app.utils import fetch_with_correct_encoding
from starlette.background import BackgroundTask

from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side
import copy
from openpyxl.drawing.image import Image

logger = logging.getLogger(__name__)
router = APIRouter(tags=["Results"])

class ShortResultRequest(BaseModel):
    username: str    # Base64-RSA
    password: str    # Base64-RSA
    details_path: str
    wynik_gosp_pol: str
    wynik_gosc_pol: str
    wynik_gosp_full: str
    wynik_gosc_full: str
    dogrywka_karne_gosp: str
    dogrywka_karne_gosc: str
    karne_ile_gosp: str
    karne_bramki_gosp: str
    karne_ile_gosc: str
    karne_bramki_gosc: str
    timeout1_gosp_ii: str
    timeout1_gosp_ss: str
    timeout2_gosp_ii: str
    timeout2_gosp_ss: str
    timeout3_gosp_ii: str
    timeout3_gosp_ss: str
    timeout1_gosc_ii: str
    timeout1_gosc_ss: str
    timeout2_gosc_ii: str
    timeout2_gosc_ss: str
    timeout3_gosc_ii: str
    timeout3_gosc_ss: str
    widzowie: Optional[str] = ""


def _decrypt_field(enc_b64: str, private_key) -> str:
    """
    Odszyfrowuje pole zaszyfrowane RSA+Base64.
    """
    try:
        cipher = base64.b64decode(enc_b64)
        plain = private_key.decrypt(
            cipher,
            padding.PKCS1v15()
        )
        return plain.decode('utf-8')
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Błąd deszyfrowania: {e}")


async def _login_and_client(user: str, pwd: str, settings: Settings) -> AsyncClient:
    client = AsyncClient(
        base_url=settings.ZPRP_BASE_URL,
        follow_redirects=True
    )
    resp_login, _ = await fetch_with_correct_encoding(
        client,
        "/login.php",
        method="POST",
        data={"login": user, "haslo": pwd, "from": "/index.php?"},
    )
    if "/index.php" not in resp_login.url.path:
        await client.aclose()
        logger.error("Logowanie nie powiodło się dla user %s", user)
        raise HTTPException(status_code=401, detail="Logowanie nie powiodło się")
    client.cookies.update(resp_login.cookies)
    return client


async def _submit_short_result(
    client: AsyncClient,
    match_id: str,
    user: str,
    overrides: Dict[str, str],
) -> bool:
    # 1) Otwórz modal 'Wynik skrócony'
    initial_data = {
        "IdZawody": match_id,
        "akcja": "WynikSkrocony",
        "user": user,
    }
    _, html = await fetch_with_correct_encoding(
        client,
        "/zawody_WynikSkrocony.php",
        method="POST",
        data=initial_data,
    )
    soup = BeautifulSoup(html, "html.parser")
    form = soup.find("form", {"name": "zawody_WynikSkrocony"})
    if not form:
        return False

    # 2) Parsowanie pól formularza
    form_fields: Dict[str, str] = {}
    for inp in form.find_all(["input", "select", "textarea"]):
        name = inp.get("name")
        if not name:
            continue
        if inp.name == "select":
            opt = inp.find("option", selected=True)
            form_fields[name] = opt.get("value", "") if opt else ""
        elif inp.name == "textarea":
            form_fields[name] = inp.text or ""
        else:
            form_fields[name] = inp.get("value", "") or ""

    # 3) Nadpisanie wybranych pól
    form_fields.update(overrides)

    # 4) Zatwierdzenie zmian
    body = urlencode(form_fields, encoding="iso-8859-2", errors="replace")
    headers = {"Content-Type": "application/x-www-form-urlencoded; charset=ISO-8859-2"}
    resp = await client.request(
        "POST",
        "/zawody_WynikSkrocony.php",
        content=body.encode("ascii"),
        headers=headers,
        cookies=client.cookies,
    )

    text = resp.content.decode("iso-8859-2", errors="replace")
    if resp.status_code != 200:
        raise RuntimeError(f"Błąd HTTP {resp.status_code}: {text[:200]}")

    # jeżeli pojawił się komunikat „Zapisano zmiany” → sukces
    return "Zapisano zmiany" in text

import unicodedata

def _norm(s: str) -> str:
    """lower + usunięcie znaków diakrytycznych, by porównania były odporne na warianty."""
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s.lower()

def _is_host_swapped(soup: BeautifulSoup) -> bool:
    """
    Elastyczne wykrywanie zmiany gospodarza:
    1) dowolny <img> ze źródłem zawierającym 'zmiana' (np. 'pliki/zmiana.png')
    2) alt/title zawierające rdzenie 'zmian' i 'gospod' (np. 'nastapila zmiana gospodarza')
    3) fallback: sam tekst strony z taką frazą (na wypadek braku obrazka)
    """
    # 1) Po nazwie pliku/ścieżce (najstabilniejsze)
    for img in soup.find_all("img"):
        src_norm = _norm(img.get("src", ""))
        if "zmiana" in src_norm:   # łapie też '.../Zmiana.png', '.../ico-zmiana.svg' itd.
            return True

        # 2) Po alt/title (luźne dopasowanie rdzeni)
        meta = _norm((img.get("alt") or "") + " " + (img.get("title") or ""))
        if ("zmian" in meta) and ("gospod" in meta):
            return True

    # 3) Fallback: tekstowy komunikat na stronie
    page_text = _norm(soup.get_text(" ", strip=True))
    if ("zmian" in page_text) and ("gospod" in page_text):
        return True

    return False


def _swap_gosp_gosc(overrides: Dict[str, str]) -> Dict[str, str]:
    """
    Zamienia wartości par kluczy *_gosp* ↔ *_gosc* w słowniku overrides.
    Działa dla wszystkich wariantów nazw (np. *_full, *_pol, *_ii, *_ss, itp.).
    """
    swapped = overrides.copy()
    visited = set()

    for k in list(overrides.keys()):
        if k in visited:
            continue
        if "gosp" in k:
            twin = k.replace("gosp", "gosc")
            if twin in overrides:
                swapped[k], swapped[twin] = overrides[twin], overrides[k]
                visited.add(k)
                visited.add(twin)
        elif "gosc" in k:
            twin = k.replace("gosc", "gosp")
            if twin in overrides:
                swapped[k], swapped[twin] = overrides[twin], overrides[k]
                visited.add(k)
                visited.add(twin)

    return swapped


@router.post(
    "/judge/results/short",
    summary="Zapisz wynik skrócony meczu",
)
async def short_result(
    req: ShortResultRequest,
    settings: Settings = Depends(get_settings),
    keys=Depends(get_rsa_keys),       # tu pobieramy (private_key, public_key)
):
    private_key, _ = keys
    # 1) odszyfruj login i hasło
    user_plain = _decrypt_field(req.username, private_key)
    pass_plain = _decrypt_field(req.password, private_key)

    try:
        client = await _login_and_client(user_plain, pass_plain, settings)
        try:
            # 2) Wejdź na stronę szczegółów meczu
            details_url = _details_path_to_url(req.details_path)
            resp, html = await fetch_with_correct_encoding(
                client,
                details_url,
                method="GET",
                cookies=client.cookies,
            )
            soup = BeautifulSoup(html, "html.parser")
            host_swapped = _is_host_swapped(soup)

            # 3) sprawdź dostępność przycisku/modalu
            if not soup.find("button", class_="przycisk3", string="Wynik skrócony"):
                return {"success": False, "error": "Wynik skrócony niedostępny"}

            # 4) Wyciągnij IdZawody
            match_id = _extract_match_id(req.details_path)

            # 5) Przygotuj overrides (w tym wynik_bramki_* z wynik_*_full)
            overrides = {
                "wynik_gosp_pol": req.wynik_gosp_pol,
                "wynik_gosc_pol": req.wynik_gosc_pol,
                "wynik_gosp_full": req.wynik_gosp_full,
                "wynik_gosc_full": req.wynik_gosc_full,
                "wynik_bramki_gosp": req.wynik_gosp_full,
                "wynik_bramki_gosc": req.wynik_gosc_full,
                "dogrywka_karne_gosp": req.dogrywka_karne_gosp,
                "dogrywka_karne_gosc": req.dogrywka_karne_gosc,
                "karne_ile_gosp": req.karne_ile_gosp,
                "karne_bramki_gosp": req.karne_bramki_gosp,
                "karne_ile_gosc": req.karne_ile_gosc,
                "karne_bramki_gosc": req.karne_bramki_gosc,
                "timeout1_gosp_ii": req.timeout1_gosp_ii,
                "timeout1_gosp_ss": req.timeout1_gosp_ss,
                "timeout2_gosp_ii": req.timeout2_gosp_ii,
                "timeout2_gosp_ss": req.timeout2_gosp_ss,
                "timeout3_gosp_ii": req.timeout3_gosp_ii,
                "timeout3_gosp_ss": req.timeout3_gosp_ss,
                "timeout1_gosc_ii": req.timeout1_gosc_ii,
                "timeout1_gosc_ss": req.timeout1_gosc_ss,
                "timeout2_gosc_ii": req.timeout2_gosc_ii,
                "timeout2_gosc_ss": req.timeout2_gosc_ss,
                "timeout3_gosc_ii": req.timeout3_gosc_ii,
                "timeout3_gosc_ss": req.timeout3_gosc_ss,
                "widzowie": req.widzowie or ""
            }

            if host_swapped:
                overrides = _swap_gosp_gosc(overrides)

            ok = await _submit_short_result(
                client,
                match_id=match_id,
                user=user_plain,
                overrides=overrides,
            )
        finally:
            await client.aclose()

        if not ok:
            return {"success": False, "error": "Zapis nie powiódł się"}
        return {"success": True}

    except HTTPException:
        raise
    except Exception as e:
        logger.error("short_result error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail=f"Nie udało się zapisać wyniku skróconego: {e}")


# =========================
# NEW - protokół pełny
# =========================

# =========================
# New model: Protocol save
# =========================

class ProtocolSaveRequest(BaseModel):
    username: str              # Base64-RSA
    password: str              # Base64-RSA
    details_path: str          # np. "a=zawody&b=protokol&Filtr_sezon=...&IdZawody=..."
    data_json: Dict[str, Any]  # cały JSON meczu (jak w przykładzie)


def _details_path_to_url(details_path: str) -> str:
    dp = (details_path or "").strip()
    if not dp:
        raise HTTPException(400, "details_path jest pusty")

    if dp.startswith("http://") or dp.startswith("https://"):
        m = re.match(r"^https?://[^/]+(?P<path>/.*)$", dp)
        return m.group("path") if m else dp

    if dp.startswith("/index.php"):
        return dp
    if dp.startswith("index.php"):
        return "/" + dp
    if dp.startswith("/"):
        return dp
    return f"/index.php?{dp}"


def _extract_match_id(details_path: str) -> str:
    q = details_path
    if "?" in q:
        q = q.split("?", 1)[1]
    params = parse_qs(q)
    match_id = params.get("IdZawody", [None])[0]
    if not match_id:
        raise HTTPException(400, "Brak parametru IdZawody w details_path")
    return str(match_id)


# ============================================================
# PROTOCOL SAVE (4 BLOCKS) + DELTA MODE
#
# Zmiana wg Twojej obserwacji:
# - Nazwa drużyny do identyfikacji bloków ma sens wyłącznie w:
#   (a) nagłówku tabeli zawodników danej drużyny
#   (b) nagłówku tabeli osób towarzyszących danej drużyny
# - Ignorujemy “zestawienie drużyn grających” na górze strony.
#
# Mapowanie:
# - zawodnicy mapowani po numerze koszulki (NrKoszulki2 value)
# - jeśli ProEl ma numer, którego nie ma w tabeli -> skip (idziemy dalej)
# - jeśli numer koszulki jest zdublowany w HTML -> bierzemy pierwszy napotkany (resztę ignorujemy)
#
# Zapis:
# - zapiszProtok  -> zawody_zapisz2.php (ad1..ad4)
# - zapiszProtok4 -> zawody_zapisz4.php (ad1..ad8)
#
# DELTA:
# - wysyłamy tylko jeśli docelowa wartość != DOM
# ============================================================

PLAYERS_FIELD_TO_KIND = {
    "bramki": "goals",
    "wyjscie": "entered",  # checkbox "W" (played/entered)
    "upomnienie": "warn",  # warning checkbox
    "2minuty": "p2",       # count of 2'
    "dyskwalifikacja": "disq",  # D
    "karne_liczba": "pk_total",
    "karne_bramki": "pk_goals",
    "karne_liczba_seria": "so_total",
    "karne_bramki_seria": "so_goals",
    # "kd": (intentionally not forced here)
}

COMP_CHECKBOX_VALUE_TO_KIND = {
    "1": "warn",  # U
    "2": "p2",    # 2'
    "3": "disq",  # D
}


def _count_nonempty_penalties(ps: Dict[str, Any]) -> int:
    c = 0
    for k in ("penalty1", "penalty2", "penalty3"):
        v = ps.get(k)
        if isinstance(v, str) and v.strip():
            c += 1
    return c


def _truthy(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v != 0
    if isinstance(v, str):
        return v.strip() != ""
    return True


def _build_event_counters(data_json: Dict[str, Any]) -> Dict[str, Dict[str, Dict[str, int]]]:
    """
    Counters from `protocol` section:
      - warning count (type == warning)
      - penalty kick totals/goals (type == penaltyKickScored / penaltyKickMissed)
      - companion 2' (type == penalty1/2/3) also counted here as p2_events
      - disqualification count (type == disqualification)
    Keys:
      team: host/guest
      player: str(number) or "A".."E"
    """
    counters: Dict[str, Dict[str, Dict[str, int]]] = {"host": {}, "guest": {}}
    prot = data_json.get("protocol") or []

    for ev in prot:
        if not isinstance(ev, dict):
            continue
        team = ev.get("team")
        if team not in ("host", "guest"):
            continue
        player = ev.get("player")
        if player is None:
            continue

        key = str(player).strip().upper() if isinstance(player, str) else str(int(player))

        if key not in counters[team]:
            counters[team][key] = {
                "warning": 0,
                "pk_total": 0,
                "pk_goals": 0,
                "p2_events": 0,     # for companions (and fallback)
                "disq": 0,
            }

        t = ev.get("type")

        if t == "warning":
            counters[team][key]["warning"] += 1
        elif t == "penaltyKickScored":
            counters[team][key]["pk_total"] += 1
            counters[team][key]["pk_goals"] += 1
        elif t == "penaltyKickMissed":
            counters[team][key]["pk_total"] += 1
        elif t in ("penalty1", "penalty2", "penalty3"):
            counters[team][key]["p2_events"] += 1
        elif t == "disqualification":
            counters[team][key]["disq"] += 1

    return counters


def _build_shootout_counters(data_json: Dict[str, Any]) -> Dict[str, Dict[str, Dict[str, int]]]:
    shoot: Dict[str, Dict[str, Dict[str, int]]] = {"host": {}, "guest": {}}
    pshots = data_json.get("penaltyShots") or {}
    for team in ("host", "guest"):
        arr = pshots.get(team) or []
        for item in arr:
            if not isinstance(item, dict):
                continue
            p = item.get("player")
            if p is None:
                continue
            key = str(int(p))
            if key not in shoot[team]:
                shoot[team][key] = {"so_total": 0, "so_goals": 0}
            shoot[team][key]["so_total"] += 1
            shoot[team][key]["so_goals"] += 1 if int(item.get("result") or 0) == 1 else 0
    return shoot


def _build_stats_map(data_json: Dict[str, Any]) -> Dict[str, Dict[str, Dict[str, Any]]]:
    """
    Output:
      out["host"]["11"] = {goals, entered, warn, p2, disq, pk_total, pk_goals, so_total, so_goals}
      out["host"]["C"]  = {warn, p2, disq} (companions A..E)
    """
    eventc = _build_event_counters(data_json)
    shootc = _build_shootout_counters(data_json)

    out: Dict[str, Dict[str, Dict[str, Any]]] = {"host": {}, "guest": {}}

    # ---- players (numbers) ----
    for team, stats_list_key in (("host", "hostPlayerStats"), ("guest", "guestPlayerStats")):
        arr = data_json.get(stats_list_key) or []
        for ps in arr:
            if not isinstance(ps, dict):
                continue
            num = ps.get("number")
            if num is None:
                continue
            k = str(int(num))

            goals = int(ps.get("goals") or 0)
            entered = bool(ps.get("entered")) if "entered" in ps else False
            p2 = _count_nonempty_penalties(ps)
            disq = _truthy(ps.get("disqualification")) or _truthy(ps.get("disqualificationDesc"))

            w = eventc.get(team, {}).get(k, {}).get("warning", 0)
            pk_total = eventc.get(team, {}).get(k, {}).get("pk_total", 0)
            pk_goals = eventc.get(team, {}).get(k, {}).get("pk_goals", 0)
            so_total = shootc.get(team, {}).get(k, {}).get("so_total", 0)
            so_goals = shootc.get(team, {}).get(k, {}).get("so_goals", 0)

            out[team][k] = {
                "goals": goals,
                "entered": entered,
                "warn": w > 0,
                "p2": p2,
                "disq": disq,
                "pk_total": pk_total,
                "pk_goals": pk_goals,
                "so_total": so_total,
                "so_goals": so_goals,
            }

    # ---- companions A..E ----
    mc = data_json.get("matchConfig") or {}
    valid_letters = {"A", "B", "C", "D", "E"}

    for team in ("host", "guest"):
        comp_key = f"{team}Companions"
        comps = mc.get(comp_key) or []

        for c in comps:
            if not isinstance(c, dict):
                continue
            cid = str(c.get("id") or "").strip().upper()
            if cid not in valid_letters:
                continue

            warn_cfg = bool(c.get("warned")) if "warned" in c else False
            warn_ev = eventc.get(team, {}).get(cid, {}).get("warning", 0) > 0
            warn = warn_cfg or warn_ev

            p2_cfg = len(c.get("penaltyTimes") or []) if isinstance(c.get("penaltyTimes"), list) else 0
            p2_ev = eventc.get(team, {}).get(cid, {}).get("p2_events", 0)
            p2 = max(p2_cfg, p2_ev)

            disq = eventc.get(team, {}).get(cid, {}).get("disq", 0) > 0

            out[team][cid] = {
                "warn": warn,
                "p2": p2,
                "disq": disq,
            }

    return out


def _split_js_args(argstr: str) -> List[str]:
    s = argstr.strip()
    out: List[str] = []
    buf = []
    in_q: Optional[str] = None
    esc = False

    for ch in s:
        if esc:
            buf.append(ch)
            esc = False
            continue
        if ch == "\\":
            buf.append(ch)
            esc = True
            continue
        if in_q:
            buf.append(ch)
            if ch == in_q:
                in_q = None
            continue
        if ch in ("'", '"'):
            buf.append(ch)
            in_q = ch
            continue
        if ch == ",":
            out.append("".join(buf).strip())
            buf = []
            continue
        buf.append(ch)

    if buf:
        out.append("".join(buf).strip())
    return out


def _unquote_js(s: str) -> str:
    s2 = s.strip()
    if len(s2) >= 2 and ((s2[0] == "'" and s2[-1] == "'") or (s2[0] == '"' and s2[-1] == '"')):
        return s2[1:-1]
    return s2


def _extract_zapisz2_args(js: str) -> Optional[List[str]]:
    if not js:
        return None
    m = re.search(r"zapiszProtok\s*\(\s*(.*?)\s*\)", js, flags=re.IGNORECASE | re.DOTALL)
    if not m:
        return None
    args = _split_js_args(m.group(1))
    if len(args) < 4:
        return None
    return args[:4]

def _extract_zapisz3_args(js: str) -> Optional[List[str]]:
    if not js:
        return None
    m = re.search(r"zapiszProtok3\s*\(\s*(.*?)\s*\)", js, flags=re.IGNORECASE | re.DOTALL)
    if not m:
        return None
    args = _split_js_args(m.group(1))
    if len(args) < 4:
        return None
    return args[:4]

def _extract_zapisz4_args(js: str) -> Optional[List[str]]:
    if not js:
        return None
    m = re.search(r"zapiszProtok4\s*\(\s*(.*?)\s*\)", js, flags=re.IGNORECASE | re.DOTALL)
    if not m:
        return None
    args = _split_js_args(m.group(1))
    if len(args) < 8:
        return None
    return args[:8]


def _js_token_eval(token: str, *, value_str: str, checked: bool) -> str:
    t = token.strip()
    if re.fullmatch(r"this\.value", t, flags=re.IGNORECASE):
        return value_str
    if re.fullmatch(r"this\.checked", t, flags=re.IGNORECASE):
        return "true" if checked else "false"
    return _unquote_js(t)


async def _save_via_zapisz2(
    client: AsyncClient,
    args4: List[str],
    *,
    value_str: str,
    checked: bool,
) -> Tuple[bool, str]:
    payload = {
        "ad1": _js_token_eval(args4[0], value_str=value_str, checked=checked),
        "ad2": _js_token_eval(args4[1], value_str=value_str, checked=checked),
        "ad3": _js_token_eval(args4[2], value_str=value_str, checked=checked),
        "ad4": _js_token_eval(args4[3], value_str=value_str, checked=checked),
        "sid": str(random.random()),
    }
    _, text = await fetch_with_correct_encoding(
        client,
        "/zawody_zapisz2.php",
        method="POST",
        data=payload,
        cookies=client.cookies,
    )
    t = (text or "").strip()
    ok = (t == "OK") or ("OK" in t and "ERROR" not in t)
    return ok, t[:200]

async def _save_via_zapisz3(
    client: AsyncClient,
    args4: List[str],
    *,
    value_str: str,
    checked: bool,
) -> Tuple[bool, str]:
    payload = {
        "ad1": _js_token_eval(args4[0], value_str=value_str, checked=checked),
        "ad2": _js_token_eval(args4[1], value_str=value_str, checked=checked),
        "ad3": _js_token_eval(args4[2], value_str=value_str, checked=checked),
        "ad4": _js_token_eval(args4[3], value_str=value_str, checked=checked),
        "sid": str(random.random()),
    }
    _, text = await fetch_with_correct_encoding(
        client,
        "/zawody_zapisz3.php",
        method="POST",
        data=payload,
        cookies=client.cookies,
    )
    t = (text or "").strip()
    ok = (t == "OK") or ("OK" in t and "ERROR" not in t)
    return ok, t[:200]

async def _save_via_zapisz4(
    client: AsyncClient,
    args8: List[str],
    *,
    value_str: str,
    checked: bool,
) -> Tuple[bool, str]:
    payload = {
        "ad1": _js_token_eval(args8[0], value_str=value_str, checked=checked),
        "ad2": _js_token_eval(args8[1], value_str=value_str, checked=checked),
        "ad3": _js_token_eval(args8[2], value_str=value_str, checked=checked),
        "ad4": _js_token_eval(args8[3], value_str=value_str, checked=checked),
        "ad5": _js_token_eval(args8[4], value_str=value_str, checked=checked),
        "ad6": _js_token_eval(args8[5], value_str=value_str, checked=checked),
        "ad7": _js_token_eval(args8[6], value_str=value_str, checked=checked),
        "ad8": _js_token_eval(args8[7], value_str=value_str, checked=checked),
        "sid": str(random.random()),
    }
    _, text = await fetch_with_correct_encoding(
        client,
        "/zawody_zapisz4.php",
        method="POST",
        data=payload,
        cookies=client.cookies,
    )
    t = (text or "").strip()
    ok = (t == "OK") or ("OK" in t and "ERROR" not in t)
    return ok, t[:200]


def _normalize_space(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())

def _sanitize_comment_text(s: str) -> str:
    """
    Minimalna, bezpieczna normalizacja:
    - zamienia CRLF na LF
    - usuwa znaki kontrolne
    - podmienia apostrofy / myślniki typograficzne na proste odpowiedniki
    """
    if not isinstance(s, str):
        return ""
    x = s.replace("\r\n", "\n").replace("\r", "\n")

    # normalizacja unicode (usuwa część "dziwnych" wariantów)
    x = unicodedata.normalize("NFKC", x)

    # typograficzne znaki na proste (częsty powód ostrzeżeń)
    x = x.replace("’", "'").replace("`", "'")
    x = x.replace("–", "-").replace("—", "-")

    # usuń znaki kontrolne poza \n i \t
    x = "".join(ch for ch in x if ch in ("\n", "\t") or ord(ch) >= 32)

    # opcjonalnie: przytnij długość, żeby nie wpakować megatekstu (możesz zmienić limit)
    return x.strip()[:2000]


def _table_text(table) -> str:
    return _normalize_space(table.get_text(" ", strip=True))


import os
import uuid

# =========================
# DEBUG helpers (Railway logs)
# =========================

def _dbg_enabled() -> bool:
    """
    W Railway ustaw env:
      RESULTS_PROTOCOL_DEBUG=1
    aby włączyć bardzo obszerny log.
    """
    v = (os.getenv("RESULTS_PROTOCOL_DEBUG") or "").strip().lower()
    return v in ("1", "true", "yes", "on")


def _dbg(msg: str, **kw):
    if not _dbg_enabled():
        return
    if kw:
        try:
            extras = " ".join([f"{k}={repr(v)[:400]}" for k, v in kw.items()])
        except Exception:
            extras = ""
        logger.warning("[protocol-debug] %s | %s", msg, extras)  # <-- WARNING
    else:
        logger.warning("[protocol-debug] %s", msg)               # <-- WARNING


def _short_html(el, limit: int = 240) -> str:
    try:
        s = str(el)
    except Exception:
        return ""
    s = re.sub(r"\s+", " ", s).strip()
    return s[:limit]


def _summarize_table_candidate(table, host_name: str, guest_name: str, team_header_min_colspan: int) -> Dict[str, Any]:
    rows = _iter_team_blocks_rows_by_order(
        table,
        team_header_min_colspan=team_header_min_colspan,
        debug_tag="candidate",
    )

    teams = {t for (t, _) in rows}
    jersey_inputs = len(table.find_all("input", attrs={"name": re.compile(r"^NrKoszulki2\d+$")}))
    has_z2 = bool(
        table.find(attrs={"onchange": re.compile(r"zapiszProtok\s*\(", re.IGNORECASE)})
        or table.find(attrs={"onclick": re.compile(r"zapiszProtok\s*\(", re.IGNORECASE)})
    )
    has_z4 = bool(table.find(attrs={"onclick": re.compile(r"zapiszProtok4\s*\(", re.IGNORECASE)}))
    # prosta "głębokość" zagnieżdżeń
    nested_tables = len(table.find_all("table"))

    return {
        "teams": sorted(list(teams)),
        "rows": len(rows),
        "jersey_inputs": jersey_inputs,
        "has_z2": has_z2,
        "has_z4": has_z4,
        "nested_tables": nested_tables,
        "sample_text": _table_text(table)[:240],
    }


# =========================
# TABLE FINDERS (with verbose logs)
# =========================

def _find_players_table(soup: BeautifulSoup, *, host_name: str, guest_name: str):
    """
    Wybiera najlepszą tabelę zawodników wg scoringu.
    Loguje kandydatów i finalny wybór.
    """
    best = None
    best_score = -1
    best_summary = None

    cand_idx = 0
    for table in soup.find_all("table"):
        if not table.find("input", attrs={"name": re.compile(r"^NrKoszulki2\d+$")}):
            continue
        if not (
            table.find(attrs={"onchange": re.compile(r"zapiszProtok\s*\(", re.IGNORECASE)})
            or table.find(attrs={"onclick": re.compile(r"zapiszProtok\s*\(", re.IGNORECASE)})
        ):
            continue

        cand_idx += 1
        summary = _summarize_table_candidate(table, host_name, guest_name, team_header_min_colspan=15)
        teams = set(summary["teams"])

        score = 0
        if "host" in teams:
            score += 1000
        if "guest" in teams:
            score += 1000

        score += summary["rows"]

        # mocna kara za "tabelę-kontener" z masą zagnieżdżeń
        score -= 50 * int(summary["nested_tables"] or 0)

        # kara za layout / menu / “API | rozgrywki…”
        sample_l = (summary.get("sample_text") or "").lower()
        if "api | rozgrywki" in sample_l:
            score -= 500


        _dbg(
            "players_table candidate",
            idx=cand_idx,
            score=score,
            teams=summary["teams"],
            rows=summary["rows"],
            jersey_inputs=summary["jersey_inputs"],
            nested_tables=summary["nested_tables"],
            has_z2=summary["has_z2"],
            sample_text=summary["sample_text"],
        )

        if score > best_score:
            best_score = score
            best = table
            best_summary = summary

    _dbg(
        "players_table chosen",
        best_score=best_score,
        best_summary=best_summary,
        found=("yes" if best else "no"),
    )
    return best


def _find_companions_table(soup: BeautifulSoup, *, host_name: str, guest_name: str):
    """
    Stabilne znalezienie tabeli 'Osoby towarzyszące' dla ZPRP.

    W realnym HTML masz:
      <td colspan="3">
        Osoby towarzyszące:
        <table> ... (nagłówki drużyn colspan=10, kolumny Osoba/Funkcja/... + checkboxy zapiszProtok4) ...
        </table>
      </td>

    Czyli: tabela jest ZAGNIEŻDŻONA wewnątrz tego samego <td>, a nie "po markerze".
    """

    def _norm_no_diacritics(s: str) -> str:
        # odporne na polskie znaki + różne encodowania
        s = (s or "").strip()
        s = unicodedata.normalize("NFKD", s)
        s = "".join(ch for ch in s if not unicodedata.combining(ch))
        s = s.lower()
        s = re.sub(r"\s+", " ", s)
        return s

    def _has_z4(table) -> bool:
        return bool(table and table.find(attrs={"onclick": re.compile(r"zapiszProtok4\s*\(", re.IGNORECASE)}))

    def _has_team_header_colspan10(table) -> bool:
        if not table:
            return False
        for tr in table.find_all("tr"):
            hn = _extract_team_header_name_from_tr(tr, team_header_min_colspan=10)
            if hn:
                return True
        return False

    def _looks_like_companions_table(table) -> bool:
        """
        Minimalne, ale trafne warunki dla Twojego HTML:
        - musi zawierać zapiszProtok4 (checkboxy U/2'/D)
        - musi mieć nagłówki drużyn w wierszach <td colspan="10"><b>...</b>
        - musi zawierać nagłówki 'Osoba' i 'Funkcja' (często są w rowspans)
        """
        if not table:
            return False
        if not _has_z4(table):
            return False
        if not _has_team_header_colspan10(table):
            return False

        txt = _norm_no_diacritics(_table_text(table))
        if ("osoba" not in txt) or ("funkcja" not in txt):
            return False

        # "Kary" zwykle występuje, ale różne encodowania mogą je psuć – nie blokujemy twardo.
        # Jeśli chcesz twardo, odkomentuj:
        # if "kary" not in txt:
        #     return False

        return True

    # ------------------------------------------------------------
    # 1) TRYB PEWNY: znajdź kontener z tekstem "Osoby towarzyszące"
    #    i weź tabelę zagnieżdżoną w środku.
    # ------------------------------------------------------------
    best = None
    best_score = -1
    best_summary = None
    cand_idx = 0

    # Szukamy tagów (np. td), w których tekst zawiera frazę.
    # Uwaga: w HTML bywa "Osoby towarzysz\u0105ce:" albo krzaki po ISO-8859-2,
    # dlatego dopasowujemy po rdzeniu "osoby towarzysz".
    for tag in soup.find_all(["td", "div", "span", "p"]):
        t = _norm_no_diacritics(tag.get_text(" ", strip=True))
        if "osoby towarzysz" not in t:
            continue

        # najczęściej tabela jest bezpośrednio w tym tagu (td colspan="3")
        inner_tables = tag.find_all("table")
        if not inner_tables:
            continue

        for tbl in inner_tables:
            if not _looks_like_companions_table(tbl):
                continue

            cand_idx += 1
            summary = _summarize_table_candidate(tbl, host_name, guest_name, team_header_min_colspan=10)
            teams = set(summary["teams"])

            # scoring: preferuj tabelę, która ma oba bloki (host+guest), ale nie wymagaj
            score = 0
            if "host" in teams:
                score += 1000
            if "guest" in teams:
                score += 1000
            score += int(summary["rows"] or 0)
            score -= 50 * int(summary["nested_tables"] or 0)

            _dbg(
                "companions_table candidate (container)",
                idx=cand_idx,
                score=score,
                teams=summary["teams"],
                rows=summary["rows"],
                nested_tables=summary["nested_tables"],
                has_z4=summary["has_z4"],
                sample_text=summary["sample_text"],
            )

            if score > best_score:
                best = tbl
                best_score = score
                best_summary = summary

    _dbg(
        "companions_table chosen (container)",
        best_score=best_score,
        best_summary=best_summary,
        found=("yes" if best else "no"),
    )
    if best:
        return best

    # ------------------------------------------------------------
    # 2) FALLBACK: globalny skan po wszystkich tabelach
    # ------------------------------------------------------------
    best = None
    best_score = -1
    best_summary = None
    cand_idx = 0

    for table in soup.find_all("table"):
        if not _looks_like_companions_table(table):
            continue

        cand_idx += 1
        summary = _summarize_table_candidate(table, host_name, guest_name, team_header_min_colspan=10)
        teams = set(summary["teams"])

        score = 0
        if "host" in teams:
            score += 1000
        if "guest" in teams:
            score += 1000
        score += int(summary["rows"] or 0)
        score -= 50 * int(summary["nested_tables"] or 0)

        _dbg(
            "companions_table candidate (fallback)",
            idx=cand_idx,
            score=score,
            teams=summary["teams"],
            rows=summary["rows"],
            nested_tables=summary["nested_tables"],
            has_z4=summary["has_z4"],
            sample_text=summary["sample_text"],
        )

        if score > best_score:
            best = table
            best_score = score
            best_summary = summary

    _dbg(
        "companions_table chosen (fallback)",
        best_score=best_score,
        best_summary=best_summary,
        found=("yes" if best else "no"),
    )
    return best


# =========================
# TEAM BLOCKS (verbose)
# =========================

def _iter_team_blocks_rows(
    table,
    *,
    host_name: str,
    guest_name: str,
    team_header_min_colspan: int,
    debug_tag: str = "main",
) -> List[Tuple[str, Any]]:
    """
    Zwraca listę (team, tr) tylko gdy current_team jest host/guest.
    Loguje wykryte nagłówki i mapping do teamów.
    """
    out: List[Tuple[str, Any]] = []
    current_team: Optional[str] = None

    if _dbg_enabled():
        _dbg(
            "iter_team_blocks_rows start",
            debug_tag=debug_tag,
            team_header_min_colspan=team_header_min_colspan,
            host_name=host_name,
            guest_name=guest_name,
            table_sample=_table_text(table)[:200],
        )

    seen_headers = 0
    for tr_i, tr in enumerate(table.find_all("tr")):
        header_name = _extract_team_header_name_from_tr(tr, team_header_min_colspan=team_header_min_colspan)
        if header_name:
            seen_headers += 1
            matched = _team_from_header_text(header_name, host_name, guest_name)
            _dbg(
                "team header detected",
                debug_tag=debug_tag,
                tr_index=tr_i,
                header_name=header_name,
                matched_team=matched,
                tr_html=_short_html(tr),
            )
            current_team = matched  # może być None
            continue

        if current_team in ("host", "guest"):
            out.append((current_team, tr))

    if _dbg_enabled():
        c_host = sum(1 for t, _ in out if t == "host")
        c_guest = sum(1 for t, _ in out if t == "guest")
        _dbg(
            "iter_team_blocks_rows end",
            debug_tag=debug_tag,
            headers_seen=seen_headers,
            rows_total=len(out),
            rows_host=c_host,
            rows_guest=c_guest,
        )

    return out

def _iter_team_blocks_rows_by_order(
    table,
    *,
    team_header_min_colspan: int,
    debug_tag: str = "main",
) -> List[Tuple[str, Any]]:
    """
    Zwraca listę (team, tr) na podstawie KOLEJNOŚCI nagłówków drużyn:
      - pierwszy wykryty nagłówek drużyny => host
      - drugi wykryty nagłówek drużyny   => guest
    Ignoruje dopasowanie po nazwie (odpornie na mojibake typu 'Zag³êbie').
    """
    out: List[Tuple[str, Any]] = []
    current_team: Optional[str] = None
    header_index = -1  # 0->host, 1->guest

    if _dbg_enabled():
        _dbg(
            "iter_team_blocks_rows_by_order start",
            debug_tag=debug_tag,
            team_header_min_colspan=team_header_min_colspan,
            table_sample=_table_text(table)[:200],
        )

    for tr_i, tr in enumerate(table.find_all("tr")):
        header_name = _extract_team_header_name_from_tr(tr, team_header_min_colspan=team_header_min_colspan)
        if header_name:
            header_index += 1
            if header_index == 0:
                current_team = "host"
            elif header_index == 1:
                current_team = "guest"
            else:
                current_team = None  # kolejne nagłówki ignorujemy

            _dbg(
                "team header detected (by_order)",
                debug_tag=debug_tag,
                tr_index=tr_i,
                header_name=header_name,
                assigned_team=current_team,
                tr_html=_short_html(tr),
            )
            continue

        if current_team in ("host", "guest"):
            out.append((current_team, tr))

    if _dbg_enabled():
        c_host = sum(1 for t, _ in out if t == "host")
        c_guest = sum(1 for t, _ in out if t == "guest")
        _dbg(
            "iter_team_blocks_rows_by_order end",
            debug_tag=debug_tag,
            headers_seen=header_index + 1,
            rows_total=len(out),
            rows_host=c_host,
            rows_guest=c_guest,
        )

    return out



# =========================
# INPUT COLLECTORS (verbose)
# =========================

def _collect_players_inputs(
    soup: BeautifulSoup,
    *,
    host_name: str,
    guest_name: str,
) -> Dict[Tuple[str, str, str], Dict[str, Any]]:
    table = _find_players_table(soup, host_name=host_name, guest_name=guest_name)
    if not table:
        _dbg("collect_players_inputs: no table found")
        return {}

    result: Dict[Tuple[str, str, str], Dict[str, Any]] = {}
    seen_jerseys: Dict[str, set] = {"host": set(), "guest": set()}

    rows = _iter_team_blocks_rows_by_order(
        table,
        team_header_min_colspan=15,
        debug_tag="players",
    )


    jersey_seen_counts = {"host": 0, "guest": 0}
    row_with_jersey_counts = {"host": 0, "guest": 0}

    for team, tr in rows:
        jersey_inp = None

        # find jersey input (field == NrKoszulki2)
        for inp in tr.find_all(["input", "select", "textarea"]):
            js = inp.get("onchange") or inp.get("onclick") or ""
            if "zapiszProtok" not in js:
                continue
            args4 = _extract_zapisz2_args(js)
            if not args4:
                continue
            field = _unquote_js(args4[1]).strip()
            if field == "NrKoszulki2":
                jersey_inp = inp
                break

        if not jersey_inp:
            continue

        jersey_val = (jersey_inp.get("value") or "").strip()
        if not re.fullmatch(r"\d{1,3}", jersey_val):
            _dbg("players row jersey invalid", team=team, jersey_val=jersey_val, tr_html=_short_html(tr))
            continue

        jersey = str(int(jersey_val))
        row_with_jersey_counts[team] += 1

        if jersey in seen_jerseys[team]:
            _dbg("players row jersey duplicate in HTML", team=team, jersey=jersey, tr_html=_short_html(tr))
            continue
        seen_jerseys[team].add(jersey)
        jersey_seen_counts[team] += 1

        # collect mapped inputs (kinds)
        for inp in tr.find_all(["input", "select", "textarea"]):
            js = inp.get("onchange") or inp.get("onclick") or ""
            if "zapiszProtok" not in js:
                continue
            args4 = _extract_zapisz2_args(js)
            if not args4:
                continue
            field = _unquote_js(args4[1]).strip()
            if field == "NrKoszulki2":
                continue
            kind = PLAYERS_FIELD_TO_KIND.get(field)
            if not kind:
                continue

            key = (team, jersey, kind)
            if key in result:
                _dbg("players input duplicate key ignored", team=team, jersey=jersey, kind=kind, field=field)
                continue

            result[key] = {"inp": inp, "field": field, "args4": args4}
            _dbg(
                "players input mapped",
                team=team,
                jersey=jersey,
                kind=kind,
                field=field,
                dom_value=_current_text_value(inp),
                dom_checked=_is_checked_dom(inp),
                args4=args4,
            )

    _dbg(
        "collect_players_inputs summary",
        total=len(result),
        host_keys=sum(1 for k in result.keys() if k[0] == "host"),
        guest_keys=sum(1 for k in result.keys() if k[0] == "guest"),
        jerseys_host=sorted(list(seen_jerseys["host"]))[:80],
        jerseys_guest=sorted(list(seen_jerseys["guest"]))[:80],
        rows_with_jersey_host=row_with_jersey_counts["host"],
        rows_with_jersey_guest=row_with_jersey_counts["guest"],
    )
    return result


def _collect_companion_inputs(
    soup: BeautifulSoup,
    *,
    host_name: str,
    guest_name: str,
) -> Dict[Tuple[str, str, str], Dict[str, Any]]:
    table = _find_companions_table(soup, host_name=host_name, guest_name=guest_name)
    if not table:
        _dbg("collect_companion_inputs: no table found")
        return {}

    letter_col = _find_letter_col_index(table)
    result: Dict[Tuple[str, str, str], Dict[str, Any]] = {}

    rows = _iter_team_blocks_rows_by_order(
        table,
        team_header_min_colspan=10,
        debug_tag="companions",
    )

    for team, tr in rows:
        tds = tr.find_all("td")
        if not tds or len(tds) < 3:
            continue

        letter = None
        if letter_col is not None and letter_col < len(tds):
            letter = _extract_letter_from_cell(tds[letter_col])
        else:
            for td in tds:
                cand = _extract_letter_from_cell(td)
                if cand:
                    letter = cand
                    break

        if not letter:
            continue

        for inp in tr.find_all("input"):
            js = inp.get("onclick") or ""
            if "zapiszProtok4" not in js:
                continue
            args8 = _extract_zapisz4_args(js)
            if not args8:
                continue

            v = (inp.get("value") or "").strip()
            kind = COMP_CHECKBOX_VALUE_TO_KIND.get(v)
            if not kind:
                continue

            key = (team, letter, kind)
            if key in result:
                _dbg("companions input duplicate key ignored", team=team, letter=letter, kind=kind)
                continue

            result[key] = {"inp": inp, "args8": args8, "checkbox_value": v}
            _dbg(
                "companions input mapped",
                team=team,
                letter=letter,
                kind=kind,
                checkbox_value=v,
                dom_checked=_is_checked_dom(inp),
                args8=args8,
            )

    _dbg(
        "collect_companion_inputs summary",
        total=len(result),
        host_keys=sum(1 for k in result.keys() if k[0] == "host"),
        guest_keys=sum(1 for k in result.keys() if k[0] == "guest"),
        letter_col=letter_col,
    )
    return result


def _collect_comment_input(soup: BeautifulSoup) -> Optional[Dict[str, Any]]:
    """
    Szuka textarea name="komentarz" z onchange="zapiszProtok3(...)".
    Zwraca meta: {inp, args4}
    """
    ta = soup.find("textarea", attrs={"name": "komentarz"})
    if not ta:
        return None

    js = ta.get("onchange") or ta.get("onclick") or ""
    if "zapiszProtok3" not in js:
        # czasem bywa w onchange, ale zostawiamy defensywnie
        return None

    args4 = _extract_zapisz3_args(js)
    if not args4:
        return None

    return {"inp": ta, "args4": args4}


def _norm_team_name(s: str) -> str:
    s = (s or "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s


def _team_from_header_text(header_txt: str, host_name: str, guest_name: str) -> Optional[str]:
    """
    Dopasowanie po nazwie drużyny z nagłówka <b>...</b>.
    Bez fallbacków: jeśli nie pasuje do host ani guest -> None.
    """
    ht = _norm_team_name(header_txt)
    h = _norm_team_name(host_name)
    g = _norm_team_name(guest_name)

    if not ht:
        return None

    has_h = bool(h and (ht == h or h in ht or ht in h))
    has_g = bool(g and (ht == g or g in ht or ht in g))

    if has_h and has_g:
        # nagłówek nie powinien zawierać obu nazw (w praktyce ignorujemy taki wiersz)
        return None
    if has_h:
        return "host"
    if has_g:
        return "guest"
    return None


def _extract_team_header_name_from_tr(tr, *, team_header_min_colspan: int) -> Optional[str]:
    """
    Prawdziwy nagłówek sekcji drużyny w ZPRP:
      - komórka td/th z dużym colspan (players ~15, companions ~10)
      - w środku <b>NAZWA DRUŻYNY</b>
    """
    for cell in tr.find_all(["td", "th"]):
        try:
            cs = int(cell.get("colspan") or 0)
        except Exception:
            cs = 0
        if cs < int(team_header_min_colspan or 0):
            continue

        b = cell.find("b")
        if not b:
            continue

        name = _normalize_space(b.get_text(" ", strip=True))
        if not name:
            continue

        # odfiltruj wiersze typu "Osoby towarzyszące:" jeśli kiedykolwiek trafią tu przez colspan
        if "osoby towarzysz" in name.lower():
            continue

        return name

    return None


def _find_letter_col_index(table) -> Optional[int]:
    for tr in table.find_all("tr"):
        cells = tr.find_all(["th", "td"])
        if not cells:
            continue
        row_text = _normalize_space(tr.get_text(" ", strip=True)).lower()
        if "kolejność" in row_text and "klik" in row_text and "sortowanie" in row_text:
            for i, c in enumerate(cells):
                t = _normalize_space(c.get_text(" ", strip=True)).lower()
                if "kolejność" in t and "sortowanie" in t:
                    return i
    return None


def _extract_letter_from_cell(td) -> Optional[str]:
    sel = td.find("select")
    if sel:
        opt = sel.find("option", selected=True)
        v = (opt.get("value") if opt else "") or ""
        v = v.strip().upper()
        if re.fullmatch(r"[A-E]", v):
            return v
    txt = _normalize_space(td.get_text(" ", strip=True)).upper()
    if re.fullmatch(r"[A-E]", txt):
        return txt
    if "ZGŁOŚ" in txt or "ZGLOS" in txt:
        return None
    return None

def _desired_value_for_player_kind(st: Dict[str, Any], kind: str) -> Any:
    if kind == "goals":
        return int(st.get("goals") or 0)
    if kind == "entered":
        return bool(st.get("entered") or False)
    if kind == "warn":
        return bool(st.get("warn") or False)
    if kind == "p2":
        return int(st.get("p2") or 0)
    if kind == "disq":
        return bool(st.get("disq") or False)
    if kind == "pk_total":
        return int(st.get("pk_total") or 0)
    if kind == "pk_goals":
        return int(st.get("pk_goals") or 0)
    if kind == "so_total":
        return int(st.get("so_total") or 0)
    if kind == "so_goals":
        return int(st.get("so_goals") or 0)
    return None


def _desired_value_for_companion_kind(st: Dict[str, Any], kind: str) -> bool:
    if kind == "warn":
        return bool(st.get("warn") or False)
    if kind == "p2":
        return int(st.get("p2") or 0) >= 1
    if kind == "disq":
        return bool(st.get("disq") or False)
    return False


# -------------------------
# DELTA helpers
# -------------------------

_NUMERIC_KINDS = {"goals", "p2", "pk_total", "pk_goals", "so_total", "so_goals"}


def _is_checked_dom(inp) -> bool:
    """
    In BeautifulSoup DOM of HTML:
      checked can be: checked="checked" / checked (attribute exists) / value in some cases.
    We treat attribute presence as True.
    """
    if inp is None:
        return False
    if inp.has_attr("checked"):
        return True
    v = (inp.get("checked") or "").strip().lower()
    return v in ("checked", "true", "1", "yes")


def _current_text_value(inp) -> str:
    if inp is None:
        return ""
    if inp.name == "textarea":
        return (inp.text or "").strip()
    if inp.name == "select":
        opt = inp.find("option", selected=True)
        return ((opt.get("value", "") if opt else "") or "").strip()
    return (inp.get("value") or "").strip()


def _norm_num_str(s: str) -> int:
    """
    Treat empty / '0' / '00' as 0.
    Non-numeric -> raises.
    """
    s2 = (s or "").strip()
    if s2 == "":
        return 0
    s2 = s2.replace("\xa0", "").strip()
    return int(s2)


def _desired_str_for_numeric(desired_int: int) -> str:
    # In ZPRP protocol, "0" is commonly represented as empty.
    return "" if int(desired_int) == 0 else str(int(desired_int))


def _delta_equal_player(inp, kind: str, desired: Any) -> bool:
    t = (inp.get("type") or "").lower()

    # checkbox kinds
    if t == "checkbox":
        cur = _is_checked_dom(inp)
        des = bool(desired)
        return cur == des

    # numeric kinds
    cur_s = _current_text_value(inp)
    if kind in _NUMERIC_KINDS:
        try:
            cur_i = _norm_num_str(cur_s)
            des_i = int(desired or 0)
            return cur_i == des_i
        except Exception:
            return cur_s.strip() == _desired_str_for_numeric(int(desired or 0)).strip()

    # fallback as text
    return cur_s.strip() == (str(desired) if desired is not None else "").strip()


def _delta_equal_companion(inp, desired_checked: bool) -> bool:
    return _is_checked_dom(inp) == bool(desired_checked)


async def _apply_protocol_updates_4blocks(
    client: AsyncClient,
    soup: BeautifulSoup,
    stats_map: Dict[str, Dict[str, Dict[str, Any]]],
    *,
    host_name: str,
    guest_name: str,
    referee_comment: str = "",
) -> Dict[str, Any]:
    """
    Wersja z bardzo obszernym loggingiem:
    - ile inputów znaleziono per team
    - jakie jersey w HTML wykryto
    - dla KAŻDEGO update: current vs desired, delta-skip, payload/args, response
    """

    # request correlation id (żebyś mógł filtrować logi jednego requestu)
    req_id = str(uuid.uuid4())[:8]
    _dbg("apply_protocol start", req_id=req_id, host_name=host_name, guest_name=guest_name)

    players_inputs = _collect_players_inputs(soup, host_name=host_name, guest_name=guest_name)
    comp_inputs = _collect_companion_inputs(soup, host_name=host_name, guest_name=guest_name)
    comment_meta = _collect_comment_input(soup)

    players_inputs_host = sum(1 for k in players_inputs.keys() if k[0] == "host")
    players_inputs_guest = sum(1 for k in players_inputs.keys() if k[0] == "guest")
    comp_inputs_host = sum(1 for k in comp_inputs.keys() if k[0] == "host")
    comp_inputs_guest = sum(1 for k in comp_inputs.keys() if k[0] == "guest")

    _dbg(
        "inputs counts",
        req_id=req_id,
        players_inputs_total=len(players_inputs),
        players_inputs_host=players_inputs_host,
        players_inputs_guest=players_inputs_guest,
        companions_inputs_total=len(comp_inputs),
        companions_inputs_host=comp_inputs_host,
        companions_inputs_guest=comp_inputs_guest,
        comment_found=bool(comment_meta),
    )

    updated = 0
    skipped = 0
    failed: List[Dict[str, Any]] = []
    missing: List[Dict[str, Any]] = []
    skipped_items: List[Dict[str, Any]] = []

    player_kinds_order = ["goals", "entered", "warn", "p2", "disq", "pk_total", "pk_goals", "so_total", "so_goals"]
    comp_kinds_order = ["warn", "p2", "disq"]

    # ---- players ----
    for team in ("host", "guest"):
        team_stats = stats_map.get(team) or {}
        _dbg("team players processing", req_id=req_id, team=team, players_in_stats=len(team_stats))

        for key, st in team_stats.items():
            if not re.fullmatch(r"\d{1,3}", str(key)):
                continue
            jersey = str(int(key))

            for kind in player_kinds_order:
                desired = _desired_value_for_player_kind(st, kind)
                if desired is None:
                    continue

                meta = players_inputs.get((team, jersey, kind))
                if not meta:
                    missing.append({"section": "players", "team": team, "player": jersey, "kind": kind})
                    _dbg("MISSING players input", req_id=req_id, team=team, jersey=jersey, kind=kind, desired=desired)
                    continue

                inp = meta["inp"]
                args4 = meta["args4"]

                cur_val = _current_text_value(inp)
                cur_checked = _is_checked_dom(inp)

                # DELTA: skip if already equal
                if _delta_equal_player(inp, kind, desired):
                    skipped += 1
                    skipped_items.append({"section": "players", "team": team, "player": jersey, "kind": kind})
                    continue

                inp_type = (inp.get("type") or "").lower()
                if inp_type == "checkbox":
                    checked = bool(desired)
                    value_str = (inp.get("value") or "1").strip()
                else:
                    checked = False
                    if kind in _NUMERIC_KINDS:
                        value_str = _desired_str_for_numeric(int(desired or 0))
                    else:
                        value_str = str(desired)

                _dbg(
                    "UPDATE players sending",
                    req_id=req_id,
                    team=team,
                    jersey=jersey,
                    kind=kind,
                    desired=desired,
                    cur_val=cur_val,
                    cur_checked=cur_checked,
                    send_value=value_str,
                    send_checked=checked,
                    args4=args4,
                )

                ok, resp_txt = await _save_via_zapisz2(client, args4, value_str=value_str, checked=checked)
                if ok:
                    updated += 1
                    _dbg(
                        "UPDATE players OK",
                        req_id=req_id,
                        team=team,
                        jersey=jersey,
                        kind=kind,
                        resp=resp_txt,
                    )
                else:
                    failed.append({
                        "section": "players",
                        "team": team,
                        "player": jersey,
                        "kind": kind,
                        "sent_value": value_str,
                        "sent_checked": checked,
                        "resp": resp_txt,
                    })
                    _dbg(
                        "UPDATE players FAIL",
                        req_id=req_id,
                        team=team,
                        jersey=jersey,
                        kind=kind,
                        resp=resp_txt,
                    )

    # ---- companions A..E ----
    for team in ("host", "guest"):
        team_stats = stats_map.get(team) or {}
        _dbg("team companions processing", req_id=req_id, team=team, items_in_stats=len(team_stats))

        for key, st in team_stats.items():
            if not re.fullmatch(r"[A-E]", str(key).upper()):
                continue
            letter = str(key).upper()

            for kind in comp_kinds_order:
                desired_checked = _desired_value_for_companion_kind(st, kind)
                meta = comp_inputs.get((team, letter, kind))
                if not meta:
                    missing.append({"section": "companions", "team": team, "player": letter, "kind": kind})
                    _dbg("MISSING companions input", req_id=req_id, team=team, letter=letter, kind=kind, desired=desired_checked)
                    continue

                inp = meta["inp"]
                args8 = meta["args8"]
                cur_checked = _is_checked_dom(inp)
                value_str = (inp.get("value") or "").strip()

                if _delta_equal_companion(inp, desired_checked):
                    skipped += 1
                    skipped_items.append({"section": "companions", "team": team, "player": letter, "kind": kind})
                    _dbg(
                        "SKIP delta companions",
                        req_id=req_id,
                        team=team,
                        letter=letter,
                        kind=kind,
                        desired=desired_checked,
                        cur_checked=cur_checked,
                        args8=args8,
                    )
                    continue

                _dbg(
                    "UPDATE companions sending",
                    req_id=req_id,
                    team=team,
                    letter=letter,
                    kind=kind,
                    desired=desired_checked,
                    cur_checked=cur_checked,
                    send_checked=desired_checked,
                    checkbox_value=value_str,
                    args8=args8,
                )

                ok, resp_txt = await _save_via_zapisz4(client, args8, value_str=value_str, checked=desired_checked)
                if ok:
                    updated += 1
                    _dbg("UPDATE companions OK", req_id=req_id, team=team, letter=letter, kind=kind, resp=resp_txt)
                else:
                    failed.append({
                        "section": "companions",
                        "team": team,
                        "player": letter,
                        "kind": kind,
                        "sent_value": value_str,
                        "sent_checked": desired_checked,
                        "resp": resp_txt,
                    })
                    _dbg("UPDATE companions FAIL", req_id=req_id, team=team, letter=letter, kind=kind, resp=resp_txt)
    # ---- referee comment (zapiszProtok3 -> zawody_zapisz3.php) ----
    if comment_meta:
        inp = comment_meta["inp"]
        args4 = comment_meta["args4"]

        desired_text = _sanitize_comment_text(referee_comment or "")
        if desired_text.strip() == "":
            _dbg("SKIP comment empty desired", req_id=req_id)
        else:
            cur_text = _current_text_value(inp)

            # DELTA
            cur_norm = _sanitize_comment_text(cur_text or "")
            des_norm = desired_text
            if cur_norm == des_norm:
                skipped += 1
                skipped_items.append({"section": "comment", "team": None, "player": None, "kind": "komentarz"})
                _dbg("SKIP delta comment", req_id=req_id, desired=desired_text[:200], cur=cur_text[:200], args4=args4)
            else:
                _dbg("UPDATE comment sending", req_id=req_id, desired=desired_text[:200], cur=cur_text[:200], args4=args4)

                # w JS jest this.checked, ale textarea nie ma sensownego checked => wysyłamy false
                ok, resp_txt = await _save_via_zapisz3(client, args4, value_str=desired_text, checked=False)

                if ok:
                    updated += 1
                    _dbg("UPDATE comment OK", req_id=req_id, resp=resp_txt)
                else:
                    failed.append({
                        "section": "comment",
                        "team": None,
                        "player": None,
                        "kind": "komentarz",
                        "sent_value": desired_text[:4000],
                        "resp": resp_txt,
                    })
                    _dbg("UPDATE comment FAIL", req_id=req_id, resp=resp_txt)
    else:
        # textarea nie istnieje na stronie protokołu
        if (referee_comment or "").strip():
            missing.append({"section": "comment", "team": None, "player": None, "kind": "komentarz"})
            _dbg("MISSING comment textarea", req_id=req_id)

    _dbg(
        "apply_protocol end",
        req_id=req_id,
        updated_cells=updated,
        skipped_cells=skipped,
        failed=len(failed),
        missing=len(missing),
        players_inputs_host=players_inputs_host,
        players_inputs_guest=players_inputs_guest,
        companions_inputs_host=comp_inputs_host,
        companions_inputs_guest=comp_inputs_guest,
    )

    return {
        "updated_cells": updated,
        "skipped_cells": skipped,
        "failed": failed,
        "missing": missing,
        "skipped": skipped_items,
        "debug": {
            "req_id": req_id,
            "host_name": host_name,
            "guest_name": guest_name,
            "players_inputs_host": players_inputs_host,
            "players_inputs_guest": players_inputs_guest,
            "companions_inputs_host": comp_inputs_host,
            "companions_inputs_guest": comp_inputs_guest,
            "players_inputs_total": len(players_inputs),
            "companions_inputs_total": len(comp_inputs),
        },
    }


@router.post(
    "/judge/results/protocol",
    summary="Zapisz protokół na baza.zprp.pl na podstawie data_json (zawodnicy + osoby towarzyszące; 4 bloki; delta)",
)
async def save_protocol_from_json(
    req: ProtocolSaveRequest,
    settings: Settings = Depends(get_settings),
    keys=Depends(get_rsa_keys),
):
    logger.warning(
        "[protocol-debug] ENTER save_protocol_from_json env=%r enabled=%s",
        os.getenv("RESULTS_PROTOCOL_DEBUG"),
        _dbg_enabled(),
    )
    private_key, _ = keys
    user_plain = _decrypt_field(req.username, private_key)
    pass_plain = _decrypt_field(req.password, private_key)

    match_id = _extract_match_id(req.details_path)

    data_json = req.data_json or {}
    if not isinstance(data_json, dict):
        raise HTTPException(400, "data_json musi być obiektem JSON")

    try:
        client = await _login_and_client(user_plain, pass_plain, settings)
        try:
            details_url = _details_path_to_url(req.details_path)
            _, html = await fetch_with_correct_encoding(
                client,
                details_url,
                method="GET",
                cookies=client.cookies,
            )
            soup = BeautifulSoup(html, "html.parser")

            stats_map = _build_stats_map(data_json)
            mc = data_json.get("matchConfig") or {}
            host_name = mc.get("hostTeamName") or ""
            guest_name = mc.get("guestTeamName") or ""

            mc = data_json.get("matchConfig") or {}
            extras = mc.get("extras") or {}
            ref_comment = extras.get("detailedRefereeNotesText") or ""

            result = await _apply_protocol_updates_4blocks(
                client,
                soup,
                stats_map,
                host_name=host_name,
                guest_name=guest_name,
                referee_comment=ref_comment,
            )
        finally:
            await client.aclose()

        failed = result.get("failed") or []
        success = (len(failed) == 0)

        return {
            "success": success,
            "match_id": match_id,
            **result,
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error("save_protocol_from_json error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail=f"Nie udało się zapisać protokołu: {e}")

# ============================================================
# EXPORT PROTOCOL PDF FROM ProEl data_json -> XLSX TEMPLATE -> PDF
# ============================================================

import asyncio
import os
import math
import shutil
import subprocess
import tempfile

from fastapi.responses import FileResponse
from openpyxl import load_workbook
from io import BytesIO
import zipfile

def _load_template_media_bytes(template_path: str) -> Dict[str, bytes]:
    """
    Czyta wszystkie pliki xl/media/* z szablonu XLSX do pamięci.
    Klucz: 'xl/media/image1.png'
    """
    media: Dict[str, bytes] = {}
    with zipfile.ZipFile(template_path, "r") as z:
        for name in z.namelist():
            if name.startswith("xl/media/"):
                media[name] = z.read(name)
    return media


def _rehydrate_images_in_workbook(wb, media: Dict[str, bytes]) -> None:
    """
    Podmienia obrazy w każdym arkuszu na takie, które trzymają dane w BytesIO,
    żeby wb.save() nie próbował czytać z zamkniętego strumienia.
    """
    for ws in wb.worksheets:
        imgs = list(getattr(ws, "_images", []) or [])
        if not imgs:
            continue

        # usuń stare
        ws._images = []

        for img in imgs:
            # openpyxl trzyma ścieżkę jako '/xl/media/imageX.png'
            path = (getattr(img, "path", "") or "").lstrip("/")
            blob = media.get(path)

            if not blob:
                # jeśli z jakiegoś powodu nie ma w media, to lepiej pominąć niż wywalić save()
                logger.warning("Protocol PDF: missing media for image path=%r", path)
                continue

            bio = BytesIO(blob)
            new_img = Image(bio)
            new_img.width = img.width
            new_img.height = img.height
            new_img.anchor = copy.deepcopy(img.anchor)

            ws.add_image(new_img)


def _copy_images_safe(src_ws, dst_ws):
    """
    Kopiuje obrazy, zakładając że src_ws ma już rehydratowane obrazy (ref=BytesIO).
    """
    for img in getattr(src_ws, "_images", []) or []:
        data: Optional[bytes] = None

        ref = getattr(img, "ref", None)
        if ref is not None and hasattr(ref, "getvalue"):
            data = ref.getvalue()

        if not data:
            # fallback (powinno już być bezpieczne po rehydratacji)
            try:
                data = img._data()
            except Exception as e:
                logger.warning("Protocol PDF: could not clone image: %s", e)
                continue

        bio = BytesIO(data)
        new_img = Image(bio)
        new_img.width = img.width
        new_img.height = img.height
        new_img.anchor = copy.deepcopy(img.anchor)

        dst_ws.add_image(new_img)


class ProtocolPdfRequest(BaseModel):
    data_json: Dict[str, Any]  # dokładnie ten sam JSON ProEl


def _ms_to_mmss(ms: Optional[int]) -> str:
    if ms is None:
        return ""
    try:
        ms_i = int(ms)
    except Exception:
        return ""
    if ms_i < 0:
        ms_i = 0
    mm = ms_i // 60000
    ss = (ms_i % 60000) // 1000
    return f"{mm:02d}:{ss:02d}"


def _event_minute_from_ms(ms: int) -> int:
    # minute numbering: 0:00 => 1, 53:12 => 54 (floor + 1)
    try:
        ms_i = int(ms)
    except Exception:
        ms_i = 0
    if ms_i < 0:
        ms_i = 0
    return (ms_i // 60000) + 1


def _safe_int(v: Any, default: int = 0) -> int:
    try:
        return int(v)
    except Exception:
        return default


def _get_match_core(data_json: Dict[str, Any]) -> Dict[str, Any]:
    mc = data_json.get("matchConfig") or {}
    return {
        "matchNumber": (mc.get("matchNumber") or "").strip(),
        "hostName": (mc.get("hostTeamName") or "").strip(),
        "guestName": (mc.get("guestTeamName") or "").strip(),
        "halfTimeMin": _safe_int(mc.get("halfTime") or 30, 30),
        "scoreHost": _safe_int(data_json.get("scoreHost"), 0),
        "scoreGuest": _safe_int(data_json.get("scoreGuest"), 0),
        "halfScoreHost": _safe_int((data_json.get("halfScore") or {}).get("host"), 0),
        "halfScoreGuest": _safe_int((data_json.get("halfScore") or {}).get("guest"), 0),
        "hostPlayers": list(mc.get("hostPlayers") or []),
        "guestPlayers": list(mc.get("guestPlayers") or []),
        "venueAddress": (mc.get("venueAddress") or "").strip(),
        "referee1": (mc.get("referee1") or "").strip(),
        "referee2": (mc.get("referee2") or "").strip(),
        "delegate": (mc.get("delegate") or "").strip(),
        "timekeeper": (mc.get("timekeeper") or "").strip(),
        "secretary": (mc.get("secretary") or "").strip(),
        "hostPlayerCards": list(mc.get("hostPlayerCards") or []),
        "guestPlayerCards": list(mc.get("guestPlayerCards") or []),
        "hostCompanions": list(mc.get("hostCompanions") or []),
        "guestCompanions": list(mc.get("guestCompanions") or []),

    }


def _place_timeouts(ws, *, team_timeouts: Dict[str, Any], half_ms: int, is_host: bool) -> None:
    """
    Wypełnia czasy zgodnie z Twoimi regułami, ale uproszczone do logiki:
    - w 1. połowie: pierwszy timeout do (row10), drugi do (row11)
    - w 2. połowie: pierwszy timeout do (row10), drugi do (row11)
    Zgadza się z opisanymi przypadkami (max 2 czasy na połowę).
    """
    t1 = team_timeouts.get("first")
    t2 = team_timeouts.get("second")
    t3 = team_timeouts.get("third")
    times = [t for t in [t1, t2, t3] if t is not None]

    # sort rosnąco po czasie (ms)
    def _as_int(x):
        try:
            return int(x)
        except Exception:
            return 10**18

    times.sort(key=_as_int)

    half1 = [t for t in times if _as_int(t) < half_ms]
    half2 = [t for t in times if _as_int(t) >= half_ms]

    # host: half1 -> AL10/AL11, half2 -> AW10/AW11
    # guest: half1 -> AU10/AU11, half2 -> BF10/BF11
    if is_host:
        h1_cells = ["AL10", "AL11"]
        h2_cells = ["AW10", "AW11"]
    else:
        h1_cells = ["AU10", "AU11"]
        h2_cells = ["BF10", "BF11"]

    # 1. połowa
    if len(half1) == 0:
        # Nie wzięto żadnego czasu w 1. połowie - wpisujemy "---" w AL10 i AL11
        ws[h1_cells[0]].value = "---"
        ws[h1_cells[1]].value = "---"
    elif len(half1) == 1:
        # Wzięto tylko 1 czas - wpisujemy w AL10 oraz AL11
        ws[h1_cells[0]].value = _ms_to_mmss(half1[0])
        ws[h1_cells[1]].value = "---"
    else:
        # Wzięto dwa czasy - wpisujemy oba
        ws[h1_cells[0]].value = _ms_to_mmss(half1[0])
        ws[h1_cells[1]].value = _ms_to_mmss(half1[1])

    # 2. połowa
    if len(half2) == 0:
        # Nie wzięto żadnego czasu w 2. połowie - wpisujemy "---" w AW10 i AW11
        ws[h2_cells[0]].value = "---"
        ws[h2_cells[1]].value = "---"
    elif len(half2) == 1:
        # Wzięto tylko 1 czas w 2. połowie - wpisujemy w AW10 oraz AW11
        ws[h2_cells[0]].value = _ms_to_mmss(half2[0])
        ws[h2_cells[1]].value = "---"
    else:
        # Wzięto dwa czasy w 2. połowie - wpisujemy oba
        ws[h2_cells[0]].value = _ms_to_mmss(half2[0])
        ws[h2_cells[1]].value = _ms_to_mmss(half2[1])


def _player_stats_map(data_json: Dict[str, Any], team: str) -> Dict[int, Dict[str, Any]]:
    key = "hostPlayerStats" if team == "host" else "guestPlayerStats"
    arr = data_json.get(key) or []
    out: Dict[int, Dict[str, Any]] = {}
    for ps in arr:
        if not isinstance(ps, dict):
            continue
        n = ps.get("number")
        if n is None:
            continue
        try:
            out[int(n)] = ps
        except Exception:
            continue
    return out

# ============================================================================
# Ptaszki badań lekarskich w protokole
# ============================================================================
#
# Szablon protocol_template.xlsx dostał JEDNĄ nową, wąską kolumnę PRZED kolumną
# A — wyłącznie na ptaszki. Cała dotychczasowa zawartość arkusza przesunęła się
# o kolumnę w prawo.
#
# Zamiast przepisywać ~170 literalnych adresów w tym pliku, zakładamy na arkusz
# nakładkę: ws["A11"] trafia do fizycznego "B11". Wycofanie zmiany to
# ustawienie PROTOCOL_COL_SHIFT na 0 i przywrócenie
# protocol_template_BACKUP_preExamCol.xlsx.

PROTOCOL_COL_SHIFT = 1

_A1_RE = re.compile(r"^(\$?)([A-Za-z]{1,3})(\$?)(\d+)$")


def _col_to_idx(letters: str) -> int:
    n = 0
    for ch in letters.upper():
        n = n * 26 + (ord(ch) - 64)
    return n


def _idx_to_col(idx: int) -> str:
    out = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        out = chr(65 + rem) + out
    return out


def shift_ref(ref: str, shift: int = PROTOCOL_COL_SHIFT) -> str:
    """"A11" -> "B11", "AL15:AV15" -> "AM15:AW15". Obsługuje też "$A$1"."""
    if not shift:
        return ref
    parts = str(ref).split(":")
    out: List[str] = []
    for part in parts:
        m = _A1_RE.match(part.strip())
        if not m:
            return ref  # nieznany kształt — lepiej nie ruszać
        d1, col, d2, row = m.groups()
        out.append("%s%s%s%s" % (d1, _idx_to_col(_col_to_idx(col) + shift), d2, row))
    return ":".join(out)


class ShiftedWS:
    """
    Przezroczysta nakładka na arkusz: adresy w kodzie zostają takie, jak przed
    dołożeniem kolumny ptaszków, a trafiają o kolumnę dalej.

    Do `wb.copy_worksheet`, `_copy_images_safe` i wstawiania samych ptaszków
    trzeba brać `.raw` — one operują na fizycznych współrzędnych.
    """

    __slots__ = ("_ws", "_shift")

    def __init__(self, ws, shift: int = PROTOCOL_COL_SHIFT):
        self._ws = ws
        self._shift = shift

    @property
    def raw(self):
        return self._ws

    def __getitem__(self, ref):
        return self._ws[shift_ref(ref, self._shift)]

    def __setitem__(self, ref, value):
        self._ws[shift_ref(ref, self._shift)] = value

    def merge_cells(self, range_string=None, **kwargs):
        if range_string is not None:
            return self._ws.merge_cells(shift_ref(range_string, self._shift), **kwargs)
        return self._ws.merge_cells(**kwargs)

    def add_image(self, img, anchor=None):
        if isinstance(anchor, str):
            anchor = shift_ref(anchor, self._shift)
        return self._ws.add_image(img, anchor)

    def __getattr__(self, name):
        return getattr(self._ws, name)


# Whitelista dozwolonych wartości pola `exam`. Zostaje jako mapa (a nie zbiór),
# bo po tych kluczach chodzi `_player_exam_map_from_cards` i testy; wartości to
# barwa wiodąca statusu, ta sama co `EXAM_COLORS` po stronie aplikacji.
EXAM_MARK_RGB = {
    "zprp": (46, 158, 91),
    "wzpr": (47, 158, 126),
    "manual": (28, 52, 122),
}

# Paleta pieczątki — port z `components/matchDetails/ExamBadge.tsx`, żeby ten
# sam status wyglądał tak samo w składzie i na wydruku. Zieleń jest o stopień
# jaśniejsza niż `EXAM_MARK_RGB`: tamta jest do obwódek, ta do wypełnienia.
_EXAM_RAMP = {
    "zprp": ("#5FE3C2", "#49D3B0", "#22A06B"),
    "wzpr": ("#5FE3C2", "#49D3B0", "#22A06B"),
    "manual": ("#5B84D8", "#3E63BE", "#24408A"),
}
_EXAM_RAMP_STOPS = (0.0, 0.46, 1.0)
_EXAM_CHECK = (247, 255, 251)
_EXAM_PLATE_BG = (15, 11, 9, 235)
_EXAM_PLATE_RIM = (95, 227, 194, 107)
_EXAM_WZPR_TEXT = (141, 240, 212)

# Podpis pod ptaszkiem WZPR. Pełne „WZPR" mieści się w 5,3 mm szerokości z
# wysokością liter ~1,2 mm (≈3,5 pt) — czytelne, bo PNG jedzie w wysokiej
# rozdzielczości, ale to granica. Gdyby na papierze okazało się za drobne,
# jedyna zmiana to podmiana tej stałej na "W".
EXAM_WZPR_CAPTION = "WZPR"

_EXAM_MARK_CACHE: Dict[str, bytes] = {}

# Rozmiar znaczka wynika z twardego ograniczenia szablonu, nie z gustu: kolumna
# ptaszków ma 20 px, bo dokładnie tyle miejsca oddaje jej dawny LEWY MARGINES
# (0,19685" przy skali 90 %). Szerzej = tabela nie mieści się na A4 i prawy skraj
# protokołu wyjeżdża na osobną stronę. Wiersz zawodnika ma 13,2 pt ≈ 17,6 px.
EXAM_MARK_W = 20
EXAM_MARK_H = 16
# Rysujemy _EXAM_SS razy większe i — w odróżnieniu od poprzedniej wersji — NIE
# skalujemy w dół. PNG zostaje w pełnej rozdzielczości, a do 20×16 px sprowadza
# go dopiero rozmiar wyświetlania w arkuszu. Wcześniejszy `resize` zapisywał
# bitmapę 20×15 px, czyli ~96 DPI — na wydruku znaczek był z tego powodu
# rozmyty niezależnie od tego, jak starannie go narysowaliśmy.
_EXAM_SS = 8
# Znaczek zajmuje całą szerokość kolumny; w pionie 1 px zapasu, żeby nie kleił
# się do linii wiersza (16 + 1 mieści się w 17,6 px wiersza).
_EXAM_MARK_OFFSET_X_PX = 0
_EXAM_MARK_OFFSET_Y_PX = 1


def _hex_rgb(value: str) -> Tuple[int, int, int]:
    v = value.lstrip("#")
    return (int(v[0:2], 16), int(v[2:4], 16), int(v[4:6], 16))


def _ramp_color(ramp: Tuple[str, ...], stops: Tuple[float, ...], t: float):
    """Kolor z trójstopniowej rampy w punkcie t∈[0,1] (interpolacja liniowa)."""
    t = 0.0 if t < 0 else (1.0 if t > 1 else t)
    cols = [_hex_rgb(c) for c in ramp]
    for i in range(len(stops) - 1):
        lo, hi = stops[i], stops[i + 1]
        if t <= hi or i == len(stops) - 2:
            k = 0.0 if hi <= lo else (t - lo) / (hi - lo)
            k = 0.0 if k < 0 else (1.0 if k > 1 else k)
            a, b = cols[i], cols[i + 1]
            return tuple(int(round(a[j] + (b[j] - a[j]) * k)) for j in range(3))
    return cols[-1]


def _exam_gradient(size, ramp, stops, start=(0.18, 0.0), end=(0.82, 1.0)):
    """Gradient liniowy w zadanym kierunku. Liczony raz na rodzaj i cache'owany."""
    w, h = size
    sx, sy = start
    ex, ey = end
    dx, dy = ex - sx, ey - sy
    denom = (dx * dx + dy * dy) or 1.0
    lut = [_ramp_color(ramp, stops, i / 255.0) + (255,) for i in range(256)]
    px = []
    for y in range(h):
        vy = ((y + 0.5) / h - sy) * dy
        for x in range(w):
            t = (((x + 0.5) / w - sx) * dx + vy) / denom
            idx = int(t * 255)
            px.append(lut[0 if idx < 0 else (255 if idx > 255 else idx)])
    img = PILImage.new("RGBA", (w, h))
    img.putdata(px)
    return img


def _exam_gloss(size):
    """Rozbłysk u góry i ciemny rant u dołu — bez tego kwadrat jest płaski."""
    w, h = size
    px = []
    for y in range(h):
        for x in range(w):
            t = (((x + 0.5) / w) + ((y + 0.5) / h)) / 2.0
            if t <= 0.46:
                px.append((255, 255, 255, int(round(76 * (1 - t / 0.46)))))
            else:
                px.append((11, 60, 40, int(round(51 * (t - 0.46) / 0.54))))
    img = PILImage.new("RGBA", (w, h))
    img.putdata(px)
    return img


def _exam_rounded_mask(size, radius: int):
    from PIL import ImageDraw

    mask = PILImage.new("L", size, 0)
    ImageDraw.Draw(mask).rounded_rectangle(
        [0, 0, size[0] - 1, size[1] - 1], radius=radius, fill=255
    )
    return mask


def _exam_draw_check(draw, box, color, stroke: int) -> None:
    """
    Ptaszek dokładnie z makiety aplikacji (viewBox 52×52, „M12 27.5 L21.5 37
    L40 15.5"). Ionicons rysuje go pod innym kątem, a chodzi o to, żeby wydruk
    i skład miały ten sam znak.
    """
    x0, y0, x1, y1 = box
    w, h = x1 - x0, y1 - y0
    pts = [
        (x0 + nx * w, y0 + ny * h)
        for nx, ny in ((12 / 52, 27.5 / 52), (21.5 / 52, 37 / 52), (40 / 52, 15.5 / 52))
    ]
    draw.line(pts, fill=color, width=stroke, joint="curve")
    # PIL nie zna zaokrąglonych końców linii — dokładamy je kółkami.
    r = stroke / 2.0
    for px_, py_ in pts:
        draw.ellipse([px_ - r, py_ - r, px_ + r, py_ + r], fill=color)


def _exam_text_strip(text: str, rgb, max_w: int, max_h: int):
    """
    Napis dopasowany do prostokąta. Renderujemy w dużej skali i dopiero potem
    skalujemy — dzięki temu wynik nie zależy od tego, czy w obrazie znalazła się
    jakakolwiek czcionka TTF (fallback `load_default` to bitmapa 11 px, która
    bez tego kroku byłaby mikroskopijna względem kanwy).
    """
    from PIL import ImageDraw, ImageFont

    font = None
    for name in ("DejaVuSans-Bold.ttf", "arialbd.ttf", "DejaVuSans.ttf", "arial.ttf"):
        try:
            font = ImageFont.truetype(name, 96)
            break
        except Exception:
            continue
    if font is None:
        font = ImageFont.load_default()

    probe = ImageDraw.Draw(PILImage.new("RGBA", (1, 1)))
    box = probe.textbbox((0, 0), text, font=font)
    tw, th = max(1, box[2] - box[0]), max(1, box[3] - box[1])

    layer = PILImage.new("RGBA", (tw, th), (0, 0, 0, 0))
    ImageDraw.Draw(layer).text(
        (-box[0], -box[1]), text, font=font, fill=tuple(rgb) + (255,)
    )
    k = min(max_w / float(tw), max_h / float(th))
    return layer.resize((max(1, int(tw * k)), max(1, int(th * k))), PILImage.LANCZOS)


def _exam_mark_png(kind: str) -> bytes:
    """
    PNG ze znaczkiem badań — pieczątka w stylu `ExamBadge` z aplikacji.
    Deterministyczny i cache'owany (trzy obrazy na cały proces).

      • zprp / manual → squircle z trójstopniowym gradientem i białym ptaszkiem,
      • wzpr          → ta sama zielona głowa z ptaszkiem, a pod nią ciemna
                        stopka z podpisem: status ma być rozpoznawalny bez
                        porównywania odcieni zieleni z sąsiednim wierszem.
    """
    kind = str(kind or "").strip().lower()
    if kind not in EXAM_MARK_RGB:
        return b""
    cached = _EXAM_MARK_CACHE.get(kind)
    if cached is not None:
        return cached
    if PILImage is None:
        return b""

    try:
        from PIL import ImageDraw
    except Exception:
        return b""

    ss = _EXAM_SS
    W, H = EXAM_MARK_W * ss, EXAM_MARK_H * ss
    img = PILImage.new("RGBA", (W, H), (0, 0, 0, 0))
    ramp = _EXAM_RAMP[kind]

    if kind == "wzpr":
        # ── pieczątka dwuczęściowa: głowa + stopka z podpisem ──
        radius = int(0.26 * W)
        head_h = int(0.60 * H)

        plate = PILImage.new("RGBA", (W, H), _EXAM_PLATE_BG)
        head = _exam_gradient((W, head_h), ramp, _EXAM_RAMP_STOPS)
        head.alpha_composite(_exam_gloss((W, head_h)))
        plate.paste(head, (0, 0), head)

        mask = _exam_rounded_mask((W, H), radius)
        img.paste(plate, (0, 0), mask)

        d = ImageDraw.Draw(img)
        side = int(0.42 * W)
        cx, cy = W / 2.0, head_h / 2.0
        _exam_draw_check(
            d,
            (cx - side / 2.0, cy - side / 2.0, cx + side / 2.0, cy + side / 2.0),
            _EXAM_CHECK,
            max(1, int(0.085 * W)),
        )

        caption = _exam_text_strip(
            EXAM_WZPR_CAPTION,
            _EXAM_WZPR_TEXT,
            int(W * 0.86),
            int((H - head_h) * 0.62),
        )
        img.alpha_composite(
            caption,
            (
                int((W - caption.width) / 2),
                int(head_h + ((H - head_h) - caption.height) / 2),
            ),
        )

        d.rounded_rectangle(
            [0, 0, W - 1, H - 1],
            radius=radius,
            outline=_EXAM_PLATE_RIM,
            width=max(1, int(0.35 * ss)),
        )
    else:
        # ── squircle z pełnym wypełnieniem ──
        side = H - 2 * ss
        radius = int(0.36 * side)
        badge = _exam_gradient((side, side), ramp, _EXAM_RAMP_STOPS)
        badge.alpha_composite(_exam_gloss((side, side)))

        d = ImageDraw.Draw(badge)
        inner = 0.60 * side
        off = (side - inner) / 2.0
        _exam_draw_check(
            d,
            (off, off, off + inner, off + inner),
            _EXAM_CHECK,
            max(1, int(0.105 * side)),
        )

        img.paste(
            badge,
            (int((W - side) / 2), int((H - side) / 2)),
            _exam_rounded_mask((side, side), radius),
        )

    bio = BytesIO()
    img.save(bio, format="PNG")
    data = bio.getvalue()
    _EXAM_MARK_CACHE[kind] = data
    return data


def _add_exam_mark(ws_raw, *, row: int, kind: str) -> bool:
    """
    Wstawia ptaszek w NOWEJ kolumnie A (fizycznej), czyli tej dołożonej przed
    dawną kolumną numeru. Dlatego bierze arkusz surowy, nie przesunięty.
    """
    data = _exam_mark_png(kind)
    if not data:
        return False
    try:
        from openpyxl.drawing.spreadsheet_drawing import (
            AnchorMarker,
            OneCellAnchor,
        )
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        from openpyxl.utils.units import pixels_to_EMU

        # Każdy obrazek MUSI mieć własny BytesIO — openpyxl trzyma referencję do
        # strumienia i współdzielenie jednego bufora psuje zapis pliku.
        img = Image(BytesIO(data))
        img.width = EXAM_MARK_W
        img.height = EXAM_MARK_H
        # Kotwica budowana ręcznie, a nie przez add_image(img, "A11"): tamta
        # postać zostawia zwykły string i nie da się ustawić przesunięcia, więc
        # ptaszek kleiłby się do lewej krawędzi komórki.
        img.anchor = OneCellAnchor(
            _from=AnchorMarker(
                col=0,  # FIZYCZNA kolumna A = kolumna ptaszków
                colOff=pixels_to_EMU(_EXAM_MARK_OFFSET_X_PX),
                row=row - 1,
                rowOff=pixels_to_EMU(_EXAM_MARK_OFFSET_Y_PX),
            ),
            ext=XDRPositiveSize2D(
                pixels_to_EMU(EXAM_MARK_W), pixels_to_EMU(EXAM_MARK_H)
            ),
        )
        ws_raw.add_image(img)
        return True
    except Exception:
        logger.warning("Nie udało się wstawić ptaszka badań (wiersz %s)", row, exc_info=True)
        return False


def _player_exam_map_from_cards(cards: List[Any]) -> Dict[int, str]:
    """
    number -> "zprp" | "wzpr" | "manual".

    Starsze wersje aplikacji nie wysyłają pola `exam` — wtedy mapa jest pusta,
    kolumna ptaszków zostaje czysta i wydruk wygląda dokładnie jak dotąd.
    """
    out: Dict[int, str] = {}
    for c in cards or []:
        if not isinstance(c, dict):
            continue
        n = c.get("number")
        if n is None:
            continue
        try:
            num = int(n)
        except Exception:
            continue
        kind = str(c.get("exam") or "").strip().lower()
        if kind in EXAM_MARK_RGB:
            out[num] = kind
    return out


def _player_fullname_map_from_cards(cards: List[Any]) -> Dict[int, str]:
    out: Dict[int, str] = {}
    for c in cards or []:
        if not isinstance(c, dict):
            continue
        n = c.get("number")
        if n is None:
            continue
        try:
            num = int(n)
        except Exception:
            continue
        name = (c.get("fullName") or "").strip()
        if name:
            out[num] = name
    return out


def _player_fullname_map_from_stats(stats_by_number: Dict[int, Dict[str, Any]]) -> Dict[int, str]:
    out: Dict[int, str] = {}
    for num, ps in (stats_by_number or {}).items():
        if not isinstance(ps, dict):
            continue
        name = (ps.get("fullName") or "").strip()
        if name:
            out[int(num)] = name
    return out


def _pick_companion_time(c: Dict[str, Any], *keys: str) -> str:
    for k in keys:
        v = c.get(k)
        if isinstance(v, str) and v.strip():
            return v.strip()
    return ""

def _companion_penalty_strings(comp_list: List[Any]) -> Dict[str, Dict[str, str]]:
    """
    Zwraca mapę:
      "A".."E" -> {"warn": "U - MM:SS" | "---", "p2": "2' - MM:SS" | "---", "disq": "D - MM:SS" | "---"}
    Źródła:
      - upomnienie: warned + warnTime/warningTime/warnedTime (jeśli istnieje)
      - 2 minuty: penaltyTimes[0] (pierwsza 2')
      - dyskwalifikacja: red + redTime
    """
    out: Dict[str, Dict[str, str]] = {}
    for c in comp_list or []:
        if not isinstance(c, dict):
            continue
        cid = str(c.get("id") or "").strip().upper()
        if cid not in ("A", "B", "C", "D", "E"):
            continue

        # --- warning ---
        warned = bool(c.get("warned")) if "warned" in c else bool(c.get("warn")) if "warn" in c else False
        warn_time = _pick_companion_time(c, "warnTime", "warningTime", "warnedTime", "warning")
        warn_str = f"U - {warn_time}" if (warned and warn_time) else ("U - __:__" if warned else "")

        # --- 2' ---
        p_times = c.get("penaltyTimes") if isinstance(c.get("penaltyTimes"), list) else []
        p2_time = ""
        if p_times:
            first = p_times[0]
            if isinstance(first, str) and first.strip():
                p2_time = first.strip()
        # czasem możesz mieć boola "twoMinutes" bez listy — wtedy wpisz placeholder
        two_min = bool(c.get("twoMinutes")) if "twoMinutes" in c else False
        p2_str = f"2' - {p2_time}" if p2_time else ("2' - __:__" if two_min else "")

        # --- disq (red) ---
        red = bool(c.get("red")) if "red" in c else bool(c.get("disq")) if "disq" in c else False
        red_time = _pick_companion_time(c, "redTime", "disqTime", "disqualificationTime")
        disq_str = f"D - {red_time}" if (red and red_time) else ("D - __:__" if red else "")

        out[cid] = {"warn": warn_str, "p2": p2_str, "disq": disq_str}

    return out

def _companion_fullname_map(comp_list: List[Any]) -> Dict[str, str]:
    """
    Zwraca mapę: "A".."E" -> "NAZWISKO Imię"
    """
    out: Dict[str, str] = {}
    for c in comp_list or []:
        if not isinstance(c, dict):
            continue
        cid = str(c.get("id") or "").strip().upper()
        if cid not in ("A", "B", "C", "D", "E"):
            continue
        name = (c.get("fullName") or "").strip()
        if name:
            out[cid] = name
    return out

def _companion_meta_map(comp_list: List[Any]) -> Dict[str, Dict[str, str]]:
    """
    Zwraca mapę: "A".."E" -> {"function": "...", "license": "..."}
    """
    out: Dict[str, Dict[str, str]] = {}
    for c in comp_list or []:
        if not isinstance(c, dict):
            continue
        cid = str(c.get("id") or "").strip().upper()
        if cid not in ("A", "B", "C", "D", "E"):
            continue

        func = (c.get("function") or "").strip()
        lic = (c.get("license") or "").strip()

        out[cid] = {"function": func, "license": lic}
    return out


def _fill_players_block(
    ws,
    *,
    players: List[Any],
    stats_by_number: Dict[int, Dict[str, Any]],
    fullnames_by_number: Dict[int, str],
    start_row: int,
    end_row: int,
    exam_by_number: Optional[Dict[int, str]] = None,
    mark_ws=None,
) -> None:
    """
    Kolumny wg Twojej specyfikacji:
      - A: numer
      - Q: wejście "W" / "-"
      - S: bramki liczba / "-"
      - U: upomnienie "[minuta]'" / "-"
      - W: 2' #1 (MM:SS) / "---"
      - Z: 2' #2 / "---"
      - AC: 2' #3 / "---"
      - AF: dyskwalifikacja lub dysq z opisem / "---"
      - AI: zawsze "---"

    Dodatkowo, jeśli podano `exam_by_number` i `mark_ws` (arkusz SUROWY), w
    nowej kolumnie A ląduje ptaszek badań lekarskich. Zawodnicy bez ważnych
    badań nie dostają nic — puste miejsce jest tu informacją.
    """
    nums: List[int] = []
    for p in players or []:
        try:
            nums.append(int(p))
        except Exception:
            continue
    nums = sorted(set(nums))

    max_rows = (end_row - start_row + 1)
    nums = nums[:max_rows]

    for i in range(max_rows):
        row = start_row + i
        ws[f"AI{row}"].value = "---"  # zawsze

        if i >= len(nums):
            # zostaw pusto jeśli mniej zawodników
            ws[f"A{row}"].value = "--"
            ws[f"C{row}"].value = "-------------------------------------------"
            ws[f"Q{row}"].value = "-"
            ws[f"S{row}"].value = "-"
            ws[f"U{row}"].value = "-"
            ws[f"W{row}"].value = "---"
            ws[f"Z{row}"].value = "---"
            ws[f"AC{row}"].value = "---"
            ws[f"AF{row}"].value = "---"
            continue

        num = nums[i]
        ps = stats_by_number.get(num) or {}

        entered = bool(ps.get("entered") or False)
        goals = _safe_int(ps.get("goals") or 0, 0)
        warning = ps.get("warning")  # w ProEl bywa "12'"
        penalty1 = (ps.get("penalty1") or "").strip()
        penalty2 = (ps.get("penalty2") or "").strip()
        penalty3 = (ps.get("penalty3") or "").strip()
        disq_time = (ps.get("disqualification") or "").strip()
        disq_desc = (ps.get("disqualificationDesc") or "").strip()
        has_red = bool(ps.get("hasRedCard") or False)

        ws[f"A{row}"].value = num
        ws[f"C{row}"].value = (fullnames_by_number.get(num) or "")

        kind = (exam_by_number or {}).get(num)
        if kind and mark_ws is not None:
            _add_exam_mark(mark_ws, row=row, kind=kind)
        ws[f"Q{row}"].value = "W" if entered else "-"
        ws[f"S{row}"].value = goals if goals > 0 else "-"
        ws[f"U{row}"].value = str(warning).strip() if isinstance(warning, str) and warning.strip() else "-"

        ws[f"W{row}"].value = penalty1 if penalty1 else "---"
        ws[f"Z{row}"].value = penalty2 if penalty2 else "---"
        ws[f"AC{row}"].value = penalty3 if penalty3 else "---"

        if disq_time or disq_desc:
            if disq_time and disq_desc:
                ws[f"AF{row}"].value = f"{disq_time} {disq_desc}"
            elif disq_time:
                ws[f"AF{row}"].value = disq_time
            elif disq_desc:
                ws[f"AF{row}"].value = disq_desc
        else:
            ws[f"AF{row}"].value = "---"


TIMELINE_START_ROW = 15
TIMELINE_END_ROW = 63
TIMELINE_SKIP_ROWS = {31, 57}

TIMELINE_ROWS = [r for r in range(TIMELINE_START_ROW, TIMELINE_END_ROW + 1) if r not in TIMELINE_SKIP_ROWS]
TIMELINE_MAX_ROWS = len(TIMELINE_ROWS)  # było 47, teraz będzie 45

def _filter_protocol_events_for_timeline(protocol: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Zwraca listę eventów do przebiegu meczu po uwzględnieniu cofniętych bramek.

    Reguła:
    - zwykłe bramki (`goal`) trafiają do przebiegu,
    - jeśli później pojawi się `goalRemoved` z extra.origTime wskazującym na tę bramkę,
      to ta wcześniejsza bramka jest usuwana z przebiegu,
    - `penaltyKickScored` / `penaltyKickMissed` zostają bez zmian.
    """
    if not isinstance(protocol, list):
        return []

    filtered: List[Dict[str, Any]] = []

    for ev in protocol:
        if not isinstance(ev, dict):
            continue

        ev_type = ev.get("type")

        # Najpierw normalne zdarzenia, które realnie mogą być pokazane w przebiegu
        if ev_type in ("goal", "penaltyKickScored", "penaltyKickMissed"):
            filtered.append(ev)
            continue

        # Cofnięcie bramki - usuń wcześniej dodany odpowiadający event "goal"
        if ev_type == "goalRemoved":
            team = ev.get("team")
            player = ev.get("player")
            extra_raw = ev.get("extra")

            orig_time = None
            penalty_flag = False

            if isinstance(extra_raw, str) and extra_raw.strip():
                try:
                    extra_obj = json.loads(extra_raw)
                    if isinstance(extra_obj, dict):
                        orig_time = extra_obj.get("origTime")
                        penalty_flag = bool(extra_obj.get("penalty", False))
                except Exception:
                    pass
            elif isinstance(extra_raw, dict):
                orig_time = extra_raw.get("origTime")
                penalty_flag = bool(extra_raw.get("penalty", False))

            # Nas interesuje tylko cofnięcie zwykłej bramki, nie ewentualnych karnych
            if penalty_flag:
                continue

            # Szukamy od końca ostatniego pasującego eventu "goal"
            for i in range(len(filtered) - 1, -1, -1):
                prev = filtered[i]
                if not isinstance(prev, dict):
                    continue
                if prev.get("type") != "goal":
                    continue
                if prev.get("team") != team:
                    continue
                if prev.get("player") != player:
                    continue

                prev_time = prev.get("time")

                # Preferuj dopasowanie po origTime, a jeśli go nie ma - po czasie goalRemoved
                if orig_time is not None:
                    if prev_time == orig_time:
                        filtered.pop(i)
                        break
                else:
                    if prev_time == ev.get("time"):
                        filtered.pop(i)
                        break

    return filtered


def _extract_timeline_events(data_json: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    Zwraca (evs1, evs2) dla przebiegu meczu po odfiltrowaniu cofniętych bramek.
    Uwzględnia typy:
      - goal
      - penaltyKickScored
      - penaltyKickMissed
    oraz usuwa z przebiegu te `goal`, które zostały później cofnięte przez `goalRemoved`.
    """
    prot = data_json.get("protocol") or []
    filtered_protocol = _filter_protocol_events_for_timeline(prot)

    evs1: List[Dict[str, Any]] = []
    evs2: List[Dict[str, Any]] = []

    for ev in filtered_protocol:
        if not isinstance(ev, dict):
            continue

        t = ev.get("type")
        if t not in ("goal", "penaltyKickScored", "penaltyKickMissed"):
            continue

        half = ev.get("half")
        if half == 1:
            evs1.append(ev)
        elif half == 2:
            evs2.append(ev)

    def _ev_ms(e: Dict[str, Any]) -> int:
        return _safe_int(e.get("time") or 0, 0)

    evs1.sort(key=_ev_ms)
    evs2.sort(key=_ev_ms)
    return evs1, evs2


def _advance_scores_for_events(
    evs: List[Dict[str, Any]],
    start_host: int,
    start_guest: int,
) -> Tuple[int, int]:
    """
    Przelicza score po zdarzeniach (liczymy tylko goal i penaltyKickScored).
    penaltyKickMissed nie zmienia wyniku.
    """
    h = _safe_int(start_host, 0)
    g = _safe_int(start_guest, 0)

    for ev in evs:
        if not isinstance(ev, dict):
            continue
        t = ev.get("type")
        team = ev.get("team")
        if t == "goal" or t == "penaltyKickScored":
            if team == "host":
                h += 1
            elif team == "guest":
                g += 1

    return h, g


def _fill_timeline_half_chunk(
    ws,
    *,
    evs: List[Dict[str, Any]],
    start_row: int,
    end_row: int,
    half_ms: int,
    start_score_host: int,
    start_score_guest: int,
    col_minute: str,
    col_host_action: str,
    col_host_score: str,
    col_guest_score: str,
    col_guest_action: str,
) -> Tuple[int, int]:
    def _ev_ms(e: Dict[str, Any]) -> int:
        return _safe_int(e.get("time") or 0, 0)

    h_score = _safe_int(start_score_host, 0)
    g_score = _safe_int(start_score_guest, 0)

    # bierzemy tylko te wiersze, które faktycznie wolno ruszać
    rows = [r for r in TIMELINE_ROWS if start_row <= r <= end_row]

    def _write_blank(r: int) -> None:
        ws[f"{col_minute}{r}"].value = "--"
        ws[f"{col_host_action}{r}"].value = "--"
        ws[f"{col_host_score}{r}"].value = "--"
        ws[f"{col_guest_score}{r}"].value = "--"
        ws[f"{col_guest_action}{r}"].value = "--"

    # 1) wypełnij eventami tyle ile się da
    i = 0
    for ev in evs:
        if i >= len(rows):
            break
        r = rows[i]

        ms = _ev_ms(ev)
        minute = _event_minute_from_ms(ms)

        team = ev.get("team")
        player = ev.get("player")
        t = ev.get("type")

        _write_blank(r)  # default

        ws[f"{col_minute}{r}"].value = str(minute)

        host_action = ""
        guest_action = ""

        if player is not None:
            ptxt = str(player).strip()
            if isinstance(t, str) and t.startswith("penaltyKick"):
                ptxt = f"{ptxt}K"
            if team == "host":
                host_action = ptxt
            elif team == "guest":
                guest_action = ptxt

        ws[f"{col_host_action}{r}"].value = host_action if host_action else "--"
        ws[f"{col_guest_action}{r}"].value = guest_action if guest_action else "--"

        if t in ("goal", "penaltyKickScored"):
            if team == "host":
                h_score += 1
            elif team == "guest":
                g_score += 1
            ws[f"{col_host_score}{r}"].value = str(h_score)
            ws[f"{col_guest_score}{r}"].value = str(g_score)
        else:
            ws[f"{col_host_score}{r}"].value = "--"
            ws[f"{col_guest_score}{r}"].value = "--"

        i += 1

    # 2) resztę dozwolonych wierszy wyczyść na "--"
    while i < len(rows):
        _write_blank(rows[i])
        i += 1

    return h_score, g_score


def _fill_timeline_pages(
    ws_page1,
    ws_page2,
    *,
    data_json: Dict[str, Any],
    half_ms: int,
    half_score_host: int,
    half_score_guest: int,
) -> bool:
    """
    Wypełnia przebieg na stronie 1 i (opcjonalnie) kontynuację na stronie 2.
    Zwraca True jeśli powstała strona 2 (overflow), inaczej False.
    """
    evs1, evs2 = _extract_timeline_events(data_json)

    overflow1 = len(evs1) > TIMELINE_MAX_ROWS
    overflow2 = len(evs2) > TIMELINE_MAX_ROWS
    needs_page2 = overflow1 or overflow2

    # podział na chunk'i
    evs1_p1 = evs1[:TIMELINE_MAX_ROWS]
    evs1_p2 = evs1[TIMELINE_MAX_ROWS:]
    evs2_p1 = evs2[:TIMELINE_MAX_ROWS]
    evs2_p2 = evs2[TIMELINE_MAX_ROWS:]

    # --- PAGE 1 ---
    # 1 połowa startuje od 0:0
    _fill_timeline_half_chunk(
        ws_page1,
        evs=evs1_p1,
        start_row=TIMELINE_START_ROW,
        end_row=TIMELINE_END_ROW,
        half_ms=half_ms,
        start_score_host=0,
        start_score_guest=0,
        col_minute="AL",
        col_host_action="AN",
        col_host_score="AP",
        col_guest_score="AS",
        col_guest_action="AU",
    )

    # 2 połowa startuje od halfScore
    _fill_timeline_half_chunk(
        ws_page1,
        evs=evs2_p1,
        start_row=TIMELINE_START_ROW,
        end_row=TIMELINE_END_ROW,
        half_ms=half_ms,
        start_score_host=_safe_int(half_score_host, 0),
        start_score_guest=_safe_int(half_score_guest, 0),
        col_minute="AW",
        col_host_action="AY",
        col_host_score="BA",
        col_guest_score="BD",
        col_guest_action="BF",
    )

    if not needs_page2:
        return False

        # --- PAGE 2 ---
    if ws_page2 is None:
        # teoretycznie nie powinno się zdarzyć, ale bezpiecznie
        return False

    # jeśli nie ma kontynuacji danej połowy -> i tak wypełniamy "--"
    # 1 połowa (kontynuacja): start score musi być po evs1_p1
    h1_after, g1_after = _advance_scores_for_events(evs1_p1, 0, 0)
    _fill_timeline_half_chunk(
        ws_page2,
        evs=evs1_p2,
        start_row=TIMELINE_START_ROW,
        end_row=TIMELINE_END_ROW,
        half_ms=half_ms,
        start_score_host=h1_after,
        start_score_guest=g1_after,
        col_minute="AL",
        col_host_action="AN",
        col_host_score="AP",
        col_guest_score="AS",
        col_guest_action="AU",
    )

    # 2 połowa (kontynuacja): start score musi być po evs2_p1 (od halfScore)
    h2_start = _safe_int(half_score_host, 0)
    g2_start = _safe_int(half_score_guest, 0)
    h2_after, g2_after = _advance_scores_for_events(evs2_p1, h2_start, g2_start)

    _fill_timeline_half_chunk(
        ws_page2,
        evs=evs2_p2,
        start_row=TIMELINE_START_ROW,
        end_row=TIMELINE_END_ROW,
        half_ms=half_ms,
        start_score_host=h2_after,
        start_score_guest=g2_after,
        col_minute="AW",
        col_host_action="AY",
        col_host_score="BA",
        col_guest_score="BD",
        col_guest_action="BF",
    )

    return True


def _convert_xlsx_to_pdf(xlsx_path: str, out_dir: str) -> str:
    """
    Konwersja przez LibreOffice:
      soffice --headless --convert-to pdf --outdir <out_dir> <xlsx_path>
    Zwraca ścieżkę do PDF.
    """
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        raise RuntimeError("Brak LibreOffice (soffice) w środowisku. Doinstaluj libreoffice w Dockerfile.")

    # LO w kontenerze często próbuje pisać do HOME — zabezpieczamy to envem
    env = os.environ.copy()
    env.setdefault("HOME", "/tmp")
    env.setdefault("XDG_CACHE_HOME", "/tmp")
    env.setdefault("XDG_CONFIG_HOME", "/tmp")

    profile_dir = os.path.join(out_dir, "lo_profile")
    os.makedirs(profile_dir, exist_ok=True)

    cmd = [
        soffice,
        "--headless",
        "--nologo",
        "--nolockcheck",
        "--nodefault",
        "--norestore",
        f"-env:UserInstallation=file://{profile_dir}",
        "--convert-to",
        "pdf",
        "--outdir",
        out_dir,
        xlsx_path,
    ]


    proc = subprocess.run(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        env=env,
        timeout=90,  # ważne: żeby request nie wisiał wiecznie
    )

    if proc.returncode != 0:
        raise RuntimeError(
            f"LibreOffice convert failed (code={proc.returncode}). "
            f"stderr={proc.stderr[:500]} stdout={proc.stdout[:200]}"
        )

    base = os.path.splitext(os.path.basename(xlsx_path))[0]
    pdf_path = os.path.join(out_dir, base + ".pdf")

    if not os.path.exists(pdf_path):
        # fallback: znajdź najbardziej pasujący PDF w out_dir
        pdfs = [p for p in os.listdir(out_dir) if p.lower().endswith(".pdf")]
        if len(pdfs) == 1:
            pdf_path = os.path.join(out_dir, pdfs[0])
        else:
            # spróbuj dopasować po base
            cand = [p for p in pdfs if os.path.splitext(p)[0] == base]
            if len(cand) == 1:
                pdf_path = os.path.join(out_dir, cand[0])
            else:
                raise RuntimeError(f"Nie znaleziono wyjściowego PDF po konwersji. Pliki: {pdfs[:10]}")

    return pdf_path

def _parse_penalty_score(penalty_score: str) -> Tuple[int, int]:
    """
    "5 - 4" / "5-4" -> (5,4)
    """
    s = (penalty_score or "").strip()
    if not s:
        return (0, 0)
    m = re.search(r"(\d+)\s*-\s*(\d+)", s)
    if not m:
        return (0, 0)
    return int(m.group(1)), int(m.group(2))


def _shootout_needed(data_json: Dict[str, Any]) -> bool:
    """
    Karne tylko gdy:
      - wynik końcowy remis
      - i jest penaltyScore albo penaltyShots
    """
    try:
        sh = int(data_json.get("scoreHost") or 0)
        sg = int(data_json.get("scoreGuest") or 0)
    except Exception:
        sh, sg = 0, 0

    if sh != sg:
        return False

    ps = (data_json.get("penaltyScore") or "").strip()
    shots = data_json.get("penaltyShots") or {}
    has_shots = bool((shots.get("host") or []) or (shots.get("guest") or []))
    return bool(ps) or has_shots


def _fill_shootout_page(ws, *, data_json: Dict[str, Any]) -> None:
    """
    Strona "RZUTY KARNE" w przebiegu (kolumny AL..AU, wiersze 15..61).
    Zaczynamy od wiersza 16.
    """
    # 1) Nagłówek: merge + tekst
    ws.merge_cells("AL15:AV15")
    ws["AL15"].value = "RZUTY KARNE"
    ws["AL15"].alignment = Alignment(horizontal="center", vertical="center")
    ws["AL15"].font = Font(bold=True)

    # 2) Wyczyść/ustaw "--" w obu blokach przebiegu żeby nie było śmieci z kopiowanego arkusza
    # lewy blok (AL..AU)
    for r in range(16, 62):
        ws[f"AL{r}"].value = "--"
        ws[f"AN{r}"].value = "--"
        ws[f"AP{r}"].value = "--"
        ws[f"AS{r}"].value = "--"
        ws[f"AU{r}"].value = "--"

    # prawy blok (AW..BF) – na stronie karnych nie używamy, więc czyścimy
    for r in range(15, 62):
        ws[f"AW{r}"].value = "--"
        ws[f"AY{r}"].value = "--"
        ws[f"BA{r}"].value = "--"
        ws[f"BD{r}"].value = "--"
        ws[f"BF{r}"].value = "--"

    # 3) Dane karnych
    shots = data_json.get("penaltyShots") or {}
    host_arr = shots.get("host") or []
    guest_arr = shots.get("guest") or []

    # ile serii (po 1 strzale na drużynę)
    series_count = max(len(host_arr), len(guest_arr))
    if series_count <= 0:
        return

    # 4) Wpisy od wiersza 16, po 2 wiersze na serię
    row = 16
    host_score = 0
    guest_score = 0

    def _shot(arr, idx) -> Optional[Dict[str, Any]]:
        if idx < 0 or idx >= len(arr):
            return None
        x = arr[idx]
        return x if isinstance(x, dict) else None

    # co 5 serii zmiana startującego:
    # serie 1-5: host first
    # serie 6-10: guest first
    # serie 11-15: host first
    # itd.
    # W tej sekcji w kodzie odpowiedzialnej za ustawianie kolejności drużyn do wykonywania karnych

    # Odczytanie drużyny rozpoczynającej karne z JSON
    penalty_starter_team = data_json.get("penaltyStarterTeam", "guest")

    # Zmienna do kontrolowania, która drużyna jako pierwsza wykonuje rzuty karne
    flip = False  # Flaga do przełączania kolejności

    # Co 5 kolejek zmieniamy drużynę zaczynającą karne
    for s in range(1, series_count + 1):
        if row > 61:
            break  # brak miejsca w szablonie

        idx = s - 1

        # Zmiana kolejności co 5 serii
        if (s - 1) // 5 % 2 == 1:
            flip = not flip

        # Jeśli 'penaltyStarterTeam' to 'guest', to 'guest' zaczyna
        first_team = penalty_starter_team if not flip else ("guest" if penalty_starter_team == "host" else "host")
        second_team = "guest" if first_team == "host" else "host"

        def write_team_shot(team: str, series_no: int):
            nonlocal row, host_score, guest_score
            if row > 61:
                return

            ws[f"AL{row}"].value = str(series_no)

            if team == "host":
                sh = _shot(host_arr, idx)
                player = sh.get("player") if sh else None
                result = int(sh.get("result") or 0) if sh else 0

                ws[f"AN{row}"].value = str(int(player)) if player is not None else "--"
                ws[f"AU{row}"].value = "--"

                if result == 1:
                    host_score += 1
                    ws[f"AP{row}"].value = str(host_score)
                    ws[f"AS{row}"].value = str(guest_score)   # <-- pokaż wynik przeciwnika
                else:
                    ws[f"AP{row}"].value = "--"
                    ws[f"AS{row}"].value = "--"


            else:
                sh = _shot(guest_arr, idx)
                player = sh.get("player") if sh else None
                result = int(sh.get("result") or 0) if sh else 0

                ws[f"AN{row}"].value = "--"
                ws[f"AU{row}"].value = str(int(player)) if player is not None else "--"

                if result == 1:
                    guest_score += 1
                    ws[f"AP{row}"].value = str(host_score)    # <-- pokaż wynik gospodarza
                    ws[f"AS{row}"].value = str(guest_score)
                else:
                    ws[f"AP{row}"].value = "--"
                    ws[f"AS{row}"].value = "--"

            row += 1

        # 1) pierwszy strzał w serii
        write_team_shot(first_team, s)
        # 2) drugi strzał w serii
        write_team_shot(second_team, s)

from datetime import datetime, date, timezone
from io import BytesIO
from typing import Callable

# jeśli masz pillow (zwykle jest), to da nam rozmiar obrazka do skalowania
try:
    from PIL import Image as PILImage
except Exception:
    PILImage = None


BACKEND_STATIC_PREFIX = "https://zprp-backend-production.up.railway.app"


def _full_static_url(rel_or_abs: str) -> str:
    """
    '/static/xxx.png' -> 'https://.../static/xxx.png'
    jeśli już jest absolutny URL -> zwraca bez zmian
    """
    s = (rel_or_abs or "").strip()
    if not s:
        return ""
    if s.startswith("http://") or s.startswith("https://"):
        return s
    if not s.startswith("/"):
        s = "/" + s
    return BACKEND_STATIC_PREFIX + s


def _fmt_date_ddmmyyyy(iso_ymd: str) -> str:
    """
    '2026-02-23' -> '23.02.2026'
    """
    s = (iso_ymd or "").strip()
    if not s:
        return ""
    try:
        d = datetime.strptime(s, "%Y-%m-%d").date()
        return d.strftime("%d.%m.%Y")
    except Exception:
        return ""


def _fmt_time_hhmm(hhmm: str) -> str:
    """
    '18:00' -> '18:00' (waliduje format)
    """
    s = (hhmm or "").strip()
    if not s:
        return ""
    try:
        t = datetime.strptime(s, "%H:%M").time()
        return t.strftime("%H:%M")
    except Exception:
        return ""


def _set_yes_no_x(ws, *, yes_cell: str, no_cell: str, value: Any, yes_when_true: bool = True) -> None:
    """
    Wstawia "X" do pary komórek (tak/nie albo brak/verte).
    Jeśli value truthy -> X do yes_cell (gdy yes_when_true=True), inaczej do no_cell.
    """
    v = bool(value) if value is not None else False
    ws[yes_cell].value = ""
    ws[no_cell].value = ""
    if yes_when_true:
        ws[yes_cell].value = "X" if v else ""
        ws[no_cell].value = "" if v else "X"
    else:
        # odwrotna logika (raczej niepotrzebna tutaj, ale zostawiam)
        ws[yes_cell].value = "" if v else "X"
        ws[no_cell].value = "X" if v else ""


async def _fetch_png_bytes(url: str) -> bytes:
    """
    Pobiera obraz PNG/JPG z URL. Zwraca bytes albo b'' gdy brak/nieprawidłowy.
    """
    u = (url or "").strip()
    if not u:
        return b""
    try:
        async with AsyncClient(follow_redirects=True, timeout=15.0) as c:
            r = await c.get(u)
            if r.status_code != 200:
                return b""
            return r.content or b""
    except Exception:
        return b""


def _add_signature_image(
    ws,
    *,
    image_bytes: bytes,
    anchor_cell: str,
    max_width_px: int = 220,
    max_height_px: int = 90,
    x_offset_px: int = 0,
    y_offset_px: int = 0,
) -> bool:
    """
    Wstawia obraz do arkusza, zakotwiczony w anchor_cell.
    Skaluje do max_width/max_height (px).
    Opcjonalnie przesuwa obraz o x/y w pikselach (przydatne do centrowania).
    """
    if not image_bytes:
        return False

    bio = BytesIO(image_bytes)
    img = Image(bio)

    # Skala (jeśli mamy PIL, weźmiemy faktyczny rozmiar)
    if PILImage is not None:
        try:
            bio2 = BytesIO(image_bytes)
            pil = PILImage.open(bio2)
            w, h = pil.size
            if w and h:
                scale = min(max_width_px / float(w), max_height_px / float(h), 1.0)
                img.width = int(w * scale)
                img.height = int(h * scale)
        except Exception:
            pass
    else:
        img.width = min(img.width or max_width_px, max_width_px)
        img.height = min(img.height or max_height_px, max_height_px)

    ws.add_image(img, anchor_cell)

    # offset (działa w nowszych openpyxl; jak nie zadziała to nic nie popsuje)
    try:
        anchor = getattr(img, "anchor", None)
        if anchor is not None and hasattr(anchor, "_from"):
            if x_offset_px:
                anchor._from.colOff = int(x_offset_px) * 9525  # px -> EMU
            if y_offset_px:
                anchor._from.rowOff = int(y_offset_px) * 9525
    except Exception:
        pass

    return True

def _extract_city_from_venue_address(venue_address: str) -> str:
    """
    Heurystyka: venueAddress zwykle ma format:
      "Hala Relax Piotrków Trybunalski, Stefana Batorego 8"
    Chcemy wyciągnąć miejscowość do nagłówka: "Piotrków Trybunalski".

    Reguła:
      - bierzemy część przed pierwszym przecinkiem
      - z niej bierzemy ostatnie 2 wyrazy (fallback: 1 wyraz)
    """
    s = (venue_address or "").strip()
    if not s:
        return ""
    left = s.split(",", 1)[0].strip()
    parts = [p for p in left.split() if p.strip()]
    if not parts:
        return ""
    if len(parts) >= 2:
        return " ".join(parts[-2:])
    return parts[-1]


def _create_detailed_notes_sheet(
    wb,
    *,
    page_no: int,
    total_pages: int,
    date_ddmmyyyy: str,
    place: str,
    notes_text: str,
    referee1_name: str,
    referee2_name: str,
    referee1_sig_bytes: bytes,
    referee2_sig_bytes: bytes,
):
    ws_notes = wb.create_sheet(title="Uwagi sędziów")

    # --- Page setup (A4) ---
    try:
        ws_notes.page_setup.paperSize = ws_notes.PAPERSIZE_A4
        ws_notes.page_setup.orientation = ws_notes.ORIENTATION_PORTRAIT
    except Exception:
        pass

    # --- FORCE: jedna strona na szerokość (eliminuje “strona 3” w prawo) ---
    try:
        ws_notes.sheet_properties.pageSetUpPr.fitToPage = True
        ws_notes.page_setup.fitToWidth = 1
        ws_notes.page_setup.fitToHeight = 1  # jeśli wolisz, możesz dać 0 (automatycznie)
        ws_notes.page_setup.scale = None
    except Exception:
        pass

    # print area (żeby LO nie brał “czegoś dalej”)
    try:
        ws_notes.print_area = "A1:N45"
    except Exception:
        pass

    # trochę większe marginesy boczne = “nie ucieka” w prawo
    try:
        ws_notes.page_margins.left = 0.7
        ws_notes.page_margins.right = 0.7
        ws_notes.page_margins.top = 0.6
        ws_notes.page_margins.bottom = 0.6
    except Exception:
        pass

    # Siatka “kartki”
    for col in ["A","B","C","D","E","F","G","H","I","J","K","L","M","N"]:
        try:
            ws_notes.column_dimensions[col].width = 6.2
        except Exception:
            pass
    try:
        ws_notes.column_dimensions["A"].width = 4.5
        ws_notes.column_dimensions["N"].width = 4.5
    except Exception:
        pass

    # --- Nagłówek ---
    ws_notes.merge_cells("A1:F1")
    ws_notes["A1"].value = f"Strona {page_no}/{total_pages}"
    ws_notes["A1"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws_notes["A1"].font = Font(size=10)

    # --- Prawy nagłówek (KROK 3) ---
    ws_notes.merge_cells("G1:M1")

    right_hdr = ""
    if date_ddmmyyyy and place:
        right_hdr = f"{date_ddmmyyyy}, {place}"
    elif date_ddmmyyyy:
        right_hdr = date_ddmmyyyy
    elif place:
        right_hdr = place

    ws_notes["G1"].value = right_hdr
    ws_notes["G1"].alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    ws_notes["G1"].font = Font(size=10)

    # --- Treść opisu ---
    # Zostawiamy opis do 30 wierszy, a podpisy zaczynamy zaraz pod nim.
    ws_notes.merge_cells("A3:N30")
    ws_notes["A3"].value = (notes_text or "").strip()
    ws_notes["A3"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws_notes["A3"].font = Font(size=12)

    for r in range(3, 31):
        try:
            ws_notes.row_dimensions[r].height = 18
        except Exception:
            pass

    # --- Podpisy (bliżej opisu, bardziej wyśrodkowane) ---
    # Blok podpisów: I..N (zamiast J..N) => całość “wchodzi” bliżej środka
    # Sędzia 1: nazwa I32..N32, podpis K33 (centrowanie przez kotwicę w K)
    ws_notes.merge_cells("I32:N32")
    ws_notes["I32"].value = (referee1_name or "").strip()
    ws_notes["I32"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_notes["I32"].font = Font(size=11, bold=True)

    _add_signature_image(
        ws_notes,
        image_bytes=referee1_sig_bytes or b"",
        anchor_cell="K33",
        max_width_px=180,
        max_height_px=65,
        x_offset_px=0,
        y_offset_px=0,
    )

    # Sędzia 2: nazwa I36..N36, podpis K37
    ws_notes.merge_cells("I36:N36")
    ws_notes["I36"].value = (referee2_name or "").strip()
    ws_notes["I36"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_notes["I36"].font = Font(size=11, bold=True)

    _add_signature_image(
        ws_notes,
        image_bytes=referee2_sig_bytes or b"",
        anchor_cell="K37",
        max_width_px=180,
        max_height_px=65,
        x_offset_px=0,
        y_offset_px=0,
    )

    # wysokości wierszy w okolicy podpisów (żeby obraz “miał miejsce”)
    for r in range(31, 39):
        try:
            ws_notes.row_dimensions[r].height = 18
        except Exception:
            pass
    try:
        ws_notes.row_dimensions[33].height = 55
        ws_notes.row_dimensions[37].height = 55
    except Exception:
        pass

    return ws_notes

def _safe_filename_from_match_number(match_number: str) -> str:
    base = (match_number or "mecz").strip().replace("/", "-")
    base = re.sub(r"[^0-9A-Za-z._-]+", "_", base)
    return f"protokol_{base}.pdf"

DOWNLOAD_DIR = "/tmp/protocol_downloads"
DOWNLOAD_TTL_SECONDS = 10 * 60  # 10 min

def _ensure_download_dir():
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)

def _cleanup_expired_downloads():
    try:
        _ensure_download_dir()
        now = time.time()
        for fn in os.listdir(DOWNLOAD_DIR):
            p = os.path.join(DOWNLOAD_DIR, fn)
            try:
                st = os.stat(p)
                if now - st.st_mtime > DOWNLOAD_TTL_SECONDS:
                    os.remove(p)
            except Exception:
                pass
    except Exception:
        pass


OFFICIAL_NAME_FALLBACK = "--------------------------"
OFFICIAL_CITY_FALLBACK = "     -------------     "
OFFICIAL_SIGN_FALLBACK = "-------"

def _fallback_text(v: Any, fallback: str) -> str:
    s = (v or "").strip() if isinstance(v, str) else str(v).strip() if v is not None else ""
    return s if s else fallback

def _set_cell_fallback(ws, cell: str, v: Any, fallback: str) -> None:
    ws[cell].value = _fallback_text(v, fallback)

def _is_blank_str(v: Any) -> bool:
    return (v is None) or (isinstance(v, str) and v.strip() == "")

def _is_empty_companion_block(
    letter: str,
    *,
    names_map: Dict[str, str],
    meta_map: Dict[str, Dict[str, str]],
    pen_map: Dict[str, Dict[str, str]],
) -> bool:
    """
    True gdy dana osoba towarzysząca (B-E) jest „pusta”:
    - brak fullName
    - brak function
    - brak license
    - oraz brak jakichkolwiek kar (warn/p2/disq)
    """
    l = (letter or "").strip().upper()
    name = (names_map.get(l) or "").strip()
    func = ((meta_map.get(l) or {}).get("function") or "").strip()
    lic  = ((meta_map.get(l) or {}).get("license") or "").strip()

    pen = pen_map.get(l) or {}
    warn = (pen.get("warn") or "").strip()
    p2   = (pen.get("p2") or "").strip()
    disq = (pen.get("disq") or "").strip()

    # Uwaga: u Ciebie pen_map daje "---" tylko gdy klucza brak,
    # ale gdy osoba jest pusta zazwyczaj będzie "" lub "---" — traktujemy oba jako puste.
    def _pen_empty(s: str) -> bool:
        return s == "" or s == "---"

    return (
        name == ""
        and func == ""
        and lic == ""
        and _pen_empty(warn)
        and _pen_empty(p2)
        and _pen_empty(disq)
    )

def _merge_and_diagonal_cross(ws, *, start_cell: str, end_cell: str) -> None:
    """
    Scala zakres i robi diagonalne przekreślenie całej scalonej komórki.
    Ustawia też alignment: left + top (z wrap).
    """
    rng = f"{start_cell}:{end_cell}"
    ws.merge_cells(rng)

    tl = ws[start_cell]  # top-left cell zakresu scalonego

    # wyczyść wartość (żeby nie zostały śmieci po wcześniejszych wpisach)
    tl.value = ""

    # wyrównanie: lewo + góra (jak chciałeś)
    tl.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # diagonalne przekreślenie (jedna przekątna przez cały merge)
    # Jeśli chcesz „X”, daj diagonalUp + diagonalDown; Ty pisałeś „diagonalne przekreślenie” -> zazwyczaj 1 linia.
    thin = Side(style="thin", color="000000")

    # zachowaj istniejące obramowania (jeśli jakieś były) i tylko dodaj diagonal
    cur = tl.border or Border()
    tl.border = Border(
        left=cur.left,
        right=cur.right,
        top=cur.top,
        bottom=cur.bottom,
        diagonal=thin,
        diagonalDown=True,   # linia z lewego-górnego do prawego-dolnego
        diagonalUp=False,
        outline=cur.outline,
        vertical=cur.vertical,
        horizontal=cur.horizontal,
    )

def _apply_companion_crossouts(ws, *, host_names, host_meta, host_pen, guest_names, guest_meta, guest_pen) -> None:
    """
    Jeśli B-E są puste dla danej drużyny => scal + diagonal.
    """
    HOST_RANGES = {
        "B": ("J29", "P32"),
        "C": ("Q29", "W32"),
        "D": ("X29", "AD32"),
        "E": ("AE29", "AK32"),
    }
    GUEST_RANGES = {
        "B": ("J55", "P58"),
        "C": ("Q55", "W58"),
        "D": ("X55", "AD58"),
        "E": ("AE55", "AK58"),
    }

    for ltr, (a, b) in HOST_RANGES.items():
        if _is_empty_companion_block(ltr, names_map=host_names, meta_map=host_meta, pen_map=host_pen):
            _merge_and_diagonal_cross(ws, start_cell=a, end_cell=b)

    for ltr, (a, b) in GUEST_RANGES.items():
        if _is_empty_companion_block(ltr, names_map=guest_names, meta_map=guest_meta, pen_map=guest_pen):
            _merge_and_diagonal_cross(ws, start_cell=a, end_cell=b)

# ============================================================================
# Nagłówek i stopka wydruku (IdZawody + kto/kiedy wygenerował)
# ============================================================================
#
# Oba napisy idą przez NAGŁÓWEK i STOPKĘ STRONY, a nie przez komórki. Powód jest
# twardy: arkusz zajmuje 71 wierszy, które przy skali 90 % wypełniają wysokość
# A4 co do milimetra. Dołożenie choćby jednego wiersza wypycha ostatnią linię
# protokołu na drugą stronę, a wstawienie wiersza NAD ramką przesunęłoby o jeden
# wszystkie ~170 literalnych adresów w tym pliku.
#
# Marginesy przestawiamy tak, żeby wysokość obszaru druku została DOKŁADNIE
# taka sama. Szablon ma górny margines 8 mm i zerowy dolny; dzielimy te same
# 8 mm na 5,2 mm u góry i 2,8 mm u dołu. Suma się zgadza co do zera, więc blok
# protokołu tylko przesuwa się o 2,8 mm w górę — i nic poza tym.
#
# Liczymy w milimetrach, a nie w zaokrąglonych calach, właśnie po to, żeby ta
# suma była równa dokładnie, a nie „prawie".
_MM_IN = 1.0 / 25.4
PROTOCOL_TOP_MARGIN_IN = 5.2 * _MM_IN
PROTOCOL_HEADER_MARGIN_IN = 0.5 * _MM_IN
PROTOCOL_BOTTOM_MARGIN_IN = 2.8 * _MM_IN
PROTOCOL_FOOTER_MARGIN_IN = 0.0

# ============================================================================
# Pola, które muszą zostać w jednym wierszu
# ============================================================================
#
# Szablon jest pod tym względem niespójny: „osoba towarzysząca A" ma
# `wrapText`, B i E mają `shrinkToFit`, C i D nie mają nic. Efekt widać na
# wydruku — długie nazwisko w rubryce A łamie się na dwa wiersze, choć obok, w
# rubryce B, to samo nazwisko mieści się w jednej linii.
#
# `shrinkToFit` zmniejsza czcionkę TYLKO wtedy, gdy treść się nie mieści.
# Krótkie miasto („Płock") zachowuje rozmiar z szablonu, a „Piotrków
# Trybunalski" zjeżdża tyle, ile trzeba. To ta sama metoda, co
# `_set_fitted_text` w protokole BAZA Beach.
#
# Dlaczego nie „niech tekst wyjedzie na sąsiednie komórki": wszystkie te pola
# to komórki SCALONE, a scalona komórka nie przelewa treści poza swój obszar —
# ani w Excelu, ani w LibreOffice. Wyjazd w bok wymagałby rozsunięcia scaleń,
# czyli przebudowy rubryk pod podpisy.
#
# Zmierzone przypadki, dla których to jest konieczne (szerokość rubryki kontra
# potrzebna szerokość tekstu, przy skali wydruku 90 %):
#   • miejscowość sędziego     17,1 mm ← „Piotrków Trybunalski"  24,7 mm
#   • nazwisko sędziego        28,6 mm ← podwójne nazwisko       40,9 mm
#   • osoba towarzysząca A     12,1 mm ← „MALINOWSKI Wojciech"   16,8 mm
#   • osoba towarzysząca D/E   12,9 mm ← to samo                 13,4 mm
#   • funkcja D/E              15,0 mm ← „OSOBA TOWARZYSZĄCA"    15,9 mm
FITTED_TEXT_CELLS: Tuple[str, ...] = (
    # sędziowie i oficjele — nazwisko i miejscowość
    "I66", "I67", "I68", "I69", "I70",
    "W66", "W67", "W68", "W69", "W70",
    # osoby towarzyszące — nazwisko (gospodarze / goście)
    "B29", "K29", "R29", "Y29", "AF29",
    "B55", "K55", "R55", "Y55", "AF55",
    # osoby towarzyszące — funkcja
    "A30", "J30", "Q30", "X30", "AE30",
    "A56", "J56", "Q56", "X56", "AE56",
    # osoby towarzyszące — licencja
    "A31", "J31", "Q31", "X31", "AE31",
    "A57", "J57", "Q57", "X57", "AE57",
)

# Adres hali (`AL4`) świadomie NIE jest na tej liście: jego rubryka ma 59,5 mm
# przy tekście 59,3 mm, czyli jest zaprojektowana pod DWA wiersze. Wyłączenie
# tam zawijania skróciłoby adres zamiast go pokazać.


def _apply_fitted_text(ws_raw) -> None:
    """Jeden wiersz zamiast łamania — bez ruszania scaleń i rozmiarów rubryk."""
    for logical in FITTED_TEXT_CELLS:
        try:
            cell = ws_raw[shift_ref(logical)]
            alignment = copy.copy(cell.alignment)
            alignment.shrinkToFit = True
            alignment.wrapText = False
            cell.alignment = alignment
        except Exception:
            logger.warning("Nie udało się ustawić dopasowania w %s", logical, exc_info=True)


_ZPRP_ID_IN_LINK = re.compile(r"[?&](?:IdZawody|Zawody|Mecz|match_id)=([0-9]+)", re.I)


def _zprp_match_id(data_json: Dict[str, Any]) -> str:
    """
    IdZawody — identyfikator meczu w systemie ZPRP, ten sam, który drukuje się
    w prawym górnym rogu oryginalnego protokołu.

    Świadomie NIE bierzemy `data_json["id"]`: to znacznik czasu nadawany przez
    aplikację przy budowaniu paczki, a nie identyfikator zawodów. Wpisanie go w
    nagłówek dałoby liczbę, która wygląda wiarygodnie i nie znaczy nic.
    """
    mc = data_json.get("matchConfig") or {}
    extras = mc.get("extras") or {} if isinstance(mc, dict) else {}
    for src in (mc, extras, data_json):
        if not isinstance(src, dict):
            continue
        for key in ("matchId", "match_id", "IdZawody", "idZawody", "zawodyId", "Id"):
            v = str(src.get(key) or "").strip()
            if v.isdigit():
                return v
        for key in ("details_path", "detailsPath", "matchLink", "href", "url"):
            m = _ZPRP_ID_IN_LINK.search(str(src.get(key) or ""))
            if m:
                return m.group(1)
    return ""


def _warsaw_now() -> datetime:
    try:
        from zoneinfo import ZoneInfo

        return datetime.now(timezone.utc).astimezone(ZoneInfo("Europe/Warsaw"))
    except Exception:
        # Bez bazy stref (okrojony obraz) lepiej wydrukować czas serwera niż nic.
        return datetime.now()


def _hf_escape(text: str) -> str:
    """W nagłówku/stopce arkusza `&` otwiera kod formatujący — trzeba go podwoić."""
    return str(text or "").replace("&", "&&")


def _apply_protocol_page_marks(
    wb,
    *,
    match_id: str,
    generated_by: str,
    generated_at: str,
) -> None:
    """
    Nakłada na WSZYSTKIE arkusze skoroszytu nagłówek z IdZawody i stopkę z
    podpisem generowania. Wołane po `copy_worksheet`, żeby strona uwag i strona
    rzutów karnych też były opisane — inaczej wyrwana kartka byłaby anonimowa.
    """
    footer = "Wygenerowano automatycznie przez użytkownika %s dnia %s" % (
        generated_by,
        generated_at,
    )
    header = "IdZawody: %s" % match_id if match_id else ""

    for ws in wb.worksheets:
        # Dopasowanie treści do rubryk — tu, razem z nagłówkiem, bo obie rzeczy
        # muszą objąć także arkusze powstałe przez `copy_worksheet`.
        _apply_fitted_text(ws)

        pm = ws.page_margins
        pm.top = PROTOCOL_TOP_MARGIN_IN
        pm.header = PROTOCOL_HEADER_MARGIN_IN
        pm.bottom = PROTOCOL_BOTTOM_MARGIN_IN
        pm.footer = PROTOCOL_FOOTER_MARGIN_IN

        if header:
            ws.oddHeader.right.text = _hf_escape(header)
            ws.oddHeader.right.size = 6
            ws.oddHeader.right.font = "Arial,Regular"
            ws.oddHeader.right.color = "808080"

        ws.oddFooter.center.text = _hf_escape(footer)
        ws.oddFooter.center.size = 6
        ws.oddFooter.center.font = "Arial,Regular"
        ws.oddFooter.center.color = "808080"


# ============================================================================
# Dziennik generowania protokołów + podpis pliku
# ============================================================================
#
# Dwie warstwy, bo żadna sama nie wystarcza:
#
#  1. DZIENNIK NA SERWERZE (`protocol_audit`) — jedyne miejsce, które odpowiada
#     na pytanie „jak wyglądał oryginał, zanim ktoś przy nim grzebał". Trzyma
#     pełny stan meczu, więc protokół da się odtworzyć co do treści.
#  2. PODPIS RSA W PLIKU — pozwala potwierdzić autora, czas i stan komuś, kto ma
#     sam PDF i klucz publiczny, bez dostępu do bazy.
#
# Znacznik w pliku NIE jest zabezpieczeniem sam w sobie: metadane kasuje się
# jednym poleceniem, a „drukuj do PDF" gubi je same z siebie. Jest wskaźnikiem
# do dziennika i dowodem integralności — dlatego dziennik zapisuje też
# `pdf_sha256` i numer meczu, czyli drogi dojścia do wpisu, gdy znacznika już
# w pliku nie ma.

# Alfabet bez I, L, O, U — kod bywa przepisywany z wydruku ręcznie.
_AUDIT_ALPHABET = "0123456789ABCDEFGHJKMNPQRSTVWXYZ"
_XMP_NS = "https://baza.zprp.pl/ns/protocol/1.0/"


def _b64u(raw: bytes) -> str:
    return base64.urlsafe_b64encode(raw).decode("ascii").rstrip("=")


def _b64u_dec(s: str) -> bytes:
    return base64.urlsafe_b64decode(s + "=" * (-len(s) % 4))


def _protocol_audit_code() -> str:
    """BZ-XXXX-XXXX. 32^8 ≈ 1,1e12 kombinacji; kolizję i tak łapie klucz główny."""
    raw = "".join(secrets.choice(_AUDIT_ALPHABET) for _ in range(8))
    return "BZ-%s-%s" % (raw[:4], raw[4:])


def _sha256_bytes(raw: bytes) -> str:
    return hashlib.sha256(raw).hexdigest()


def _sha256_file(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 256), b""):
            h.update(chunk)
    return h.hexdigest()


def _canonical_json(obj: Any) -> bytes:
    """Bajty do podpisu. Sortowane klucze i brak spacji — inaczej ten sam stan
    dałby dwa różne skróty w zależności od kolejności pól w JSON-ie."""
    return json.dumps(
        obj, separators=(",", ":"), sort_keys=True, ensure_ascii=False
    ).encode("utf-8")


def _sign_protocol_claim(claim: Dict[str, Any]) -> str:
    """
    Podpis RSA-PSS SHA-256 kluczem serwera (`RSA_PRIVATE_KEY`, ten sam, którym
    szyfrujemy IdZawody przy wysyłce protokołu). Token ma postać
    `payload_b64.sig_b64`, więc czytelny jest bez klucza — a wiarygodny dopiero
    z kluczem publicznym.
    """
    from cryptography.hazmat.primitives import hashes
    from cryptography.hazmat.primitives.asymmetric import padding

    payload = _canonical_json(claim)
    private_key, _ = get_rsa_keys()
    sig = private_key.sign(
        payload,
        padding.PSS(mgf=padding.MGF1(hashes.SHA256()), salt_length=padding.PSS.MAX_LENGTH),
        hashes.SHA256(),
    )
    return "%s.%s" % (_b64u(payload), _b64u(sig))


def _verify_protocol_signature(token: str) -> Tuple[bool, Dict[str, Any]]:
    """(czy podpis jest nasz, odczytany claim). Claim zwracamy także dla podpisu
    fałszywego — po to, żeby admin widział, co ktoś próbował podstawić."""
    from cryptography.exceptions import InvalidSignature
    from cryptography.hazmat.primitives import hashes
    from cryptography.hazmat.primitives.asymmetric import padding

    claim: Dict[str, Any] = {}
    try:
        payload_b64, sig_b64 = str(token or "").split(".", 1)
        payload = _b64u_dec(payload_b64)
        claim = json.loads(payload.decode("utf-8"))
    except Exception:
        return False, {}

    try:
        _, public_key = get_rsa_keys()
        public_key.verify(
            _b64u_dec(sig_b64),
            payload,
            padding.PSS(mgf=padding.MGF1(hashes.SHA256()), salt_length=padding.PSS.MAX_LENGTH),
            hashes.SHA256(),
        )
        return True, claim
    except (InvalidSignature, Exception):  # noqa: B014 — celowo szeroko
        return False, claim


def _xml_escape(v: Any) -> str:
    return (
        str(v or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _protocol_xmp(fields: Dict[str, Any]) -> str:
    body = "".join(
        "<baza:%s>%s</baza:%s>" % (k, _xml_escape(v), k)
        for k, v in fields.items()
        if str(v or "").strip()
    )
    return (
        '<?xpacket begin="﻿" id="W5M0MpCehiHzreSzNTczkc9d"?>'
        '<x:xmpmeta xmlns:x="adobe:ns:meta/">'
        '<rdf:RDF xmlns:rdf="http://www.w3.org/1999/02/22-rdf-syntax-ns#">'
        '<rdf:Description rdf:about="" xmlns:baza="%s">%s</rdf:Description>'
        "</rdf:RDF></x:xmpmeta>"
        '<?xpacket end="w"?>' % (_XMP_NS, body)
    )


def _stamp_pdf_metadata(
    pdf_path: str,
    *,
    code: str,
    signature: str,
    match_id: str,
    match_number: str,
    state_sha256: str,
    generated_by: str,
    generated_at_iso: str,
) -> bool:
    """
    Wpisuje znacznik w metadane pliku — niewidocznie na stronie.

    Dwa miejsca naraz, bo przeżywają różne rzeczy: standardowe pola /Info
    (`subject`, `keywords`) zachowuje większość narzędzi, XMP przeżywa część
    konwersji, które /Info gubią. Awaria tego kroku NIE może przewrócić
    generowania — sędzia ma dostać protokół, a ślad i tak został w dzienniku.
    """
    try:
        import fitz  # pymupdf
    except Exception:
        logger.warning("Brak pymupdf — PDF bez znacznika audytu (kod %s)", code)
        return False

    try:
        doc = fitz.open(pdf_path)
        try:
            doc.set_metadata(
                {
                    "title": "Protokół zawodów %s" % (match_number or ""),
                    "author": generated_by,
                    "subject": "IdZawody %s | %s" % (match_id or "-", code),
                    "keywords": "BazaCode=%s; BazaState=%s; BazaSig=%s"
                    % (code, state_sha256, signature),
                    "creator": "BAZA",
                    "producer": "BAZA protocol generator",
                }
            )
            doc.set_xml_metadata(
                _protocol_xmp(
                    {
                        "code": code,
                        "matchId": match_id,
                        "matchNumber": match_number,
                        "generatedBy": generated_by,
                        "generatedAt": generated_at_iso,
                        "stateSha256": state_sha256,
                        "signature": signature,
                    }
                )
            )
            # Zapis obok i podmiana, a nie `saveIncr()`: zapis przyrostowy
            # potrafi odmówić (linearyzacja, dziwny xref), a wtedy plik zostaje
            # bez znacznika mimo braku wyjątku.
            stamped = pdf_path + ".stamped"
            doc.save(stamped)
        finally:
            doc.close()
        os.replace(stamped, pdf_path)
        return True
    except Exception:
        logger.warning("Nie udało się opisać PDF-a (kod %s)", code, exc_info=True)
        return False


def _extract_pdf_text(data: bytes) -> str:
    """
    Tekst wszystkich stron PDF-a, wiersz po wierszu.

    Zapisujemy go przy generowaniu, żeby przy weryfikacji dało się powiedzieć
    nie tylko „plik został zmieniony", ale też CO. Sam skrót pliku wykrywa
    każdą edycję, lecz o zmienionej wartości nie mówi nic — a właśnie ona jest
    pytaniem admina.
    """
    try:
        import fitz  # pymupdf
    except Exception:
        return ""
    try:
        doc = fitz.open(stream=data, filetype="pdf")
    except Exception:
        return ""
    try:
        return "\n".join(page.get_text() or "" for page in doc)
    except Exception:
        return ""
    finally:
        try:
            doc.close()
        except Exception:
            pass


def _normalize_pdf_lines(text: str) -> List[str]:
    """Wiersze bez pustych i bez różnic w białych znakach — inaczej diff
    pokazywałby przesunięcia układu zamiast zmian treści."""
    out: List[str] = []
    for raw in str(text or "").splitlines():
        line = " ".join(raw.split())
        if line:
            out.append(line)
    return out


#: Powyżej tylu różnic przestajemy wyliczać — plik przepuszczony przez inny
#: program potrafi mieć przebudowany cały strumień tekstu i lista zmian robi
#: się bezużyteczna. Lepiej powiedzieć „przebudowany", niż zasypać admina.
PDF_DIFF_LIMIT = 40


def _diff_pdf_text(before: str, after: str) -> Dict[str, Any]:
    """
    Co się zmieniło między wydrukiem z chwili generowania a wgranym plikiem.

    Zmiany typu „podmiana wartości" (852 → 853) raportujemy parami, bo tak
    wygląda realna fałszywka. Dopisane i usunięte wiersze idą osobno.
    """
    import difflib

    old = _normalize_pdf_lines(before)
    new = _normalize_pdf_lines(after)

    changes: List[Dict[str, str]] = []
    truncated = False
    for tag, i1, i2, j1, j2 in difflib.SequenceMatcher(
        None, old, new, autojunk=False
    ).get_opcodes():
        if tag == "equal":
            continue
        if len(changes) >= PDF_DIFF_LIMIT:
            truncated = True
            break
        if tag == "replace":
            for k in range(max(i2 - i1, j2 - j1)):
                changes.append(
                    {
                        "kind": "changed",
                        "before": old[i1 + k] if i1 + k < i2 else "",
                        "after": new[j1 + k] if j1 + k < j2 else "",
                    }
                )
        elif tag == "delete":
            changes.extend(
                {"kind": "removed", "before": old[k], "after": ""} for k in range(i1, i2)
            )
        elif tag == "insert":
            changes.extend(
                {"kind": "added", "before": "", "after": new[k]} for k in range(j1, j2)
            )

    return {
        "changes": changes[:PDF_DIFF_LIMIT],
        "truncated": truncated or len(changes) > PDF_DIFF_LIMIT,
    }


async def _soft_verify_actor(judge_id: str, installation_id: str) -> bool:
    """
    Czy para (urządzenie, sędzia) zgadza się z `push_tokens`.

    W odróżnieniu od `proel_actor` NIE rzuca 401 przy niezgodności — tutaj to
    tylko etykieta w dzienniku. Sędzia, który odmówił zgody na powiadomienia,
    nigdy nie trafia do `push_tokens`, więc `False` znaczy „nie potwierdzono",
    a nie „na pewno oszust".
    """
    if not judge_id or not installation_id:
        return False
    try:
        from sqlalchemy import select as _select

        from app.db import database, push_tokens

        row = await database.fetch_one(
            _select(push_tokens.c.judge_id).where(
                push_tokens.c.installation_id == installation_id
            )
        )
        if row is None:
            return False
        return str(row["judge_id"] or "").strip() == judge_id
    except Exception:
        logger.warning("protocol_audit: weryfikacja urządzenia nieudana", exc_info=True)
        return False


async def _record_protocol_audit(row: Dict[str, Any]) -> None:
    """
    Wpis do dziennika. Świadomie nie przewraca żądania: sędzia stojący przy
    stoliku ma dostać protokół nawet wtedy, gdy baza akurat nie odpowiada —
    stratą jest brak wpisu, a nie brak dokumentu. Awaria idzie do logów jako
    ostrzeżenie, bo to realna dziura w audycie.
    """
    try:
        from app.db import database, protocol_audit

        await database.execute(protocol_audit.insert().values(**row))
    except Exception:
        logger.warning(
            "protocol_audit: nie udało się zapisać wpisu %s", row.get("code"), exc_info=True
        )


@router.post(
    "/judge/results/protocol/pdf",
    summary="Generuj PDF z protokołu na podstawie data_json (ProEl) i szablonu XLSX",
)
async def generate_protocol_pdf(
    req: ProtocolPdfRequest,
    request: Request,
    x_judge_id: Optional[str] = Header(None, alias="X-Judge-Id"),
    x_installation_id: Optional[str] = Header(None, alias="X-Installation-Id"),
    x_actor_name: Optional[str] = Header(None, alias="X-Actor-Name"),
    x_app_version: Optional[str] = Header(None, alias="X-App-Version"),
):
    data_json = req.data_json or {}
    if not isinstance(data_json, dict):
        raise HTTPException(400, "data_json musi być obiektem JSON")

    # Tożsamość generującego czytana MIĘKKO, w odróżnieniu od `proel_actor`,
    # który przy braku nagłówków rzuca 401. Wersje aplikacji sprzed tej zmiany
    # ich nie wysyłają, a protokół musi się wygenerować także im — wtedy stopka
    # mówi wprost, że autor jest nieznany, zamiast zmyślać nazwisko.
    # `header_text`, a nie `strip()`: nazwisko jedzie nagłówkiem HTTP, a te są
    # ze specyfikacji latin-1 — bez naprawy „Radosław" ląduje w stopce
    # protokołu jako „RadosÅ‚aw".
    actor_name = header_text(x_actor_name)
    actor_judge_id = header_text(x_judge_id)
    actor_install = header_text(x_installation_id)
    actor_verified = await _soft_verify_actor(actor_judge_id, actor_install)
    if actor_name:
        generated_by = actor_name
    elif actor_judge_id:
        generated_by = "nr %s" % actor_judge_id
    else:
        generated_by = "nieznanego"
    generated_dt = _warsaw_now()
    generated_at = generated_dt.strftime("%d.%m.%Y %H:%M:%S")
    generated_at_iso = generated_dt.isoformat()
    zprp_match_id = _zprp_match_id(data_json)
    match_number = str((data_json.get("matchConfig") or {}).get("matchNumber") or "").strip()

    # Ślad audytu liczony PRZED generowaniem: skrót stanu ma opisywać to, co
    # sędzia wysłał, a nie to, co po drodze zrobił z tym generator.
    audit_code = _protocol_audit_code()
    state_bytes = _canonical_json(data_json)
    state_sha256 = _sha256_bytes(state_bytes)
    try:
        signature = _sign_protocol_claim(
            {
                "v": 1,
                "c": audit_code,
                "m": zprp_match_id,
                "n": match_number,
                "j": actor_judge_id,
                "a": actor_name,
                "t": generated_at_iso,
                "s": state_sha256,
            }
        )
    except Exception:
        # Brak/zły RSA_PRIVATE_KEY nie może zatrzymać protokołu — dziennik i tak
        # zapisze stan, tylko bez dowodu działającego bez dostępu do bazy.
        signature = ""
        logger.warning("Nie udało się podpisać protokołu %s", audit_code, exc_info=True)

    # --- locate template ---
    template_path = SysPath(__file__).resolve().parent / "templates" / "protocol_template.xlsx"
    if not template_path.exists():
        raise HTTPException(
            500,
            f"Brak szablonu XLSX: {template_path}. Umieść plik w app/templates/protocol_template.xlsx i dodaj do repo.",
        )

    core = _get_match_core(data_json)
    half_ms = core["halfTimeMin"] * 60 * 1000

    # penalties totals
    pen = data_json.get("penaltyStats") or {}
    pen_h = pen.get("host") or {}
    pen_g = pen.get("guest") or {}
    pk_host_total = _safe_int(pen_h.get("total"), 0)
    pk_host_goals = _safe_int(pen_h.get("goals"), 0)
    pk_guest_total = _safe_int(pen_g.get("total"), 0)
    pk_guest_goals = _safe_int(pen_g.get("goals"), 0)

    # timeouts
    tt = data_json.get("teamTimeouts") or {}
    tt_host = tt.get("host") or {}
    tt_guest = tt.get("guest") or {}

    # players stats
    host_stats = _player_stats_map(data_json, "host")
    guest_stats = _player_stats_map(data_json, "guest")
    host_names = _player_fullname_map_from_cards(core.get("hostPlayerCards") or [])
    guest_names = _player_fullname_map_from_cards(core.get("guestPlayerCards") or [])

    if not host_names:
        host_names = _player_fullname_map_from_stats(host_stats)
    if not guest_names:
        guest_names = _player_fullname_map_from_stats(guest_stats)

    # Badania lekarskie — pusta mapa dla starszych wersji aplikacji.
    host_exams = _player_exam_map_from_cards(core.get("hostPlayerCards") or [])
    guest_exams = _player_exam_map_from_cards(core.get("guestPlayerCards") or [])

    host_comp_names = _companion_fullname_map(core.get("hostCompanions") or [])
    guest_comp_names = _companion_fullname_map(core.get("guestCompanions") or [])

    host_comp_meta = _companion_meta_map(core.get("hostCompanions") or [])
    guest_comp_meta = _companion_meta_map(core.get("guestCompanions") or [])

    host_comp_pen = _companion_penalty_strings(core.get("hostCompanions") or [])
    guest_comp_pen = _companion_penalty_strings(core.get("guestCompanions") or [])

    # winner (A/B) – przy remisie rozstrzygamy z penaltyScore
    winner = ""
    if core["scoreHost"] > core["scoreGuest"]:
        winner = "A"
    elif core["scoreGuest"] > core["scoreHost"]:
        winner = "B"
    else:
        ph, pg = _parse_penalty_score(data_json.get("penaltyScore") or "")
        if ph > pg:
            winner = "A"
        elif pg > ph:
            winner = "B"
        else:
            winner = ""  # jeśli brak/niepoprawny penaltyScore


    try:
        td = tempfile.mkdtemp(prefix="protocol_")  # ✅ nie usuwa się samo
        safe_code = re.sub(r"[^0-9A-Za-z_-]+", "_", (core.get("matchNumber") or "mecz"))
        filled_xlsx = os.path.join(td, f"protocol_{safe_code}.xlsx")

        wb = load_workbook(str(template_path))

        # 🔥 kluczowe: zanim cokolwiek zapiszesz / skopiujesz arkusze, rehydratacja obrazów
        media = _load_template_media_bytes(str(template_path))
        _rehydrate_images_in_workbook(wb, media)

        # Szablon ma dołożoną kolumnę ptaszków przed dawną kolumną A, więc
        # adresujemy arkusz przez nakładkę. `ws_raw` jest potrzebny wszędzie
        # tam, gdzie liczą się współrzędne fizyczne: kopiowanie arkuszy,
        # przenoszenie obrazków i wstawianie samych ptaszków.
        ws_raw = wb.active
        ws = ShiftedWS(ws_raw)

        # --- extras (NOWE POLA Z data_json) ---
        mc = data_json.get("matchConfig") or {}
        extras = mc.get("extras") or {}

        # --- detailed referee notes (last page) ---
        detailed_notes = bool(extras.get("detailedRefereeNotes")) if extras.get("detailedRefereeNotes") is not None else False
        detailed_notes_text = (extras.get("detailedRefereeNotesText") or "").strip()

        needs_detailed_notes_page = bool(detailed_notes and detailed_notes_text != "")

        # data/godzina
        ws["AB8"].value = _fmt_date_ddmmyyyy(extras.get("matchDate"))
        ws["AH8"].value = _fmt_time_hhmm(extras.get("matchTime"))

        # medyk
        medic = extras.get("medic") or {}
        ws["U61"].value = (medic.get("fullName") or "").strip()
        ws["U62"].value = (medic.get("number") or "").strip()

        # widzowie / pojemność
        ws["G62"].value = extras.get("spectatorsCount") if extras.get("spectatorsCount") is not None else ""
        ws["Q62"].value = extras.get("venueCapacity") if extras.get("venueCapacity") is not None else ""

        # szczegółowe uwagi sędziów: brak -> O61, verte -> S61
        ws["O61"].value = "X" if not detailed_notes else ""
        ws["S61"].value = "X" if detailed_notes else ""

        # rejestracja zawodów: tak -> O63, nie -> S63
        event_reg = bool(extras.get("eventRegistration")) if extras.get("eventRegistration") is not None else False
        ws["O63"].value = "X" if event_reg else ""
        ws["S63"].value = "X" if not event_reg else ""

        # dodatkowy raport: tak -> O64, nie -> S64
        extra_report = bool(extras.get("extraReport")) if extras.get("extraReport") is not None else False
        ws["O64"].value = "X" if extra_report else ""
        ws["S64"].value = "X" if not extra_report else ""

        # miejscowości sędziów (W66..W70)
        officials = extras.get("officials") or {}

        _set_cell_fallback(ws, "W66", (officials.get("referee1") or {}).get("city"), OFFICIAL_CITY_FALLBACK)
        _set_cell_fallback(ws, "W67", (officials.get("referee2") or {}).get("city"), OFFICIAL_CITY_FALLBACK)
        _set_cell_fallback(ws, "W68", (officials.get("secretary") or {}).get("city"), OFFICIAL_CITY_FALLBACK)
        _set_cell_fallback(ws, "W69", (officials.get("timekeeper") or {}).get("city"), OFFICIAL_CITY_FALLBACK)
        _set_cell_fallback(ws, "W70", (officials.get("delegate") or {}).get("city"), OFFICIAL_CITY_FALLBACK)

                # --- SIGNATURES (PNG z backendu) ---
        SIGN_ANCHORS = {
            "hostTeamSignature": "F29",
            "guestTeamSignature": "F55",
            "medic": "Z63",
            "referee1": "AI66",
            "referee2": "AI67",
            "secretary": "AI68",
            "timekeeper": "AI69",
            "delegate": "AI70",
        }

        # 1) podpisy drużyn
        host_sig_url = _full_static_url(extras.get("hostTeamSignature") or "")
        guest_sig_url = _full_static_url(extras.get("guestTeamSignature") or "")

        host_sig_bytes = await _fetch_png_bytes(host_sig_url)
        guest_sig_bytes = await _fetch_png_bytes(guest_sig_url)

        _add_signature_image(
            ws,
            image_bytes=host_sig_bytes,
            anchor_cell=SIGN_ANCHORS["hostTeamSignature"],
            max_width_px=80,
            max_height_px=45,
        )
        _add_signature_image(
            ws,
            image_bytes=guest_sig_bytes,
            anchor_cell=SIGN_ANCHORS["guestTeamSignature"],
            max_width_px=80,
            max_height_px=45,
        )

        # 2) podpis medyka
        medic_sig_url = _full_static_url((medic.get("signature") or "").strip())
        medic_sig_bytes = await _fetch_png_bytes(medic_sig_url)
        _add_signature_image(
            ws,
            image_bytes=medic_sig_bytes,
            anchor_cell=SIGN_ANCHORS["medic"],
            max_width_px=120,
            max_height_px=40,
        )

        # 3) podpisy officials
        def _off_sig_url(key: str) -> str:
            return _full_static_url((((officials.get(key) or {}).get("signature")) or "").strip())

        official_sig_bytes: Dict[str, bytes] = {}

        for key in ("referee1", "referee2", "secretary", "timekeeper", "delegate"):
            url = _off_sig_url(key)
            blob = await _fetch_png_bytes(url)
            official_sig_bytes[key] = blob or b""

            ok = _add_signature_image(
                ws,
                image_bytes=blob,
                anchor_cell=SIGN_ANCHORS[key],
                max_width_px=70,
                max_height_px=22,
            )

            if not ok:
                ws[SIGN_ANCHORS[key]].value = OFFICIAL_SIGN_FALLBACK


        # --- header mapping ---
        ws["AY1"].value = core["matchNumber"]
        ws["AL4"].value = core.get("venueAddress") or ""
        ws["C4"].value = core["hostName"]
        ws["D9"].value = core["hostName"]
        ws["C7"].value = core["guestName"]
        ws["D35"].value = core["guestName"]

        ws["AL6"].value = str(core["scoreHost"])
        ws["AQ6"].value = str(core["scoreGuest"])
        ws["AU6"].value = str(core["halfScoreHost"])
        ws["AY6"].value = str(core["halfScoreGuest"])
        ws["BB6"].value = winner

        # --- timeouts mapping ---
        _place_timeouts(ws, team_timeouts=tt_host, half_ms=half_ms, is_host=True)
        _place_timeouts(ws, team_timeouts=tt_guest, half_ms=half_ms, is_host=False)

        # --- penalties totals ---
        ws["AN65"].value = str(pk_host_total)
        ws["AR65"].value = str(pk_host_goals)
        ws["AY65"].value = str(pk_guest_total)
        ws["BC65"].value = str(pk_guest_goals)

        # --- players numbers + stats ---
        # UWAGA: ptaszki wstawiamy TU, czyli PRZED wb.copy_worksheet niżej —
        # dzięki temu _copy_images_safe przeniesie je na stronę 2 i na stronę
        # rzutów karnych. Przeniesienie tego bloku poniżej kopiowania sprawi,
        # że ptaszki znikną z kolejnych stron.
        _fill_players_block(
            ws,
            players=core["hostPlayers"],
            stats_by_number=host_stats,
            fullnames_by_number=host_names,
            start_row=11,
            end_row=28,
            exam_by_number=host_exams,
            mark_ws=ws_raw,
        )
        _fill_players_block(
            ws,
            players=core["guestPlayers"],
            stats_by_number=guest_stats,
            fullnames_by_number=guest_names,
            start_row=37,
            end_row=54,
            exam_by_number=guest_exams,
            mark_ws=ws_raw,
        )

        # Osoby towarzyszące gospodarzy
        ws["B29"].value  = host_comp_names.get("A", "")
        ws["K29"].value  = host_comp_names.get("B", "")
        ws["R29"].value  = host_comp_names.get("C", "")
        ws["Y29"].value  = host_comp_names.get("D", "")
        ws["AF29"].value = host_comp_names.get("E", "")

        # Osoby towarzyszące gości
        ws["B55"].value  = guest_comp_names.get("A", "")
        ws["K55"].value  = guest_comp_names.get("B", "")
        ws["R55"].value  = guest_comp_names.get("C", "")
        ws["Y55"].value  = guest_comp_names.get("D", "")
        ws["AF55"].value = guest_comp_names.get("E", "")

        # =========================
        # FUNKCJA + LICENCJA osób towarzyszących
        # (wg Twojego mapowania komórek)
        # =========================

        # GOSPODARZE:
        # A: function A30, license A31
        ws["A30"].value  = host_comp_meta.get("A", {}).get("function", "")
        ws["A31"].value  = host_comp_meta.get("A", {}).get("license", "")

        # B: function J30, license J31
        ws["J30"].value  = host_comp_meta.get("B", {}).get("function", "")
        ws["J31"].value  = host_comp_meta.get("B", {}).get("license", "")

        # C: function Q30, license Q31
        ws["Q30"].value  = host_comp_meta.get("C", {}).get("function", "")
        ws["Q31"].value  = host_comp_meta.get("C", {}).get("license", "")

        # D: function X30, license X31
        ws["X30"].value  = host_comp_meta.get("D", {}).get("function", "")
        ws["X31"].value  = host_comp_meta.get("D", {}).get("license", "")

        # E: function AE30, license AE31
        ws["AE30"].value = host_comp_meta.get("E", {}).get("function", "")
        ws["AE31"].value = host_comp_meta.get("E", {}).get("license", "")


        # GOŚCIE:
        # A: function A56, license A57
        ws["A56"].value  = guest_comp_meta.get("A", {}).get("function", "")
        ws["A57"].value  = guest_comp_meta.get("A", {}).get("license", "")

        # B: function J56, license J57
        ws["J56"].value  = guest_comp_meta.get("B", {}).get("function", "")
        ws["J57"].value  = guest_comp_meta.get("B", {}).get("license", "")

        # C: function Q56, license Q57
        ws["Q56"].value  = guest_comp_meta.get("C", {}).get("function", "")
        ws["Q57"].value  = guest_comp_meta.get("C", {}).get("license", "")

        # D: function X56, license X57
        ws["X56"].value  = guest_comp_meta.get("D", {}).get("function", "")
        ws["X57"].value  = guest_comp_meta.get("D", {}).get("license", "")

        # E: function AE56, license AE57
        ws["AE56"].value = guest_comp_meta.get("E", {}).get("function", "")
        ws["AE57"].value = guest_comp_meta.get("E", {}).get("license", "")


        # --- Kary osób towarzyszących (format: U/2'/D - MM:SS) ---

        # HOST A..E (row 31)
        ws["A32"].value  = host_comp_pen.get("A", {}).get("warn", "---")
        ws["D32"].value  = host_comp_pen.get("A", {}).get("p2", "---")
        ws["G32"].value  = host_comp_pen.get("A", {}).get("disq", "---")

        ws["J32"].value  = host_comp_pen.get("B", {}).get("warn", "---")
        ws["L32"].value  = host_comp_pen.get("B", {}).get("p2", "---")
        ws["O32"].value  = host_comp_pen.get("B", {}).get("disq", "---")

        ws["Q32"].value  = host_comp_pen.get("C", {}).get("warn", "---")
        ws["S32"].value  = host_comp_pen.get("C", {}).get("p2", "---")
        ws["V32"].value  = host_comp_pen.get("C", {}).get("disq", "---")

        ws["X32"].value  = host_comp_pen.get("D", {}).get("warn", "---")
        ws["Z32"].value  = host_comp_pen.get("D", {}).get("p2", "---")
        ws["AC32"].value = host_comp_pen.get("D", {}).get("disq", "---")

        ws["AE32"].value = host_comp_pen.get("E", {}).get("warn", "---")
        ws["AH32"].value = host_comp_pen.get("E", {}).get("p2", "---")
        ws["AJ32"].value = host_comp_pen.get("E", {}).get("disq", "---")

        # GUEST A..E (row 56)
        ws["A58"].value  = guest_comp_pen.get("A", {}).get("warn", "---")
        ws["D58"].value  = guest_comp_pen.get("A", {}).get("p2", "---")
        ws["G58"].value  = guest_comp_pen.get("A", {}).get("disq", "---")

        ws["J58"].value  = guest_comp_pen.get("B", {}).get("warn", "---")
        ws["L58"].value  = guest_comp_pen.get("B", {}).get("p2", "---")
        ws["O58"].value  = guest_comp_pen.get("B", {}).get("disq", "---")

        ws["Q58"].value  = guest_comp_pen.get("C", {}).get("warn", "---")
        ws["S58"].value  = guest_comp_pen.get("C", {}).get("p2", "---")
        ws["V58"].value  = guest_comp_pen.get("C", {}).get("disq", "---")

        ws["X58"].value  = guest_comp_pen.get("D", {}).get("warn", "---")
        ws["Z58"].value  = guest_comp_pen.get("D", {}).get("p2", "---")
        ws["AC58"].value = guest_comp_pen.get("D", {}).get("disq", "---")

        ws["AE58"].value = guest_comp_pen.get("E", {}).get("warn", "---")
        ws["AH58"].value = guest_comp_pen.get("E", {}).get("p2", "---")
        ws["AJ58"].value = guest_comp_pen.get("E", {}).get("disq", "---")

        # --- Companion crossouts (B-E) if empty ---
        _apply_companion_crossouts(
            ws,
            host_names=host_comp_names,
            host_meta=host_comp_meta,
            host_pen=host_comp_pen,
            guest_names=guest_comp_names,
            guest_meta=guest_comp_meta,
            guest_pen=guest_comp_pen,
        )

        # Sędziowie
        _set_cell_fallback(ws, "I66", core.get("referee1"), OFFICIAL_NAME_FALLBACK)
        _set_cell_fallback(ws, "I67", core.get("referee2"), OFFICIAL_NAME_FALLBACK)
        _set_cell_fallback(ws, "I68", core.get("secretary"), OFFICIAL_NAME_FALLBACK)
        _set_cell_fallback(ws, "I69", core.get("timekeeper"), OFFICIAL_NAME_FALLBACK)
        _set_cell_fallback(ws, "I70", core.get("delegate"), OFFICIAL_NAME_FALLBACK)

        # --- timeline (match events) + optional pages (overflow + shootout) ---
        evs1, evs2 = _extract_timeline_events(data_json)
        needs_timeline_page2 = (len(evs1) > TIMELINE_MAX_ROWS) or (len(evs2) > TIMELINE_MAX_ROWS)

        needs_shootout_page = _shootout_needed(data_json)

        # Tworzymy listę stron (arkuszy) w kolejności: 1, (2 - overflow), (shootout), (UWAGI - zawsze ostatnie jeśli warunek)
        pages = [ws]

        ws2 = None
        if needs_timeline_page2:
            ws2_raw = wb.copy_worksheet(ws_raw)
            _copy_images_safe(ws_raw, ws2_raw)
            try:
                ws2_raw.title = "Strona 2"
            except Exception:
                pass
            ws2 = ShiftedWS(ws2_raw)
            pages.append(ws2)

        ws_shoot = None
        if needs_shootout_page:
            ws_shoot_raw = wb.copy_worksheet(ws_raw)
            _copy_images_safe(ws_raw, ws_shoot_raw)
            try:
                ws_shoot_raw.title = "Rzuty karne"
            except Exception:
                pass
            ws_shoot = ShiftedWS(ws_shoot_raw)
            pages.append(ws_shoot)

        ws_notes = None
        if needs_detailed_notes_page:
            date_ddmmyyyy = _fmt_date_ddmmyyyy(extras.get("matchDate") or "")
            place = _extract_city_from_venue_address(core.get("venueAddress") or "")

            # preferuj dane z extras.officials (bo tam masz fullName + signature)
            ref1_name = ((officials.get("referee1") or {}).get("fullName") or core.get("referee1") or "").strip()
            ref2_name = ((officials.get("referee2") or {}).get("fullName") or core.get("referee2") or "").strip()

            ref1_sig = official_sig_bytes.get("referee1", b"")
            ref2_sig = official_sig_bytes.get("referee2", b"")

            # tymczasowo dodamy jako ostatnią, ale numer strony policzymy po ustaleniu total_pages
            ws_notes = _create_detailed_notes_sheet(
                wb,
                page_no=0,               # podmienimy po wyliczeniu total_pages
                total_pages=0,           # podmienimy po wyliczeniu total_pages
                date_ddmmyyyy=date_ddmmyyyy,
                place=place,
                notes_text=detailed_notes_text,
                referee1_name=ref1_name,
                referee2_name=ref2_name,
                referee1_sig_bytes=ref1_sig,
                referee2_sig_bytes=ref2_sig,
            )
            pages.append(ws_notes)

        # 1) Ustaw numerację stron na wszystkich stronach protokołu (AQ2),
        #    a na stronie uwag ustawiamy nagłówek "Strona X/X" w A1 (już w helperze).
        total_pages = len(pages)

        # Arkusze protokołu: AQ2 = STRONA X/N (jak dotychczas)
        if total_pages > 1:
            for i, p in enumerate(pages, start=1):
                if p is ws_notes:
                    continue
                p["AQ2"].value = f"STRONA {i}/{total_pages}"
        else:
            ws["AQ2"].value = ""

        # Uzupełnij poprawnie page_no/total_pages na stronie uwag (bo helper dostał 0/0)
        if ws_notes is not None:
            page_no = pages.index(ws_notes) + 1  # 1-based
            ws_notes["A1"].value = f"Strona {page_no}/{total_pages}"
            # prawy nagłówek (G1) helper już wypełnił – nie ruszamy

        # 2) Wypełnij przebieg meczu na stronach 1 oraz (opcjonalnie) 2
        if needs_timeline_page2 and ws2 is not None:
            _fill_timeline_pages(
                ws,
                ws2,
                data_json=data_json,
                half_ms=half_ms,
                half_score_host=core["halfScoreHost"],
                half_score_guest=core["halfScoreGuest"],
            )
        else:
            _fill_timeline_pages(
                ws,
                None,
                data_json=data_json,
                half_ms=half_ms,
                half_score_host=core["halfScoreHost"],
                half_score_guest=core["halfScoreGuest"],
            )

        # 3) Jeśli jest strona karnych – podmień przebieg na "RZUTY KARNE"
        if ws_shoot is not None:
            _fill_shootout_page(ws_shoot, data_json=data_json)


        # Nagłówek/stopka DOPIERO tutaj — po wszystkich `copy_worksheet`, żeby
        # objęły też stronę uwag i stronę rzutów karnych.
        _apply_protocol_page_marks(
            wb,
            match_id=zprp_match_id,
            generated_by=generated_by,
            generated_at=generated_at,
        )

        wb.save(filled_xlsx)

        # --- convert to PDF ---
        # `_convert_xlsx_to_pdf` woła `subprocess.run` na LibreOffice, co blokuje
        # wątek na kilka sekund. Ta funkcja jest `async def`, a serwer startuje
        # jako `uvicorn main:app` BEZ `--workers`, więc jedna pętla zdarzeń
        # obsługuje całą flotę: bez `to_thread` generowanie jednego protokołu
        # zawiesza KAŻDE inne żądanie w tym czasie.
        pdf_path = await asyncio.to_thread(_convert_xlsx_to_pdf, filled_xlsx, td)

        # Znacznik w metadanych, a potem skrót — w tej kolejności, bo liczy się
        # suma kontrolna pliku, który dostanie użytkownik.
        _stamp_pdf_metadata(
            pdf_path,
            code=audit_code,
            signature=signature,
            match_id=zprp_match_id,
            match_number=match_number,
            state_sha256=state_sha256,
            generated_by=generated_by,
            generated_at_iso=generated_at_iso,
        )
        pdf_sha256 = _sha256_file(pdf_path)
        # Treść wydruku zapisujemy TERAZ, na gotowym pliku — to jedyny moment,
        # w którym mamy oryginał w ręku.
        with open(pdf_path, "rb") as _fh:
            pdf_text = _extract_pdf_text(_fh.read())

        await _record_protocol_audit(
            {
                "code": audit_code,
                "match_id": zprp_match_id or None,
                "match_number": match_number or None,
                "judge_id": actor_judge_id or None,
                "actor_name": actor_name or None,
                "installation_id": actor_install or None,
                "verified": actor_verified,
                "state_sha256": state_sha256,
                "pdf_sha256": pdf_sha256,
                "state_gzip": gzip.compress(state_bytes),
                "pdf_text_gzip": (
                    gzip.compress(pdf_text.encode("utf-8")) if pdf_text else None
                ),
                "signature": signature or None,
                "app_version": str(x_app_version or "").strip() or None,
                "client_ip": (request.client.host if request and request.client else None),
                "created_at": generated_dt,
            }
        )

        # przygotuj plik do pobrania po tokenie
        _cleanup_expired_downloads()
        _ensure_download_dir()

        token = str(uuid.uuid4())
        filename = _safe_filename_from_match_number(core.get("matchNumber") or "mecz")

        # zapisujemy finalny plik w /tmp (nie usuwamy go BackgroundTask od razu)
        download_path = os.path.join(DOWNLOAD_DIR, f"{token}.pdf")
        shutil.copyfile(pdf_path, download_path)

        # sprzątnij roboczy katalog po konwersji (xlsx + profile LO)
        shutil.rmtree(td, ignore_errors=True)

        # zwróć link do pobrania
        return {
            "success": True,
            "token": token,
            "filename": filename,
            "download_url": f"/judge/results/protocol/pdf/download/{token}?filename={filename}",
            # Kod audytu wraca do aplikacji świadomie: pozwala sędziemu podać go
            # przy zgłoszeniu („wygenerowałem BZ-…, coś jest nie tak"), zanim
            # ktokolwiek zacznie szukać po numerze meczu i godzinie.
            "audit_code": audit_code,
        }


    except HTTPException:
        raise
    except Exception as e:
        logger.error("generate_protocol_pdf error: %s", e, exc_info=True)
        raise HTTPException(status_code=500, detail=f"Nie udało się wygenerować PDF: {e}")


# ─────────────────────── dziennik: wgląd dla admina ───────────────────────


async def _require_protocol_admin(actor) -> None:
    from app.proel_auth import is_admin

    if not await is_admin(actor.judge_id):
        raise HTTPException(403, detail={"code": "ADMIN_REQUIRED", "message": "Brak uprawnień"})


def _audit_public(row: Any) -> Dict[str, Any]:
    """
    Wpis dziennika w postaci nadającej się do JSON-a.

    Spakowane blob-y wypadają w całości: stan wydajemy osobnym endpointem i
    tylko na żądanie, a treść wydruku służy wyłącznie do porównania po stronie
    serwera.

    Zamiatanie po TYPIE, nie po nazwie, jest tu celowe. Wersja wymieniająca
    kolumny z nazwy przepuściła `pdf_text_gzip` dodane później i wywracała
    KAŻDĄ odpowiedź dziennika błędem 500: FastAPI serializuje `bytes` przez
    `.decode()`, a gzip zaczyna się od bajtu, którego UTF-8 nie zna. Padał
    wtedy nie tylko podgląd wpisu, ale i lista — czyli całe okno dziennika.
    """
    d = dict(row)
    for key, value in list(d.items()):
        if isinstance(value, (bytes, bytearray, memoryview)):
            d.pop(key, None)
    created = d.get("created_at")
    if hasattr(created, "isoformat"):
        d["created_at"] = created.isoformat()
    return d


@router.get(
    "/judge/results/protocol/audit/{code}",
    summary="[admin] Wpis dziennika generowania protokołu (kto, kiedy, jaki stan)",
)
async def protocol_audit_entry(
    code: str = ApiPath(...),
    actor=Depends(proel_actor),
):
    await _require_protocol_admin(actor)
    from sqlalchemy import select as _select

    from app.db import database, protocol_audit

    row = await database.fetch_one(
        _select(protocol_audit).where(protocol_audit.c.code == code.strip().upper())
    )
    if row is None:
        raise HTTPException(404, "Nie ma takiego kodu w dzienniku")

    out = _audit_public(row)
    ok, claim = _verify_protocol_signature(str(row["signature"] or ""))
    out["signature_valid"] = ok
    out["signature_claim"] = claim
    return out


@router.get(
    "/judge/results/protocol/audit",
    summary="[admin] Historia generowania protokołów (po meczu lub sędzim)",
)
async def protocol_audit_list(
    match: Optional[str] = Query(None, description="IdZawody albo numer meczu (dokładnie)"),
    judge_id: Optional[str] = Query(None),
    pdf_sha256: Optional[str] = Query(None, description="Skrót pliku, gdy metadane wyczyszczono"),
    q: Optional[str] = Query(None, description="Szukaj w kodzie, meczu, nazwisku"),
    limit: int = Query(50, ge=1, le=500),
    actor=Depends(proel_actor),
):
    await _require_protocol_admin(actor)
    from sqlalchemy import func as _func, or_, select as _select

    from app.db import database, protocol_audit

    query = _select(protocol_audit)
    m = str(match or "").strip()
    if m:
        query = query.where(
            or_(protocol_audit.c.match_id == m, protocol_audit.c.match_number == m)
        )
    if judge_id:
        query = query.where(protocol_audit.c.judge_id == str(judge_id).strip())
    if pdf_sha256:
        query = query.where(protocol_audit.c.pdf_sha256 == str(pdf_sha256).strip().lower())

    # Szukanie tekstowe po stronie bazy, a nie filtrowanie doładowanej strony:
    # inaczej „znajdź mi wszystko od tego sędziego" działałoby wyłącznie w
    # obrębie ostatnich 50 wpisów i po cichu gubiło starsze.
    text_q = str(q or "").strip().lower()
    if text_q:
        like = "%%%s%%" % text_q
        query = query.where(
            or_(
                _func.lower(protocol_audit.c.code).like(like),
                _func.lower(protocol_audit.c.match_number).like(like),
                _func.lower(protocol_audit.c.match_id).like(like),
                _func.lower(protocol_audit.c.actor_name).like(like),
                _func.lower(protocol_audit.c.judge_id).like(like),
            )
        )

    rows = await database.fetch_all(
        query.order_by(protocol_audit.c.created_at.desc()).limit(limit)
    )
    return {"count": len(rows), "items": [_audit_public(r) for r in rows]}


@router.get(
    "/judge/results/protocol/audit/{code}/state",
    summary="[admin] Stan meczu, z którego powstał ten protokół (do odtworzenia)",
)
async def protocol_audit_state(
    code: str = ApiPath(...),
    actor=Depends(proel_actor),
):
    """
    Zwraca dokładnie ten `data_json`, który przyszedł z aplikacji. Wysłany z
    powrotem na `POST /judge/results/protocol/pdf` odtwarza protokół co do
    treści — stopka i kod audytu będą inne, i tak ma być: kopia ma dać się
    odróżnić od oryginału.
    """
    await _require_protocol_admin(actor)
    from sqlalchemy import select as _select

    from app.db import database, protocol_audit

    row = await database.fetch_one(
        _select(protocol_audit).where(protocol_audit.c.code == code.strip().upper())
    )
    if row is None:
        raise HTTPException(404, "Nie ma takiego kodu w dzienniku")

    raw = gzip.decompress(row["state_gzip"])
    return {
        "code": row["code"],
        "state_sha256": row["state_sha256"],
        "matches_stored_hash": _sha256_bytes(raw) == row["state_sha256"],
        "data_json": json.loads(raw.decode("utf-8")),
    }


class ProtocolVerifyRequest(BaseModel):
    signature: str
    pdf_sha256: Optional[str] = None


@router.post(
    "/judge/results/protocol/verify",
    summary="[admin] Sprawdź podpis protokołu (i czy plik jest tym z dziennika)",
)
async def protocol_verify(
    req: ProtocolVerifyRequest,
    actor=Depends(proel_actor),
):
    await _require_protocol_admin(actor)
    ok, claim = _verify_protocol_signature(req.signature)

    out: Dict[str, Any] = {"signature_valid": ok, "claim": claim}
    code = str(claim.get("c") or "").strip()
    if not code:
        return out

    from sqlalchemy import select as _select

    from app.db import database, protocol_audit

    row = await database.fetch_one(
        _select(protocol_audit).where(protocol_audit.c.code == code)
    )
    if row is None:
        out["known_in_ledger"] = False
        return out

    out["known_in_ledger"] = True
    out["entry"] = _audit_public(row)
    # Rozróżnienie, na którym wszystko stoi: podpis mówi „to nasz plik dla tego
    # stanu", a skrót — „to TEN egzemplarz, nikt go po drodze nie tknął".
    out["state_matches_ledger"] = str(claim.get("s") or "") == str(row["state_sha256"] or "")
    if req.pdf_sha256:
        out["pdf_matches_ledger"] = (
            str(req.pdf_sha256).strip().lower() == str(row["pdf_sha256"] or "").lower()
        )
    return out


_PDF_MARK_RE = {
    "code": re.compile(r"BazaCode=([^;\s]+)"),
    "signature": re.compile(r"BazaSig=([^;\s]+)"),
    "state_sha256": re.compile(r"BazaState=([0-9a-fA-F]+)"),
}
_XMP_MARK_RE = re.compile(r"<baza:(\w+)>(.*?)</baza:\1>", re.S)


def _extract_pdf_marks(data: bytes) -> Dict[str, str]:
    """
    Wyciąga znacznik z metadanych PDF-a — z `/Info` i z XMP naraz.

    Dwa miejsca, bo przeżywają różne rzeczy: `keywords` zachowuje większość
    narzędzi, XMP bywa jedynym, co zostaje po konwersji, która `/Info` gubi.
    Bierzemy pierwszą niepustą wartość z każdego pola.
    """
    out: Dict[str, str] = {}
    try:
        import fitz  # pymupdf
    except Exception:
        return out

    try:
        doc = fitz.open(stream=data, filetype="pdf")
    except Exception:
        return out

    try:
        blob = " ".join(
            str(v or "") for v in (doc.metadata or {}).values()
        )
        for key, rx in _PDF_MARK_RE.items():
            m = rx.search(blob)
            if m:
                out[key] = m.group(1)

        try:
            xmp = doc.get_xml_metadata() or ""
        except Exception:
            xmp = ""
        for key, value in _XMP_MARK_RE.findall(xmp):
            mapped = {"stateSha256": "state_sha256"}.get(key, key)
            if mapped in ("code", "signature", "state_sha256") and not out.get(mapped):
                out[mapped] = value.strip()
    finally:
        try:
            doc.close()
        except Exception:
            pass
    return out


@router.post(
    "/judge/results/protocol/verify-file",
    summary="[admin] Sprawdź plik PDF protokołu — znacznik, podpis i zgodność z dziennikiem",
)
async def protocol_verify_file(
    file: UploadFile = File(..., description="Plik PDF protokołu"),
    actor=Depends(proel_actor),
):
    """
    Weryfikacja od strony pliku, a nie wklejonego tokenu.

    Robi trzy rzeczy naraz, których wklejanie podpisu z ręki nie potrafi:
      • czyta znacznik samo, więc nikt nie musi grzebać we właściwościach PDF-a,
      • liczy skrót TEGO egzemplarza i porównuje go z dziennikiem, czyli
        odpowiada nie tylko „czy nasz", ale „czy nietknięty",
      • gdy metadane wyczyszczono, próbuje trafić do wpisu po samym skrócie —
        bo skasowanie znacznika nie zmienia bajtów strony.
    """
    await _require_protocol_admin(actor)

    data = await file.read()
    if not data[:5] == b"%PDF-":
        raise HTTPException(400, "To nie jest plik PDF")

    marks = _extract_pdf_marks(data)
    pdf_sha256 = _sha256_bytes(data)

    out: Dict[str, Any] = {
        "pdf_sha256": pdf_sha256,
        "has_marks": bool(marks),
        "code": marks.get("code") or None,
    }

    signature = marks.get("signature") or ""
    if signature:
        ok, claim = _verify_protocol_signature(signature)
        out["signature_valid"] = ok
        out["claim"] = claim
    else:
        out["signature_valid"] = None
        out["claim"] = {}

    from sqlalchemy import select as _select

    from app.db import database, protocol_audit

    row = None
    if marks.get("code"):
        row = await database.fetch_one(
            _select(protocol_audit).where(protocol_audit.c.code == marks["code"])
        )
        if row is not None:
            out["found_by"] = "code"
    if row is None:
        row = await database.fetch_one(
            _select(protocol_audit).where(protocol_audit.c.pdf_sha256 == pdf_sha256)
        )
        if row is not None:
            out["found_by"] = "hash"

    if row is None:
        out["known_in_ledger"] = False
        return out

    out["known_in_ledger"] = True
    out["entry"] = _audit_public(row)
    # Trzy RÓŻNE pytania, często mylone:
    #  • podpis           — czy znacznik wystawił nasz serwer,
    #  • state_sha256     — czy znacznik wskazuje TEN wpis w dzienniku
    #                       (czyli czy nikt nie przekleił znacznika z innego
    #                       protokołu); NIC nie mówi o treści strony,
    #  • pdf_sha256       — czy bajty pliku są te, które wydaliśmy.
    # Edycja liczby widzów w gotowym PDF-ie zmienia wyłącznie to ostatnie —
    # dlatego „stan" potrafi się zgadzać przy ewidentnie podrobionym wydruku.
    out["pdf_matches_ledger"] = pdf_sha256 == str(row["pdf_sha256"] or "")
    if marks.get("state_sha256"):
        out["marker_points_to_entry"] = marks["state_sha256"] == str(
            row["state_sha256"] or ""
        )
        # Stara nazwa zostaje dla zgodności z wcześniejszą wersją aplikacji.
        out["state_matches_ledger"] = out["marker_points_to_entry"]

    # Co konkretnie się zmieniło — tylko gdy plik faktycznie odbiega od
    # oryginału i mamy z czym porównać.
    if not out["pdf_matches_ledger"]:
        stored = row["pdf_text_gzip"]
        if stored:
            try:
                before = gzip.decompress(stored).decode("utf-8")
                out["text_diff"] = _diff_pdf_text(before, _extract_pdf_text(data))
            except Exception:
                logger.warning("Nie udało się porównać treści PDF", exc_info=True)
                out["text_diff"] = None
        else:
            # Wpis sprzed wersji, która zapisuje treść wydruku — wiemy, że plik
            # jest inny, ale nie mamy oryginału do porównania.
            out["text_diff"] = None
    return out


@router.get(
    "/judge/results/protocol/public-key",
    summary="Klucz publiczny do weryfikacji podpisu protokołu (PEM)",
)
async def protocol_public_key():
    """Bez autoryzacji — klucz publiczny jest publiczny z definicji, a bez niego
    dowód „to plik z naszego serwera" działałby wyłącznie przez nasz serwer."""
    from cryptography.hazmat.primitives import serialization

    _, public_key = get_rsa_keys()
    pem = public_key.public_bytes(
        encoding=serialization.Encoding.PEM,
        format=serialization.PublicFormat.SubjectPublicKeyInfo,
    ).decode("ascii")
    return {"algorithm": "RSA-PSS/SHA-256", "public_key_pem": pem}


from fastapi.responses import FileResponse

@router.get(
    "/judge/results/protocol/pdf/download/{token}",
    summary="Pobierz wygenerowany PDF protokołu (attachment)",
)
async def download_protocol_pdf(
    token: str = ApiPath(...),
    filename: str = Query("protokol.pdf"),
):
    _ensure_download_dir()
    file_path = os.path.join(DOWNLOAD_DIR, f"{token}.pdf")
    if not os.path.exists(file_path):
        raise HTTPException(404, "Plik wygasł lub nie istnieje")

    # nagłówki jak w excelu
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Cache-Control": "no-store",
    }

    # po pobraniu: możesz sprzątnąć (opcjonalnie)
    # UWAGA: czasem system pobierania może dociągać plik chwilę,
    # ale FileResponse(background=...) sprząta po zakończeniu odpowiedzi.
    return FileResponse(
        path=file_path,
        media_type="application/pdf",
        filename=filename,
        headers=headers,
        background=BackgroundTask(lambda: os.remove(file_path) if os.path.exists(file_path) else None),
    )