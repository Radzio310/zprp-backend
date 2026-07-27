"""MP elimination appearance verification and protocol-list snapshots.

The module deliberately keeps two concepts separate:
* positive evidence from a completed ProEl match;
* a revisioned fallback snapshot of the tournament "Do protokołu" list.

This lets the application enforce the final-entry rule without pretending that
an administrative list is proof of physical participation.
"""
from __future__ import annotations

import asyncio
import base64
import json
import logging
import os
import tempfile
import unicodedata
import urllib.parse
import uuid
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Literal, Optional, Sequence, Tuple
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field
from sqlalchemy import and_, func, insert, select, update
from starlette.background import BackgroundTask

from app.beach.activity_log import get_actor_name, log_activity
from app.beach.capabilities import resolve_user_capabilities
from app.db import (
    beach_admins,
    beach_mp_eligibility_settings,
    beach_proel_matches,
    beach_teams,
    beach_tournament_protocol_snapshots,
    beach_tournaments,
    beach_users,
    database,
)
from app.deps import beach_get_current_user_id

logger = logging.getLogger(__name__)
router = APIRouter(
    prefix="/beach/mp-appearances",
    tags=["Beach: MP appearances"],
)

WARSAW = ZoneInfo("Europe/Warsaw")
DEFAULT_SEASON_ID = "8"
DEFAULT_ENABLED_CATEGORIES = ["Senior"]
VALID_CATEGORIES = ["Senior", "Junior", "Junior mł.", "Młodzik", "Dzieci"]

STATUS_PRIORITY = {
    "approved_proel": 50,
    "finished_proel": 40,
    "frozen_confirmed": 30,
    "frozen_warning": 20,
    "missing": 0,
    "insufficient": -10,
}

MP_EXPORT_DIR = Path(tempfile.gettempdir()) / "mp_appearance_exports"
MP_TEMPLATE_DIR = Path(__file__).resolve().parent.parent / "templates"
MP_REPORT_TEMPLATE = "wystepy_mp.html"
MP_CATEGORY_COLORS = {
    "Senior": "#E85A30",
    "Junior": "#3A7FBF",
    "Junior mł.": "#2BA8A0",
    "Młodzik": "#7A5FC7",
    "Dzieci": "#D48A25",
}
MP_STATUS_LABELS = {
    "approved_proel": "ProEl zatwierdzony",
    "finished_proel": "ProEl zakończony",
    "frozen_confirmed": "Z listy zgłoszeniowej do turnieju",
    "frozen_warning": "Z listy zgłoszeniowej do turnieju — sprawdź",
    "missing": "Brak potwierdzenia",
    "insufficient": "Niewystarczające dane",
}
MP_STATUS_COLORS = {
    "approved_proel": "#34A853",
    "finished_proel": "#20AFA8",
    "frozen_confirmed": "#347FD1",
    "frozen_warning": "#E58A00",
    "missing": "#D9443E",
    "insufficient": "#7A7F89",
}
MP_STATUS_SHORT_LABELS = {
    "approved_proel": "ProEl zatwierdzony",
    "finished_proel": "ProEl zakończony",
    "frozen_confirmed": "Lista zgłoszeniowa",
    "frozen_warning": "Lista — sprawdź",
    "missing": "Brak potwierdzenia",
    "insufficient": "Za mało danych",
}


def _rollout_at() -> datetime:
    raw = os.getenv("MP_SNAPSHOT_ROLLOUT_AT", "2026-07-27T00:00:00+00:00")
    try:
        parsed = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        return (
            parsed.replace(tzinfo=timezone.utc)
            if parsed.tzinfo is None
            else parsed.astimezone(timezone.utc)
        )
    except ValueError:
        return datetime(2026, 7, 27, tzinfo=timezone.utc)


class MpSettingsUpdate(BaseModel):
    enabled_categories: List[str] = Field(default_factory=list)


class SnapshotRefreshRequest(BaseModel):
    reason: str = Field(..., min_length=3, max_length=500)
    context_tournament_id: Optional[int] = None


class MpReportExportRequest(BaseModel):
    format: Literal["pdf", "xlsx"]
    status: Optional[
        Literal[
            "approved_proel",
            "finished_proel",
            "frozen_confirmed",
            "frozen_warning",
            "missing",
            "insufficient",
        ]
    ] = None
    gender: Literal["all", "K", "M"] = "all"
    only_issues: bool = False
    query: str = Field(default="", max_length=120)


def _json(raw: Any, fallback: Any = None) -> Any:
    if raw is None:
        return {} if fallback is None else fallback
    if isinstance(raw, (dict, list)):
        return raw
    try:
        return json.loads(raw)
    except Exception:
        return {} if fallback is None else fallback


def _ascii(value: Any) -> str:
    # NFKD does not decompose Polish ł/Ł, while historical tournament names
    # commonly use "Finał".  Transliterate it explicitly before stripping
    # the remaining diacritics.
    normalized_value = str(value or "").translate(
        str.maketrans({"ł": "l", "Ł": "L"})
    )
    return "".join(
        c for c in unicodedata.normalize("NFKD", normalized_value.lower())
        if not unicodedata.combining(c)
    )


def infer_mp_phase(name: Any, competition_type: Any, explicit: Any = None) -> Optional[str]:
    if explicit in ("elimination", "final"):
        return str(explicit)
    if str(competition_type or "").strip().upper() != "MP":
        return None
    normalized = _ascii(name)
    return "final" if "final" in normalized else "elimination"


def tournament_season_id(row: Dict[str, Any]) -> str:
    data = _json(row.get("data_json"))
    explicit = row.get("season_id") or data.get("season_id")
    if explicit:
        return str(explicit)
    return f"unassigned:{_parse_event_date(row.get('event_date')).year}"


def tournament_phase(row: Dict[str, Any]) -> Optional[str]:
    return infer_mp_phase(
        row.get("name"),
        row.get("competition_type"),
        row.get("mp_phase"),
    )


def season_year_label(season_id: Any) -> str:
    try:
        return str(2018 + int(season_id))
    except (TypeError, ValueError):
        return str(season_id or "")


def _numeric_ids(values: Any) -> List[int]:
    result: List[int] = []
    seen: set[int] = set()
    for value in values if isinstance(values, list) else []:
        try:
            parsed = int(value)
        except (TypeError, ValueError):
            continue
        if parsed > 0 and parsed not in seen:
            seen.add(parsed)
            result.append(parsed)
    return result


def effective_protocol_ids(squad_entry: Dict[str, Any]) -> List[int]:
    # Presence of an explicitly empty protocol list is meaningful.  Fall back
    # only for legacy entries which predate protocol_players.
    if "protocol_players" in squad_entry:
        return _numeric_ids(squad_entry.get("protocol_players"))
    return _numeric_ids(squad_entry.get("default_players"))


def _regular_team(match_team: Any) -> Optional[Tuple[int, str]]:
    if not isinstance(match_team, dict):
        return None
    raw_id = match_team.get("id")
    try:
        team_id = int(raw_id)
    except (TypeError, ValueError):
        return None
    if team_id <= 0 or str(raw_id).startswith("ct_"):
        return None
    return team_id, str(match_team.get("name") or "")


def _is_match_cancelled(match: Dict[str, Any]) -> bool:
    flags = (
        match.get("cancelled"),
        match.get("canceled"),
        match.get("walkover"),
        match.get("isWalkover"),
        match.get("isCancelled"),
    )
    return any(bool(value) for value in flags)


def _schedule_matches(data: Dict[str, Any]) -> List[Dict[str, Any]]:
    schedule = data.get("schedule")
    if not isinstance(schedule, dict):
        return []
    return [
        match
        for match in (schedule.get("matches") or [])
        if isinstance(match, dict) and match.get("kind", "match") == "match"
    ]


def _parse_event_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.astimezone(WARSAW).date()
    text = str(value or "")[:10]
    try:
        return date.fromisoformat(text)
    except ValueError:
        return datetime.now(WARSAW).date()


def _parse_clock(value: Any) -> time:
    text = str(value or "00:00").strip()
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            continue
    return time(0, 0)


def scheduled_match_at(
    tournament_row: Dict[str, Any],
    tournament_data: Dict[str, Any],
    match: Dict[str, Any],
) -> datetime:
    schedule = tournament_data.get("schedule")
    config = schedule.get("config") if isinstance(schedule, dict) else {}
    days = config.get("days") if isinstance(config, dict) else []
    try:
        day_index = max(0, int(match.get("dayIndex") or 0))
    except (TypeError, ValueError):
        day_index = 0
    match_date: Optional[date] = None
    if isinstance(days, list) and day_index < len(days) and isinstance(days[day_index], dict):
        raw_day = days[day_index].get("date")
        if raw_day:
            try:
                match_date = date.fromisoformat(str(raw_day)[:10])
            except ValueError:
                pass
    if match_date is None:
        match_date = _parse_event_date(tournament_row.get("event_date")) + timedelta(days=day_index)
    clock = _parse_clock(match.get("originalTime") or match.get("startTime"))
    return datetime.combine(match_date, clock, tzinfo=WARSAW).astimezone(timezone.utc)


def first_team_matches(
    tournament_row: Dict[str, Any],
    tournament_data: Dict[str, Any],
) -> Dict[int, Tuple[datetime, str]]:
    result: Dict[int, Tuple[datetime, str]] = {}
    for match in _schedule_matches(tournament_data):
        when = scheduled_match_at(tournament_row, tournament_data, match)
        for key in ("teamA", "teamB"):
            parsed = _regular_team(match.get(key))
            if not parsed:
                continue
            team_id, team_name = parsed
            current = result.get(team_id)
            if current is None or when < current[0]:
                result[team_id] = (when, team_name)
    return result


async def _is_admin(user_id: int) -> bool:
    return bool(
        await database.fetch_one(
            select(beach_admins.c.user_id).where(beach_admins.c.user_id == user_id)
        )
    )


def _is_head_judge(data: Dict[str, Any], user_id: int) -> bool:
    try:
        return int(data.get("head_judge_id")) == int(user_id)
    except (TypeError, ValueError):
        return False


async def _can_view_report(tournament: Dict[str, Any], user_id: int) -> bool:
    if await _is_admin(user_id):
        return True
    data = _json(tournament.get("data_json"))
    if _is_head_judge(data, user_id):
        return True
    for host in data.get("hosts") or []:
        if not isinstance(host, dict):
            continue
        try:
            if int(host.get("id")) == int(user_id):
                return True
        except (TypeError, ValueError):
            pass
    for judge in data.get("judges") or []:
        if not isinstance(judge, dict) or judge.get("role") == "table":
            continue
        try:
            if int(judge.get("user_id") or judge.get("id")) == int(user_id):
                return True
        except (TypeError, ValueError):
            pass
    user = await database.fetch_one(
        select(beach_users.c.badges, beach_users.c.roles).where(
            beach_users.c.id == user_id
        )
    )
    invited_team_ids = {
        int(value)
        for value in data.get("invited_team_ids") or []
        if isinstance(value, int) or str(value).isdigit()
    }
    roles = _json(user["roles"], []) if user else []
    if isinstance(roles, list) and any(
        isinstance(role, dict)
        and role.get("type") == "coach"
        and (
            isinstance(role.get("team_id"), int)
            or str(role.get("team_id")).isdigit()
        )
        and int(role.get("team_id")) in invited_team_ids
        for role in roles
    ):
        return True
    capabilities = await resolve_user_capabilities(user["badges"] if user else [])
    return bool(
        {"tournament.docs.use", "tournament.actAsHostEverywhere"}
        & set(capabilities)
    )


async def enabled_categories_for_season(season_id: str) -> List[str]:
    if str(season_id).startswith("unassigned:"):
        return []
    row = await database.fetch_one(
        select(beach_mp_eligibility_settings.c.enabled_categories).where(
            beach_mp_eligibility_settings.c.season_id == str(season_id)
        )
    )
    if not row:
        return (
            list(DEFAULT_ENABLED_CATEGORIES)
            if str(season_id) == DEFAULT_SEASON_ID
            else []
        )
    values = _json(row["enabled_categories"], [])
    return [value for value in values if value in VALID_CATEGORIES]


async def _active_snapshot(tournament_id: int, team_id: int) -> Optional[Dict[str, Any]]:
    row = await database.fetch_one(
        select(beach_tournament_protocol_snapshots)
        .where(
            and_(
                beach_tournament_protocol_snapshots.c.tournament_id == tournament_id,
                beach_tournament_protocol_snapshots.c.team_id == team_id,
                beach_tournament_protocol_snapshots.c.is_active.is_(True),
            )
        )
        .order_by(beach_tournament_protocol_snapshots.c.revision.desc())
    )
    return dict(row) if row else None


async def create_protocol_snapshot(
    tournament_row: Dict[str, Any],
    team_id: int,
    *,
    source: str,
    reason: Optional[str] = None,
    actor_user_id: Optional[int] = None,
    force_revision: bool = False,
    now: Optional[datetime] = None,
) -> Optional[Dict[str, Any]]:
    tournament_id = int(tournament_row["id"])
    data = _json(tournament_row.get("data_json"))
    first = first_team_matches(tournament_row, data).get(int(team_id))
    if not first:
        return None
    first_match_at, scheduled_team_name = first
    now_utc = (now or datetime.now(timezone.utc)).astimezone(timezone.utc)
    active = await _active_snapshot(tournament_id, int(team_id))
    if active and not force_revision:
        return active

    squads = data.get("team_squads") if isinstance(data.get("team_squads"), dict) else {}
    squad_entry = squads.get(str(team_id)) if isinstance(squads.get(str(team_id)), dict) else {}
    player_ids = effective_protocol_ids(squad_entry)
    team_row = await database.fetch_one(
        select(
            beach_teams.c.team_name,
            beach_teams.c.gender,
        ).where(beach_teams.c.id == int(team_id))
    )
    team_name = str(team_row["team_name"]) if team_row else scheduled_team_name
    gender = str(team_row["gender"]) if team_row and team_row["gender"] else None
    revision = (int(active["revision"]) + 1) if active else 1
    actor_name = await get_actor_name(actor_user_id) if actor_user_id else None

    async with database.transaction():
        if active:
            await database.execute(
                update(beach_tournament_protocol_snapshots)
                .where(beach_tournament_protocol_snapshots.c.id == int(active["id"]))
                .values(is_active=False)
            )
        new_id = await database.execute(
            insert(beach_tournament_protocol_snapshots).values(
                tournament_id=tournament_id,
                team_id=int(team_id),
                season_id=tournament_season_id(tournament_row),
                category=tournament_row.get("category"),
                gender=gender,
                team_name=team_name,
                revision=revision,
                first_match_at=first_match_at,
                frozen_at=now_utc,
                protocol_player_ids=player_ids,
                source=source,
                reason=reason,
                frozen_by_id=actor_user_id,
                frozen_by_name=actor_name,
                supersedes_snapshot_id=int(active["id"]) if active else None,
                is_active=True,
            )
        )
    row = await database.fetch_one(
        select(beach_tournament_protocol_snapshots).where(
            beach_tournament_protocol_snapshots.c.id == int(new_id)
        )
    )
    return dict(row) if row else None


async def ensure_due_snapshot(
    tournament_row: Dict[str, Any],
    team_id: int,
    *,
    now: Optional[datetime] = None,
) -> Optional[Dict[str, Any]]:
    if str(tournament_row.get("competition_type") or "").upper() != "MP":
        return None
    if tournament_phase(tournament_row) != "elimination":
        return None
    data = _json(tournament_row.get("data_json"))
    first = first_team_matches(tournament_row, data).get(int(team_id))
    if not first:
        return None
    now_utc = (now or datetime.now(timezone.utc)).astimezone(timezone.utc)
    if first[0] > now_utc:
        return None
    source = "legacy_reconstructed" if first[0] < _rollout_at() else "auto"
    return await create_protocol_snapshot(
        tournament_row,
        int(team_id),
        source=source,
        reason=(
            "Rekonstrukcja danych historycznych po wdrożeniu mechanizmu zamrażania"
            if source == "legacy_reconstructed"
            else None
        ),
        now=now_utc,
    )


async def assert_protocol_edit_allowed(
    tournament_row: Dict[str, Any],
    team_id: int,
    current_user_id: int,
) -> None:
    """Freeze the old source before a post-cutoff edit and reject non-authorities."""
    season_id = tournament_season_id(tournament_row)
    if tournament_row.get("category") not in await enabled_categories_for_season(
        season_id
    ):
        return
    snapshot = await ensure_due_snapshot(tournament_row, team_id)
    if not snapshot:
        return
    data = _json(tournament_row.get("data_json"))
    if await _is_admin(current_user_id) or _is_head_judge(data, current_user_id):
        return
    raise HTTPException(
        status_code=423,
        detail={
            "code": "MP_PROTOCOL_SNAPSHOT_LOCKED",
            "message": (
                "Lista „Do protokołu” została utrwalona z chwilą pierwszego meczu. "
                "Może ją odświeżyć wyłącznie sędzia główny lub administrator."
            ),
            "snapshot_revision": snapshot.get("revision"),
            "frozen_at": snapshot.get("frozen_at"),
        },
    )


def _played_schedule_matches(data: Dict[str, Any]) -> List[Dict[str, Any]]:
    result: List[Dict[str, Any]] = []
    for match in _schedule_matches(data):
        if str(match.get("status") or "").lower() != "finished":
            continue
        if _is_match_cancelled(match):
            continue
        if not (_regular_team(match.get("teamA")) and _regular_team(match.get("teamB"))):
            continue
        result.append(match)
    return result


def _proel_config(data_json: Any) -> Dict[str, Any]:
    data = _json(data_json)
    config = data.get("matchConfig")
    return config if isinstance(config, dict) else {}


def _proel_player_ids(config: Dict[str, Any], team_id: int) -> List[int]:
    host_id = config.get("hostTeamId")
    guest_id = config.get("guestTeamId")
    try:
        is_host = int(host_id) == int(team_id)
    except (TypeError, ValueError):
        is_host = False
    try:
        is_guest = int(guest_id) == int(team_id)
    except (TypeError, ValueError):
        is_guest = False
    players = config.get("hostPlayers") if is_host else config.get("guestPlayers") if is_guest else []
    return _numeric_ids(
        [
            player.get("player_id")
            for player in players if isinstance(player, dict) and player.get("selected") is not False
        ]
    )


def _match_team_ids(match: Dict[str, Any]) -> List[int]:
    return [
        parsed[0]
        for parsed in (_regular_team(match.get("teamA")), _regular_team(match.get("teamB")))
        if parsed
    ]


async def _source_evidence(
    season_id: str,
    category: str,
) -> Tuple[Dict[str, Dict[int, Dict[str, Any]]], List[Dict[str, Any]]]:
    rows = await database.fetch_all(
        select(beach_tournaments).where(
            and_(
                func.upper(beach_tournaments.c.competition_type) == "MP",
                beach_tournaments.c.category == category,
            )
        )
    )
    evidence_by_gender: Dict[str, Dict[int, Dict[str, Any]]] = {}
    source_summaries: List[Dict[str, Any]] = []

    for raw in rows:
        tournament = dict(raw)
        if tournament_season_id(tournament) != str(season_id):
            continue
        if tournament_phase(tournament) != "elimination":
            continue
        data = _json(tournament.get("data_json"))
        played = _played_schedule_matches(data)
        if not played:
            continue
        proel_rows = await database.fetch_all(
            select(beach_proel_matches).where(
                beach_proel_matches.c.tournament_id == int(tournament["id"])
            )
        )
        proel_by_match: Dict[str, Dict[str, Any]] = {}
        for raw_proel in proel_rows:
            item = dict(raw_proel)
            schedule_id = item.get("schedule_match_id")
            if schedule_id:
                proel_by_match[str(schedule_id)] = item

        team_ids = sorted({team_id for match in played for team_id in _match_team_ids(match)})
        team_rows = await database.fetch_all(
            select(
                beach_teams.c.id,
                beach_teams.c.team_name,
                beach_teams.c.gender,
            ).where(beach_teams.c.id.in_(team_ids))
        ) if team_ids else []
        team_meta = {int(row["id"]): dict(row) for row in team_rows}

        for team_id in team_ids:
            team_matches = [m for m in played if team_id in _match_team_ids(m)]
            completed = 0
            proel_sources: List[Tuple[str, Dict[str, Any]]] = []
            for match in team_matches:
                match_id = str(match.get("id") or "")
                proel = proel_by_match.get(match_id)
                status = str(proel.get("status") if proel else "").lower()
                if status not in ("approved", "finished"):
                    continue
                completed += 1
                config = _proel_config(proel.get("data_json"))
                for player_id in _proel_player_ids(config, team_id):
                    proel_sources.append((status, {
                        "tournament_id": int(tournament["id"]),
                        "tournament_name": tournament.get("name"),
                        "team_id": team_id,
                        "team_name": team_meta.get(team_id, {}).get("team_name"),
                        "match_id": match_id,
                        "match_number": config.get("matchNumber") or match.get("matchNumber"),
                        "source": "proel",
                        "proel_status": status,
                    }))
                    gender = str(team_meta.get(team_id, {}).get("gender") or "")
                    bucket = evidence_by_gender.setdefault(gender, {})
                    candidate_status = "approved_proel" if status == "approved" else "finished_proel"
                    _merge_evidence(bucket, player_id, candidate_status, proel_sources[-1][1])

            coverage_complete = bool(team_matches) and completed == len(team_matches)
            snapshot = None
            if not coverage_complete:
                snapshot = await _active_snapshot(int(tournament["id"]), team_id)
                if snapshot:
                    frozen_ids = _numeric_ids(_json(snapshot.get("protocol_player_ids"), []))
                    frozen_status = (
                        "frozen_confirmed"
                        if len(frozen_ids) <= 10
                        and snapshot.get("source") != "legacy_reconstructed"
                        else "frozen_warning"
                    )
                    gender = str(snapshot.get("gender") or team_meta.get(team_id, {}).get("gender") or "")
                    bucket = evidence_by_gender.setdefault(gender, {})
                    for player_id in frozen_ids:
                        _merge_evidence(bucket, player_id, frozen_status, {
                            "tournament_id": int(tournament["id"]),
                            "tournament_name": tournament.get("name"),
                            "team_id": team_id,
                            "team_name": snapshot.get("team_name") or team_meta.get(team_id, {}).get("team_name"),
                            "source": "frozen_list",
                            "snapshot_id": snapshot.get("id"),
                            "snapshot_revision": snapshot.get("revision"),
                            "snapshot_source": snapshot.get("source"),
                            "frozen_at": snapshot.get("frozen_at"),
                            "players_count": len(frozen_ids),
                        })
            source_summaries.append({
                "tournament_id": int(tournament["id"]),
                "tournament_name": tournament.get("name"),
                "team_id": team_id,
                "team_name": team_meta.get(team_id, {}).get("team_name"),
                "played_matches": len(team_matches),
                "completed_proel_matches": completed,
                "proel_coverage_complete": coverage_complete,
                "fallback_snapshot_used": bool(snapshot),
            })
    return evidence_by_gender, source_summaries


def _merge_evidence(
    bucket: Dict[int, Dict[str, Any]],
    player_id: int,
    status: str,
    source: Dict[str, Any],
) -> None:
    current = bucket.get(player_id)
    if current is None:
        bucket[player_id] = {"status": status, "sources": [source]}
        return
    current["sources"].append(source)
    if STATUS_PRIORITY[status] > STATUS_PRIORITY[current["status"]]:
        current["status"] = status


def _roster_players(raw: Any) -> List[Dict[str, Any]]:
    result: List[Dict[str, Any]] = []
    for player in _json(raw, []):
        if not isinstance(player, dict):
            continue
        try:
            player_id = int(player.get("player_id"))
        except (TypeError, ValueError):
            continue
        result.append({
            "player_id": player_id,
            "first_name": player.get("first_name") or "",
            "last_name": player.get("last_name") or "",
            "jersey_number": player.get("jersey_number"),
            "photo_url": player.get("photo_url"),
        })
    return result


async def build_tournament_report(tournament_id: int) -> Dict[str, Any]:
    row = await database.fetch_one(
        select(beach_tournaments).where(beach_tournaments.c.id == tournament_id)
    )
    if not row:
        raise HTTPException(404, "Nie znaleziono turnieju")
    tournament = dict(row)
    data = _json(tournament.get("data_json"))
    season_id = tournament_season_id(tournament)
    category = str(tournament.get("category") or "")
    is_mp = str(tournament.get("competition_type") or "").upper() == "MP"
    is_final = tournament_phase(tournament) == "final"
    category_enabled = category in await enabled_categories_for_season(season_id)
    enforcement_active = is_mp and is_final and category_enabled

    evidence_by_gender, source_summaries = await _source_evidence(season_id, category)
    target_ids: set[int] = set()
    for raw_id in data.get("invited_team_ids") or []:
        try:
            target_ids.add(int(raw_id))
        except (TypeError, ValueError):
            pass
    for match in _schedule_matches(data):
        target_ids.update(_match_team_ids(match))
    # A newly-created final may not have invited teams yet.  The report must
    # still show the historical elimination pool instead of an empty screen.
    target_ids.update(
        int(source["team_id"])
        for source in source_summaries
        if source.get("team_id") is not None
    )
    team_rows = await database.fetch_all(
        select(
            beach_teams.c.id,
            beach_teams.c.team_name,
            beach_teams.c.gender,
            beach_teams.c.roster_json,
        ).where(beach_teams.c.id.in_(sorted(target_ids)))
    ) if target_ids else []

    squads = data.get("team_squads") if isinstance(data.get("team_squads"), dict) else {}
    teams: List[Dict[str, Any]] = []
    counts = {key: 0 for key in STATUS_PRIORITY}
    for raw_team in team_rows:
        team = dict(raw_team)
        team_id = int(team["id"])
        gender = str(team.get("gender") or "")
        entry = squads.get(str(team_id)) if isinstance(squads.get(str(team_id)), dict) else {}
        selected_ids = set(effective_protocol_ids(entry))
        if not selected_ids:
            selected_ids.update(_numeric_ids(entry.get("default_players")))
        players: List[Dict[str, Any]] = []
        for player in _roster_players(team.get("roster_json")):
            evidence = evidence_by_gender.get(gender, {}).get(player["player_id"])
            status = (
                evidence["status"]
                if evidence
                else "missing" if source_summaries else "insufficient"
            )
            counts[status] += 1
            players.append({
                **player,
                "selected_for_protocol": player["player_id"] in selected_ids,
                "status": status,
                "eligible": status in (
                    "approved_proel",
                    "finished_proel",
                    "frozen_confirmed",
                    "frozen_warning",
                ),
                "requires_warning_acceptance": status == "frozen_warning",
                "sources": evidence["sources"] if evidence else [],
            })
        teams.append({
            "team_id": team_id,
            "team_name": team.get("team_name"),
            "gender": gender,
            "players": players,
        })
    return {
        "tournament_id": tournament_id,
        "tournament_name": tournament.get("name"),
        "season_id": season_id,
        "category": category,
        "is_mp": is_mp,
        "is_final": is_final,
        "category_enabled": category_enabled,
        "enforcement_active": enforcement_active,
        "preview_only": is_mp and category_enabled and not is_final,
        "disclaimer": (
            "Lista zgłoszeniowa „Do protokołu” jest dowodem administracyjnym, "
            "a nie stuprocentowym potwierdzeniem fizycznego udziału w meczu."
        ),
        "counts": counts,
        "teams": sorted(teams, key=lambda item: str(item.get("team_name") or "")),
        "sources": source_summaries,
    }


def _filtered_export_teams(
    report: Dict[str, Any],
    body: MpReportExportRequest,
) -> List[Dict[str, Any]]:
    query = body.query.strip().casefold()
    result: List[Dict[str, Any]] = []
    for team in report.get("teams") or []:
        if body.gender != "all" and str(team.get("gender") or "") != body.gender:
            continue
        players = []
        for player in team.get("players") or []:
            status = str(player.get("status") or "")
            if body.status and status != body.status:
                continue
            if body.only_issues and status not in (
                "missing",
                "insufficient",
                "frozen_warning",
            ):
                continue
            searchable = " ".join(
                (
                    str(player.get("first_name") or ""),
                    str(player.get("last_name") or ""),
                    str(team.get("team_name") or ""),
                )
            ).casefold()
            if query and query not in searchable:
                continue
            players.append(dict(player))
        if players:
            result.append({**dict(team), "players": players})
    return result


def _safe_export_name(value: Any) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    cleaned = "".join(
        char if char.isalnum() or char in ("-", "_") else "_"
        for char in ascii_value
    )
    return "_".join(part for part in cleaned.split("_") if part)[:60]


def _filtered_export_counts(teams: Sequence[Dict[str, Any]]) -> Dict[str, int]:
    counts = {status: 0 for status in MP_STATUS_LABELS}
    for team in teams:
        for player in team.get("players") or []:
            status = str(player.get("status") or "")
            if status in counts:
                counts[status] += 1
    return counts


def _player_export_name(player: Dict[str, Any]) -> str:
    return " ".join(
        part
        for part in (
            str(player.get("last_name") or "").upper(),
            str(player.get("first_name") or ""),
        )
        if part
    )


def _player_source_label(player: Dict[str, Any]) -> str:
    labels: List[str] = []
    for source in player.get("sources") or []:
        if source.get("source") == "proel":
            labels.append(
                "ProEl: "
                + str(
                    source.get("match_number")
                    or source.get("match_id")
                    or "mecz"
                )
            )
        else:
            labels.append(
                "Lista zgłoszeniowa: "
                + str(source.get("tournament_name") or "turniej eliminacyjny")
            )
    return " | ".join(dict.fromkeys(labels))


def _export_filter_label(body: MpReportExportRequest) -> str:
    values: List[str] = []
    if body.status:
        values.append(MP_STATUS_LABELS[body.status])
    if body.only_issues:
        values.append("tylko uwagi")
    if body.gender != "all":
        values.append("kobiety" if body.gender == "K" else "mężczyźni")
    if body.query.strip():
        values.append(f'szukaj: „{body.query.strip()}”')
    return " · ".join(values) or "pełny raport"


def _write_mp_report_xlsx(
    path: Path,
    report: Dict[str, Any],
    teams: List[Dict[str, Any]],
    body: MpReportExportRequest,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Występy w MP"
    accent = "20C6BE"
    dark = "17212B"
    muted = "64748B"
    soft = "E8F8F7"
    border = Border(bottom=Side(style="thin", color="D9E2E8"))

    sheet.merge_cells("A1:G1")
    sheet["A1"] = "Sprawdź występy w MP"
    sheet["A1"].font = Font(size=18, bold=True, color=dark)
    sheet.merge_cells("A2:G2")
    sheet["A2"] = (
        f'{report.get("tournament_name") or "Turniej"} · '
        f'{report.get("category") or ""} · '
        f'{season_year_label(report.get("season_id"))}'
    )
    sheet["A2"].font = Font(size=11, bold=True, color=muted)
    sheet.merge_cells("A3:G3")
    sheet["A3"] = f"Zakres: {_export_filter_label(body)}"
    sheet["A3"].font = Font(size=10, italic=True, color=muted)

    headers = [
        "Sekcja",
        "Zawodnik",
        "Nr",
        "Do protokołu",
        "Status",
        "Kryterium",
        "Źródło",
    ]
    row_index = 5
    for cell_index, value in enumerate(headers, 1):
        cell = sheet.cell(row_index, cell_index, value)
        cell.fill = PatternFill("solid", fgColor=accent)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[row_index].height = 24

    for team in teams:
        row_index += 1
        sheet.merge_cells(
            start_row=row_index,
            start_column=1,
            end_row=row_index,
            end_column=7,
        )
        team_cell = sheet.cell(row_index, 1, str(team.get("team_name") or "Drużyna"))
        team_cell.fill = PatternFill("solid", fgColor=soft)
        team_cell.font = Font(size=12, bold=True, color=dark)
        team_cell.alignment = Alignment(vertical="center")
        sheet.row_dimensions[row_index].height = 23

        protocol = [
            player
            for player in team.get("players") or []
            if player.get("selected_for_protocol")
        ]
        remaining = [
            player
            for player in team.get("players") or []
            if not player.get("selected_for_protocol")
        ]
        for section_label, players in (
            ("Skład „Do protokołu”", protocol),
            ("Pozostali zawodnicy", remaining),
        ):
            if not players:
                continue
            for player_index, player in enumerate(players):
                row_index += 1
                values = [
                    section_label if player_index == 0 else "",
                    _player_export_name(player),
                    player.get("jersey_number"),
                    "TAK" if player.get("selected_for_protocol") else "NIE",
                    MP_STATUS_LABELS.get(
                        str(player.get("status") or ""),
                        str(player.get("status") or ""),
                    ),
                    "Spełnia" if player.get("eligible") else "Nie spełnia",
                    _player_source_label(player),
                ]
                for column, value in enumerate(values, 1):
                    cell = sheet.cell(row_index, column, value)
                    cell.border = border
                    cell.alignment = Alignment(
                        vertical="top",
                        wrap_text=column in (1, 5, 7),
                    )
                    if column == 6:
                        cell.font = Font(
                            bold=True,
                            color="16803A" if player.get("eligible") else "C9382A",
                        )

    if not teams:
        row_index += 1
        sheet.merge_cells(
            start_row=row_index,
            start_column=1,
            end_row=row_index,
            end_column=7,
        )
        sheet.cell(row_index, 1, "Brak wyników dla wybranych filtrów")

    widths = {
        "A": 25,
        "B": 30,
        "C": 8,
        "D": 15,
        "E": 32,
        "F": 14,
        "G": 48,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = "A6"
    sheet.sheet_view.showGridLines = False

    sources_sheet = workbook.create_sheet("Źródła eliminacyjne")
    source_headers = [
        "Turniej",
        "Drużyna",
        "Mecze rozegrane",
        "ProEl zakończony",
        "Pełne pokrycie ProEl",
        "Lista zgłoszeniowa",
    ]
    for column, value in enumerate(source_headers, 1):
        cell = sources_sheet.cell(1, column, value)
        cell.fill = PatternFill("solid", fgColor=accent)
        cell.font = Font(bold=True, color="FFFFFF")
    for row_number, source in enumerate(report.get("sources") or [], 2):
        values = [
            source.get("tournament_name"),
            source.get("team_name"),
            source.get("played_matches"),
            source.get("completed_proel_matches"),
            "TAK" if source.get("proel_coverage_complete") else "NIE",
            "TAK" if source.get("fallback_snapshot_used") else "NIE",
        ]
        for column, value in enumerate(values, 1):
            sources_sheet.cell(row_number, column, value)
    for column, width in zip("ABCDEF", (38, 30, 17, 18, 20, 22)):
        sources_sheet.column_dimensions[column].width = width
    sources_sheet.freeze_panes = "A2"
    sources_sheet.sheet_view.showGridLines = False
    workbook.save(path)


def _load_mp_report_logo() -> str:
    logo_path = MP_TEMPLATE_DIR / "baza_beach_logo.png"
    if not logo_path.exists():
        return ""
    try:
        return base64.b64encode(logo_path.read_bytes()).decode("ascii")
    except Exception:
        logger.exception("Could not load BAZA Beach logo for MP report")
        return ""


def _mp_pdf_player_source(source: Dict[str, Any]) -> Dict[str, str]:
    if source.get("source") == "proel":
        match_label = (
            source.get("match_number")
            or source.get("match_id")
            or "mecz"
        )
        return {
            "kind": "proel",
            "label": (
                f'{source.get("tournament_name") or "Turniej eliminacyjny"}'
                f" · ProEl {match_label}"
            ),
        }
    return {
        "kind": "list",
        "label": (
            f'{source.get("tournament_name") or "Turniej eliminacyjny"}'
            " · lista zgłoszeniowa do turnieju"
        ),
    }


def _build_mp_pdf_context(
    report: Dict[str, Any],
    teams: List[Dict[str, Any]],
    body: MpReportExportRequest,
) -> Dict[str, Any]:
    counts = _filtered_export_counts(teams)
    category = str(report.get("category") or "")
    accent = MP_CATEGORY_COLORS.get(category, "#20AFA8")
    team_context: List[Dict[str, Any]] = []

    for team in teams:
        protocol = [
            player
            for player in team.get("players") or []
            if player.get("selected_for_protocol")
        ]
        remaining = [
            player
            for player in team.get("players") or []
            if not player.get("selected_for_protocol")
        ]

        def player_context(player: Dict[str, Any]) -> Dict[str, Any]:
            status = str(player.get("status") or "")
            return {
                "name": _player_export_name(player),
                "jersey_number": player.get("jersey_number"),
                "eligible": bool(player.get("eligible")),
                "status_label": MP_STATUS_LABELS.get(status, status),
                "status_color": MP_STATUS_COLORS.get(status, "#7A7F89"),
                "sources": [
                    _mp_pdf_player_source(dict(source))
                    for source in player.get("sources") or []
                    if isinstance(source, dict)
                ],
            }

        protocol_context = [player_context(player) for player in protocol]
        remaining_context = [player_context(player) for player in remaining]
        sections = []
        if protocol_context:
            sections.append({
                "label": "Skład „Do protokołu”",
                "primary": True,
                "players": protocol_context,
            })
        if remaining_context:
            sections.append({
                "label": "Pozostali zawodnicy",
                "primary": False,
                "players": remaining_context,
            })
        gender = str(team.get("gender") or "")
        team_context.append({
            "team_name": team.get("team_name") or "Drużyna",
            "gender_label": (
                "Kobiety"
                if gender == "K"
                else "Mężczyźni" if gender == "M" else "Drużyna"
            ),
            "gender_color": (
                "#E85A78"
                if gender == "K"
                else "#2BA8A0" if gender == "M" else accent
            ),
            "protocol": protocol_context,
            "remaining": remaining_context,
            "sections": sections,
        })

    sources = []
    for source in report.get("sources") or []:
        if not isinstance(source, dict):
            continue
        complete = bool(source.get("proel_coverage_complete"))
        sources.append({
            "team_name": source.get("team_name") or "Drużyna",
            "tournament_name": source.get("tournament_name") or "Turniej eliminacyjny",
            "played_matches": int(source.get("played_matches") or 0),
            "completed_proel_matches": int(
                source.get("completed_proel_matches") or 0
            ),
            "fallback_snapshot_used": bool(
                source.get("fallback_snapshot_used")
            ),
            "color": "#20AFA8" if complete else "#347FD1",
        })

    return {
        "tournament_name": report.get("tournament_name") or "Turniej",
        "category": category,
        "season_year": season_year_label(report.get("season_id")),
        "accent": accent,
        "filter_label": _export_filter_label(body),
        "players_count": sum(
            len(team.get("players") or []) for team in teams
        ),
        "teams_count": len(teams),
        "status_cards": [
            {
                "status": status,
                "label": MP_STATUS_SHORT_LABELS[status],
                "color": MP_STATUS_COLORS[status],
                "count": counts[status],
            }
            for status in MP_STATUS_LABELS
        ],
        "teams": team_context,
        "sources": sources,
        "disclaimer": report.get("disclaimer") or "",
        "generated_at": datetime.now(WARSAW).strftime("%d.%m.%Y %H:%M"),
        "logo_b64": _load_mp_report_logo(),
    }


def _write_mp_report_pdf(
    path: Path,
    report: Dict[str, Any],
    teams: List[Dict[str, Any]],
    body: MpReportExportRequest,
) -> None:
    from jinja2 import Environment, FileSystemLoader, select_autoescape
    import weasyprint

    template_path = MP_TEMPLATE_DIR / MP_REPORT_TEMPLATE
    if not template_path.exists():
        raise FileNotFoundError(f"Brak szablonu: {MP_REPORT_TEMPLATE}")
    environment = Environment(
        loader=FileSystemLoader(str(MP_TEMPLATE_DIR)),
        autoescape=select_autoescape(("html", "xml")),
    )
    template = environment.get_template(MP_REPORT_TEMPLATE)
    document = template.render(**_build_mp_pdf_context(report, teams, body))
    weasyprint.HTML(
        string=document,
        base_url=str(MP_TEMPLATE_DIR),
    ).write_pdf(str(path))


async def validate_final_player_ids(
    tournament_row: Dict[str, Any],
    team_id: int,
    player_ids: Sequence[int],
    accepted_warning_ids: Sequence[int] = (),
) -> None:
    if str(tournament_row.get("competition_type") or "").upper() != "MP":
        return
    season_id = tournament_season_id(tournament_row)
    if tournament_phase(tournament_row) != "final":
        return
    if tournament_row.get("category") not in await enabled_categories_for_season(season_id):
        return
    report = await build_tournament_report(int(tournament_row["id"]))
    target_team = next(
        (team for team in report["teams"] if int(team["team_id"]) == int(team_id)),
        None,
    )
    by_id = {
        int(player["player_id"]): player
        for player in (target_team.get("players") if target_team else [])
    }
    missing: List[int] = []
    warning: List[int] = []
    accepted = set(_numeric_ids(list(accepted_warning_ids)))
    for player_id in _numeric_ids(list(player_ids)):
        player = by_id.get(player_id)
        if not player or not player.get("eligible"):
            missing.append(player_id)
        elif player.get("requires_warning_acceptance") and player_id not in accepted:
            warning.append(player_id)
    if missing:
        raise HTTPException(
            409,
            detail={
                "code": "MP_APPEARANCE_REQUIRED",
                "message": "Zawodnik nie ma potwierdzonego występu w eliminacjach MP w tym sezonie.",
                "player_ids": missing,
            },
        )
    if warning:
        raise HTTPException(
            409,
            detail={
                "code": "MP_APPEARANCE_WARNING_ACCEPTANCE_REQUIRED",
                "message": (
                    "Podstawa kwalifikacji pochodzi z listy zgłoszeniowej do turnieju. "
                    "Potwierdź ostrzeżenie, aby dodać zawodnika."
                ),
                "player_ids": warning,
            },
        )


async def assert_no_unidentified_final_players(
    tournament_row: Dict[str, Any],
    extra_players: Sequence[Any],
) -> None:
    selected_extras = [
        player
        for player in extra_players
        if not isinstance(player, dict) or player.get("selected") is not False
    ]
    if not selected_extras:
        return
    if str(tournament_row.get("competition_type") or "").upper() != "MP":
        return
    if tournament_phase(tournament_row) != "final":
        return
    season_id = tournament_season_id(tournament_row)
    if tournament_row.get("category") not in await enabled_categories_for_season(season_id):
        return
    raise HTTPException(
        409,
        detail={
            "code": "MP_IDENTIFIED_PLAYER_REQUIRED",
            "message": (
                "W finale MP z aktywną weryfikacją można dodać wyłącznie "
                "zidentyfikowanego zawodnika z bazy ZPRP."
            ),
        },
    )


def _embedded_proel_link(data_json: Any) -> Tuple[Optional[int], Optional[str]]:
    data = _json(data_json)
    candidates: List[Dict[str, Any]] = []
    config = data.get("matchConfig")
    if isinstance(config, dict) and isinstance(config.get("extras"), dict):
        candidates.append(config["extras"])
    if isinstance(data.get("extras"), dict):
        candidates.append(data["extras"])
    candidates.append(data)
    tournament_id: Optional[int] = None
    schedule_match_id: Optional[str] = None
    for source in candidates:
        raw_tournament_id = (
            source.get("tournamentId")
            if source.get("tournamentId") is not None
            else source.get("tournament_id")
        )
        if tournament_id is None and raw_tournament_id is not None:
            try:
                tournament_id = int(raw_tournament_id)
            except (TypeError, ValueError):
                pass
        raw_schedule_id = (
            source.get("scheduleMatchId")
            if source.get("scheduleMatchId") is not None
            else source.get("schedule_match_id")
        )
        if schedule_match_id is None and raw_schedule_id is not None:
            schedule_match_id = str(raw_schedule_id)
    return tournament_id, schedule_match_id


def _normalized_match_number(value: Any) -> str:
    return str(value or "").strip().casefold()


async def run_mp_historical_backfill(*, force: bool = False) -> Dict[str, Any]:
    """One-time reconstruction for tournaments created before MP snapshots.

    It never claims that a reconstructed list was frozen at the historical
    cutoff.  Such revisions remain labelled ``legacy_reconstructed``.
    """
    marker = "mp_historical_backfill_v2"
    await database.execute(
        """CREATE TABLE IF NOT EXISTS beach_migrations (
               name VARCHAR PRIMARY KEY,
               ran_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
           )"""
    )
    if not force:
        completed = await database.fetch_one(
            "SELECT name FROM beach_migrations WHERE name = :name",
            {"name": marker},
        )
        if completed:
            return {"success": True, "skipped": True, "marker": marker}

    tournament_rows = [
        dict(row) for row in await database.fetch_all(select(beach_tournaments))
    ]
    tournaments_updated = 0
    schedule_links: Dict[
        str, List[Tuple[int, str]]
    ] = {}
    schedule_links_by_tournament: Dict[
        Tuple[int, str], Tuple[int, str]
    ] = {}
    schedule_links_by_id: Dict[str, List[Tuple[int, str]]] = {}

    for tournament in tournament_rows:
        values: Dict[str, Any] = {}
        if not tournament.get("season_id"):
            event_year = _parse_event_date(tournament.get("event_date")).year
            derived_season_id = event_year - 2018
            if derived_season_id > 0:
                values["season_id"] = str(derived_season_id)
                tournament["season_id"] = str(derived_season_id)
        inferred_phase = infer_mp_phase(
            tournament.get("name"),
            tournament.get("competition_type"),
            tournament.get("mp_phase"),
        )
        if inferred_phase and not tournament.get("mp_phase"):
            values["mp_phase"] = inferred_phase
            tournament["mp_phase"] = inferred_phase
        if values:
            await database.execute(
                update(beach_tournaments)
                .where(beach_tournaments.c.id == int(tournament["id"]))
                .values(**values)
            )
            tournaments_updated += 1

        data = _json(tournament.get("data_json"))
        for match in _schedule_matches(data):
            match_id = match.get("id")
            if match_id is None:
                continue
            link = (int(tournament["id"]), str(match_id))
            schedule_links_by_id.setdefault(str(match_id), []).append(link)
            match_number = _normalized_match_number(
                match.get("matchNumber") or match.get("match_number")
            )
            if not match_number:
                continue
            schedule_links.setdefault(match_number, []).append(link)
            schedule_links_by_tournament[
                (int(tournament["id"]), match_number)
            ] = link

    proel_links_updated = 0
    proel_rows = await database.fetch_all(select(beach_proel_matches))
    for raw_proel in proel_rows:
        proel = dict(raw_proel)
        embedded_tournament_id, embedded_match_id = _embedded_proel_link(
            proel.get("data_json")
        )
        # The immutable link embedded in the historical match state is more
        # authoritative than columns populated by an earlier heuristic.
        tournament_id = embedded_tournament_id or proel.get("tournament_id")
        schedule_match_id = embedded_match_id or proel.get("schedule_match_id")
        config = _proel_config(proel.get("data_json"))
        match_number = _normalized_match_number(
            config.get("matchNumber") or proel.get("match_number")
        )

        if tournament_id is not None and not schedule_match_id and match_number:
            scoped = schedule_links_by_tournament.get(
                (int(tournament_id), match_number)
            )
            if scoped:
                schedule_match_id = scoped[1]
        if tournament_id is None and match_number:
            candidates = schedule_links.get(match_number) or []
            if len(candidates) == 1:
                tournament_id, schedule_match_id = candidates[0]
        if schedule_match_id and tournament_id is None:
            candidates = schedule_links_by_id.get(str(schedule_match_id)) or []
            if len(candidates) == 1:
                tournament_id = candidates[0][0]

        values = {}
        if tournament_id is not None and proel.get("tournament_id") != int(
            tournament_id
        ):
            values["tournament_id"] = int(tournament_id)
        if schedule_match_id is not None and str(
            proel.get("schedule_match_id") or ""
        ) != str(schedule_match_id):
            values["schedule_match_id"] = str(schedule_match_id)
        if values:
            await database.execute(
                update(beach_proel_matches)
                .where(
                    beach_proel_matches.c.match_number
                    == proel["match_number"]
                )
                .values(**values)
            )
            proel_links_updated += 1

    snapshots_created = 0
    snapshots_rebuilt = 0
    now_utc = datetime.now(timezone.utc)
    for tournament in tournament_rows:
        if str(tournament.get("competition_type") or "").upper() != "MP":
            continue
        if tournament_phase(tournament) != "elimination":
            continue
        data = _json(tournament.get("data_json"))
        squads = (
            data.get("team_squads")
            if isinstance(data.get("team_squads"), dict)
            else {}
        )
        for team_id, (first_match_at, _) in first_team_matches(
            tournament, data
        ).items():
            if first_match_at > now_utc:
                continue
            active = await _active_snapshot(int(tournament["id"]), team_id)
            entry = (
                squads.get(str(team_id))
                if isinstance(squads.get(str(team_id)), dict)
                else {}
            )
            current_ids = effective_protocol_ids(entry)
            if not active:
                created = await create_protocol_snapshot(
                    tournament,
                    team_id,
                    source="legacy_reconstructed",
                    reason=(
                        "Jednorazowa rekonstrukcja historycznych turniejów MP"
                    ),
                    now=now_utc,
                )
                if created:
                    snapshots_created += 1
                continue
            active_ids = _numeric_ids(
                _json(active.get("protocol_player_ids"), [])
            )
            if not active_ids and current_ids:
                rebuilt = await create_protocol_snapshot(
                    tournament,
                    team_id,
                    source="legacy_reconstructed",
                    reason=(
                        "Odbudowa pustego historycznego snapshotu z zachowanej "
                        "listy „Do protokołu”"
                    ),
                    force_revision=True,
                    now=now_utc,
                )
                if rebuilt:
                    snapshots_rebuilt += 1

    await database.execute(
        """INSERT INTO beach_migrations (name, ran_at)
           VALUES (:name, NOW())
           ON CONFLICT (name) DO UPDATE SET ran_at = EXCLUDED.ran_at""",
        {"name": marker},
    )
    return {
        "success": True,
        "skipped": False,
        "marker": marker,
        "tournaments_updated": tournaments_updated,
        "proel_links_updated": proel_links_updated,
        "snapshots_created": snapshots_created,
        "snapshots_rebuilt": snapshots_rebuilt,
    }


async def freeze_due_protocol_snapshots(now: Optional[datetime] = None) -> int:
    rows = await database.fetch_all(
        select(beach_tournaments).where(
            func.upper(beach_tournaments.c.competition_type) == "MP"
        )
    )
    frozen = 0
    for raw in rows:
        tournament = dict(raw)
        if tournament_phase(tournament) != "elimination":
            continue
        data = _json(tournament.get("data_json"))
        for team_id in first_team_matches(tournament, data):
            before = await _active_snapshot(int(tournament["id"]), team_id)
            after = await ensure_due_snapshot(tournament, team_id, now=now)
            if after and not before:
                frozen += 1
    return frozen


async def run_mp_snapshot_scheduler() -> None:
    while True:
        try:
            count = await freeze_due_protocol_snapshots()
            if count:
                logger.info("MP protocol snapshots frozen: %d", count)
        except asyncio.CancelledError:
            raise
        except Exception:
            logger.exception("MP protocol snapshot scheduler failed")
        await asyncio.sleep(30)


@router.get("/settings", response_model=dict)
async def get_mp_settings(
    current_user_id: int = Depends(beach_get_current_user_id),
):
    del current_user_id
    team_seasons = await database.fetch_all(
        select(beach_teams.c.season_id, beach_teams.c.season)
        .where(beach_teams.c.season_id.is_not(None))
        .distinct()
    )
    tournament_seasons = await database.fetch_all(
        select(beach_tournaments.c.season_id)
        .where(beach_tournaments.c.season_id.is_not(None))
        .distinct()
    )
    setting_rows = await database.fetch_all(select(beach_mp_eligibility_settings))
    configured = {str(row["season_id"]): dict(row) for row in setting_rows}
    seasons: Dict[str, Dict[str, Any]] = {
        str(row["season_id"]): {
            "season_id": str(row["season_id"]),
            "label": season_year_label(row["season_id"]),
        }
        for row in team_seasons if row["season_id"] is not None
    }
    for row in tournament_seasons:
        season_id = str(row["season_id"])
        seasons.setdefault(
            season_id,
            {"season_id": season_id, "label": season_year_label(season_id)},
        )
    seasons.setdefault(DEFAULT_SEASON_ID, {
        "season_id": DEFAULT_SEASON_ID,
        "label": season_year_label(DEFAULT_SEASON_ID),
    })
    return {
        "categories": VALID_CATEGORIES,
        "seasons": [
            {
                **meta,
                "enabled_categories": (
                    _json(configured[season_id]["enabled_categories"], [])
                    if season_id in configured
                    else (
                        list(DEFAULT_ENABLED_CATEGORIES)
                        if season_id == DEFAULT_SEASON_ID
                        else []
                    )
                ),
                "updated_at": configured.get(season_id, {}).get("updated_at"),
            }
            for season_id, meta in sorted(seasons.items(), reverse=True)
        ],
    }


@router.put("/settings/{season_id}", response_model=dict)
async def update_mp_settings(
    season_id: str,
    body: MpSettingsUpdate,
    current_user_id: int = Depends(beach_get_current_user_id),
):
    if not await _is_admin(current_user_id):
        raise HTTPException(403, "Tylko administrator może zmienić kryteria MP")
    invalid = [value for value in body.enabled_categories if value not in VALID_CATEGORIES]
    if invalid:
        raise HTTPException(400, f"Nieprawidłowe kategorie: {', '.join(invalid)}")
    categories = [value for value in VALID_CATEGORIES if value in set(body.enabled_categories)]
    actor_name = await get_actor_name(current_user_id)
    existing = await database.fetch_one(
        select(beach_mp_eligibility_settings).where(
            beach_mp_eligibility_settings.c.season_id == season_id
        )
    )
    if existing:
        await database.execute(
            update(beach_mp_eligibility_settings)
            .where(beach_mp_eligibility_settings.c.season_id == season_id)
            .values(
                enabled_categories=categories,
                updated_by_id=current_user_id,
                updated_by_name=actor_name,
                updated_at=datetime.now(timezone.utc),
            )
        )
    else:
        await database.execute(
            insert(beach_mp_eligibility_settings).values(
                season_id=season_id,
                enabled_categories=categories,
                updated_by_id=current_user_id,
                updated_by_name=actor_name,
            )
        )
    await log_activity(
        area="tournament",
        action="mp_appearances.settings_updated",
        actor_user_id=current_user_id,
        actor_name=actor_name,
        target_id=season_id,
        target_label=f"Sezon {season_id}",
        details={"enabled_categories": categories},
    )
    return {"season_id": season_id, "enabled_categories": categories}


@router.post("/historical-backfill", response_model=dict)
async def trigger_historical_backfill(
    current_user_id: int = Depends(beach_get_current_user_id),
):
    if not await _is_admin(current_user_id):
        raise HTTPException(
            403,
            "Jednorazowy backfill historyczny może uruchomić tylko administrator",
        )
    result = await run_mp_historical_backfill(force=True)
    await log_activity(
        area="tournament",
        action="mp_appearances.historical_backfill",
        actor_user_id=current_user_id,
        actor_name=await get_actor_name(current_user_id),
        target_id="all",
        target_label="Historyczne turnieje MP",
        details=result,
    )
    return result


@router.get("/tournaments/{tournament_id}/report", response_model=dict)
async def get_tournament_report(
    tournament_id: int,
    current_user_id: int = Depends(beach_get_current_user_id),
):
    row = await database.fetch_one(
        select(beach_tournaments).where(beach_tournaments.c.id == tournament_id)
    )
    if not row:
        raise HTTPException(404, "Nie znaleziono turnieju")
    if not await _can_view_report(dict(row), current_user_id):
        raise HTTPException(403, "Brak dostępu do dokumentów tego turnieju")
    return await build_tournament_report(tournament_id)


@router.post("/tournaments/{tournament_id}/export", response_model=dict)
async def generate_tournament_report_export(
    tournament_id: int,
    body: MpReportExportRequest,
    current_user_id: int = Depends(beach_get_current_user_id),
):
    row = await database.fetch_one(
        select(beach_tournaments).where(beach_tournaments.c.id == tournament_id)
    )
    if not row:
        raise HTTPException(404, "Nie znaleziono turnieju")
    if not await _can_view_report(dict(row), current_user_id):
        raise HTTPException(403, "Brak dostępu do dokumentów tego turnieju")

    report = await build_tournament_report(tournament_id)
    teams = _filtered_export_teams(report, body)
    MP_EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    token = str(uuid.uuid4())
    export_path = MP_EXPORT_DIR / f"{token}.{body.format}"
    try:
        if body.format == "pdf":
            _write_mp_report_pdf(export_path, report, teams, body)
        else:
            _write_mp_report_xlsx(export_path, report, teams, body)
    except Exception as error:
        export_path.unlink(missing_ok=True)
        logger.exception("MP appearance report export failed")
        raise HTTPException(500, f"Nie udało się wygenerować raportu: {error}")

    safe_tournament = _safe_export_name(report.get("tournament_name")) or "turniej"
    filename = (
        f"wystepy_mp_{safe_tournament}_{season_year_label(report.get('season_id'))}"
        f".{body.format}"
    )
    return {
        "success": True,
        "format": body.format,
        "download_url": (
            f"/beach/mp-appearances/export/download/{token}"
            f"?format={body.format}&filename={urllib.parse.quote(filename)}"
        ),
        "teams_count": len(teams),
        "players_count": sum(len(team.get("players") or []) for team in teams),
    }


@router.get("/export/download/{token}")
async def download_tournament_report_export(
    token: str,
    format: Literal["pdf", "xlsx"] = Query(...),
    filename: str = Query("wystepy_mp"),
):
    try:
        uuid.UUID(token)
    except ValueError:
        raise HTTPException(400, "Nieprawidłowy token")
    export_path = MP_EXPORT_DIR / f"{token}.{format}"
    if not export_path.exists():
        raise HTTPException(404, "Plik wygasł lub nie istnieje")
    media_type = (
        "application/pdf"
        if format == "pdf"
        else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    download_name = Path(filename).name
    if not download_name.lower().endswith(f".{format}"):
        download_name = f"{download_name}.{format}"
    return FileResponse(
        path=export_path,
        media_type=media_type,
        filename=download_name,
        background=BackgroundTask(lambda: export_path.unlink(missing_ok=True)),
    )


@router.post(
    "/tournaments/{tournament_id}/teams/{team_id}/snapshot/refresh",
    response_model=dict,
)
async def refresh_protocol_snapshot(
    tournament_id: int,
    team_id: int,
    body: SnapshotRefreshRequest,
    current_user_id: int = Depends(beach_get_current_user_id),
):
    row = await database.fetch_one(
        select(beach_tournaments).where(beach_tournaments.c.id == tournament_id)
    )
    if not row:
        raise HTTPException(404, "Nie znaleziono turnieju")
    tournament = dict(row)
    data = _json(tournament.get("data_json"))
    can_refresh = await _is_admin(current_user_id) or _is_head_judge(
        data, current_user_id
    )
    if not can_refresh and body.context_tournament_id:
        context_row = await database.fetch_one(
            select(beach_tournaments).where(
                beach_tournaments.c.id == body.context_tournament_id
            )
        )
        if context_row:
            context = dict(context_row)
            context_data = _json(context.get("data_json"))
            can_refresh = (
                _is_head_judge(context_data, current_user_id)
                and str(context.get("competition_type") or "").upper() == "MP"
                and tournament_phase(context) == "final"
                and tournament_season_id(context)
                == tournament_season_id(tournament)
                and context.get("category") == tournament.get("category")
            )
    if not can_refresh:
        raise HTTPException(
            403,
            "Listę zgłoszeniową może odświeżyć tylko administrator lub właściwy sędzia główny",
        )
    snapshot = await create_protocol_snapshot(
        tournament,
        team_id,
        source="manual_refresh",
        reason=body.reason.strip(),
        actor_user_id=current_user_id,
        force_revision=True,
    )
    if not snapshot:
        raise HTTPException(400, "Drużyna nie ma zaplanowanego meczu w tym turnieju")
    await log_activity(
        area="tournament",
        action="mp_appearances.snapshot_refreshed",
        actor_user_id=current_user_id,
        actor_name=await get_actor_name(current_user_id),
        target_id=str(tournament_id),
        target_label=str(tournament.get("name") or ""),
        details={
            "team_id": team_id,
            "revision": snapshot.get("revision"),
            "reason": body.reason.strip(),
            "context_tournament_id": body.context_tournament_id,
            "players": len(_json(snapshot.get("protocol_player_ids"), [])),
        },
    )
    return {
        "success": True,
        "snapshot_id": snapshot.get("id"),
        "revision": snapshot.get("revision"),
        "frozen_at": snapshot.get("frozen_at"),
    }
