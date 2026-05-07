"""
Generuje PDF terminarza turnieju beach handball z szablonu xlsx.

Szablon (app/templates/szablon_terminarz.xlsx):
  C2  – nazwa turnieju
  C3  – daty (np. "14–16.05.2026")
  C4  – lokalizacja
  E8  – nagłówek dnia, np. "DZIEŃ I | 14.05.2026 (poniedziałek)"
  Wiersz 10+ – mecze:
      A=Godzina, B=Boisko, C=Kategoria, D=Etap/Grupa,
      E=Drużyna A, F=Drużyna B, G=Wynik
"""
import copy
import logging
import os
import shutil
import subprocess
import tempfile
import urllib.parse
import uuid
import zipfile
from collections import defaultdict
from datetime import date
from io import BytesIO
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, HTTPException, Path as ApiPath, Query
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from pydantic import BaseModel
from sqlalchemy import select
from starlette.background import BackgroundTask

from app.db import database, beach_tournaments, beach_users

logger = logging.getLogger(__name__)

router = APIRouter(tags=["Beach Schedule PDF"])

TEMPLATE_PATH = os.path.normpath(
    os.path.join(os.path.dirname(__file__), "..", "templates", "szablon_terminarz.xlsx")
)

DOWNLOAD_DIR = "/tmp/schedule_pdf_downloads"

FIRST_MATCH_ROW = 10   # row where first match goes
MAX_ROWS_PER_SHEET = 30  # matches per sheet before starting an overflow sheet

WEEKDAYS_PL = [
    "poniedziałek", "wtorek", "środa",
    "czwartek", "piątek", "sobota", "niedziela",
]
_ROMAN = {
    1: "I", 2: "II", 3: "III", 4: "IV", 5: "V",
    6: "VI", 7: "VII", 8: "VIII", 9: "IX", 10: "X",
}


# ─── request model ────────────────────────────────────────────────────────────

class SchedulePdfRequest(BaseModel):
    schedule: Dict[str, Any]
    tournament_name: str = ""
    tournament_location: str = ""
    tournament_dates: str = ""  # optional override; computed from days if empty
    tournament_id: Optional[int] = None   # used to look up judge/host emails
    exclude_user_id: Optional[int] = None  # current user — excluded from recipients


# ─── helpers ──────────────────────────────────────────────────────────────────

def _roman(n: int) -> str:
    return _ROMAN.get(n, str(n))


def _day_header(day_index: int, date_str: Optional[str]) -> str:
    roman = _roman(day_index + 1)
    if date_str:
        try:
            d = date.fromisoformat(date_str)
            weekday = WEEKDAYS_PL[d.weekday()]
            formatted = d.strftime("%d.%m.%Y")
            return f"DZIEŃ {roman} | {formatted} ({weekday})"
        except Exception:
            pass
    return f"DZIEŃ {roman}"


def _compute_date_range(days: List[Dict[str, Any]]) -> str:
    dates = sorted(d.get("date", "") for d in days if d.get("date"))
    if not dates:
        return ""
    try:
        d0 = date.fromisoformat(dates[0])
        d1 = date.fromisoformat(dates[-1])
        if d0 == d1:
            return d0.strftime("%d.%m.%Y")
        return f"{d0.strftime('%d.%m.%Y')} \u2013 {d1.strftime('%d.%m.%Y')}"
    except Exception:
        return f"{dates[0]} \u2013 {dates[-1]}" if len(dates) > 1 else dates[0]


def _stage_label(m: Dict[str, Any]) -> str:
    stage = m.get("stage", "")
    group = m.get("group") or ""
    if stage == "group":
        return "Każdy z każdym"
    return {
        "quarterfinal": "Ćwierćfinał",
        "semifinal": "Półfinał",
        "final": "Finał",
        "third_place": "3. miejsce",
        "fifth_place": "5. miejsce",
        "seventh_place": "7. miejsce",
        "fifth_semifinal": "Półfinał o 5.",
    }.get(stage, stage)


def _category_label(m: Dict[str, Any]) -> str:
    g = m.get("gender", "")
    return "M" if g == "M" else "K" if g == "K" else ""


def _team_name(team: Optional[Dict[str, Any]]) -> str:
    if team and team.get("name"):
        return str(team["name"])
    return ""


def _score_str(m: Dict[str, Any]) -> str:
    a, b = m.get("scoreA"), m.get("scoreB")
    if a is None or b is None:
        return ""
    base = f"{a}:{b}"
    sets = m.get("sets") or []
    if sets:
        parts = [
            f"{s.get('ptA', '')}:{s.get('ptB', '')}"
            for s in sets
            if isinstance(s, dict)
        ]
        if parts:
            return f"{base} ({', '.join(parts)})"
    return base


def _load_media(template_path: str) -> Dict[str, bytes]:
    """Extract xl/media/* entries from the xlsx zip for image rehydration."""
    media: Dict[str, bytes] = {}
    try:
        with zipfile.ZipFile(template_path, "r") as z:
            for name in z.namelist():
                if name.startswith("xl/media/"):
                    media[name] = z.read(name)
    except Exception:
        pass
    return media


def _rehydrate_images(wb, media: Dict[str, bytes]) -> None:
    """Re-attach images that openpyxl loaded but lost the raw data for."""
    for ws in wb.worksheets:
        imgs = list(getattr(ws, "_images", []) or [])
        ws._images = []
        for img in imgs:
            path = (getattr(img, "path", "") or "").lstrip("/")
            blob = media.get(path)
            if not blob:
                continue
            bio = BytesIO(blob)
            new_img = Image(bio)
            new_img.width = img.width
            new_img.height = img.height
            new_img.anchor = copy.deepcopy(img.anchor)
            ws.add_image(new_img)


def _fill_header(
    ws, *, name: str, dates: str, location: str, day_hdr: str
) -> None:
    ws["C2"] = name
    ws["C3"] = dates
    ws["C4"] = location
    ws["E8"] = day_hdr


def _fill_match_row(ws, row: int, m: Dict[str, Any]) -> None:
    ws.cell(row=row, column=1).value = m.get("startTime") or ""   # A – Godzina
    ws.cell(row=row, column=2).value = m.get("court") or ""        # B – Boisko
    ws.cell(row=row, column=3).value = _category_label(m)          # C – Kategoria
    ws.cell(row=row, column=4).value = _stage_label(m)             # D – Etap/Grupa
    ws.cell(row=row, column=5).value = _team_name(m.get("teamA"))  # E – Drużyna A
    ws.cell(row=row, column=6).value = _team_name(m.get("teamB"))  # F – Drużyna B
    ws.cell(row=row, column=7).value = _score_str(m)               # G – Wynik


def _convert_xlsx_to_pdf(xlsx_path: str, out_dir: str) -> str:
    """Convert xlsx → pdf using LibreOffice. Returns the pdf path."""
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        raise RuntimeError(
            "Brak LibreOffice (soffice) w środowisku. Doinstaluj libreoffice w Dockerfile."
        )

    env = os.environ.copy()
    env.setdefault("HOME", "/tmp")
    env.setdefault("XDG_CACHE_HOME", "/tmp")
    env.setdefault("XDG_CONFIG_HOME", "/tmp")

    profile_dir = os.path.join(out_dir, "lo_profile")
    os.makedirs(profile_dir, exist_ok=True)

    cmd = [
        soffice,
        "--headless", "--nologo", "--nolockcheck",
        "--nodefault", "--norestore",
        f"-env:UserInstallation=file://{profile_dir}",
        "--convert-to", "pdf",
        "--outdir", out_dir,
        xlsx_path,
    ]

    proc = subprocess.run(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        env=env,
        timeout=90,
    )

    if proc.returncode != 0:
        raise RuntimeError(
            f"LibreOffice convert failed (code={proc.returncode}). "
            f"stderr={proc.stderr[:500]}"
        )

    base = os.path.splitext(os.path.basename(xlsx_path))[0]
    pdf_path = os.path.join(out_dir, base + ".pdf")

    if not os.path.exists(pdf_path):
        pdfs = [p for p in os.listdir(out_dir) if p.lower().endswith(".pdf")]
        if len(pdfs) == 1:
            pdf_path = os.path.join(out_dir, pdfs[0])
        elif pdfs:
            cand = [p for p in pdfs if os.path.splitext(p)[0] == base]
            pdf_path = os.path.join(out_dir, (cand or pdfs)[0])
        else:
            raise RuntimeError("PDF nie znaleziony po konwersji LibreOffice.")

    return pdf_path


# ─── endpoint ─────────────────────────────────────────────────────────────────

def _ensure_download_dir():
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)


async def _get_judge_host_emails(
    tournament_id: int,
    exclude_user_id: Optional[int],
) -> List[str]:
    """Return emails of all judges + hosts for the tournament, excluding exclude_user_id."""
    try:
        row = await database.fetch_one(
            select(beach_tournaments.c.data_json).where(
                beach_tournaments.c.id == tournament_id
            )
        )
        if not row:
            return []
        data_json = row["data_json"]
        if isinstance(data_json, str):
            import json as _json
            data_json = _json.loads(data_json)
        if not isinstance(data_json, dict):
            return []

        host_ids = {
            int(h["id"])
            for h in (data_json.get("hosts") or [])
            if isinstance(h, dict) and h.get("id") is not None
        }
        judge_ids = {
            int(j["id"])
            for j in (data_json.get("judges") or [])
            if isinstance(j, dict) and j.get("id") is not None
        }
        head_judge_id = data_json.get("head_judge_id")
        if isinstance(head_judge_id, int):
            judge_ids.add(head_judge_id)

        all_ids = host_ids | judge_ids
        if exclude_user_id is not None:
            all_ids.discard(int(exclude_user_id))

        if not all_ids:
            return []

        rows = await database.fetch_all(
            select(beach_users.c.email).where(
                beach_users.c.id.in_(list(all_ids))
            )
        )
        return [r["email"] for r in rows if r["email"]]
    except Exception:
        logger.warning("Could not fetch judge/host emails for tournament %s", tournament_id, exc_info=True)
        return []


@router.post("/beach/schedule/pdf", summary="Generuj PDF terminarza turnieju")
async def generate_schedule_pdf(req: SchedulePdfRequest):
    if not os.path.exists(TEMPLATE_PATH):
        raise HTTPException(500, detail=f"Brak szablonu terminarza: {TEMPLATE_PATH}")

    schedule = req.schedule
    matches: List[Dict[str, Any]] = schedule.get("matches") or []
    config: Dict[str, Any] = schedule.get("config") or {}
    days: List[Dict[str, Any]] = config.get("days") or []

    tournament_name = req.tournament_name.strip()
    date_range = req.tournament_dates.strip() or _compute_date_range(days)
    location = req.tournament_location.strip()

    # ── group matches by day, sort by time ──
    by_day: Dict[int, List[Dict[str, Any]]] = defaultdict(list)
    for m in matches:
        by_day[int(m.get("dayIndex") or 0)].append(m)
    for idx in by_day:
        by_day[idx].sort(
            key=lambda m: (m.get("startTime") or "99:99", m.get("order") or 0)
        )

    day_indices = sorted(by_day.keys()) or [0]

    # ── load template ──
    media = _load_media(TEMPLATE_PATH)
    wb = load_workbook(TEMPLATE_PATH)
    _rehydrate_images(wb, media)
    template_ws = wb.active

    # ── build (day_idx, chunk_idx, total_chunks, match_list) tuples ──
    all_chunks: List[tuple] = []
    for day_idx in day_indices:
        day_matches = by_day.get(day_idx, [])
        if day_matches:
            chunks = [
                day_matches[i: i + MAX_ROWS_PER_SHEET]
                for i in range(0, len(day_matches), MAX_ROWS_PER_SHEET)
            ]
        else:
            chunks = [[]]  # empty day still gets one sheet
        for ci, chunk in enumerate(chunks):
            all_chunks.append((day_idx, ci, len(chunks), chunk))

    # ── create sheets: reuse template_ws as first, copy for the rest ──
    created: List[tuple] = []  # (ws, day_idx, ci, total, chunk)
    for i, (day_idx, ci, total, chunk) in enumerate(all_chunks):
        ws = template_ws if i == 0 else wb.copy_worksheet(template_ws)
        created.append((ws, day_idx, ci, total, chunk))

    # ── fill sheets ──
    for ws, day_idx, ci, total, chunk in created:
        day_cfg = days[day_idx] if day_idx < len(days) else {}
        day_hdr = _day_header(day_idx, day_cfg.get("date"))
        if total > 1:
            day_hdr += f" cz. {ci + 1}"

        roman = _roman(day_idx + 1)
        ws.title = f"Dzień {roman}" + (f" ({ci + 1})" if total > 1 else "")

        _fill_header(
            ws,
            name=tournament_name,
            dates=date_range,
            location=location,
            day_hdr=day_hdr,
        )
        # repeat header rows on each printed page
        ws.print_title_rows = f"1:{FIRST_MATCH_ROW - 1}"

        for row_i, m in enumerate(chunk):
            _fill_match_row(ws, FIRST_MATCH_ROW + row_i, m)

    # ── save xlsx and convert to pdf ──
    tmp_dir = tempfile.mkdtemp()
    try:
        xlsx_path = os.path.join(tmp_dir, "terminarz.xlsx")
        wb.save(xlsx_path)
        pdf_path = _convert_xlsx_to_pdf(xlsx_path, tmp_dir)

        safe_name = (
            "".join(c if c.isalnum() or c in " _-" else "_" for c in tournament_name)[:40]
            or "terminarz"
        )
        download_name = f"terminarz_{safe_name}.pdf"

        # ── save to download dir with a one-time token ──
        _ensure_download_dir()
        token = str(uuid.uuid4())
        download_path = os.path.join(DOWNLOAD_DIR, f"{token}.pdf")
        shutil.copyfile(pdf_path, download_path)
        shutil.rmtree(tmp_dir, ignore_errors=True)  # clean working dir

        encoded_name = urllib.parse.quote(download_name)

        # ── fetch judge/host emails if tournament_id provided ──
        judge_host_emails: List[str] = []
        if req.tournament_id:
            judge_host_emails = await _get_judge_host_emails(
                req.tournament_id, req.exclude_user_id
            )

        return {
            "success": True,
            "download_url": f"/beach/schedule/pdf/download/{token}?filename={encoded_name}",
            "judge_host_emails": judge_host_emails,
        }
    except HTTPException:
        raise
    except Exception as e:
        shutil.rmtree(tmp_dir, ignore_errors=True)
        logger.exception("Schedule PDF generation failed")
        raise HTTPException(500, detail=str(e))


@router.get(
    "/beach/schedule/pdf/download/{token}",
    summary="Pobierz wygenerowany PDF terminarza (attachment)",
)
async def download_schedule_pdf(
    token: str = ApiPath(...),
    filename: str = Query("terminarz.pdf"),
):
    _ensure_download_dir()
    # Validate token is a UUID to prevent path traversal
    try:
        uuid.UUID(token)
    except ValueError:
        raise HTTPException(400, "Nieprawidłowy token")
    file_path = os.path.join(DOWNLOAD_DIR, f"{token}.pdf")
    if not os.path.exists(file_path):
        raise HTTPException(404, "Plik wygasł lub nie istnieje")
    return FileResponse(
        path=file_path,
        media_type="application/pdf",
        filename=filename,
        background=BackgroundTask(
            lambda: os.remove(file_path) if os.path.exists(file_path) else None
        ),
    )
