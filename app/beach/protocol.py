"""
Generuje surowe protokoły meczowe beach handball z szablonu xlsx.

Szablon (app/templates/beach_protocol_template.xlsx):
  Kategoria (krzyżyk "X"):
    Senior M → M5,  Senior K → P5
    Junior M → M6,  Junior K → P6
    Junior mł. M → M7,  Junior mł. K → P7
    Młodzik M → M8,  Młodzik K → P8
  S6  – numer meczu
  B10 – nazwa drużyny gospodarzy
  G10 – nazwa drużyny gości
  B12 – miejsce zawodów (miasto)
  C12 – data (DD.MM.YYYY)
  F12 – godzina meczu
  Zawodnicy gospodarzy: wiersze 15-30  (A=numer, B=NAZWISKO Imię)  max 16
  Zawodnicy gości:      wiersze 38-53  (A=numer, B=NAZWISKO Imię)  max 16
  Osoby towarzyszące gospodarzy: wiersze 31-34 (B=pełne imię i nazwisko)
  Osoby towarzyszące gości:      wiersze 54-57 (B=pełne imię i nazwisko)
  Sędzia boiskowy 1:  B61 (Nazwisko Imię), C61 (miasto)
  Sędzia boiskowy 2:  B62 (Nazwisko Imię), C62 (miasto)
  Sekretarz:           B64 (Nazwisko Imię), C64 (miasto)
  Mierzący czas:       B65 (Nazwisko Imię), C65 (miasto)
  Sędzia główny/delegat: B67 (Nazwisko Imię), C67 (miasto)
"""
import json as _json
import logging
import math
import os
import shutil
import tempfile
import urllib.parse
import uuid
import zipfile
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, HTTPException, Path as ApiPath, Query
from fastapi.responses import FileResponse
from httpx import AsyncClient
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from sqlalchemy import select
from starlette.background import BackgroundTask

from app.db import database, beach_teams, beach_tournaments, beach_users, beach_proel_matches
from app.beach.schedule_pdf import _convert_xlsx_to_pdf

try:
    from PIL import Image as PILImage
except ImportError:
    PILImage = None

logger = logging.getLogger(__name__)

router = APIRouter(tags=["Beach Protocol"])

TEMPLATE_PATH = os.path.normpath(
    os.path.join(os.path.dirname(__file__), "..", "templates", "beach_protocol_template.xlsx")
)

DOWNLOAD_DIR = "/tmp/beach_protocol_downloads"

# ── category → cell mapping ──────────────────────────────────────────────────

CATEGORY_CELL: Dict[Tuple[str, str], str] = {
    ("Senior", "M"):       "M5",
    ("Senior", "K"):       "P5",
    ("Junior", "M"):       "M6",
    ("Junior", "K"):       "P6",
    ("Junior mł.", "M"):   "M7",
    ("Junior mł.", "K"):   "P7",
    ("Młodzik", "M"):      "M8",
    ("Młodzik", "K"):      "P8",
}

# ── host player rows 15-30, guest 38-53 ──
HOST_PLAYER_ROWS = list(range(15, 31))     # 15..30 inclusive (16 slots)
GUEST_PLAYER_ROWS = list(range(38, 54))    # 38..53 inclusive (16 slots)
HOST_COMPANION_ROWS = list(range(31, 35))  # 31..34 inclusive (4 slots)
GUEST_COMPANION_ROWS = list(range(54, 58)) # 54..57 inclusive (4 slots)

# ── companion role → row offset (A=0, B=1, C=2, D=3) ──
COMPANION_ROLE_OFFSET = {"A": 0, "B": 1, "C": 2, "D": 3}

# ── Signature helpers (shared with results.py) ────────────────────────────────

BACKEND_STATIC_PREFIX = "https://zprp-backend-production.up.railway.app"


def _full_static_url(rel_or_abs: str) -> str:
    s = (rel_or_abs or "").strip()
    if not s:
        return ""
    if s.startswith("http://") or s.startswith("https://"):
        return s
    if not s.startswith("/"):
        s = "/" + s
    return BACKEND_STATIC_PREFIX + s


async def _fetch_png_bytes(url: str) -> bytes:
    u = (url or "").strip()
    if not u:
        return b""
    try:
        async with AsyncClient(follow_redirects=True, timeout=15.0) as c:
            r = await c.get(u)
            if r.status_code != 200:
                return b""
            return r.content or b""
    except Exception:
        return b""


def _add_signature_image(
    ws,
    *,
    image_bytes: bytes,
    anchor_cell: str,
    max_width_px: int = 220,
    max_height_px: int = 90,
    offset_x_px: int = 0,
    offset_y_px: int = 0,
) -> bool:
    if not image_bytes:
        return False
    bio = BytesIO(image_bytes)
    img = XlImage(bio)
    if PILImage is not None:
        try:
            pil = PILImage.open(BytesIO(image_bytes))
            w, h = pil.size
            if w and h:
                scale = min(max_width_px / float(w), max_height_px / float(h), 1.0)
                img.width = int(w * scale)
                img.height = int(h * scale)
        except Exception:
            pass
    else:
        img.width = min(img.width or max_width_px, max_width_px)
        img.height = min(img.height or max_height_px, max_height_px)

    if offset_x_px or offset_y_px:
        from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
        from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
        from openpyxl.utils.units import pixels_to_EMU

        col_letter, row_num = coordinate_from_string(anchor_cell)
        col_idx = column_index_from_string(col_letter) - 1  # 0-based
        row_idx = row_num - 1  # 0-based

        marker = AnchorMarker(
            col=col_idx,
            colOff=pixels_to_EMU(offset_x_px),
            row=row_idx,
            rowOff=pixels_to_EMU(offset_y_px),
        )
        anchor = OneCellAnchor(_from=marker, ext=None)
        anchor.ext.cx = pixels_to_EMU(img.width)
        anchor.ext.cy = pixels_to_EMU(img.height)
        img.anchor = anchor
        ws.add_image(img)
    else:
        ws.add_image(img, anchor_cell)
    return True


# ── request models ────────────────────────────────────────────────────────────

class ProtocolBulkRequest(BaseModel):
    schedule: Dict[str, Any]
    tournament_name: str = ""
    tournament_location: str = ""
    tournament_id: Optional[int] = None
    category: str = ""               # "Senior", "Junior", "Junior mł", "Młodzik"
    format: str = "xlsx"             # "xlsx" | "pdf"


class ProtocolSingleRequest(BaseModel):
    match: Dict[str, Any]
    schedule_config: Dict[str, Any]  # config with days
    tournament_name: str = ""
    tournament_location: str = ""
    tournament_id: Optional[int] = None
    category: str = ""
    format: str = "xlsx"


class FilledProtocolSingleRequest(BaseModel):
    match_number: str
    tournament_id: Optional[int] = None
    format: str = "pdf"


class FilledProtocolBulkRequest(BaseModel):
    match_numbers: List[str]
    tournament_id: Optional[int] = None
    tournament_name: str = ""
    format: str = "pdf"


# ── helpers ───────────────────────────────────────────────────────────────────

def _ensure_download_dir():
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)


def _parse_data_json(raw: Any) -> dict:
    if isinstance(raw, dict):
        return raw
    if isinstance(raw, str):
        try:
            return _json.loads(raw)
        except Exception:
            return {}
    return {}


async def _fetch_tournament_data(tournament_id: int) -> dict:
    """Fetch data_json from beach_tournaments."""
    row = await database.fetch_one(
        select(beach_tournaments.c.data_json).where(
            beach_tournaments.c.id == tournament_id
        )
    )
    if not row:
        return {}
    return _parse_data_json(row["data_json"])


async def _fetch_team_rosters(team_ids: List[int]) -> Dict[int, dict]:
    """Fetch roster_json, companions_json, jersey_overrides, team_name for given team IDs."""
    if not team_ids:
        return {}
    rows = await database.fetch_all(
        select(
            beach_teams.c.id,
            beach_teams.c.team_name,
            beach_teams.c.roster_json,
            beach_teams.c.companions_json,
            beach_teams.c.jersey_overrides,
        ).where(beach_teams.c.id.in_(team_ids))
    )
    return {
        r["id"]: {
            "team_name": r["team_name"],
            "roster_json": r["roster_json"] or [],
            "companions_json": r["companions_json"] or [],
            "jersey_overrides": r["jersey_overrides"] or {},
        }
        for r in rows
    }


async def _fetch_user_cities(user_ids: List[int]) -> Dict[int, Tuple[str, str]]:
    """Fetch full_name and city for users by IDs. Returns {id: (full_name, city)}."""
    if not user_ids:
        return {}
    rows = await database.fetch_all(
        select(
            beach_users.c.id,
            beach_users.c.full_name,
            beach_users.c.city,
        ).where(beach_users.c.id.in_(user_ids))
    )
    return {
        r["id"]: (r["full_name"] or "", r["city"] or "")
        for r in rows
    }


def _format_name_protocol(full_name: str) -> str:
    """Format name as 'Nazwisko Imię' for referee cells."""
    parts = full_name.strip().split()
    if len(parts) >= 2:
        return f"{parts[-1]} {' '.join(parts[:-1])}"
    return full_name


def _format_player_name(last_name: str, first_name: str) -> str:
    """Format as 'NAZWISKO Imię'."""
    ln = (last_name or "").strip().upper()
    fn = (first_name or "").strip()
    return f"{ln} {fn}".strip()


def _safe_filename(s: str, max_len: int = 40) -> str:
    return "".join(c if c.isalnum() or c in " _-" else "_" for c in s)[:max_len].strip("_") or "protokol"


def _strikethrough_empty_rows(
    ws,
    rows: List[int],
    filled_count: int,
    *,
    is_companion: bool = False,
) -> None:
    """Diagonal cross on empty rows.

    Col B: dashes "-----------------------------------------------"
    Col C: empty (no diagonal)
    Companions: col D also empty (no diagonal)
    Remaining cols: diagonal border.
    """
    thin = Side(style="thin", color="000000")
    start_col = 2 if is_companion else 1  # B for companions, A for players
    end_col = 8  # H
    # Columns that stay empty (no diagonal): C=3, and D=4 only for companions
    skip_cols = {3, 4} if is_companion else {3}

    for idx in range(filled_count, len(rows)):
        row = rows[idx]
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            if col == 2:  # B – dashes
                cell.value = "-----------------------------------------------"
            elif col in skip_cols:  # C (and D for companions) – empty
                cell.value = ""
            else:  # diagonal cross
                cell.value = ""
                cur = cell.border or Border()
                cell.border = Border(
                    left=cur.left,
                    right=cur.right,
                    top=cur.top,
                    bottom=cur.bottom,
                    diagonal=thin,
                    diagonalDown=True,
                    diagonalUp=False,
                )


def _match_sort_key(m: Dict[str, Any]) -> tuple:
    return (m.get("dayIndex") or 0, m.get("startTime") or "99:99", m.get("order") or 0)


def _match_file_label(m: Dict[str, Any], idx: int) -> str:
    """Generate a filename label for a match protocol."""
    mn = m.get("matchNumber")
    if mn:
        return mn.replace("/", "_").replace("\\", "_")
    # fallback: day + order
    day = (m.get("dayIndex") or 0) + 1
    order = m.get("order", idx)
    team_a = (m.get("teamA") or {}).get("name", "")
    team_b = (m.get("teamB") or {}).get("name", "")
    if team_a and team_b:
        return _safe_filename(f"{team_a}_vs_{team_b}", 50)
    return f"mecz_{idx + 1:03d}"


# ── core: fill a single protocol sheet ────────────────────────────────────────

def _fill_protocol_sheet(
    ws,
    match: Dict[str, Any],
    *,
    tournament_name: str,
    tournament_location: str,
    category: str,
    days: List[Dict[str, Any]],
    team_squads: Dict[str, Any],
    custom_teams: List[Dict[str, Any]],
    team_rosters: Dict[int, dict],
    user_cities: Dict[int, Tuple[str, str]],
    head_judge_id: Optional[int],
) -> None:
    """Fill all protocol cells for a single match."""

    gender = match.get("gender", "")

    # ── Category checkbox ──
    # Normalize category: strip whitespace, ensure "Junior mł." has dot
    cat_norm = category.strip()
    if cat_norm.lower().startswith("junior m") and "mł" in cat_norm.lower():
        cat_norm = "Junior mł."
    cell_key = (cat_norm, gender)
    cell_ref = CATEGORY_CELL.get(cell_key)
    if not cell_ref:
        # fallback: try without dot
        cell_ref = CATEGORY_CELL.get((cat_norm.rstrip("."), gender))
    if cell_ref:
        ws[cell_ref] = "X"

    # ── Match number ──
    match_number = match.get("matchNumber") or ""
    ws["S6"] = match_number

    # ── Team names ──
    team_a = match.get("teamA") or {}
    team_b = match.get("teamB") or {}
    ws["B10"] = team_a.get("name", "")
    ws["G10"] = team_b.get("name", "")

    # ── Venue, date, time ──
    ws["B12"] = tournament_location

    day_index = match.get("dayIndex") or 0
    day_cfg = days[day_index] if day_index < len(days) else {}
    date_str = day_cfg.get("date", "")
    if date_str:
        # Convert YYYY-MM-DD → DD.MM.YYYY
        try:
            parts = date_str.split("-")
            ws["C12"] = f"{parts[2]}.{parts[1]}.{parts[0]}"
        except (IndexError, ValueError):
            ws["C12"] = date_str
    else:
        ws["C12"] = ""

    ws["F12"] = match.get("startTime") or ""

    # ── Players & companions ──
    _fill_team_squad(
        ws,
        team_ref=team_a,
        player_rows=HOST_PLAYER_ROWS,
        companion_rows=HOST_COMPANION_ROWS,
        match_id=match.get("id", ""),
        team_squads=team_squads,
        custom_teams=custom_teams,
        team_rosters=team_rosters,
    )
    _fill_team_squad(
        ws,
        team_ref=team_b,
        player_rows=GUEST_PLAYER_ROWS,
        companion_rows=GUEST_COMPANION_ROWS,
        match_id=match.get("id", ""),
        team_squads=team_squads,
        custom_teams=custom_teams,
        team_rosters=team_rosters,
    )

    # ── Referees ──
    referees = match.get("referees") or {}
    _fill_referee(ws, "B61", "C61", referees.get("fieldA"), user_cities)
    _fill_referee(ws, "B62", "C62", referees.get("fieldB"), user_cities)
    _fill_referee(ws, "B64", "C64", referees.get("tableSecretary"), user_cities)
    _fill_referee(ws, "B65", "C65", referees.get("tableTimer"), user_cities)

    # ── Head judge / delegat ──
    if head_judge_id and head_judge_id in user_cities:
        name, city = user_cities[head_judge_id]
        ws["B67"] = _format_name_protocol(name)
        ws["C67"] = city


def _fill_referee(
    ws, name_cell: str, city_cell: str,
    ref: Optional[Dict[str, Any]],
    user_cities: Dict[int, Tuple[str, str]],
) -> None:
    if not ref:
        return
    ref_id = ref.get("id")
    if ref_id and ref_id in user_cities:
        name, city = user_cities[ref_id]
        ws[name_cell] = _format_name_protocol(name)
        ws[city_cell] = city
    elif ref.get("name"):
        ws[name_cell] = _format_name_protocol(ref["name"])


def _fill_team_squad(
    ws,
    *,
    team_ref: Dict[str, Any],
    player_rows: List[int],
    companion_rows: List[int],
    match_id: str,
    team_squads: Dict[str, Any],
    custom_teams: List[Dict[str, Any]],
    team_rosters: Dict[int, dict],
) -> None:
    """Fill player and companion rows for one team."""
    team_id = team_ref.get("id")
    if not team_id:
        return  # TBD team in knockout — leave empty

    team_id_str = str(team_id)

    # Check if it's a custom team (string ID starting with "ct_" or negative int)
    is_custom = isinstance(team_id, str) and team_id.startswith("ct_")
    if not is_custom:
        try:
            is_custom = int(team_id) < 0
        except (ValueError, TypeError):
            pass

    if is_custom:
        _fill_custom_team_squad(ws, team_id_str, player_rows, companion_rows, custom_teams)
    else:
        _fill_regular_team_squad(
            ws, int(team_id), team_id_str, player_rows, companion_rows,
            match_id, team_squads, team_rosters,
        )


def _fill_custom_team_squad(
    ws,
    team_id: str,
    player_rows: List[int],
    companion_rows: List[int],
    custom_teams: List[Dict[str, Any]],
) -> None:
    """Fill squad from custom_teams data (for manually created teams)."""
    ct = None
    for t in custom_teams:
        if str(t.get("id")) == team_id:
            ct = t
            break
    if not ct:
        return

    # Players: filter by defaultPlayers selection
    all_players = ct.get("players") or []
    default_ids = set(ct.get("defaultPlayers") or [])
    if default_ids:
        selected = [p for p in all_players if p.get("id") in default_ids]
    else:
        selected = all_players

    for i, row in enumerate(player_rows):
        if i < len(selected):
            p = selected[i]
            jersey = p.get("jerseyNumber", "")
            ws.cell(row=row, column=1).value = jersey
            ws.cell(row=row, column=2).value = _format_player_name(
                p.get("lastName", ""), p.get("firstName", "")
            )
    _strikethrough_empty_rows(ws, player_rows, len(selected))

    # Companions: filter by defaultCompanions selection
    all_companions = ct.get("companions") or []
    default_comp_ids = set(ct.get("defaultCompanions") or [])
    if default_comp_ids:
        selected_comp = [c for c in all_companions if c.get("id") in default_comp_ids]
    else:
        selected_comp = all_companions

    for i, row in enumerate(companion_rows):
        if i < len(selected_comp):
            c = selected_comp[i]
            ln = (c.get("lastName") or "").strip()
            fn = (c.get("firstName") or "").strip()
            ws.cell(row=row, column=2).value = f"{ln} {fn}".strip()
    _strikethrough_empty_rows(ws, companion_rows, len(selected_comp), is_companion=True)


def _fill_regular_team_squad(
    ws,
    team_id_int: int,
    team_id_str: str,
    player_rows: List[int],
    companion_rows: List[int],
    match_id: str,
    team_squads: Dict[str, Any],
    team_rosters: Dict[int, dict],
) -> None:
    """Fill squad from team_squads selection + roster_json data.

    Analogous to BeachNewMatchScreen.tsx buildPlayers/buildCompanions:
    1. Load ALL players from roster_json
    2. Apply jersey_overrides
    3. If squad selection exists (default_players / match_overrides) → only selected
    4. If no selection → all players from roster (capped at row slots)
    """
    squad_entry = team_squads.get(team_id_str) or {}
    roster_data = team_rosters.get(team_id_int)
    if not roster_data:
        return  # no roster data at all

    roster = roster_data.get("roster_json") or []
    jersey_overrides = roster_data.get("jersey_overrides") or {}
    companions = roster_data.get("companions_json") or []

    # Determine selected player IDs (match override → default → fallback all)
    match_overrides = squad_entry.get("match_overrides") or {}
    match_override = match_overrides.get(match_id) or {}
    selected_player_ids = match_override.get("players") or squad_entry.get("default_players") or []
    selected_companion_ids = match_override.get("companions") or squad_entry.get("default_companions") or []

    # ── Players ──
    if selected_player_ids:
        # Use only selected players, preserving selection order (only those in_squad)
        selected_id_set = set(selected_player_ids)
        id_to_player = {}
        for p in roster:
            pid = p.get("player_id")
            if pid is not None and pid in selected_id_set and pid not in id_to_player and p.get("in_squad", True):
                id_to_player[pid] = p
        ordered_players = [id_to_player[pid] for pid in selected_player_ids if pid in id_to_player]
    else:
        # No selection → fill all in-squad roster players (like BeachNewMatchScreen shows all)
        ordered_players = [p for p in roster if p.get("in_squad", True)]

    for i, row in enumerate(player_rows):
        if i < len(ordered_players):
            p = ordered_players[i]
            pid_str = str(p.get("player_id", ""))
            jersey = jersey_overrides.get(pid_str) or p.get("jersey_number") or ""
            ws.cell(row=row, column=1).value = jersey
            ws.cell(row=row, column=2).value = _format_player_name(
                p.get("last_name", ""), p.get("first_name", "")
            )
    _strikethrough_empty_rows(ws, player_rows, len(ordered_players))

    # ── Companions ──
    if selected_companion_ids:
        selected_comp_set = set(selected_companion_ids)
        id_to_comp = {}
        for c in companions:
            cid = c.get("person_id")
            if cid is not None and cid in selected_comp_set and cid not in id_to_comp:
                id_to_comp[cid] = c
        ordered_comps = [id_to_comp[cid] for cid in selected_companion_ids if cid in id_to_comp]
    else:
        # No selection → fill all companions
        ordered_comps = list(companions)

    for i, row in enumerate(companion_rows):
        if i < len(ordered_comps):
            c = ordered_comps[i]
            ws.cell(row=row, column=2).value = c.get("full_name", "")
    _strikethrough_empty_rows(ws, companion_rows, len(ordered_comps), is_companion=True)


# ── collect all referee IDs from matches ──────────────────────────────────────

def _collect_referee_ids(matches: List[Dict[str, Any]], head_judge_id: Optional[int]) -> List[int]:
    ids = set()
    for m in matches:
        refs = m.get("referees") or {}
        for slot in ("fieldA", "fieldB", "tableSecretary", "tableTimer"):
            ref = refs.get(slot)
            if ref and ref.get("id"):
                ids.add(int(ref["id"]))
    if head_judge_id:
        ids.add(int(head_judge_id))
    return list(ids)


def _collect_team_ids(matches: List[Dict[str, Any]]) -> List[int]:
    """Collect numeric team IDs (skip custom teams with string/negative IDs)."""
    ids = set()
    for m in matches:
        for slot in ("teamA", "teamB"):
            team = m.get(slot) or {}
            tid = team.get("id")
            if tid is None:
                continue
            try:
                tid_int = int(tid)
                if tid_int > 0:
                    ids.add(tid_int)
            except (ValueError, TypeError):
                pass
    return list(ids)


# ══════════════════════════════════════════════════════════════════════════════
# FILLED PROTOCOL — fill match results from ProEl data_json
# ══════════════════════════════════════════════════════════════════════════════

_THIN = Side(style="thin", color="000000")


def _diagonal_cross(ws, cell_ref: str) -> None:
    """Apply a diagonal-down border to a cell (visual cross-out)."""
    cell = ws[cell_ref]
    cur = cell.border or Border()
    cell.border = Border(
        left=cur.left, right=cur.right, top=cur.top, bottom=cur.bottom,
        diagonal=_THIN, diagonalDown=True, diagonalUp=False,
    )


def _diagonal_cross_rc(ws, row: int, col: int) -> None:
    """Apply a diagonal cross to a cell by row/col numbers."""
    cell = ws.cell(row=row, column=col)
    cur = cell.border or Border()
    cell.border = Border(
        left=cur.left, right=cur.right, top=cur.top, bottom=cur.bottom,
        diagonal=_THIN, diagonalDown=True, diagonalUp=False,
    )


def _check_mark(ws, cell_ref: str) -> None:
    """Write a check (X) into a cell."""
    ws[cell_ref] = "X"


def _fmt_timeout_minute(time_remaining_ms: int, set_time_sec: int = 600) -> str:
    """Convert timeRemainingMs to elapsed minute string like "8'" for the protocol."""
    elapsed_ms = (set_time_sec * 1000) - time_remaining_ms
    minute = max(1, math.ceil(elapsed_ms / 60_000))
    return f"{minute}'"


# ── Column mapping for set progressions ──────────────────────────────────────
# Each set has two column groups (main + overflow) for host and guest.
# Within each group: start_row and column letter.
# Main group covers pts 0-34 (rows 15-49), overflow covers pts 35-68 (rows 16-49).

_SET_COLUMNS = {
    1: {
        "host":  [("J", 15, 49), ("M", 16, 49)],  # main col, start_row, end_row
        "guest": [("L", 15, 49), ("O", 16, 49)],
    },
    2: {
        "host":  [("Q", 15, 49), ("T", 16, 49)],
        "guest": [("S", 15, 49), ("V", 16, 49)],
    },
}

_SET_RESULT_CELLS = {
    1: {"pts": "M51", "win": "M53"},
    2: {"pts": "T51", "win": "T53"},
}

# Shootout columns K(11) through T(20), rows by team:
_SHOOTOUT_COLS = list(range(11, 21))  # K=11 .. T=20 (10 rounds)
_SHOOTOUT_HOST_ROWS = {"shooter": 59, "gk": 60, "penalty": 61, "pts": 62}
_SHOOTOUT_GUEST_ROWS = {"shooter": 64, "gk": 65, "penalty": 66, "pts": 67}


def _build_player_points(protocol: List[Dict[str, Any]]) -> Dict[str, Dict[int, int]]:
    """Build {team: {jersey: total_points}} from protocol events.

    goal1pt = 1, goal2pt = 2, penaltyKickScored = 2.
    goalRemoved events are NOT subtracted (they indicate the goal was removed from the UI,
    but the protocol already shows them differently).
    """
    points: Dict[str, Dict[int, int]] = {"host": {}, "guest": {}}
    for ev in protocol:
        t = ev.get("type", "")
        team = ev.get("team", "")
        player = ev.get("player")
        if player is None or team not in points:
            continue
        if t == "goal1pt":
            points[team][player] = points[team].get(player, 0) + 1
        elif t in ("goal2pt", "penaltyKickScored"):
            points[team][player] = points[team].get(player, 0) + 2
    return points


def _build_exclusion_counts(protocol: List[Dict[str, Any]]) -> Dict[str, Dict[int, int]]:
    """Build {team: {jersey: count_of_exclusions}} from protocol events."""
    counts: Dict[str, Dict[int, int]] = {"host": {}, "guest": {}}
    for ev in protocol:
        if ev.get("type") != "exclusion":
            continue
        team = ev.get("team", "")
        player = ev.get("player")
        if player is None or team not in counts:
            continue
        counts[team][player] = counts[team].get(player, 0) + 1
    return counts


def _build_disqualified_set(protocol: List[Dict[str, Any]]) -> Dict[str, set]:
    """Build {team: set_of_jerseys_disqualified}."""
    result: Dict[str, set] = {"host": set(), "guest": set()}
    for ev in protocol:
        if ev.get("type") != "disqualification":
            continue
        team = ev.get("team", "")
        player = ev.get("player")
        if player is not None and team in result:
            result[team].add(player)
    return result


def _build_companion_penalties(
    protocol: List[Dict[str, Any]],
    companions_by_team: Dict[str, List[Dict[str, Any]]],
) -> Dict[str, Dict[str, Dict[str, bool]]]:
    """Build companion penalty info from both protocol events AND companion objects.

    Returns {team: {role_letter: {"exclusion": bool, "disqualification": bool}}}.

    Protocol events for companions have player=None and extra="companion:A" (etc).
    Companion objects can have exclusion1 and isDisqualified fields.
    """
    result: Dict[str, Dict[str, Dict[str, bool]]] = {"host": {}, "guest": {}}

    # From protocol events
    for ev in protocol:
        team = ev.get("team", "")
        player = ev.get("player")
        extra = ev.get("extra") or ""
        if player is not None or team not in result:
            continue
        if not extra.startswith("companion:"):
            continue
        role = extra.split(":", 1)[1].strip().upper()
        if role not in COMPANION_ROLE_OFFSET:
            continue
        entry = result[team].setdefault(role, {"exclusion": False, "disqualification": False})
        if ev.get("type") == "exclusion":
            entry["exclusion"] = True
        elif ev.get("type") == "disqualification":
            entry["disqualification"] = True

    # From companion objects (fallback / supplement)
    for team_key, comp_key in [("host", "hostCompanions"), ("guest", "guestCompanions")]:
        for comp in companions_by_team.get(team_key, []):
            role = (comp.get("role") or "").strip().upper()
            if role not in COMPANION_ROLE_OFFSET:
                continue
            entry = result[team_key].setdefault(role, {"exclusion": False, "disqualification": False})
            if comp.get("exclusion1"):
                entry["exclusion"] = True
            if comp.get("isDisqualified"):
                entry["disqualification"] = True

    return result


def _get_goal_events_for_set(protocol: List[Dict[str, Any]], set_num: int) -> List[Dict[str, Any]]:
    """Get all scoring events for a given set in order, excluding goalRemoved."""
    scoring_types = {"goal1pt", "goal2pt", "penaltyKickScored"}
    return [
        ev for ev in protocol
        if ev.get("set") == set_num
        and ev.get("type") in scoring_types
        and ev.get("player") is not None
    ]


def _fill_set_progression(ws, set_num: int, protocol: List[Dict[str, Any]]) -> None:
    """Fill the goal-by-goal progression columns for a set (1 or 2)."""
    goals = _get_goal_events_for_set(protocol, set_num)
    cols = _SET_COLUMNS[set_num]

    # Running score trackers — these track the row position in each column
    team_score = {"host": 0, "guest": 0}

    # For each team, track which column group we're in and current row
    team_cursors: Dict[str, Dict[str, Any]] = {}
    for team in ("host", "guest"):
        grp = cols[team]
        team_cursors[team] = {
            "groups": grp,
            "group_idx": 0,
            "pts": 0,  # cumulative points for this team
        }

    for ev in goals:
        team = ev.get("team", "")
        if team not in team_cursors:
            continue

        player = ev.get("player")
        ev_type = ev.get("type", "")
        extra = ev.get("extra") or ""

        pts = 1 if ev_type == "goal1pt" else 2
        cursor = team_cursors[team]
        cursor["pts"] += pts

        # Determine the cell: pts maps to row offset
        current_pts = cursor["pts"]

        # Find the right column group
        gi = cursor["group_idx"]
        grp = cursor["groups"]
        col_letter, start_row, end_row = grp[gi]
        max_pts_in_group = end_row - start_row + 1 if gi == 0 else end_row - start_row + 1

        # For group 0: pts 1 → row start_row (which represents score=1)
        #   The cell at start_row represents the team reaching score 1.
        #   Pts go 1..35 in group 0 (rows 15..49 = 35 rows)
        #   Pts go 36..69 in group 1 (rows 16..49 = 34 rows)
        if gi == 0:
            threshold = end_row - start_row + 1  # 35
            if current_pts > threshold and len(grp) > 1:
                gi = 1
                cursor["group_idx"] = 1
                col_letter, start_row, end_row = grp[gi]
                row = start_row + (current_pts - threshold - 1)
            else:
                row = start_row + (current_pts - 1)
        else:
            threshold_prev = grp[0][2] - grp[0][1] + 1
            row = start_row + (current_pts - threshold_prev - 1)

        if row > end_row:
            continue  # safety: exceed max rows

        label = str(player)
        if extra.lower() == "gg":
            label += " GG"

        ws[f"{col_letter}{row}"] = label

    # Cross out the cell after the last entry for each team
    for team in ("host", "guest"):
        cursor = team_cursors[team]
        current_pts = cursor["pts"]
        if current_pts == 0:
            continue  # no goals, nothing to cross

        gi = cursor["group_idx"]
        grp = cursor["groups"]
        col_letter, start_row, end_row = grp[gi]

        if gi == 0:
            next_row = start_row + current_pts
        else:
            threshold_prev = grp[0][2] - grp[0][1] + 1
            next_row = start_row + (current_pts - threshold_prev)

        if next_row <= end_row:
            _diagonal_cross(ws, f"{col_letter}{next_row}")


def _fill_shootout(ws, data_json: Dict[str, Any]) -> None:
    """Fill the set-3 shootout table."""
    shots: List[Dict[str, Any]] = data_json.get("shootoutShots") or []
    if not shots:
        return

    # Separate shots by team, maintaining order
    host_shots: List[Dict[str, Any]] = []
    guest_shots: List[Dict[str, Any]] = []
    for s in shots:
        if s.get("team") == "host":
            host_shots.append(s)
        elif s.get("team") == "guest":
            guest_shots.append(s)

    def _write_team_shots(team_shots: List[Dict[str, Any]], rows: Dict[str, int]) -> int:
        total_pts = 0
        for i, shot in enumerate(team_shots):
            if i >= len(_SHOOTOUT_COLS):
                break
            col = _SHOOTOUT_COLS[i]
            player = shot.get("player")
            shot_type = shot.get("shotType", "normal")
            secondary = shot.get("secondaryPlayer")
            result = shot.get("result", 0)
            total_pts += result

            # Shooter row
            ws.cell(row=rows["shooter"], column=col).value = player

            # GK row — only if goalkeeper scored
            if shot_type == "goalkeeper" and secondary is not None:
                ws.cell(row=rows["gk"], column=col).value = secondary
            else:
                _diagonal_cross_rc(ws, rows["gk"], col)

            # Penalty row — only if penalty kick
            if shot_type == "penalty" and secondary is not None:
                ws.cell(row=rows["penalty"], column=col).value = secondary
            else:
                _diagonal_cross_rc(ws, rows["penalty"], col)

            # Points row
            ws.cell(row=rows["pts"], column=col).value = result

        return total_pts

    host_total = _write_team_shots(host_shots, _SHOOTOUT_HOST_ROWS)
    guest_total = _write_team_shots(guest_shots, _SHOOTOUT_GUEST_ROWS)

    # Cross out unused columns (after last shot through T=col20)
    host_used = min(len(host_shots), len(_SHOOTOUT_COLS))
    guest_used = min(len(guest_shots), len(_SHOOTOUT_COLS))

    for i in range(host_used, len(_SHOOTOUT_COLS)):
        col = _SHOOTOUT_COLS[i]
        for r_key in ("shooter", "gk", "penalty", "pts"):
            _diagonal_cross_rc(ws, _SHOOTOUT_HOST_ROWS[r_key], col)

    for i in range(guest_used, len(_SHOOTOUT_COLS)):
        col = _SHOOTOUT_COLS[i]
        for r_key in ("shooter", "gk", "penalty", "pts"):
            _diagonal_cross_rc(ws, _SHOOTOUT_GUEST_ROWS[r_key], col)

    # Totals
    ws["U59"] = host_total   # col U = 21
    ws["U64"] = guest_total

    # Shootout result in pts
    ws["T56"] = f"{host_total}:{guest_total}"

    # Set 3 win
    if host_total > guest_total:
        ws["T57"] = "1:0"
    elif guest_total > host_total:
        ws["T57"] = "0:1"
    else:
        ws["T57"] = f"{host_total}:{guest_total}"


async def _fill_completed_protocol_sheet(
    ws,
    data_json: Dict[str, Any],
) -> None:
    """Fill a protocol sheet with completed match results from ProEl data_json.

    This should be called AFTER _fill_protocol_sheet() which fills base data
    (teams, referees, players, companions).
    """
    match_config = data_json.get("matchConfig") or {}
    protocol: List[Dict[str, Any]] = data_json.get("protocol") or []
    set_results: List[Dict[str, Any]] = data_json.get("setResults") or []
    timeouts: Dict[str, Any] = data_json.get("timeouts") or {}
    host_stats: List[Dict[str, Any]] = data_json.get("hostStats") or []
    guest_stats: List[Dict[str, Any]] = data_json.get("guestStats") or []
    sets_won_host = data_json.get("setsWonHost", 0)
    sets_won_guest = data_json.get("setsWonGuest", 0)
    set_time_sec = (match_config.get("setTime") or 10) * 60  # minutes → seconds

    # ── 1. Match result in sets ──
    ws["S10"] = f"{sets_won_host}:{sets_won_guest}"

    # ── 2. Timeouts ──
    for set_num_str, host_cell, guest_cell in [("set1", "G35", "G58"), ("set2", "H35", "H58")]:
        set_timeouts = timeouts.get(set_num_str) or {}
        host_to = set_timeouts.get("host")
        guest_to = set_timeouts.get("guest")

        if host_to and host_to.get("timeRemainingMs") is not None:
            ws[host_cell] = _fmt_timeout_minute(host_to["timeRemainingMs"], set_time_sec)
        else:
            _diagonal_cross(ws, host_cell)

        if guest_to and guest_to.get("timeRemainingMs") is not None:
            ws[guest_cell] = _fmt_timeout_minute(guest_to["timeRemainingMs"], set_time_sec)
        else:
            _diagonal_cross(ws, guest_cell)

    # ── 3. Player stats (goals, exclusions, disqualifications) ──
    player_points = _build_player_points(protocol)
    exclusion_counts = _build_exclusion_counts(protocol)
    disqualified = _build_disqualified_set(protocol)

    def _fill_player_stats(stats_list: List[Dict[str, Any]], player_rows: List[int], team: str):
        for i, row_num in enumerate(player_rows):
            if i >= len(stats_list):
                # Empty row — cross out D through H
                for col in range(4, 9):  # D=4, E=5, F=6, G=7, H=8
                    _diagonal_cross_rc(ws, row_num, col)
                continue

            ps = stats_list[i]
            jersey = ps.get("number")
            if jersey is None:
                for col in range(4, 9):
                    _diagonal_cross_rc(ws, row_num, col)
                continue

            pts = player_points.get(team, {}).get(jersey, 0)
            excl = exclusion_counts.get(team, {}).get(jersey, 0)
            is_disq = jersey in disqualified.get(team, set())

            # Col D: total points
            if pts > 0:
                ws.cell(row=row_num, column=4).value = pts
            else:
                _diagonal_cross_rc(ws, row_num, 4)

            # Col E: 1st exclusion
            if excl >= 1:
                ws.cell(row=row_num, column=5).value = "X"
            else:
                _diagonal_cross_rc(ws, row_num, 5)

            # Col F: 2nd exclusion
            if excl >= 2:
                ws.cell(row=row_num, column=6).value = "X"
            else:
                _diagonal_cross_rc(ws, row_num, 6)

            # Col G: disqualification
            if is_disq:
                ws.cell(row=row_num, column=7).value = "X"
            else:
                _diagonal_cross_rc(ws, row_num, 7)

            # Col H: always crossed
            _diagonal_cross_rc(ws, row_num, 8)

    _fill_player_stats(host_stats, HOST_PLAYER_ROWS, "host")
    _fill_player_stats(guest_stats, GUEST_PLAYER_ROWS, "guest")

    # ── 4. Companion stats ──
    host_companions = match_config.get("hostCompanions") or []
    guest_companions = match_config.get("guestCompanions") or []
    companions_by_team = {"host": host_companions, "guest": guest_companions}
    companion_penalties = _build_companion_penalties(protocol, companions_by_team)

    def _fill_companion_stats(
        companions_list: List[Dict[str, Any]],
        companion_rows: List[int],
        team: str,
    ):
        team_penalties = companion_penalties.get(team, {})

        for i, row_num in enumerate(companion_rows):
            if i >= len(companions_list):
                # Empty companion row — cross cols E through H
                for col in range(5, 9):  # E=5, F=6, G=7, H=8
                    _diagonal_cross_rc(ws, row_num, col)
                continue

            comp = companions_list[i]
            role = (comp.get("role") or "").strip().upper()
            penalties = team_penalties.get(role, {"exclusion": False, "disqualification": False})

            # Col E: exclusion
            if penalties["exclusion"]:
                ws.cell(row=row_num, column=5).value = "X"
            else:
                _diagonal_cross_rc(ws, row_num, 5)

            # Col F: always crossed (companions can't get 2nd exclusion)
            _diagonal_cross_rc(ws, row_num, 6)

            # Col G: disqualification
            if penalties["disqualification"]:
                ws.cell(row=row_num, column=7).value = "X"
            else:
                _diagonal_cross_rc(ws, row_num, 7)

            # Col H: always crossed
            _diagonal_cross_rc(ws, row_num, 8)

    _fill_companion_stats(host_companions, HOST_COMPANION_ROWS, "host")
    _fill_companion_stats(guest_companions, GUEST_COMPANION_ROWS, "guest")

    # ── 5. Set progressions (goal by goal) ──
    for set_num in (1, 2):
        _fill_set_progression(ws, set_num, protocol)

    # ── 6. Set results ──
    for idx, (set_num, cells) in enumerate(_SET_RESULT_CELLS.items()):
        if idx < len(set_results):
            sr = set_results[idx]
            pts_h = sr.get("ptsHost", 0)
            pts_g = sr.get("ptsGuest", 0)
            ws[cells["pts"]] = f"{pts_h}:{pts_g}"
            if pts_h > pts_g:
                ws[cells["win"]] = "1:0"
            elif pts_g > pts_h:
                ws[cells["win"]] = "0:1"
            else:
                ws[cells["win"]] = "0:0"

    # ── 7. Set 3 (shootout) ──
    shootout_shots = data_json.get("shootoutShots") or []
    has_set3 = len(shootout_shots) > 0 or len(set_results) >= 3
    if has_set3:
        _fill_shootout(ws, data_json)
        # Shootout "set result"
        if shootout_shots:
            host_pts = sum(s.get("result", 0) for s in shootout_shots if s.get("team") == "host")
            guest_pts = sum(s.get("result", 0) for s in shootout_shots if s.get("team") == "guest")
            ws["T56"] = f"{host_pts}:{guest_pts}"
            ws["T57"] = "1:0" if host_pts > guest_pts else "0:1"
    else:
        # No 3rd set — cross out shootout table and set-3 result cells
        _diagonal_cross(ws, "T56")  # set-3 pts
        _diagonal_cross(ws, "T57")  # set-3 win
        # Cross out all shootout cells (host rows 59-62, guest rows 64-67, cols K-T)
        for col in _SHOOTOUT_COLS:
            for r_key in ("shooter", "gk", "penalty", "pts"):
                _diagonal_cross_rc(ws, _SHOOTOUT_HOST_ROWS[r_key], col)
                _diagonal_cross_rc(ws, _SHOOTOUT_GUEST_ROWS[r_key], col)
        # Cross out totals column U
        _diagonal_cross(ws, "U59")
        _diagonal_cross(ws, "U64")

    # ── 8. Signatures ──
    extras = match_config.get("extras") or {}
    signatures = extras.get("signatures") or {}

    sig_mapping = [
        # (key, anchor, max_w, max_h, offset_x_px, offset_y_px)
        ("hostTeamSignature",   "B35", 120, 50, 50, -4),
        ("guestTeamSignature",  "B58", 120, 50, 50, -4),
        ("fieldASignature",     "F61",  90, 35, 10, -3),
        ("fieldBSignature",     "F62",  90, 35, 10, -3),
        ("tableSigSignature",   "F64",  90, 35, 10, -3),
        ("tableTimerSignature", "F65",  90, 35, 10, -3),
    ]

    for sig_key, anchor, max_w, max_h, off_x, off_y in sig_mapping:
        url = _full_static_url(signatures.get(sig_key) or "")
        if url:
            img_bytes = await _fetch_png_bytes(url)
            _add_signature_image(ws, image_bytes=img_bytes, anchor_cell=anchor,
                                 max_width_px=max_w, max_height_px=max_h,
                                 offset_x_px=off_x, offset_y_px=off_y)

    # Head judge signature (F67) — check if headJudge has a signature
    # Head judge is not in the standard signatures dict; skip if not present


async def _generate_filled_protocol(
    data_json: Dict[str, Any],
    match: Dict[str, Any],
    *,
    tournament_name: str,
    tournament_location: str,
    category: str,
    days: List[Dict[str, Any]],
    team_squads: Dict[str, Any],
    custom_teams: List[Dict[str, Any]],
    team_rosters: Dict[int, dict],
    user_cities: Dict[int, Tuple[str, str]],
    head_judge_id: Optional[int],
    output_format: str,
    out_dir: str,
    file_label: str,
) -> str:
    """Generate a filled (completed match) protocol file. Returns file path."""
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # 1. Fill base data (teams, referees, players, companions)
    _fill_protocol_sheet(
        ws, match,
        tournament_name=tournament_name,
        tournament_location=tournament_location,
        category=category,
        days=days,
        team_squads=team_squads,
        custom_teams=custom_teams,
        team_rosters=team_rosters,
        user_cities=user_cities,
        head_judge_id=head_judge_id,
    )

    # 2. Fill match results from ProEl data
    await _fill_completed_protocol_sheet(ws, data_json)

    safe_label = _safe_filename(file_label, 60)
    xlsx_name = f"protokol_koncowy_{safe_label}.xlsx"
    xlsx_path = os.path.join(out_dir, xlsx_name)
    wb.save(xlsx_path)

    if output_format == "pdf":
        pdf_path = _convert_xlsx_to_pdf(xlsx_path, out_dir)
        return pdf_path
    return xlsx_path


async def _resolve_schedule_match(
    tournament_data: dict,
    match_number: str,
) -> Optional[Dict[str, Any]]:
    """Find the schedule match object by matchNumber from tournament data."""
    schedule = tournament_data.get("schedule") or {}
    matches = schedule.get("matches") or []
    for m in matches:
        if m.get("matchNumber") == match_number:
            return m
    return None


# ── generate protocol for a single match ──────────────────────────────────────

async def _generate_single_protocol(
    match: Dict[str, Any],
    *,
    tournament_name: str,
    tournament_location: str,
    category: str,
    days: List[Dict[str, Any]],
    team_squads: Dict[str, Any],
    custom_teams: List[Dict[str, Any]],
    team_rosters: Dict[int, dict],
    user_cities: Dict[int, Tuple[str, str]],
    head_judge_id: Optional[int],
    output_format: str,
    out_dir: str,
    file_label: str,
) -> str:
    """Generate a single protocol file (xlsx or pdf). Returns file path."""
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    _fill_protocol_sheet(
        ws,
        match,
        tournament_name=tournament_name,
        tournament_location=tournament_location,
        category=category,
        days=days,
        team_squads=team_squads,
        custom_teams=custom_teams,
        team_rosters=team_rosters,
        user_cities=user_cities,
        head_judge_id=head_judge_id,
    )

    safe_label = _safe_filename(file_label, 60)
    xlsx_name = f"protokol_{safe_label}.xlsx"
    xlsx_path = os.path.join(out_dir, xlsx_name)
    wb.save(xlsx_path)

    if output_format == "pdf":
        pdf_path = _convert_xlsx_to_pdf(xlsx_path, out_dir)
        return pdf_path
    return xlsx_path


# ── endpoints ─────────────────────────────────────────────────────────────────

@router.post("/beach/protocol/bulk", summary="Generuj surowe protokoły dla wszystkich meczy turnieju")
async def generate_bulk_protocols(req: ProtocolBulkRequest):
    if not os.path.exists(TEMPLATE_PATH):
        raise HTTPException(500, detail=f"Brak szablonu protokołu: {TEMPLATE_PATH}")

    output_format = req.format.lower()
    if output_format not in ("xlsx", "pdf"):
        raise HTTPException(400, detail="format musi być 'xlsx' lub 'pdf'")

    schedule = req.schedule
    matches: List[Dict[str, Any]] = schedule.get("matches") or []
    config: Dict[str, Any] = schedule.get("config") or {}
    days: List[Dict[str, Any]] = config.get("days") or []

    if not matches:
        raise HTTPException(400, detail="Brak meczy w terminarzu")

    # Sort matches
    matches = sorted(matches, key=_match_sort_key)

    # Fetch tournament data
    data_json: dict = {}
    if req.tournament_id:
        data_json = await _fetch_tournament_data(req.tournament_id)

    team_squads = data_json.get("team_squads") or {}
    custom_teams = data_json.get("custom_teams") or []
    head_judge_id = data_json.get("head_judge_id")

    # Batch-fetch team rosters and referee cities
    team_ids = _collect_team_ids(matches)
    referee_ids = _collect_referee_ids(matches, head_judge_id)

    team_rosters = await _fetch_team_rosters(team_ids)
    user_cities = await _fetch_user_cities(referee_ids)

    tournament_name = req.tournament_name.strip()
    tournament_location = req.tournament_location.strip()
    category = req.category.strip()

    tmp_dir = tempfile.mkdtemp()
    try:
        protocols_dir = os.path.join(tmp_dir, "protocols")
        os.makedirs(protocols_dir)

        file_paths: List[Tuple[str, str]] = []  # (path, filename)

        for idx, m in enumerate(matches):
            file_label = _match_file_label(m, idx)
            path = await _generate_single_protocol(
                m,
                tournament_name=tournament_name,
                tournament_location=tournament_location,
                category=category,
                days=days,
                team_squads=team_squads,
                custom_teams=custom_teams,
                team_rosters=team_rosters,
                user_cities=user_cities,
                head_judge_id=head_judge_id,
                output_format=output_format,
                out_dir=protocols_dir,
                file_label=file_label,
            )
            file_paths.append((path, os.path.basename(path)))

        # Pack into ZIP
        safe_name = _safe_filename(tournament_name, 40) or "protokoly"
        zip_name = f"protokoly_{safe_name}.zip"
        zip_path = os.path.join(tmp_dir, zip_name)

        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for fpath, fname in file_paths:
                zf.write(fpath, fname)

        # Store with download token
        _ensure_download_dir()
        token = str(uuid.uuid4())
        ext = "zip"
        download_path = os.path.join(DOWNLOAD_DIR, f"{token}.{ext}")
        shutil.copyfile(zip_path, download_path)

        encoded_name = urllib.parse.quote(zip_name)
        return {
            "success": True,
            "download_url": f"/beach/protocol/download/{token}?filename={encoded_name}&ext={ext}",
            "match_count": len(file_paths),
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Bulk protocol generation failed")
        raise HTTPException(500, detail=str(e))
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


@router.post("/beach/protocol/single", summary="Generuj surowy protokół dla jednego meczu")
async def generate_single_protocol(req: ProtocolSingleRequest):
    if not os.path.exists(TEMPLATE_PATH):
        raise HTTPException(500, detail=f"Brak szablonu protokołu: {TEMPLATE_PATH}")

    output_format = req.format.lower()
    if output_format not in ("xlsx", "pdf"):
        raise HTTPException(400, detail="format musi być 'xlsx' lub 'pdf'")

    match = req.match
    days: List[Dict[str, Any]] = (req.schedule_config or {}).get("days") or []

    # Fetch tournament data
    data_json: dict = {}
    if req.tournament_id:
        data_json = await _fetch_tournament_data(req.tournament_id)

    team_squads = data_json.get("team_squads") or {}
    custom_teams = data_json.get("custom_teams") or []
    head_judge_id = data_json.get("head_judge_id")

    # Fetch rosters and cities
    team_ids = _collect_team_ids([match])
    referee_ids = _collect_referee_ids([match], head_judge_id)

    team_rosters = await _fetch_team_rosters(team_ids)
    user_cities = await _fetch_user_cities(referee_ids)

    tmp_dir = tempfile.mkdtemp()
    try:
        file_label = _match_file_label(match, 0)
        path = await _generate_single_protocol(
            match,
            tournament_name=req.tournament_name.strip(),
            tournament_location=req.tournament_location.strip(),
            category=req.category.strip(),
            days=days,
            team_squads=team_squads,
            custom_teams=custom_teams,
            team_rosters=team_rosters,
            user_cities=user_cities,
            head_judge_id=head_judge_id,
            output_format=output_format,
            out_dir=tmp_dir,
            file_label=file_label,
        )

        ext = "pdf" if output_format == "pdf" else "xlsx"
        download_name = os.path.basename(path)

        _ensure_download_dir()
        token = str(uuid.uuid4())
        download_path = os.path.join(DOWNLOAD_DIR, f"{token}.{ext}")
        shutil.copyfile(path, download_path)

        encoded_name = urllib.parse.quote(download_name)
        return {
            "success": True,
            "download_url": f"/beach/protocol/download/{token}?filename={encoded_name}&ext={ext}",
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Single protocol generation failed")
        raise HTTPException(500, detail=str(e))
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


# ── filled protocol endpoints ─────────────────────────────────────────────────

@router.post("/beach/protocol/filled/single", summary="Generuj wypełniony protokół dla jednego meczu ProEl")
async def generate_filled_single(req: FilledProtocolSingleRequest):
    if not os.path.exists(TEMPLATE_PATH):
        raise HTTPException(500, detail=f"Brak szablonu protokołu: {TEMPLATE_PATH}")

    output_format = req.format.lower()
    if output_format not in ("xlsx", "pdf"):
        raise HTTPException(400, detail="format musi być 'xlsx' lub 'pdf'")

    if not req.match_number:
        raise HTTPException(400, detail="Brak match_number")

    # Fetch ProEl match from DB
    row = await database.fetch_one(
        select(beach_proel_matches).where(
            beach_proel_matches.c.match_number == req.match_number
        )
    )
    if not row:
        raise HTTPException(404, detail=f"Nie znaleziono meczu ProEl: {req.match_number}")

    data_json = _parse_data_json(row["data_json"])
    if not data_json:
        raise HTTPException(400, detail="Brak danych meczu (data_json)")

    # Determine tournament_id from request or from match extras
    tournament_id = req.tournament_id
    if not tournament_id:
        extras = (data_json.get("matchConfig") or {}).get("extras") or {}
        tournament_id = extras.get("tournamentId")

    # Fetch tournament data
    tournament_data: dict = {}
    if tournament_id:
        tournament_data = await _fetch_tournament_data(tournament_id)

    schedule = tournament_data.get("schedule") or {}
    config = schedule.get("config") or {}
    days: List[Dict[str, Any]] = config.get("days") or []
    team_squads = tournament_data.get("team_squads") or {}
    custom_teams = tournament_data.get("custom_teams") or []
    head_judge_id = tournament_data.get("head_judge_id")
    tournament_name = tournament_data.get("name") or ""
    tournament_location = tournament_data.get("venue_address") or ""
    category = tournament_data.get("category") or ""

    # Find schedule match for this match_number
    schedule_match = await _resolve_schedule_match(tournament_data, req.match_number)
    if not schedule_match:
        # Build a minimal match from data_json matchConfig
        mc = data_json.get("matchConfig") or {}
        schedule_match = {
            "matchNumber": req.match_number,
            "teamA": {"name": mc.get("hostTeamName", ""), "id": mc.get("hostTeamId")},
            "teamB": {"name": mc.get("guestTeamName", ""), "id": mc.get("guestTeamId")},
            "startTime": mc.get("matchTime", ""),
            "dayIndex": 0,
            "referees": mc.get("referees") or {},
            "gender": (mc.get("extras") or {}).get("gender") or "",
        }

    # Fetch rosters and cities
    team_ids = _collect_team_ids([schedule_match])
    referee_ids = _collect_referee_ids([schedule_match], head_judge_id)
    team_rosters = await _fetch_team_rosters(team_ids)
    user_cities = await _fetch_user_cities(referee_ids)

    tmp_dir = tempfile.mkdtemp()
    try:
        file_label = req.match_number.replace("/", "_").replace("\\", "_")
        path = await _generate_filled_protocol(
            data_json,
            schedule_match,
            tournament_name=tournament_name,
            tournament_location=tournament_location,
            category=category,
            days=days,
            team_squads=team_squads,
            custom_teams=custom_teams,
            team_rosters=team_rosters,
            user_cities=user_cities,
            head_judge_id=head_judge_id,
            output_format=output_format,
            out_dir=tmp_dir,
            file_label=file_label,
        )

        ext = "pdf" if output_format == "pdf" else "xlsx"
        download_name = os.path.basename(path)

        _ensure_download_dir()
        token = str(uuid.uuid4())
        download_path = os.path.join(DOWNLOAD_DIR, f"{token}.{ext}")
        shutil.copyfile(path, download_path)

        encoded_name = urllib.parse.quote(download_name)
        return {
            "success": True,
            "download_url": f"/beach/protocol/download/{token}?filename={encoded_name}&ext={ext}",
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Filled single protocol generation failed")
        raise HTTPException(500, detail=str(e))
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


@router.post("/beach/protocol/filled/bulk", summary="Generuj wypełnione protokoły dla wielu meczów ProEl")
async def generate_filled_bulk(req: FilledProtocolBulkRequest):
    if not os.path.exists(TEMPLATE_PATH):
        raise HTTPException(500, detail=f"Brak szablonu protokołu: {TEMPLATE_PATH}")

    output_format = req.format.lower()
    if output_format not in ("xlsx", "pdf"):
        raise HTTPException(400, detail="format musi być 'xlsx' lub 'pdf'")

    if not req.match_numbers:
        raise HTTPException(400, detail="Brak match_numbers")

    # Fetch all specified ProEl matches from DB
    rows = await database.fetch_all(
        select(beach_proel_matches).where(
            beach_proel_matches.c.match_number.in_(req.match_numbers)
        )
    )
    if not rows:
        raise HTTPException(404, detail="Nie znaleziono żadnych meczów ProEl")

    matches_by_number = {
        r["match_number"]: _parse_data_json(r["data_json"])
        for r in rows
    }

    # Determine tournament_id
    tournament_id = req.tournament_id
    if not tournament_id:
        for dj in matches_by_number.values():
            extras = (dj.get("matchConfig") or {}).get("extras") or {}
            tid = extras.get("tournamentId")
            if tid:
                tournament_id = tid
                break

    # Fetch tournament data once
    tournament_data: dict = {}
    if tournament_id:
        tournament_data = await _fetch_tournament_data(tournament_id)

    schedule = tournament_data.get("schedule") or {}
    config = schedule.get("config") or {}
    days: List[Dict[str, Any]] = config.get("days") or []
    team_squads = tournament_data.get("team_squads") or {}
    custom_teams = tournament_data.get("custom_teams") or []
    head_judge_id = tournament_data.get("head_judge_id")
    tournament_name = req.tournament_name or tournament_data.get("name") or ""
    tournament_location = tournament_data.get("venue_address") or ""
    category = tournament_data.get("category") or ""

    # Build schedule match lookup
    all_schedule_matches = schedule.get("matches") or []
    schedule_match_by_number = {m.get("matchNumber"): m for m in all_schedule_matches if m.get("matchNumber")}

    # Collect all team & referee IDs for batch fetch
    all_match_objs = []
    for mn, dj in matches_by_number.items():
        sm = schedule_match_by_number.get(mn)
        if not sm:
            mc = dj.get("matchConfig") or {}
            sm = {
                "matchNumber": mn,
                "teamA": {"name": mc.get("hostTeamName", ""), "id": mc.get("hostTeamId")},
                "teamB": {"name": mc.get("guestTeamName", ""), "id": mc.get("guestTeamId")},
                "startTime": mc.get("matchTime", ""),
                "dayIndex": 0,
                "referees": mc.get("referees") or {},
                "gender": (mc.get("extras") or {}).get("gender") or "",
            }
        all_match_objs.append((mn, sm, dj))

    team_ids = _collect_team_ids([m for _, m, _ in all_match_objs])
    referee_ids = _collect_referee_ids([m for _, m, _ in all_match_objs], head_judge_id)
    team_rosters = await _fetch_team_rosters(team_ids)
    user_cities = await _fetch_user_cities(referee_ids)

    tmp_dir = tempfile.mkdtemp()
    try:
        protocols_dir = os.path.join(tmp_dir, "protocols")
        os.makedirs(protocols_dir)

        file_paths: List[Tuple[str, str]] = []

        for mn, sm, dj in all_match_objs:
            file_label = mn.replace("/", "_").replace("\\", "_")
            path = await _generate_filled_protocol(
                dj, sm,
                tournament_name=tournament_name,
                tournament_location=tournament_location,
                category=category,
                days=days,
                team_squads=team_squads,
                custom_teams=custom_teams,
                team_rosters=team_rosters,
                user_cities=user_cities,
                head_judge_id=head_judge_id,
                output_format=output_format,
                out_dir=protocols_dir,
                file_label=file_label,
            )
            file_paths.append((path, os.path.basename(path)))

        # Pack into ZIP
        safe_name = _safe_filename(tournament_name, 40) or "protokoly_koncowe"
        zip_name = f"protokoly_koncowe_{safe_name}.zip"
        zip_path = os.path.join(tmp_dir, zip_name)

        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for fpath, fname in file_paths:
                zf.write(fpath, fname)

        _ensure_download_dir()
        token = str(uuid.uuid4())
        ext = "zip"
        download_path = os.path.join(DOWNLOAD_DIR, f"{token}.{ext}")
        shutil.copyfile(zip_path, download_path)

        encoded_name = urllib.parse.quote(zip_name)
        return {
            "success": True,
            "download_url": f"/beach/protocol/download/{token}?filename={encoded_name}&ext={ext}",
            "match_count": len(file_paths),
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Filled bulk protocol generation failed")
        raise HTTPException(500, detail=str(e))
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


@router.get(
    "/beach/protocol/download/{token}",
    summary="Pobierz wygenerowany protokół/y (attachment)",
)
async def download_protocol(
    token: str = ApiPath(...),
    filename: str = Query("protokol.xlsx"),
    ext: str = Query("xlsx"),
):
    _ensure_download_dir()
    try:
        uuid.UUID(token)
    except ValueError:
        raise HTTPException(400, "Nieprawidłowy token")

    if ext not in ("xlsx", "pdf", "zip"):
        ext = "xlsx"

    file_path = os.path.join(DOWNLOAD_DIR, f"{token}.{ext}")
    if not os.path.exists(file_path):
        raise HTTPException(404, "Plik wygasł lub nie istnieje")

    media_types = {
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "pdf": "application/pdf",
        "zip": "application/zip",
    }

    return FileResponse(
        path=file_path,
        media_type=media_types.get(ext, "application/octet-stream"),
        filename=filename,
        background=BackgroundTask(
            lambda: os.remove(file_path) if os.path.exists(file_path) else None
        ),
    )
