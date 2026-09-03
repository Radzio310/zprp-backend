"""Jednorazowa, deterministyczna przebudowa obu szablonow protokolu.

Szablon ma sklad i przebieg obok siebie, dlatego usuniecie calego wiersza
arkusza zmienia oba bloki naraz. Usuwamy po dwa ostatnie miejsca zawodnikow
z kazdej druzyny (stare wiersze 27-28 oraz 53-54), co jednoczesnie zmniejsza
przebieg z 47 do 43 pozycji na kazda polowe.

Pozostale 16 wierszy kazdego skladu rosnie z 13,2 do 14,85 pkt. Dzieki temu
kazdy blok skladu zachowuje dokladnie te sama wysokosc:

    18 * 13,2 == 16 * 14,85 == 237,6 pkt

Pola pod skladami zostaja wiec w tych samych miejscach na wydruku, calkowita
wysokosc arkusza sie nie zmienia, a dotychczasowa skala A4 pozostaje aktualna.

Modyfikujemy XML wewnatrz XLSX zamiast zapisywac skoroszyt przez openpyxl.
To zachowuje bez zmian style, ustawienia drukarki i osadzone grafiki Excela.
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
import re
import zipfile

from lxml import etree
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TEMPLATES = ROOT / "app" / "templates"
FILES = (
    TEMPLATES / "protocol_template.xlsx",
    TEMPLATES / "protocol_template_2.xlsx",
)
LEGACY_LAST_ROW = {
    "protocol_template.xlsx": 71,
    "protocol_template_2.xlsx": 72,
}
COMPACT_LAST_ROW = {
    "protocol_template.xlsx": 67,
    "protocol_template_2.xlsx": 68,
}

DELETED_ROWS = frozenset((27, 28, 53, 54))
NEW_PLAYER_ROWS = frozenset((*range(11, 27), *range(35, 51)))
PLAYER_ROW_HEIGHT_PT = 14.85

SHEET_XML = "xl/worksheets/sheet1.xml"
WORKBOOK_XML = "xl/workbook.xml"
CALC_CHAIN_XML = "xl/calcChain.xml"
MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

CELL_RE = re.compile(r"^(\$?)([A-Z]{1,3})(\$?)(\d+)$")


def map_row(old_row: int) -> int | None:
    if old_row in DELETED_ROWS:
        return None
    shift = 0
    if old_row > 28:
        shift += 2
    if old_row > 54:
        shift += 2
    return old_row - shift


def _split_cell(ref: str) -> tuple[str, str, str, int]:
    match = CELL_RE.fullmatch(ref)
    if not match:
        raise ValueError(f"Nieobslugiwany adres komorki: {ref!r}")
    abs_col, col, abs_row, row = match.groups()
    return abs_col, col, abs_row, int(row)


def map_cell(ref: str) -> str | None:
    abs_col, col, abs_row, old_row = _split_cell(ref)
    new_row = map_row(old_row)
    if new_row is None:
        return None
    return f"{abs_col}{col}{abs_row}{new_row}"


def map_range(ref: str, *, reject_deleted_inside: bool = False) -> str | None:
    if ":" not in ref:
        return map_cell(ref)

    left, right = ref.split(":", 1)
    left_parts = _split_cell(left)
    right_parts = _split_cell(right)
    lo, hi = sorted((left_parts[3], right_parts[3]))
    if reject_deleted_inside and any(row in DELETED_ROWS for row in range(lo, hi + 1)):
        return None

    mapped_left = map_cell(left)
    mapped_right = map_cell(right)
    if mapped_left is None or mapped_right is None:
        return None
    return f"{mapped_left}:{mapped_right}"


def map_sqref(value: str) -> str:
    mapped = [map_range(token) for token in value.split()]
    return " ".join(token for token in mapped if token)


def _parse_xml(data: bytes):
    return etree.fromstring(data, parser=etree.XMLParser(remove_blank_text=False))


def _xml_bytes(root) -> bytes:
    return etree.tostring(
        root,
        encoding="UTF-8",
        xml_declaration=True,
        standalone=True,
    )


def _transform_sheet(data: bytes) -> bytes:
    root = _parse_xml(data)
    ns = {"x": MAIN_NS}
    sheet_data = root.find("x:sheetData", namespaces=ns)
    if sheet_data is None:
        raise RuntimeError("Szablon nie ma sheetData")

    rows = list(sheet_data.findall("x:row", namespaces=ns))
    old_numbers = {int(row.get("r")) for row in rows}
    if not DELETED_ROWS.issubset(old_numbers):
        raise RuntimeError("Szablon nie ma oczekiwanych wierszy 27, 28, 53 i 54")

    for row in rows:
        old_row = int(row.get("r"))
        new_row = map_row(old_row)
        if new_row is None:
            sheet_data.remove(row)
            continue

        row.set("r", str(new_row))
        if new_row in NEW_PLAYER_ROWS:
            row.set("ht", f"{PLAYER_ROW_HEIGHT_PT:g}")
            row.set("customHeight", "1")

        for cell in row.findall("x:c", namespaces=ns):
            old_ref = cell.get("r")
            new_ref = map_cell(old_ref)
            if new_ref is None:
                raise RuntimeError(f"Komorka {old_ref} przezyla usuniety wiersz")
            cell.set("r", new_ref)

    # Scalenia na usuwanych wierszach znikaja; pozostale przesuwaja sie razem
    # z komorkami. Zaden zakres w aktualnych szablonach nie przecina granicy
    # usuwanego wiersza - taki przypadek celowo konczy migracje bledem.
    merge_cells = root.find("x:mergeCells", namespaces=ns)
    if merge_cells is not None:
        for merged in list(merge_cells):
            old_ref = merged.get("ref")
            new_ref = map_range(old_ref, reject_deleted_inside=True)
            if new_ref is None:
                merge_cells.remove(merged)
            else:
                merged.set("ref", new_ref)
        merge_cells.set("count", str(len(merge_cells)))

    # Pozostale adresy arkusza: zaznaczenie, widoczny lewy gorny rog,
    # walidacje, filtr i deklarowany wymiar danych.
    for element in root.iter():
        local_name = etree.QName(element).localname
        if local_name in {"c", "mergeCell"}:
            continue
        for attr in ("activeCell", "topLeftCell"):
            value = element.get(attr)
            if value and CELL_RE.fullmatch(value):
                mapped = map_cell(value)
                if mapped:
                    element.set(attr, mapped)
        value = element.get("sqref")
        if value:
            element.set("sqref", map_sqref(value))
        value = element.get("ref")
        if value and all(CELL_RE.fullmatch(part) for part in value.split(":", 1)):
            mapped = map_range(value)
            if mapped:
                element.set("ref", mapped)

    return _xml_bytes(root)


def _transform_workbook(data: bytes) -> bytes:
    root = _parse_xml(data)
    ns = {"x": MAIN_NS}
    for defined_name in root.findall(".//x:definedName", namespaces=ns):
        text = defined_name.text or ""
        if "!" not in text:
            continue
        sheet, ref = text.rsplit("!", 1)
        if all(CELL_RE.fullmatch(part) for part in ref.split(":", 1)):
            mapped = map_range(ref)
            if mapped:
                defined_name.text = f"{sheet}!{mapped}"
    return _xml_bytes(root)


def _transform_calc_chain(data: bytes) -> bytes:
    root = _parse_xml(data)
    for cell in root.iter():
        if etree.QName(cell).localname != "c":
            continue
        ref = cell.get("r")
        if not ref or not CELL_RE.fullmatch(ref):
            continue
        mapped = map_cell(ref)
        if mapped is None:
            raise RuntimeError(f"Formula kalkulacyjna wskazuje usuniety wiersz: {ref}")
        cell.set("r", mapped)
    return _xml_bytes(root)


def transform_xlsx(path: Path) -> bytes:
    output = BytesIO()
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(output, "w") as target:
        for info in source.infolist():
            payload = source.read(info.filename)
            if info.filename == SHEET_XML:
                payload = _transform_sheet(payload)
            elif info.filename == WORKBOOK_XML:
                payload = _transform_workbook(payload)
            elif info.filename == CALC_CHAIN_XML:
                payload = _transform_calc_chain(payload)
            target.writestr(info, payload)
    return output.getvalue()


def _sheet_height(ws) -> float:
    default = ws.sheet_format.defaultRowHeight or 15
    return sum(
        ws.row_dimensions[row].height or default
        for row in range(1, ws.max_row + 1)
    )


def validate(path: Path, data: bytes) -> None:
    old_ws = load_workbook(path).active
    expected_rows = old_ws.max_row - len(DELETED_ROWS)
    old_height = _sheet_height(old_ws)

    workbook = load_workbook(BytesIO(data))
    ws = workbook.active
    if ws.max_row != expected_rows:
        raise RuntimeError(f"{path.name}: ma {ws.max_row} wierszy zamiast {expected_rows}")
    if abs(_sheet_height(ws) - old_height) > 0.001:
        raise RuntimeError(f"{path.name}: zmienila sie calkowita wysokosc arkusza")
    if len([row for row in range(15, 60) if row not in {29, 53}]) != 43:
        raise RuntimeError("Nowy przebieg nie ma 43 pozycji")
    if ws["E33"].value != "=D7":
        raise RuntimeError(f"{path.name}: naglowek druzyny B nie przesunal sie poprawnie")
    if ws["B27"].value != "A" or ws["B51"].value != "A":
        raise RuntimeError(f"{path.name}: bloki osob towarzyszacych sa przesuniete")
    expected_area = f"A1:BH{expected_rows}"
    if expected_area not in str(ws.print_area).replace("$", ""):
        raise RuntimeError(f"{path.name}: bledny obszar wydruku {ws.print_area}")


def template_state(path: Path) -> str:
    """Chroni przed ponownym usunieciem czterech wierszy z gotowego pliku."""
    ws = load_workbook(path).active
    if ws.max_row == COMPACT_LAST_ROW[path.name]:
        if (
            ws["E33"].value == "=D7"
            and ws["B27"].value == "A"
            and ws["B51"].value == "A"
            and ws.row_dimensions[11].height == PLAYER_ROW_HEIGHT_PT
        ):
            return "compact"
        raise RuntimeError(f"{path.name}: ma kompaktowa wysokosc, ale obcy uklad")

    if ws.max_row == LEGACY_LAST_ROW[path.name]:
        if ws["E35"].value == "=D7" and ws["B29"].value == "A" and ws["B55"].value == "A":
            return "legacy"
        raise RuntimeError(f"{path.name}: ma stara wysokosc, ale obcy uklad")

    raise RuntimeError(f"{path.name}: nieznany uklad ({ws.max_row} wierszy)")


def main() -> None:
    transformed: dict[Path, bytes] = {}
    for path in FILES:
        if template_state(path) == "compact":
            print(f"Pominieto gotowy {path.relative_to(ROOT)}")
            continue
        data = transform_xlsx(path)
        validate(path, data)
        transformed[path] = data

    for path, data in transformed.items():
        path.write_bytes(data)
        print(f"Przebudowano {path.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
