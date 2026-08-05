# -*- coding: utf-8 -*-
"""Sprawdza, że nowy szablon to dokładnie stary przesunięty o jedną kolumnę."""
import os
import re
import zipfile

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string

TPL = r"c:\Users\radek\Desktop\BAZA_ALL\zprp-backend\app\templates"
OLD = os.path.join(TPL, "protocol_template_BACKUP_preExamCol.xlsx")
NEW = os.path.join(TPL, "protocol_template.xlsx")

SHIFT = 1


def shift(ref: str) -> str:
    col, row = coordinate_from_string(ref.replace("$", ""))
    return "%s%d" % (get_column_letter(column_index_from_string(col) + SHIFT), row)


wb_old = load_workbook(OLD)
wb_new = load_workbook(NEW)
ws_old, ws_new = wb_old.active, wb_new.active

problems = []

# 1) wartości komórek
checked = 0
for row in ws_old.iter_rows():
    for cell in row:
        if cell.value is None:
            continue
        checked += 1
        tgt = shift(cell.coordinate)
        got = ws_new[tgt].value
        want = cell.value
        # Formuła musi wskazywać na PRZESUNIĘTĄ komórkę — inaczej =C4 pokazywałoby
        # sąsiada zamiast nazwy drużyny.
        if isinstance(want, str) and want.startswith("="):
            want = "=" + shift(want[1:])
        if got != want:
            problems.append("wartość %s -> %s: %r != %r" % (cell.coordinate, tgt, want, got))

# 2) nowa kolumna A musi być pusta
for r in range(1, ws_new.max_row + 1):
    if ws_new.cell(row=r, column=1).value is not None:
        problems.append("nowa kolumna A niepusta w wierszu %d" % r)

# 3) scalenia
old_m = sorted(str(m) for m in ws_old.merged_cells.ranges)
new_m = sorted(str(m) for m in ws_new.merged_cells.ranges)
if len(old_m) != len(new_m):
    problems.append("liczba scaleń %d != %d" % (len(old_m), len(new_m)))
else:
    exp = sorted(":".join(shift(p) for p in m.split(":")) for m in old_m)
    if exp != new_m:
        diff = [(a, b) for a, b in zip(exp, new_m) if a != b][:5]
        problems.append("scalenia się różnią, np. %r" % diff)

# 4) szerokości kolumn
for letter, dim in ws_old.column_dimensions.items():
    if dim.width is None:
        continue
    tgt = get_column_letter(column_index_from_string(letter) + SHIFT)
    got = ws_new.column_dimensions[tgt].width
    if got is None or abs(got - dim.width) > 1e-6:
        problems.append("szerokość %s -> %s: %r != %r" % (letter, tgt, dim.width, got))

new_a = ws_new.column_dimensions["A"].width
if new_a is None or new_a < 2:
    problems.append("nowa kolumna A ma szerokość %r" % new_a)

# 5) wysokości wierszy
for r, dim in ws_old.row_dimensions.items():
    got = ws_new.row_dimensions[r].height
    if dim.height is not None and (got is None or abs(got - dim.height) > 1e-6):
        problems.append("wysokość wiersza %s: %r != %r" % (r, dim.height, got))

# 6) obrazki
if len(ws_old._images) != len(ws_new._images):
    problems.append("liczba obrazków %d != %d" % (len(ws_old._images), len(ws_new._images)))

# 7) pozycja treści na papierze — najważniejszy test wizualny
#
# Arkusz drukuje się wyśrodkowany, więc sama zgodność komórek nie wystarcza:
# dołożona kolumna przesuwa środek i cała tabela jedzie w prawo. Liczymy, gdzie
# ląduje LEWY BRZEG TREŚCI (czyli dawna kolumna A) w obu wersjach.
MDW = 7


def col_widths(path):
    """
    Szerokości kolumn 1:1 z XML-a. openpyxl trzyma zakres `<col min max>` pod
    literą samego `min`, więc odpytanie go o środek zakresu zwraca None i cicho
    podstawia domyślne 8.43 — przy 60 kolumnach to błąd rzędu metra papieru.
    """
    with zipfile.ZipFile(path) as z:
        xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8", errors="replace")
    out = {}
    for col in re.finditer(r"<col ([^>]*)/>", xml):
        attrs = dict(re.findall(r'(\w+)="([^"]*)"', col.group(1)))
        for idx in range(int(attrs["min"]), int(attrs["max"]) + 1):
            out[idx] = float(attrs.get("width", 8.43))
    return out


def col_px(widths, first, last):
    return sum(
        int(((256 * widths.get(i, 8.43) + int(128.0 / MDW)) / 256.0) * MDW)
        for i in range(first, last + 1)
    )


def content_left_mm(ws, widths, first_col, last_col, body_first_col):
    """Odległość lewego brzegu TREŚCI od krawędzi kartki, w milimetrach."""
    scale = (ws.page_setup.scale or 100) / 100.0
    left_in = float(ws.page_margins.left)
    right_in = float(ws.page_margins.right)
    page_in = 210.0 / 25.4  # A4 pionowo
    block_in = col_px(widths, first_col, last_col) / 96.0 * scale
    usable_in = page_in - left_in - right_in
    block_left_in = left_in + (usable_in - block_in) / 2.0  # horizontalCentered
    lead_in = (
        col_px(widths, first_col, body_first_col - 1) / 96.0 * scale
        if body_first_col > first_col
        else 0.0
    )
    return (block_left_in + lead_in) * 25.4, block_in * 25.4


# stary: A..BG, treść zaczyna się w A; nowy: A..BH, treść zaczyna się w B
old_left, old_block = content_left_mm(ws_old, col_widths(OLD), 1, 59, 1)
new_left, new_block = content_left_mm(ws_new, col_widths(NEW), 1, 60, 2)
print(
    "szerokość wydruku: stary %.1f mm, nowy %.1f mm (A4 = 210 mm)"
    % (old_block, new_block)
)
print("lewy brzeg treści: stary %.2f mm, nowy %.2f mm" % (old_left, new_left))


def usable_mm(ws):
    return 210.0 - (float(ws.page_margins.left) + float(ws.page_margins.right)) * 25.4


old_usable = usable_mm(ws_old)
new_usable = usable_mm(ws_new)
print(
    "obszar wydruku: stary %.1f mm, nowy %.1f mm | zapas: %.1f mm -> %.1f mm"
    % (old_usable, new_usable, old_usable - old_block, new_usable - new_block)
)
# Klucz: zapas NIE MOŻE się zmniejszyć. Kolumna ptaszków bierze miejsce z lewego
# marginesu, więc obszar wydruku rośnie dokładnie o tyle, o ile rośnie tabela.
# Gdyby miejsce szło z prawej, zapas zmalałby o szerokość kolumny i prawy skraj
# protokołu wyjechałby na osobną stronę.
if (new_usable - new_block) < (old_usable - old_block) - 0.1:
    problems.append(
        "zapas na szerokość zmalał z %.1f mm do %.1f mm — kolumna ptaszków musi "
        "brać miejsce z LEWEGO marginesu, nie z prawego"
        % (old_usable - old_block, new_usable - new_block)
    )
if float(ws_new.page_margins.left) < 0:
    problems.append("lewy margines wyszedł ujemny — kolumna ptaszków za szeroka")
if abs(old_left - new_left) > 0.5:
    problems.append(
        "treść przesunęła się na papierze o %.2f mm — popraw kompensację "
        "prawego marginesu w shift_template.py" % (new_left - old_left)
    )

# 8) obszar wydruku
print("print_area stary:", ws_old.print_area, " nowy:", ws_new.print_area)
print("scale stary:", ws_old.page_setup.scale, " nowy:", ws_new.page_setup.scale)

print("sprawdzonych komórek:", checked)
print("scaleń:", len(new_m))
print("obrazków:", len(ws_new._images))
print("nowa kolumna A width:", new_a)

if problems:
    print("\n!!! PROBLEMY (%d):" % len(problems))
    for p in problems[:25]:
        print("  -", p)
    raise SystemExit(1)
print("\nOK — nowy szablon = stary przesunięty o 1 kolumnę.")
