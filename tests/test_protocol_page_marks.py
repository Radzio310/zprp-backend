"""
Nagłówek (IdZawody) i stopka („wygenerowano przez…") wydruku protokołu.

Najważniejszy jest `test_print_area_height_unchanged`: oba napisy mieszkają w
marginesach strony, a te są jedyną rzeczą, która dzieli je od treści. Arkusz
wypełnia wysokość A4 co do milimetra, więc każde zabranie miejsca obszarowi
druku wypycha ostatni wiersz protokołu na osobną stronę — i nic tego nie
zgłosi, PDF po prostu zrobi się dwustronicowy.
"""
import os
import re
import zipfile

import pytest
from openpyxl import load_workbook

from app.results import (
    PROTOCOL_BOTTOM_MARGIN_IN,
    PROTOCOL_FOOTER_MARGIN_IN,
    PROTOCOL_HEADER_MARGIN_IN,
    PROTOCOL_TOP_MARGIN_IN,
    _apply_protocol_page_marks,
    _hf_escape,
    _zprp_match_id,
)

TEMPLATE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "app",
    "templates",
    "protocol_template.xlsx",
)


# ─────────────────────── IdZawody ───────────────────────

def test_zprp_match_id_from_config():
    assert (
        _zprp_match_id({"matchConfig": {"matchId": "190756", "matchNumber": "PP/846"}})
        == "190756"
    )


def test_zprp_match_id_ignores_packet_id():
    """
    `data_json["id"]` to znacznik czasu nadawany przy budowaniu paczki. Wygląda
    jak identyfikator i nim nie jest — wpisany w nagłówek dałby liczbę, która
    wygląda wiarygodnie i nie znaczy nic.
    """
    assert _zprp_match_id({"id": "1755266053", "matchConfig": {"matchNumber": "PP/846"}}) == ""


def test_zprp_match_id_from_link():
    got = _zprp_match_id(
        {"matchConfig": {"extras": {"matchLink": "?a=zawody&b=protokol&IdZawody=123456"}}}
    )
    assert got == "123456"


def test_zprp_match_id_missing_is_empty():
    assert _zprp_match_id({}) == ""
    assert _zprp_match_id({"matchConfig": {"matchId": "brak"}}) == ""


# ─────────────────────── znaki formatujące ───────────────────────

def test_hf_escape_doubles_ampersand():
    """W nagłówku arkusza `&` otwiera kod formatujący — nazwisko z „&" nie może
    zamienić się w polecenie."""
    assert _hf_escape("Kowalski & Syn") == "Kowalski && Syn"


# ─────────────────────── nakładanie na skoroszyt ───────────────────────

def _prepared():
    wb = load_workbook(TEMPLATE)
    wb.copy_worksheet(wb.active)  # udajemy stronę uwag / rzutów karnych
    _apply_protocol_page_marks(
        wb,
        match_id="190756",
        generated_by="WITKOWICZ Radosław",
        generated_at="15.08.2026 15:54:13",
    )
    return wb


def test_marks_land_on_every_sheet():
    """Wyrwana kartka też musi być opisana — inaczej strona 2 jest anonimowa."""
    wb = _prepared()
    assert len(wb.worksheets) == 2
    for ws in wb.worksheets:
        assert ws.oddHeader.right.text == "IdZawody: 190756"
        assert ws.oddFooter.center.text == (
            "Wygenerowano automatycznie przez użytkownika WITKOWICZ Radosław "
            "dnia 15.08.2026 15:54:13"
        )


def test_print_area_height_unchanged():
    ws_before = load_workbook(TEMPLATE).active
    before = float(ws_before.page_margins.top) + float(ws_before.page_margins.bottom)
    after = PROTOCOL_TOP_MARGIN_IN + PROTOCOL_BOTTOM_MARGIN_IN
    assert abs(after - before) < 1e-9, (
        "suma marginesów pion zmieniła się o %.4f cala — protokół wyjedzie na "
        "drugą stronę" % (after - before)
    )


def test_header_and_footer_have_room():
    """Napisy siedzą MIĘDZY marginesem nagłówka/stopki a treścią. Zero miejsca
    = LibreOffice po prostu ich nie narysuje."""
    assert PROTOCOL_TOP_MARGIN_IN - PROTOCOL_HEADER_MARGIN_IN >= 0.10
    assert PROTOCOL_BOTTOM_MARGIN_IN - PROTOCOL_FOOTER_MARGIN_IN >= 0.10


def test_left_right_margins_untouched(tmp_path):
    """
    Szerokość jest wrażliwsza niż wysokość: kolumna ptaszków bierze miejsce z
    lewego marginesu, a zapas na prawym decyduje o tym, czy skraj protokołu nie
    wyjedzie na osobną stronę.
    """
    src = load_workbook(TEMPLATE).active
    wb = _prepared()
    assert float(wb.worksheets[0].page_margins.left) == float(src.page_margins.left)
    assert float(wb.worksheets[0].page_margins.right) == float(src.page_margins.right)


def test_marks_survive_save(tmp_path):
    """Nagłówek/stopka muszą wylądować w XML — LibreOffice czyta plik, nie obiekt."""
    out = tmp_path / "protocol.xlsx"
    _prepared().save(out)
    with zipfile.ZipFile(out) as z:
        xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8", errors="replace")
    assert "<headerFooter>" in xml
    assert "IdZawody: 190756" in xml
    m = re.search(r'<pageMargins[^>]*top="([0-9.]+)"[^>]*', xml)
    assert m and abs(float(m.group(1)) - PROTOCOL_TOP_MARGIN_IN) < 1e-9
