"""
Rubryka Kd (kara dodatkowa) w protokole.

Szablon ma nagłówek „Kd" w FIZYCZNEJ kolumnie AJ (AJ10 gospodarze, AJ34
goście). Kod pisze pod adresem „AI" przez nakładkę ShiftedWS, więc rozjazd
szablonu i PROTOCOL_COL_SHIFT wpisałby czas kary o kolumnę obok - bez żadnego
wyjątku. Ten plik pilnuje obu stron: że nagłówek stoi tam, gdzie kod celuje,
i że czas z `penaltyExtra` ląduje w komórce zawodnika, a nie u sąsiada.
"""
import os

from openpyxl import load_workbook

from app.results import ShiftedWS, _fill_players_block

TEMPLATE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "app",
    "templates",
    "protocol_template.xlsx",
)

TEMPLATE_2 = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "app",
    "templates",
    "protocol_template_2.xlsx",
)


def test_kd_header_sits_in_physical_aj():
    """Oba szablony trzymają nagłówek Kd w AJ - tam celuje kod przez shift."""
    for path in (TEMPLATE, TEMPLATE_2):
        ws = load_workbook(path).active
        assert ws["AJ10"].value == "Kd", path
        assert ws["AJ34"].value == "Kd", path


def test_penalty_extra_lands_in_kd_cell():
    wb = load_workbook(TEMPLATE)
    raw = wb.active
    ws = ShiftedWS(raw)

    _fill_players_block(
        ws,
        players=[7, 61],
        stats_by_number={
            7: {"entered": True, "goals": 1, "penaltyExtra": "12:34"},
            61: {"entered": True},
        },
        fullnames_by_number={7: "KOZAK Mikołaj", 61: "RAJCA Oliwier"},
        start_row=11,
        end_row=26,
        exam_by_number={},
        mark_ws=raw,
    )

    # Kod „AI" + shift 1 = fizyczne AJ, czyli rubryka pod nagłówkiem Kd.
    assert raw["AJ11"].value == "12:34"
    # Zawodnik bez kary dodatkowej dostaje kreski jak w każdej pustej rubryce.
    assert raw["AJ12"].value == "---"
    # Kd nie może wyciec do sąsiednich rubryk (D+N po lewej).
    assert raw["AG11"].value == "---"


def test_old_payload_without_penalty_extra_keeps_dashes():
    """Zapis sprzed tej rubryki nie niesie `penaltyExtra` - wydruk bez zmian."""
    wb = load_workbook(TEMPLATE)
    raw = wb.active
    ws = ShiftedWS(raw)

    _fill_players_block(
        ws,
        players=[7],
        stats_by_number={7: {"entered": True}},
        fullnames_by_number={7: "KOZAK Mikołaj"},
        start_row=11,
        end_row=26,
        exam_by_number={},
        mark_ws=raw,
    )

    assert raw["AJ11"].value == "---"
    # Pusty wiersz składu też trzyma kreskę w rubryce Kd.
    assert raw["AJ12"].value in ("---", None)
