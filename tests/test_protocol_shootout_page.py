"""Strona „RZUTY KARNE" w przebiegu zawodów.

Strona karnych powstaje z KOPII arkusza protokołu, więc każdy wiersz przebiegu,
którego nie nadpiszemy, zostaje pusty na wydruku. Pętle czyszczące kończyły się
na wierszu 61, a tabela sięga 63 - dwa ostatnie wiersze wychodziły z drukarki
puste, w środku tabeli pełnej „--".

To jest dokładnie ten rodzaj usterki, który nie rzuca wyjątku i który widać
dopiero na papierze, więc ma test.
"""
import os

from openpyxl import load_workbook

from app.results import (
    ShiftedWS,
    TIMELINE_END_ROW,
    TIMELINE_START_ROW,
    _fill_shootout_page,
)

TEMPLATE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "app",
    "templates",
    "protocol_template.xlsx",
)

#: Lewy blok startuje od 16 - wiersz 15 zajmuje scalony nagłówek „RZUTY KARNE".
LEFT_COLS = ("AL", "AN", "AP", "AS", "AU")
RIGHT_COLS = ("AW", "AY", "BA", "BD", "BF")


def _shootout_sheet(data_json):
    wb = load_workbook(TEMPLATE)
    # Tak jak w produkcji: strona karnych to kopia arkusza protokołu.
    copy = wb.copy_worksheet(wb.active)
    ws = ShiftedWS(copy)
    _fill_shootout_page(ws, data_json=data_json)
    return copy, ws


def test_oba_bloki_wypelnione_do_konca_tabeli():
    """Żaden wiersz przebiegu nie może zostać pusty - łącznie z dwoma ostatnimi."""
    raw, ws = _shootout_sheet({})

    puste = []
    for r in range(16, TIMELINE_END_ROW + 1):
        for c in LEFT_COLS:
            if ws[f"{c}{r}"].value in (None, ""):
                puste.append(f"{c}{r}")
    for r in range(TIMELINE_START_ROW, TIMELINE_END_ROW + 1):
        for c in RIGHT_COLS:
            if ws[f"{c}{r}"].value in (None, ""):
                puste.append(f"{c}{r}")

    assert puste == []


def test_ostatnie_dwa_wiersze_tez_maja_mysliniki():
    """Regresja wprost: 62 i 63 wypadały poza zakres czyszczenia."""
    raw, ws = _shootout_sheet({})

    for r in (TIMELINE_END_ROW - 1, TIMELINE_END_ROW):
        for c in LEFT_COLS + RIGHT_COLS:
            assert ws[f"{c}{r}"].value == "--", f"{c}{r}"


def test_naglowek_i_seria_karnych():
    data = {
        "penaltyStarterTeam": "host",
        "penaltyShots": {
            "host": [{"player": 7, "result": 1}],
            "guest": [{"player": 13, "result": 0}],
        },
    }
    raw, ws = _shootout_sheet(data)

    assert ws["AL15"].value == "RZUTY KARNE"
    # Pierwsza seria: gospodarz zaczyna, więc jego strzał ląduje w wierszu 16.
    assert ws["AL16"].value == "1"
    assert ws["AN16"].value == "7"
    assert ws["AL17"].value == "1"
    assert ws["AU17"].value == "13"
