"""Geometria kompaktowego protokolu: 16 zawodnikow i 43 wpisy przebiegu."""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.results import (
    GUEST_PLAYER_END_ROW,
    GUEST_PLAYER_START_ROW,
    HOST_PLAYER_END_ROW,
    HOST_PLAYER_START_ROW,
    PROTOCOL_PLAYER_ROWS,
    TIMELINE_END_ROW,
    TIMELINE_MAX_ROWS,
    TIMELINE_ROWS,
    TIMELINE_SKIP_ROWS,
    ShiftedWS,
    _fill_players_block,
    _fill_timeline_pages,
)


TEMPLATES = Path(__file__).resolve().parents[1] / "app" / "templates"
CASES = (
    ("protocol_template.xlsx", 67, 922.35),
    ("protocol_template_2.xlsx", 68, 935.55),
)


def _height_points(ws) -> float:
    default = ws.sheet_format.defaultRowHeight or 15
    return sum(
        ws.row_dimensions[row].height or default
        for row in range(1, ws.max_row + 1)
    )


@pytest.mark.parametrize("name,last_row,expected_height", CASES)
def test_compact_templates_keep_original_a4_height(name, last_row, expected_height):
    ws = load_workbook(TEMPLATES / name).active

    assert ws.max_row == last_row
    assert f"A1:BH{last_row}" in str(ws.print_area).replace("$", "")
    assert str(ws.page_setup.paperSize) == "9"
    assert ws.page_setup.orientation == "portrait"
    assert ws.page_setup.scale == 90
    assert _height_points(ws) == pytest.approx(expected_height)


@pytest.mark.parametrize("name,_,__", CASES)
def test_each_team_has_exactly_sixteen_taller_player_rows(name, _, __):
    ws = load_workbook(TEMPLATES / name).active

    assert PROTOCOL_PLAYER_ROWS == 16
    assert HOST_PLAYER_END_ROW - HOST_PLAYER_START_ROW + 1 == PROTOCOL_PLAYER_ROWS
    assert GUEST_PLAYER_END_ROW - GUEST_PLAYER_START_ROW + 1 == PROTOCOL_PLAYER_ROWS
    for row in (
        *range(HOST_PLAYER_START_ROW, HOST_PLAYER_END_ROW + 1),
        *range(GUEST_PLAYER_START_ROW, GUEST_PLAYER_END_ROW + 1),
    ):
        assert ws.row_dimensions[row].height == pytest.approx(14.85)


def test_timeline_has_43_slots_per_half_and_avoids_both_form_seams():
    assert TIMELINE_END_ROW == 59
    assert TIMELINE_SKIP_ROWS == {29, 53}
    assert TIMELINE_MAX_ROWS == 43
    assert len(TIMELINE_ROWS) == 43
    assert TIMELINE_ROWS[0] == 15
    assert TIMELINE_ROWS[-1] == 59


@pytest.mark.parametrize("name,_,__", CASES)
def test_timeline_seams_remain_merged_after_compaction(name, _, __):
    ws = load_workbook(TEMPLATES / name).active
    merges = {str(item) for item in ws.merged_cells.ranges}

    for ref in (
        "AS28:AS29",
        "BD28:BD29",
        "AS52:AS53",
        "BD52:BD53",
    ):
        assert ref in merges


def test_seventeenth_and_eighteenth_players_cannot_overwrite_companions():
    raw = load_workbook(TEMPLATES / "protocol_template.xlsx").active
    ws = ShiftedWS(raw)

    _fill_players_block(
        ws,
        players=list(range(1, 19)),
        stats_by_number={},
        fullnames_by_number={},
        start_row=HOST_PLAYER_START_ROW,
        end_row=HOST_PLAYER_END_ROW,
        exam_by_number={},
        mark_ws=None,
    )

    assert raw["B11"].value == 1
    assert raw["B26"].value == 16
    # Fizyczne B27 to etykieta osoby towarzyszacej A, nie 17. zawodnik.
    assert raw["B27"].value == "A"


def test_forty_fourth_event_starts_the_overflow_page():
    workbook = load_workbook(TEMPLATES / "protocol_template.xlsx")
    page1_raw = workbook.active
    page2_raw = workbook.copy_worksheet(page1_raw)
    events = [
        {
            "time": second * 1000,
            "half": 1,
            "type": "goal",
            "team": "host",
            "player": second,
        }
        for second in range(1, 45)
    ]

    assert _fill_timeline_pages(
        ShiftedWS(page1_raw),
        ShiftedWS(page2_raw),
        data_json={"protocol": events},
        half_ms=30 * 60 * 1000,
        half_score_host=0,
        half_score_guest=0,
    ) is True

    # 43. wpis konczy pierwsza strone; 44. zaczyna druga.
    assert page1_raw["AO59"].value == "43"
    assert page2_raw["AO15"].value == "44"
    # Pierwszy wiersz pod przebiegiem nadal jest naglowkiem rzutow karnych.
    assert page1_raw["AM60"].value == "A"
