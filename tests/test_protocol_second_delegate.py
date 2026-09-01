"""Wariant protokołu dla meczu z dwoma delegatami."""
from pathlib import Path

from openpyxl import load_workbook

from app.results import (
    SIGN_ANCHORS,
    _apply_fitted_text,
    _configure_protocol_page,
    _get_match_core,
    _protocol_template_name,
    shift_ref,
)


TEMPLATE = (
    Path(__file__).resolve().parents[1]
    / "app"
    / "templates"
    / "protocol_template_2.xlsx"
)
STANDARD_TEMPLATE = TEMPLATE.with_name("protocol_template.xlsx")


def _sheet_height_points(ws) -> float:
    return sum(
        (ws.row_dimensions[row].height or 15)
        for row in range(1, ws.max_row + 1)
    )


def test_second_delegate_selects_separate_template():
    assert _protocol_template_name({"delegate2": "ZIELIŃSKA Ewa"}) == (
        "protocol_template_2.xlsx"
    )
    assert _protocol_template_name({"delegate2": ""}) == "protocol_template.xlsx"


def test_second_delegate_name_survives_extras_only_shape():
    core = _get_match_core(
        {
            "matchConfig": {
                "extras": {
                    "officials": {
                        "delegate2": {"fullName": "ZIELIŃSKA Ewa"}
                    }
                }
            }
        }
    )
    assert core["delegate2"] == "ZIELIŃSKA Ewa"
    assert _protocol_template_name(core) == "protocol_template_2.xlsx"


def test_second_template_has_dedicated_row_and_a4_print_area():
    ws = load_workbook(TEMPLATE).active
    assert ws.max_row == 72
    assert "A1:BH72" in str(ws.print_area).replace("$", "")
    anchors = {str(r).split(":")[0] for r in ws.merged_cells.ranges}
    # Kotwicę podpisu bierzemy z kodu, a nie z literału: rubryki podpisów już
    # raz przesunięto (AJ..AL -> AH..AL) i wtedy ten test był jedynym miejscem,
    # które trzeba było poprawić ręcznie.
    for logical in ("I71", "W71", SIGN_ANCHORS["delegate2"]):
        assert shift_ref(logical) in anchors


def test_second_template_keeps_one_page_a4_settings():
    ws = load_workbook(TEMPLATE).active
    _configure_protocol_page(ws, has_second_delegate=True)
    assert str(ws.page_setup.paperSize) == "9"
    assert ws.page_setup.orientation == "portrait"
    assert ws.page_setup.scale == 88

    # Po przeskalowaniu wariant 72-wierszowy jest nawet odrobinę niższy niż
    # dotychczasowy, sprawdzony wydruk 71-wierszowy przy 90%. Nie ma więc
    # powodu, by LibreOffice przerzucił stopkę na drugą kartkę A4.
    standard = load_workbook(STANDARD_TEMPLATE).active
    assert _sheet_height_points(ws) * 0.88 <= (
        _sheet_height_points(standard) * 0.90
    )


def test_second_delegate_name_and_city_are_shrunk_not_wrapped():
    ws = load_workbook(TEMPLATE).active
    _apply_fitted_text(ws)
    for logical in ("I71", "W71"):
        alignment = ws[shift_ref(logical)].alignment
        assert alignment.shrinkToFit is True
        assert not alignment.wrapText
