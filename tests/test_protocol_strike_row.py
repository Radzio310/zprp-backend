"""Skreślenie wiersza na protokole — brak badań i wiersz bez zawodnika.

Ta sama kreska obsługuje dwa przypadki i oba muszą zostać rozróżnialne
w testach, bo psują się osobno:

  • zawodnik BEZ WAŻNYCH BADAŃ zostaje na wydruku - był zgłoszony, więc numer
    i nazwisko muszą być widoczne - ale jego wiersz jest przekreślony,
  • wiersz BEZ ZAWODNIKA też dostaje kreskę zamiast rzędu myślników w każdej
    rubryce; myślniki czytały się jak wypełnione pola z wartością „brak".

Pilnujemy przy tym:
  • zasięgu kreski (fizycznie od B do AL, czyli wszystko poza kolumną ptaszków),
  • tego, że kreska trafia w ŚRODEK wiersza, a nie na jego krawędź,
  • i najważniejszego wyjątku: pusta mapa badań znaczy „zapis nic nie wie
    o badaniach", a nie „nikt ich nie ma". Skreślenie całego składu na
    protokole sprzed wprowadzenia tej kolumny byłoby nieprawdą.
"""
import os

from openpyxl import load_workbook

from app.results import (
    _STRIKE_COL_FROM,
    _STRIKE_COL_TO,
    ShiftedWS,
    _fill_players_block,
)

TEMPLATE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "app",
    "templates",
    "protocol_template.xlsx",
)

#: Wiersze bloku użytego w testach: dwaj zawodnicy, reszta pusta.
START_ROW = 11
END_ROW = 26
ROW_PLAYER_1 = 11  # numer 7
ROW_PLAYER_2 = 12  # numer 61
EMPTY_ROWS = list(range(13, END_ROW + 1))


def _sheet():
    wb = load_workbook(TEMPLATE)
    raw = wb.active
    return raw, ShiftedWS(raw)


def _fill(raw, ws, exam_by_number):
    _fill_players_block(
        ws,
        players=[7, 61],
        stats_by_number={},
        fullnames_by_number={7: "KOZAK Mikołaj", 61: "RAJCA Oliwier"},
        start_row=START_ROW,
        end_row=END_ROW,
        exam_by_number=exam_by_number,
        mark_ws=raw,
    )


def _strikes(raw, base: int):
    """Kreski (kotwica DWUKOMÓRKOWA) dodane po `base`, jako zbiór numerów wierszy.

    Po kształcie kotwicy odróżniamy je od ptaszków badań, bez polegania na
    kolejności wstawiania.
    """
    out = {}
    for im in raw._images[base:]:
        anchor = im.anchor
        if getattr(anchor, "to", None) is None:
            continue
        out[anchor._from.row + 1] = anchor  # kotwica jest 0-based
    return out


def _marks(raw, base: int) -> int:
    """Ile ptaszków badań (kotwica JEDNOKOMÓRKOWA) doszło po `base`."""
    return sum(
        1 for im in raw._images[base:] if getattr(im.anchor, "to", None) is None
    )


def test_brak_badan_skresla_wiersz():
    raw, ws = _sheet()
    base = len(raw._images)

    # 7 ma badania, 61 nie - przy niepustej mapie to znaczy „sprawdzone i brak".
    _fill(raw, ws, {7: "zprp"})

    strikes = _strikes(raw, base)
    assert ROW_PLAYER_2 in strikes
    assert ROW_PLAYER_1 not in strikes
    assert _marks(raw, base) == 1  # ptaszek tylko dla 7


def test_kreska_idzie_od_B_do_AL():
    raw, ws = _sheet()
    base = len(raw._images)
    _fill(raw, ws, {7: "zprp"})

    anchor = _strikes(raw, base)[ROW_PLAYER_2]

    # Fizyczna kolumna B (0-based 1) do początku AM, czyli prawej krawędzi AL.
    assert anchor._from.col == _STRIKE_COL_FROM == 1
    assert anchor.to.col == _STRIKE_COL_TO == 38


def test_kreska_jest_w_srodku_wiersza():
    raw, ws = _sheet()
    base = len(raw._images)
    _fill(raw, ws, {7: "zprp"})

    anchor = _strikes(raw, base)[ROW_PLAYER_2]

    assert anchor._from.row == anchor.to.row == ROW_PLAYER_2 - 1
    # Cienki pasek w połowie wysokości, a nie na całą komórkę.
    assert anchor.to.rowOff > anchor._from.rowOff
    grubosc = anchor.to.rowOff - anchor._from.rowOff
    # Granica nie jest okrągła z przypadku: 9525 EMU to 0,75 pt, czyli cienka
    # krawędź tabeli w Excelu. Skreślenie MUSI być cieńsze - grubsza kreska
    # czyta się jak druga krawędź wiersza i rozmywa granicę między zawodnikami.
    assert 0 < grubosc < 9_525


def test_pusta_mapa_badan_nie_skresla_zawodnikow():
    """Sedno wyjątku: protokół sprzed kolumny ptaszków ma wyglądać jak dotąd.

    Wiersze BEZ zawodnika kreskę dostają zawsze - to inna sprawa niż badania.
    """
    raw, ws = _sheet()
    base = len(raw._images)

    _fill(raw, ws, {})

    strikes = _strikes(raw, base)
    assert ROW_PLAYER_1 not in strikes
    assert ROW_PLAYER_2 not in strikes
    assert _marks(raw, base) == 0


def test_brak_mapy_badan_nie_skresla_zawodnikow():
    raw, ws = _sheet()
    base = len(raw._images)

    _fill(raw, ws, None)

    strikes = _strikes(raw, base)
    assert ROW_PLAYER_1 not in strikes
    assert ROW_PLAYER_2 not in strikes


def test_wszyscy_z_badaniami_bez_kresek():
    raw, ws = _sheet()
    base = len(raw._images)

    _fill(raw, ws, {7: "zprp", 61: "manual"})

    strikes = _strikes(raw, base)
    assert ROW_PLAYER_1 not in strikes
    assert ROW_PLAYER_2 not in strikes
    assert _marks(raw, base) == 2


def test_numer_i_nazwisko_zostaja_mimo_skreslenia():
    """Skreślamy wiersz, nie kasujemy zawodnika - to dwie różne rzeczy."""
    raw, ws = _sheet()
    _fill(raw, ws, {7: "zprp"})

    assert raw["B12"].value == 61
    assert raw["D12"].value == "RAJCA Oliwier"


# ── Wiersze bez zawodnika ───────────────────────────────────────────────────


def test_puste_wiersze_dostaja_kreske():
    raw, ws = _sheet()
    base = len(raw._images)

    _fill(raw, ws, {7: "zprp", 61: "manual"})

    strikes = _strikes(raw, base)
    assert set(EMPTY_ROWS).issubset(strikes)


def test_pusty_wiersz_nie_ma_juz_mysinikow():
    """Kreska zastępuje myślniki, a nie dokłada się do nich."""
    raw, ws = _sheet()
    _fill(raw, ws, {7: "zprp", 61: "manual"})

    # Logiczne A/C/Q → fizyczne B/D/R (kolumna ptaszków przesuwa o jedną).
    assert raw["B13"].value is None
    assert raw["D13"].value is None
    assert raw["R13"].value is None


def test_puste_wiersze_dostaja_kreske_takze_bez_mapy_badan():
    """Brak informacji o badaniach nie może zostawić pustych rubryk pustymi."""
    raw, ws = _sheet()
    base = len(raw._images)

    _fill(raw, ws, None)

    strikes = _strikes(raw, base)
    assert set(EMPTY_ROWS).issubset(strikes)


def test_bez_arkusza_surowego_zostaja_mysliniki():
    """Bez `mark_ws` kreski nie ma czym narysować - wiersz nie może zostać pusty."""
    wb = load_workbook(TEMPLATE)
    raw = wb.active
    ws = ShiftedWS(raw)

    _fill_players_block(
        ws,
        players=[7],
        stats_by_number={},
        fullnames_by_number={7: "KOZAK Mikołaj"},
        start_row=START_ROW,
        end_row=END_ROW,
        exam_by_number={7: "zprp"},
        mark_ws=None,
    )

    assert raw["B13"].value == "--"
    assert raw["D13"].value.startswith("---")
