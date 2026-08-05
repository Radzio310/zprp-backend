"""
Testy kolumny ptaszków badań lekarskich w protokole.

Najważniejszy jest `test_template_matches_code`: łapie rozjazd między
szablonem protocol_template.xlsx a stałą PROTOCOL_COL_SHIFT. Taki rozjazd nie
rzuca żadnego wyjątku — protokół wygeneruje się do końca i będzie wyglądał
wiarygodnie, tylko każda wartość wyląduje o kolumnę obok.
"""
import os
import re
import zipfile

import pytest
from openpyxl import load_workbook

from app.results import (
    PROTOCOL_COL_SHIFT,
    EXAM_MARK_RGB,
    ShiftedWS,
    _exam_mark_png,
    _fill_players_block,
    _player_exam_map_from_cards,
    shift_ref,
)

TEMPLATE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "app",
    "templates",
    "protocol_template.xlsx",
)


# ─────────────────────── przesuwanie adresów ───────────────────────

@pytest.mark.parametrize(
    "src,expected",
    [
        ("A1", "B1"),
        ("B11", "C11"),
        ("Z63", "AA63"),
        ("AZ12", "BA12"),
        ("AL15:AV15", "AM15:AW15"),
        ("$A$1", "$B$1"),
        ("A1:B2", "B1:C2"),
    ],
)
def test_shift_ref(src, expected):
    assert shift_ref(src, 1) == expected


def test_shift_ref_zero_is_identity():
    assert shift_ref("AL15:AV15", 0) == "AL15:AV15"


def test_shift_ref_leaves_unknown_shapes_alone():
    # Nie chcemy, żeby dziwny kształt cicho zmienił się w coś innego.
    assert shift_ref("nonsens", 1) == "nonsens"


def test_shifted_ws_maps_reads_and_writes():
    wb = load_workbook(TEMPLATE)
    raw = wb.active
    ws = ShiftedWS(raw, 1)

    ws["A11"].value = 61
    assert raw["B11"].value == 61
    assert ws["A11"].value == 61
    assert raw["A11"].value is None  # kolumna ptaszków zostaje pusta


# ─────────────────────── mapa badań z kart ───────────────────────

def test_player_exam_map_from_cards():
    cards = [
        {"number": 1, "exam": "zprp"},
        {"number": 2, "exam": "wzpr"},
        {"number": 3, "exam": "manual"},
        {"number": 4, "exam": "none"},      # brak badań — bez ptaszka
        {"number": 5},                       # stara wersja aplikacji
        {"number": 6, "exam": "cokolwiek"},  # nieznana wartość
        {"number": "7", "exam": "zprp"},     # numer jako string
        "śmieć",
    ]
    assert _player_exam_map_from_cards(cards) == {
        1: "zprp",
        2: "wzpr",
        3: "manual",
        7: "zprp",
    }


def test_player_exam_map_empty_for_old_clients():
    assert _player_exam_map_from_cards([]) == {}
    assert _player_exam_map_from_cards(None) == {}


# ─────────────────────── rysunki ptaszków ───────────────────────

@pytest.mark.parametrize("kind", sorted(EXAM_MARK_RGB.keys()))
def test_exam_mark_png_is_a_png(kind):
    data = _exam_mark_png(kind)
    assert data[:8] == b"\x89PNG\r\n\x1a\n"


def test_exam_mark_png_unknown_kind_is_empty():
    assert _exam_mark_png("brak") == b""
    assert _exam_mark_png("") == b""


def test_exam_mark_manual_is_deterministic():
    # Ślad „długopisem” losuje rozjazd punktów, ale z ziarnem na sztywno —
    # inaczej każdy worker rysowałby inny ptaszek.
    from app import results

    results._EXAM_MARK_CACHE.clear()
    first = _exam_mark_png("manual")
    results._EXAM_MARK_CACHE.clear()
    second = _exam_mark_png("manual")
    assert first == second


def test_exam_marks_differ_between_kinds():
    assert _exam_mark_png("zprp") != _exam_mark_png("manual")
    assert _exam_mark_png("zprp") != _exam_mark_png("wzpr")


# ─────────────────────── szablon kontra kod ───────────────────────

def test_template_has_empty_lead_column():
    """Nowa kolumna A istnieje, jest wąska i pusta w całym arkuszu."""
    wb = load_workbook(TEMPLATE)
    ws = wb.active

    width = ws.column_dimensions["A"].width
    assert width is not None and width >= 2, "kolumna ptaszków za wąska: %r" % width

    for row in range(1, ws.max_row + 1):
        assert ws.cell(row=row, column=1).value is None, (
            "kolumna ptaszków nie jest pusta w wierszu %d" % row
        )


def test_template_keeps_layout_on_paper():
    """
    Kolumna ptaszków bierze miejsce z LEWEGO MARGINESU i nic więcej.

    Sprawdzamy dwie rzeczy, obie niewidoczne w wartościach komórek:
     1. lewy brzeg treści na papierze się nie ruszył (arkusz drukuje się
        wyśrodkowany, więc sama dołożona kolumna przesuwała tabelę w prawo),
     2. zapas na szerokość NIE zmalał — gdyby miejsce szło z prawego marginesu,
        obszar wydruku by się skurczył i prawy skraj protokołu wyjechałby na
        osobną stronę. Dokładnie to się stało przy pierwszej próbie.

    Liczby bezwzględne zależą od tego, jak generator PDF-a przelicza szerokości
    kolumn na piksele, więc testujemy RÓŻNICE między szablonami — one są od tego
    przelicznika niezależne.
    """
    old = os.path.join(os.path.dirname(TEMPLATE), "protocol_template_BACKUP_preExamCol.xlsx")
    assert os.path.exists(old), "brak backupu szablonu sprzed dodania kolumny"

    mdw = 7

    def widths(path):
        with zipfile.ZipFile(path) as z:
            xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8", errors="replace")
        assert 'horizontalCentered="1"' in xml, "arkusz przestał być wyśrodkowany"
        out = {}
        for col in re.finditer(r"<col ([^>]*)/>", xml):
            attrs = dict(re.findall(r'(\w+)="([^"]*)"', col.group(1)))
            for idx in range(int(attrs["min"]), int(attrs["max"]) + 1):
                out[idx] = float(attrs.get("width", 8.43))
        return out

    def px(w, first, last):
        return sum(
            int(((256 * w.get(i, 8.43) + int(128.0 / mdw)) / 256.0) * mdw)
            for i in range(first, last + 1)
        )

    def geometry_mm(path, last_col, body_first_col):
        ws = load_workbook(path).active
        w = widths(path)
        scale = (ws.page_setup.scale or 100) / 100.0
        left = float(ws.page_margins.left)
        right = float(ws.page_margins.right)
        assert left >= 0, "lewy margines wyszedł ujemny — kolumna za szeroka"
        block = px(w, 1, last_col) / 96.0 * scale
        usable = 210.0 / 25.4 - left - right
        lead = px(w, 1, body_first_col - 1) / 96.0 * scale if body_first_col > 1 else 0.0
        content_left = (left + (usable - block) / 2.0 + lead) * 25.4
        return content_left, (usable - block) * 25.4

    old_left, old_slack = geometry_mm(old, 59, 1)
    new_left, new_slack = geometry_mm(TEMPLATE, 60, 2)

    assert abs(new_left - old_left) < 0.5, (
        "treść protokołu przesunęła się na papierze o %.2f mm" % (new_left - old_left)
    )
    assert new_slack > old_slack - 0.5, (
        "zapas na szerokość zmalał z %.1f mm do %.1f mm — kolumna ptaszków musi "
        "brać miejsce z LEWEGO marginesu, nie z prawego" % (old_slack, new_slack)
    )


def test_template_matches_code():
    """
    Numer zawodnika musi wylądować w FIZYCZNEJ kolumnie B, a nazwisko w D.
    Gdyby szablon i PROTOCOL_COL_SHIFT się rozjechały, ten test padnie —
    a wygenerowany protokół nie.
    """
    assert PROTOCOL_COL_SHIFT == 1

    wb = load_workbook(TEMPLATE)
    raw = wb.active
    ws = ShiftedWS(raw)
    # Szablon ma już jeden osadzony obrazek (logo ZPRP) — liczymy przyrost.
    base_images = len(raw._images)

    _fill_players_block(
        ws,
        players=[7, 61],
        stats_by_number={7: {"entered": True, "goals": 3}},
        fullnames_by_number={7: "KOZAK Mikołaj", 61: "RAJCA Oliwier"},
        start_row=11,
        end_row=28,
        exam_by_number={7: "zprp", 61: "manual"},
        mark_ws=raw,
    )

    # numer i nazwisko przesunięte o jedną kolumnę
    assert raw["B11"].value == 7
    assert raw["D11"].value == "KOZAK Mikołaj"
    assert raw["B12"].value == 61
    assert raw["D12"].value == "RAJCA Oliwier"

    # statystyki też: dawne Q/S -> R/T
    assert raw["R11"].value == "W"
    assert raw["T11"].value == 3

    # kolumna ptaszków zostaje bez wartości — ptaszek to obrazek
    assert raw["A11"].value is None
    assert raw["A12"].value is None

    # dwa ptaszki, zakotwiczone w fizycznej kolumnie A
    marks = raw._images[base_images:]
    assert len(marks) == 2
    for img in marks:
        assert img.anchor._from.col == 0


def test_players_without_exams_get_no_marks():
    wb = load_workbook(TEMPLATE)
    raw = wb.active
    ws = ShiftedWS(raw)
    base_images = len(raw._images)

    _fill_players_block(
        ws,
        players=[7, 61],
        stats_by_number={},
        fullnames_by_number={},
        start_row=11,
        end_row=28,
        exam_by_number={},
        mark_ws=raw,
    )
    assert len(raw._images) == base_images


def test_marks_survive_worksheet_copy():
    """
    Strona 2 i strona rzutów karnych powstają przez wb.copy_worksheet, a obrazki
    dokłada _copy_images_safe. Ptaszki muszą się na nie przenieść — czyli blok
    zawodników musi lecieć PRZED kopiowaniem arkusza.
    """
    from app.results import _copy_images_safe

    wb = load_workbook(TEMPLATE)
    raw = wb.active
    ws = ShiftedWS(raw)
    base_images = len(raw._images)

    _fill_players_block(
        ws,
        players=[7, 61],
        stats_by_number={},
        fullnames_by_number={},
        start_row=11,
        end_row=28,
        exam_by_number={7: "zprp", 61: "wzpr"},
        mark_ws=raw,
    )
    assert len(raw._images) == base_images + 2

    page2 = wb.copy_worksheet(raw)
    _copy_images_safe(raw, page2)

    assert len(page2._images) == len(raw._images)
    assert page2["B11"].value == 7


def test_old_client_payload_produces_no_marks():
    """Stara aplikacja nie wysyła `exam` — wydruk ma wyglądać jak dotychczas."""
    wb = load_workbook(TEMPLATE)
    raw = wb.active
    ws = ShiftedWS(raw)
    base_images = len(raw._images)

    _fill_players_block(
        ws,
        players=[7],
        stats_by_number={},
        fullnames_by_number={7: "KOZAK Mikołaj"},
        start_row=11,
        end_row=28,
        exam_by_number=_player_exam_map_from_cards([{"number": 7}]),
        mark_ws=raw,
    )
    assert raw["B11"].value == 7
    assert len(raw._images) == base_images
