"""Kotwice podpisów kontra rubryki w szablonie protokołu.

Adresy w kodzie są LOGICZNE - `ShiftedWS` dokłada im jedną kolumnę przy
zapisie. Kotwica obrazu musi trafiać w lewą górną komórkę scalonej rubryki:
zakotwiczona w środku scalenia przesuwa podpis w prawo i wyprowadza go poza
rubrykę, a wtedy nikt tego nie zauważy przy generowaniu - dopiero na wydruku.

Tak się to zepsuło ostatnio: rubryki podpisów oficjeli poszerzono z AJ..AL na
AH..AL kosztem miejscowości (X..AE -> X..AC), a kotwice zostały na starym
miejscu. Ten test wyłapuje każdą kolejną przebudowę tych rubryk.
"""

from pathlib import Path

import openpyxl
import pytest
from openpyxl.utils import column_index_from_string, get_column_letter

from app.results import (
    MEDIC_ROLE_CELL,
    OFFICIAL_SIGN_MAX_H_PX,
    OFFICIAL_SIGN_MAX_W_PX,
    SIGN_ANCHORS,
    shift_ref,
)

TEMPLATES = Path(__file__).resolve().parents[1] / "app" / "templates"

#: Oficjele mają rubryki w wierszach 62..66; wariant z drugim delegatem dokłada 67.
OFFICIAL_KEYS = ["referee1", "referee2", "secretary", "timekeeper", "delegate"]

#: Szerokość kolumny rubryki podpisu w znakach - wszystkie X..AL są równe.
#: 16 px na kolumnę to ta sama arytmetyka, którą trzyma ramka medyka
#: (AF..AL = 7 kolumn = 112 px w kodzie).
COL_PX = 16


def _load(name: str):
    wb = openpyxl.load_workbook(TEMPLATES / name)
    return wb.active


def _merged_range_at(ws, coord: str):
    """Scalenie, którego częścią jest ta komórka (albo None)."""
    for rng in ws.merged_cells.ranges:
        if coord in rng:
            return rng
    return None


def _split(coord: str):
    letters = "".join(ch for ch in coord if ch.isalpha())
    digits = "".join(ch for ch in coord if ch.isdigit())
    return letters, int(digits)


@pytest.mark.parametrize(
    "template,keys",
    [
        ("protocol_template.xlsx", OFFICIAL_KEYS),
        ("protocol_template_2.xlsx", OFFICIAL_KEYS + ["delegate2"]),
    ],
)
def test_kotwica_podpisu_to_lewy_gorny_rog_rubryki(template, keys):
    ws = _load(template)
    for key in keys:
        raw = shift_ref(SIGN_ANCHORS[key])
        rng = _merged_range_at(ws, raw)
        assert rng is not None, f"{template}: {key} ({raw}) nie leży w żadnej scalonej rubryce"
        top_left = f"{get_column_letter(rng.min_col)}{rng.min_row}"
        assert raw == top_left, (
            f"{template}: kotwica {key} to {raw}, a rubryka zaczyna się w {top_left} "
            f"({rng.coord}) - podpis wyjedzie poza nią"
        )


@pytest.mark.parametrize(
    "template,keys",
    [
        ("protocol_template.xlsx", OFFICIAL_KEYS),
        ("protocol_template_2.xlsx", OFFICIAL_KEYS + ["delegate2"]),
    ],
)
def test_ramka_podpisu_nie_wychodzi_poza_rubryke(template, keys):
    """Szerokość ramki musi się zmieścić w scaleniu, co do piksela.

    Za rubryką oficjela (kolumna AM) zaczyna się blok RZUTÓW KARNYCH - podpis,
    który tam wjedzie, przykryje cudzy tekst.
    """
    ws = _load(template)
    for key in keys:
        rng = _merged_range_at(ws, shift_ref(SIGN_ANCHORS[key]))
        assert rng is not None
        width_px = (rng.max_col - rng.min_col + 1) * COL_PX
        assert OFFICIAL_SIGN_MAX_W_PX <= width_px, (
            f"{template}: ramka {OFFICIAL_SIGN_MAX_W_PX} px nie mieści się w rubryce "
            f"{rng.coord} ({width_px} px)"
        )


@pytest.mark.parametrize(
    "template", ["protocol_template.xlsx", "protocol_template_2.xlsx"]
)
def test_ramka_podpisu_nie_wchodzi_w_wiersz_nizej(template):
    """Wysokość ramki kontra wysokość wiersza oficjela (pt -> px przy 96 dpi)."""
    ws = _load(template)
    default_pt = ws.sheet_format.defaultRowHeight or 13.2
    for key in OFFICIAL_KEYS:
        _, row = _split(SIGN_ANCHORS[key])
        dim = ws.row_dimensions.get(row)
        pt = dim.height if dim and dim.height else default_pt
        px = pt * 96.0 / 72.0
        assert OFFICIAL_SIGN_MAX_H_PX <= px + 1, (
            f"{template}: ramka {OFFICIAL_SIGN_MAX_H_PX} px jest wyższa niż wiersz "
            f"{row} ({px:.1f} px)"
        )


@pytest.mark.parametrize(
    "template", ["protocol_template.xlsx", "protocol_template_2.xlsx"]
)
def test_medyk_rola_i_podpis_maja_swoje_rubryki(template):
    """Rola medyka i jego podpis to dwa SĄSIADUJĄCE scalenia (V..AE i AF..AL).

    Gdy rubryki podpisu przesunięto w lewo, kotwica została na starym miejscu
    i podpis lądował w połowie rubryki roli. Tu pilnujemy obu naraz: rola musi
    zaczynać swoje scalenie, podpis swoje, i muszą do siebie przylegać.
    """
    ws = _load(template)

    role_raw = shift_ref(MEDIC_ROLE_CELL)
    role_rng = _merged_range_at(ws, role_raw)
    assert role_rng is not None, f"{template}: rola medyka ({role_raw}) poza scaleniem"
    assert role_raw == f"{get_column_letter(role_rng.min_col)}{role_rng.min_row}"

    sign_raw = shift_ref(SIGN_ANCHORS["medic"])
    sign_rng = _merged_range_at(ws, sign_raw)
    assert sign_rng is not None, f"{template}: podpis medyka ({sign_raw}) poza scaleniem"
    assert sign_raw == f"{get_column_letter(sign_rng.min_col)}{sign_rng.min_row}"

    assert sign_rng.min_col == role_rng.max_col + 1, (
        f"{template}: między rolą ({role_rng.coord}) a podpisem ({sign_rng.coord}) "
        "jest luka albo zachodzą na siebie"
    )

    # Ramka podpisu medyka (112 px w kodzie) kontra jego rubryka.
    width_px = (sign_rng.max_col - sign_rng.min_col + 1) * COL_PX
    assert width_px >= 112, (
        f"{template}: rubryka podpisu medyka {sign_rng.coord} ma {width_px} px, "
        "a kod rysuje 112 px"
    )


@pytest.mark.parametrize(
    "template", ["protocol_template.xlsx", "protocol_template_2.xlsx"]
)
def test_miejscowosc_trafia_w_swoja_rubryke(template):
    """Miejscowość oficjela ma własne scalenie i kod musi celować w jego początek."""
    ws = _load(template)
    rows = [62, 63, 64, 65, 66]
    if template.endswith("_2.xlsx"):
        rows.append(67)
    for row in rows:
        raw = shift_ref(f"W{row}")
        rng = _merged_range_at(ws, raw)
        assert rng is not None, f"{template}: W{row} ({raw}) poza scaleniem"
        top_left = f"{get_column_letter(rng.min_col)}{rng.min_row}"
        assert raw == top_left, (
            f"{template}: miejscowość {raw} nie jest początkiem rubryki {rng.coord}"
        )
        # Rubryka miejscowości sąsiaduje z etykietą „podpis:" - gdyby ktoś ją
        # rozciągnął z powrotem, etykieta znalazłaby się w środku scalenia.
        label_col = get_column_letter(rng.max_col + 1)
        label = ws[f"{label_col}{row}"].value
        assert label == "podpis:", (
            f"{template}: za rubryką miejscowości ({rng.coord}) spodziewamy się "
            f"etykiety podpisu, a jest {label!r}"
        )
