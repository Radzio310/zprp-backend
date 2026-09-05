"""Strona szczegółowych uwag sędziów - układ dopasowany do długości opisu.

Wcześniej ta strona miała układ sztywny: opis w scalonej komórce A3:N30 i
podpisy zawsze w wierszach 32-37. Krótka uwaga zostawiała pół pustej kartki
między tekstem a podpisami, a długa nie miała gdzie się zmieścić.

Testy pilnują trzech rzeczy, których nie widać po samym otwarciu pliku:
  • podpisy idą ZARAZ pod opisem, a nie na sztywnej pozycji,
  • przy długim opisie strona łamie się sama, a blok podpisów przechodzi na
    następną kartkę W CAŁOŚCI (nazwisko bez podpisu to nie jest podpis),
  • każda strona uwag dostaje własny nagłówek, więc numeracja „Strona X/Y"
    ma gdzie usiąść.
"""
import pytest
from openpyxl import Workbook

from app.results import (
    _NOTES_LINE_BUDGET,
    _NOTES_PAGE_BUDGET_PT,
    _NOTES_SHEET_PT,
    _NOTES_SIGN_BLOCK_PT,
    _create_detailed_notes_sheet,
    _text_width_em1000,
    _wrap_notes_text,
)

REF1 = "WITKOWICZ Radosław"
REF2 = "WITKOWICZ Krzysztof"


def _build(text: str, *, place: str = "Mysłowice", date_ddmmyyyy: str = "20.08.2026"):
    wb = Workbook()
    ws, header_rows = _create_detailed_notes_sheet(
        wb,
        date_ddmmyyyy=date_ddmmyyyy,
        place=place,
        notes_text=text,
        referee1_name=REF1,
        referee2_name=REF2,
        referee1_sig_bytes=b"",
        referee2_sig_bytes=b"",
    )
    return ws, header_rows


def _rows_with(ws, value: str):
    return [
        c.row
        for row in ws.iter_rows()
        for c in row
        if isinstance(c.value, str) and c.value.strip() == value
    ]


def test_krotki_opis_podpisy_tuz_pod_tekstem():
    ws, header_rows = _build(
        "Zawodnik drużyny gospodarzy z numerem 34 otrzymał karę dyskwalifikacji."
    )

    assert len(header_rows) == 1  # wszystko na jednej kartce

    last_text_row = max(
        c.row
        for row in ws.iter_rows()
        for c in row
        if isinstance(c.value, str) and "dyskwalifikacji" in c.value
    )
    first_name_row = _rows_with(ws, REF1)[0]

    # Jeden wiersz odstępu - nie kilkanaście jak w sztywnym układzie.
    assert first_name_row - last_text_row == 2


def test_podpisy_sedziow_blisko_siebie():
    ws, _ = _build("Krótka uwaga.")

    r1 = _rows_with(ws, REF1)[0]
    r2 = _rows_with(ws, REF2)[0]

    # nazwisko + podpis + przerwa + nazwisko = 3 wiersze odstępu.
    assert r2 - r1 == 3


def test_dlugi_opis_lamie_strone_a_podpisy_ida_w_calosci():
    # Tyle tekstu, że sam opis przekracza jedną kartkę.
    text = " ".join(f"zdanie numer {i} o przebiegu zawodów" for i in range(220))
    ws, header_rows = _build(text)

    assert len(header_rows) >= 2
    assert ws.row_breaks.count >= 1

    r1 = _rows_with(ws, REF1)[0]
    r2 = _rows_with(ws, REF2)[0]
    last_page_start = header_rows[-1]

    # Cały blok podpisów na tej samej, ostatniej stronie.
    assert r1 > last_page_start
    assert r2 > last_page_start


def test_kazda_strona_ma_wlasny_naglowek():
    text = " ".join(f"wiersz {i} opisu zdarzenia" for i in range(220))
    ws, header_rows = _build(text)

    for row in header_rows:
        assert ws.cell(row=row, column=1).value == "Strona"
        assert ws.cell(row=row, column=7).value == "Mysłowice, 20.08.2026"
        # Data kończy się na tej samej krawędzi co opis, czyli na kolumnie N.
        assert any(
            str(rng) == "G%d:N%d" % (row, row) for rng in ws.merged_cells.ranges
        )


@pytest.mark.parametrize("place,date_ddmmyyyy,expected", [
    ("Mysłowice", "20.08.2026", "Mysłowice, 20.08.2026"),
    ("Mysłowice", "", "Mysłowice"),
    ("", "20.08.2026", "20.08.2026"),
    ("", "", ""),
])
def test_naglowek_miejscowosc_przed_data_bez_zbednego_przecinka(place, date_ddmmyyyy, expected):
    ws, header_rows = _build("Krótka uwaga.", place=place, date_ddmmyyyy=date_ddmmyyyy)
    assert ws.cell(row=header_rows[0], column=7).value == expected


def test_blok_podpisow_nie_wisi_na_krawedzi_strony():
    """Opis dobrany tak, by kończył się tuż przed dolną krawędzią kartki."""
    # Szukamy długości, przy której podpisy MUSZĄ zejść na kolejną stronę.
    text = " ".join(f"linia {i} treści" for i in range(150))
    ws, header_rows = _build(text)

    r1 = _rows_with(ws, REF1)[0]
    if len(header_rows) > 1:
        assert r1 > header_rows[-1]
    # Blok podpisów nigdy nie jest wyższy niż budżet strony - inaczej nie
    # zmieściłby się nigdzie i pętla szukania miejsca nie miałaby końca.
    assert _NOTES_SIGN_BLOCK_PT < _NOTES_PAGE_BUDGET_PT


def test_zawijanie_zachowuje_akapity():
    lines = _wrap_notes_text("pierwszy akapit\n\ndrugi akapit", _NOTES_LINE_BUDGET)
    assert lines[0] == "pierwszy akapit"
    assert lines[1] == ""
    assert lines[2] == "drugi akapit"


def test_zawijanie_tnie_slowo_dluzsze_niz_wiersz():
    long_word = "x" * 400
    lines = _wrap_notes_text(long_word, _NOTES_LINE_BUDGET)
    assert all(_text_width_em1000(l) <= _NOTES_LINE_BUDGET for l in lines)
    assert "".join(lines) == long_word


def test_zawijanie_nie_gubi_slow():
    text = " ".join(f"slowo{i}" for i in range(400))
    lines = _wrap_notes_text(text, _NOTES_LINE_BUDGET)
    assert " ".join(lines).split() == text.split()


def test_zadna_linia_nie_wyjezdza_poza_rubryke():
    """Rubryka opisu to komórka SCALONA - za długa linia nie zawija się ani nie
    wylewa na sąsiednie, tylko znika z wydruku. Także wtedy, gdy sędzia pisze
    samymi wersalikami, a te są najszersze."""
    text = " ".join(["WWWWWWWWWW", "MMMMMMMMMM", "ZAWODNIK", "DYSKWALIFIKACJA"] * 30)
    for line in _wrap_notes_text(text, _NOTES_LINE_BUDGET):
        assert _text_width_em1000(line) <= _NOTES_LINE_BUDGET


def test_waskie_litery_wypelniaja_linie():
    """Sedno przejścia z liczenia znaków na pomiar: „il" zajmuje trzy razy mniej
    miejsca niż „WM", więc takich znaków musi się w linii zmieścić znacznie
    więcej niż dawne 64."""
    lines = _wrap_notes_text(" ".join(["il"] * 200), _NOTES_LINE_BUDGET)
    assert len(lines[0]) > 100


def test_linia_zostawia_zapas_do_krawedzi_arkusza():
    """Budżet linii jest węższy od arkusza - metryka kroju po podstawieniu go
    przez konwerter nie musi zgadzać się co do promila."""
    assert _NOTES_LINE_BUDGET * 12 / 1000.0 < _NOTES_SHEET_PT


OPIS_208136 = (
    "Zawodnik drużyny gospodarzy z numerem 11 Balicki Grzegorz otrzymał karę "
    "dyskwalifikacji w 37 min. 33 sekundzie meczu za niebezpieczne, lekkomyślne "
    "spowodowanie upadku zawodnika znajdującego się w ataku na podstawie "
    "przepisu 8:5a."
)


def test_opis_z_protokolu_208136_wypelnia_linie_i_nie_gubi_slow():
    """Ten opis wyszedł najpierw w połowie szerokości kartki, a po poprawce
    z uciętymi końcówkami wyrazów („otrzymał karę dyskwa")."""
    lines = _wrap_notes_text(OPIS_208136, _NOTES_LINE_BUDGET)

    # Nic nie wystaje poza rubrykę - czyli nic nie zostanie ucięte.
    assert all(_text_width_em1000(l) <= _NOTES_LINE_BUDGET for l in lines)
    # Ani jedno słowo nie zostało rozcięte.
    assert " ".join(lines).split() == OPIS_208136.split()
    # Pełne linie naprawdę dochodzą do krawędzi rubryki, a nie kończą się
    # w trzech czwartych szerokości.
    assert all(
        _text_width_em1000(l) > 0.85 * _NOTES_LINE_BUDGET for l in lines[:-1]
    )


def test_krój_opisu_jest_ten_sam_co_zmierzona_tablica_szerokości():
    """Cała arytmetyka szerokości stoi na metryce JEDNEGO kroju.

    Podmiana czcionki opisu bez podmiany tablicy nie daje żadnego objawu poza
    wydrukiem z uciętymi słowami - dokładnie tym, który już raz wyszedł, gdy
    LibreOffice podstawił DejaVu Serif za nieistniejący w kontenerze Times.
    """
    from app.results import _NOTES_BODY_FONT, _NOTES_CHAR_WIDTHS

    assert _NOTES_BODY_FONT == "DejaVu Serif"
    # Kilka znaków kontrolnych wprost z pliku kroju.
    assert _NOTES_CHAR_WIDTHS["W"] == 1028
    assert _NOTES_CHAR_WIDTHS["i"] == 320
    assert _NOTES_CHAR_WIDTHS[" "] == 318


def test_nazwiska_sedziow_bezszeryfowo_i_bez_pogrubienia():
    ws, _ = _build("Krótka uwaga.")

    row = _rows_with(ws, REF1)[0]
    cell = ws.cell(row=row, column=9)
    assert cell.font.name == "DejaVu Sans"
    assert not cell.font.bold
    # Podwójne nazwisko nie ma dokąd się wylać z komórki scalonej.
    assert cell.alignment.shrinkToFit


def test_podpis_miesci_sie_w_szerokosci_kartki():
    """Guard przed pustą stroną: obrazek nie ma prawa wystawać poza kolumnę N.

    Renderer nie przycina tego, co wystaje w prawo - wypycha to na dodatkową
    stronę. Objawem był jeden zawijas samotnie stojący na czwartej kartce.
    """
    from openpyxl.utils import column_index_from_string

    from app.results import (
        _NOTES_SIGN_ANCHOR_COL,
        _NOTES_SIGN_MAX_W_PX,
        _excel_col_px,
    )

    ws, _ = _build("Krótka uwaga.")

    first = column_index_from_string(_NOTES_SIGN_ANCHOR_COL)
    last = column_index_from_string("N")
    available = sum(
        _excel_col_px(ws.column_dimensions[chr(ord("A") + c - 1)].width)
        for c in range(first, last + 1)
    )

    assert _NOTES_SIGN_MAX_W_PX < available


def test_arkusz_nie_wychodzi_poza_szerokosc_A4():
    """Sedno „pustej czwartej strony".

    Wszystko, co nie mieści się w szerokości obszaru druku, renderer albo
    przenosi na kolejną kartkę (pusta strona z pionowym paskiem papieru), albo
    pomniejsza całą stronę - a wtedy rozjeżdża się rachunek wysokości liczony
    w punktach.

    Przelicznik jest ZMIERZONY na gotowym PDF (patrz `_NOTES_COL_PX_PER_UNIT`);
    wcześniej stała tu wartość ze wzoru dla Excela, o 7% za mała, i test
    przepuszczał arkusz szerszy od kartki.
    """
    from app.results import _excel_col_px

    ws, _ = _build("Krótka uwaga.")

    szerokosc_px = sum(
        _excel_col_px(ws.column_dimensions[chr(ord("A") + i)].width)
        for i in range(14)
    )
    obszar_druku_cali = 8.268 - ws.page_margins.left - ws.page_margins.right

    assert szerokosc_px / 96.0 <= obszar_druku_cali


def test_linia_opisu_ma_zapas_do_krawedzi_arkusza():
    """Rubryka opisu sięga krawędzi arkusza, więc cały zapas na obciętą literę
    musi siedzieć w budżecie linii."""
    from app.results import _NOTES_SHEET_PT

    szerokosc_linii_pt = _NOTES_LINE_BUDGET * 12 / 1000.0
    assert szerokosc_linii_pt < _NOTES_SHEET_PT * 0.95
