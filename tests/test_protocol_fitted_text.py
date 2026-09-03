"""
Pola protokołu, które muszą zostać w jednym wierszu.

Szablon jest pod tym względem niespójny sam ze sobą: „osoba towarzysząca A"
ma `wrapText`, B i E mają `shrinkToFit`, C i D nie mają nic. Dlatego to samo
nazwisko łamie się na dwa wiersze w rubryce A, a w rubryce obok mieści się w
jednej linii. Te testy pilnują, żeby po naszym przebiegu wszystkie rubryki
zachowywały się tak samo.
"""
import os
import re
import zipfile

import pytest
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from app.results import (
    FITTED_TEXT_CELLS,
    _apply_fitted_text,
    _apply_protocol_page_marks,
    shift_ref,
)

TEMPLATE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "app",
    "templates",
    "protocol_template.xlsx",
)


def _ws():
    return load_workbook(TEMPLATE).active


# ─────────────────────── lista kontra szablon ───────────────────────

def test_every_fitted_cell_exists_in_template():
    """Literówka w adresie nie rzuci wyjątku — po prostu nic nie zrobi. Ten
    test jest jedyną rzeczą, która by to zauważyła."""
    ws = _ws()
    anchors = {str(r).split(":")[0] for r in ws.merged_cells.ranges}
    for logical in FITTED_TEXT_CELLS:
        phys = shift_ref(logical)
        assert phys in anchors, (
            "%s (fizycznie %s) nie jest lewym górnym rogiem żadnego scalenia — "
            "ustawienie wyrównania nie zadziała" % (logical, phys)
        )


def test_fitted_cells_are_unique():
    assert len(FITTED_TEXT_CELLS) == len(set(FITTED_TEXT_CELLS))


def test_covers_referee_cities_and_names():
    """Miejscowość sędziego to przypadek, od którego się zaczęło."""
    for row in range(62, 67):
        assert "W%d" % row in FITTED_TEXT_CELLS
        assert "I%d" % row in FITTED_TEXT_CELLS


def test_covers_all_companion_letters_both_teams():
    """Braki w tej liście objawiają się dokładnie tym, co naprawiamy: jedna
    rubryka łamie tekst, sąsiednia nie."""
    for col in ("B", "K", "R", "Y", "AF"):
        assert "%s27" % col in FITTED_TEXT_CELLS  # nazwisko, gospodarze
        assert "%s51" % col in FITTED_TEXT_CELLS  # nazwisko, goście
    for col in ("A", "J", "Q", "X", "AE"):
        for row in (28, 29, 52, 53):  # funkcja i licencja, obie drużyny
            assert "%s%d" % (col, row) in FITTED_TEXT_CELLS


# ─────────────────────── działanie ───────────────────────

def test_apply_sets_shrink_and_clears_wrap():
    ws = _ws()
    # W szablonie „osoba A" ma zawijanie — to źródło łamania na dwa wiersze.
    assert ws[shift_ref("B27")].alignment.wrapText is True

    _apply_fitted_text(ws)

    for logical in FITTED_TEXT_CELLS:
        al = ws[shift_ref(logical)].alignment
        assert al.shrinkToFit is True, logical
        assert not al.wrapText, logical


def test_apply_keeps_font_and_horizontal_alignment():
    """`shrinkToFit` ma zmniejszać tylko to, co się nie mieści — reszta rubryki
    zostaje taka, jak zaprojektowano."""
    ws = _ws()
    before = {
        logical: (
            ws[shift_ref(logical)].font.size,
            ws[shift_ref(logical)].alignment.horizontal,
            ws[shift_ref(logical)].alignment.vertical,
        )
        for logical in FITTED_TEXT_CELLS
    }
    _apply_fitted_text(ws)
    for logical, (size, horiz, vert) in before.items():
        cell = ws[shift_ref(logical)]
        assert cell.font.size == size, logical
        assert cell.alignment.horizontal == horiz, logical
        assert cell.alignment.vertical == vert, logical


def test_hall_address_keeps_wrapping():
    """
    Adres hali NIE może trafić na tę listę: jego rubryka ma 59,5 mm przy
    tekście 59,3 mm, czyli jest zaprojektowana pod dwa wiersze. Wyłączenie
    zawijania skróciłoby adres zamiast go pokazać.
    """
    ws = _ws()
    _apply_fitted_text(ws)
    assert ws[shift_ref("AL4")].alignment.wrapText is True
    assert "AL4" not in FITTED_TEXT_CELLS


def test_applied_to_every_sheet():
    """Strona uwag i strona rzutów karnych powstają przez `copy_worksheet` —
    dopasowanie musi objąć także je."""
    wb = load_workbook(TEMPLATE)
    wb.copy_worksheet(wb.active)
    _apply_protocol_page_marks(
        wb, match_id="190756", generated_by="X", generated_at="15.08.2026 21:45:00"
    )
    assert len(wb.worksheets) == 2
    for ws in wb.worksheets:
        assert ws[shift_ref("W65")].alignment.shrinkToFit is True
        assert not ws[shift_ref("B27")].alignment.wrapText


def test_survives_save(tmp_path):
    """LibreOffice czyta plik, nie obiekt w pamięci."""
    wb = load_workbook(TEMPLATE)
    _apply_protocol_page_marks(
        wb, match_id="", generated_by="X", generated_at="15.08.2026 21:45:00"
    )
    out = tmp_path / "p.xlsx"
    wb.save(out)

    ws = load_workbook(out).active
    assert ws[shift_ref("W65")].alignment.shrinkToFit is True
    assert not ws[shift_ref("B27")].alignment.wrapText


# ─────────────────────── geometria (dlaczego to w ogóle trzeba) ───────────────────────

def _column_widths():
    with zipfile.ZipFile(TEMPLATE) as z:
        xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8", errors="replace")
    widths = {}
    for col in re.finditer(r"<col ([^>]*)/>", xml):
        attrs = dict(re.findall(r'(\w+)="([^"]*)"', col.group(1)))
        for idx in range(int(attrs["min"]), int(attrs["max"]) + 1):
            widths[idx] = float(attrs.get("width", 8.43))
    return widths


def _cell_mm(ws, widths, logical: str) -> float:
    mdw = 7
    phys = shift_ref(logical)
    rng = next((r for r in ws.merged_cells.ranges if phys in r), None)
    cols = (
        range(rng.min_col, rng.max_col + 1)
        if rng
        else [column_index_from_string(re.match(r"([A-Z]+)", phys).group(1))]
    )
    px = sum(
        int(((256 * widths.get(i, 8.43) + int(128.0 / mdw)) / 256.0) * mdw) for i in cols
    )
    scale = (ws.page_setup.scale or 100) / 100.0
    return px / 96.0 * 25.4 * scale


@pytest.mark.parametrize(
    "logical,sample",
    [
        ("W65", "Piotrków Trybunalski"),
        ("B27", "MALINOWSKI Wojciech"),
    ],
)
def test_these_really_do_not_fit(logical, sample):
    """
    Dokumentuje powód istnienia tej listy liczbami, a nie opinią. Gdyby ktoś
    kiedyś poszerzył rubryki i test przestał przechodzić — `shrinkToFit` można
    będzie z tych pól zdjąć.
    """
    ws = _ws()
    widths = _column_widths()
    available = _cell_mm(ws, widths, logical)
    pt = ws[shift_ref(logical)].font.size or 11
    needed = len(sample) * pt * 0.5 / 72 * 25.4
    assert needed > available, (
        "%s: %.1f mm tekstu w %.1f mm rubryki — mieści się, dopasowanie zbędne"
        % (logical, needed, available)
    )
