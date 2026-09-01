"""Rzut karny niewykorzystany ma w przebiegu inne oznaczenie niż trafiony.

Do tej pory oba szły jako „7K" i odróżniała je wyłącznie kolumna wyniku
(trafiony podnosi wynik, niewykorzystany zostawia „--"). Teraz niewykorzystany
dostaje „K w kółku" (U+24C0).

Tego znaku nie ma ani w DejaVu, ani w Times New Roman z szablonu, ani
w zastępniku Liberation, którym LibreOffice go podmienia - dlatego obraz
dokłada font Noto Sans Symbols (patrz Dockerfile), a testy niżej sprawdzają,
że font leży w repozytorium, faktycznie ten znak zawiera i że Dockerfile go
instaluje. Bez tego kompletu protokół drukuje pustą ramkę zamiast oznaczenia.
"""

import os

from fontTools.ttLib import TTFont
from openpyxl import load_workbook

from app.results import (
    PENALTY_MISS_MARK,
    SYMBOL_FONT_PATH,
    TIMELINE_END_ROW,
    TIMELINE_ROWS,
    TIMELINE_START_ROW,
    ShiftedWS,
    _fill_timeline_half_chunk,
)

TEMPLATE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "app",
    "templates",
    "protocol_template.xlsx",
)


def _sheet():
    wb = load_workbook(TEMPLATE)
    return ShiftedWS(wb.active)


def _fill(ws, evs):
    return _fill_timeline_half_chunk(
        ws,
        evs=evs,
        start_row=TIMELINE_START_ROW,
        end_row=TIMELINE_END_ROW,
        half_ms=30 * 60 * 1000,
        period_end_ms=30 * 60 * 1000,
        start_score_host=0,
        start_score_guest=0,
        col_minute="AL",
        col_host_action="AN",
        col_host_score="AP",
        col_guest_score="AS",
        col_guest_action="AU",
    )


def _ev(ms, typ, team="host", player=7):
    return {"time": ms, "type": typ, "team": team, "player": player}


ROWS = [r for r in TIMELINE_ROWS if TIMELINE_START_ROW <= r <= TIMELINE_END_ROW]


def test_znacznik_to_K_w_kolku():
    assert PENALTY_MISS_MARK == "Ⓚ"


def test_repozytorium_niesie_font_z_tym_znakiem():
    """Bez tego fontu w obrazie protokół drukuje pustą ramkę.

    Sprawdzamy PLIK, a nie deklarację w Dockerfile: zawartość pakietów apt
    z fontami zmienia się między wydaniami, a ten glif jest rzadki - Noto Sans
    Symbols v1 był jedynym sprawdzonym fontem, który go ma (nie mają go DejaVu,
    Times New Roman, Liberation, GNU FreeFont ani Noto Sans/Serif/Symbols2).
    """
    assert os.path.exists(SYMBOL_FONT_PATH), (
        "brak app/fonts/NotoSansSymbols-Regular.ttf"
    )
    cmap = TTFont(SYMBOL_FONT_PATH).getBestCmap()
    for ch in PENALTY_MISS_MARK:
        assert ord(ch) in cmap, f"font nie ma znaku {hex(ord(ch))}"


def test_dockerfile_instaluje_ten_font():
    """Font w repozytorium bez wpisu w obrazie to tylko martwy plik."""
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    with open(os.path.join(root, "Dockerfile"), encoding="utf-8") as fh:
        dockerfile = fh.read()
    assert "NotoSansSymbols-Regular.ttf" in dockerfile
    assert "fc-cache" in dockerfile


def test_niewykorzystany_karny_dostaje_przekreslone_K():
    ws = _sheet()
    _fill(ws, [_ev(60_000, "penaltyKickMissed", player=9)])
    assert ws[f"AN{ROWS[0]}"].value == f"9{PENALTY_MISS_MARK}"
    # Niewykorzystany nie rusza wyniku - to zostaje bez zmian.
    assert ws[f"AP{ROWS[0]}"].value == "--"


def test_trafiony_karny_zostaje_przy_zwyklym_K():
    ws = _sheet()
    _fill(ws, [_ev(60_000, "penaltyKickScored", player=9)])
    assert ws[f"AN{ROWS[0]}"].value == "9K"
    assert ws[f"AP{ROWS[0]}"].value == "1"


def test_zwykla_bramka_nie_dostaje_zadnego_K():
    ws = _sheet()
    _fill(ws, [_ev(60_000, "goal", player=9)])
    assert ws[f"AN{ROWS[0]}"].value == "9"


def test_oba_karne_w_jednym_przebiegu_daja_rozne_wpisy():
    """Sedno zmiany: dwa wiersze tego samego zawodnika mają się różnić."""
    ws = _sheet()
    _fill(
        ws,
        [
            _ev(60_000, "penaltyKickScored", player=9),
            _ev(120_000, "penaltyKickMissed", player=9),
        ],
    )
    trafiony = ws[f"AN{ROWS[0]}"].value
    niewykorzystany = ws[f"AN{ROWS[1]}"].value
    assert trafiony != niewykorzystany
    assert trafiony == "9K"
    assert niewykorzystany == f"9{PENALTY_MISS_MARK}"


def test_znacznik_zajmuje_jedno_pole():
    """Rubryka przebiegu to dwie wąskie kolumny, ~40 px przy czcionce 9 pt.

    „17Ⓚ" mieści się z zapasem, ale gdyby ktoś podmienił znacznik na „(K)",
    wpis dwucyfrowego numeru urósłby z trzech pól na pięć i przestałby się
    mieścić.
    """
    import unicodedata

    widths = [c for c in PENALTY_MISS_MARK if unicodedata.category(c) != "Mn"]
    assert len(widths) == 1, f"znacznik zajmuje {len(widths)} pól zamiast jednego"


def test_gosc_tez_dostaje_znacznik():
    ws = _sheet()
    _fill(ws, [_ev(60_000, "penaltyKickMissed", team="guest", player=3)])
    assert ws[f"AU{ROWS[0]}"].value == f"3{PENALTY_MISS_MARK}"
    assert ws[f"AN{ROWS[0]}"].value == "--"
